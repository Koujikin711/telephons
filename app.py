"""Магазин телефонов — касса, склад, реализация, финансовые отчёты."""

from __future__ import annotations

import csv
import io
import json
import re
import logging
import os
import sqlite3
from contextlib import asynccontextmanager, contextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Literal

from fastapi import FastAPI, File, Header, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from pydantic_settings import BaseSettings, SettingsConfigDict

logger = logging.getLogger("telephons")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

STATIC_DIR = Path(__file__).resolve().parent / "static"
DATA_DIR = Path("/data") if Path("/data").exists() else Path(__file__).resolve().parent / "data"
DB_PATH = DATA_DIR / "store.db"
UPLOADS_DIR = DATA_DIR / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
MAX_IMAGE_BYTES = 5 * 1024 * 1024

OwnershipType = Literal["own", "consignment"]
ReportScope = Literal["all", "own", "consignment"]
ProductCondition = Literal["new", "used", "refurbished", "partnership"]
UserRole = Literal["owner", "warehouse", "cashier", "accessories"]
DEFAULT_WAREHOUSE_NAME = "Основной склад"

ROLE_PAGES: dict[str, list[str]] = {
    "owner": [
        "dashboard", "pos", "sales", "warehouses", "products-own",
        "products-consignment", "accessories", "reports", "analytics", "debtors", "creditors",
        "stocktake", "settings",
    ],
    "warehouse": [
        "dashboard", "warehouses", "products-own", "products-consignment",
        "accessories", "stocktake",
    ],
    "cashier": ["dashboard", "pos", "sales", "debtors"],
    "accessories": ["accessories"],
}

# Упрощённый интерфейс: касса, склад, взаиморасчёты — без отдельного раздела «Продажи».
SIMPLE_ROLE_PAGES: dict[str, list[str]] = {
    "owner": ["pos", "warehouses", "debtors", "stocktake", "accessories", "reports", "settings"],
    "warehouse": ["warehouses", "stocktake"],
    "cashier": ["pos", "debtors"],
    "accessories": ["accessories"],
}

ROLE_LEVEL = {"cashier": 1, "accessories": 2, "warehouse": 2, "owner": 3}


def pages_for_role(role: str, *, simple: bool) -> list[str]:
    if simple:
        return list(SIMPLE_ROLE_PAGES.get(role, ["pos"]))
    return list(ROLE_PAGES.get(role, ["dashboard"]))


class Settings(BaseSettings):
    model_config = SettingsConfigDict(env_file=".env", extra="ignore")

    store_pin: str = ""
    store_name: str = "Магазин телефонов"
    port: int = 80
    simple_ui: bool = True


settings = Settings()
settings.port = int(os.environ.get("PORT", settings.port))


def utc_now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def row_to_dict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    if row is None:
        return None
    return dict(row)



def normalize_search_q(q: str) -> str:
    return q.strip()


def only_digits(s: str) -> str:
    return re.sub(r"\D", "", s)


def is_code_query(q: str) -> bool:
    """Digits-only query (IMEI fragment, barcode)."""
    d = only_digits(q)
    return len(d) >= 5 and len(d) == len(q.replace(" ", "").replace("-", ""))


def product_search_sql(q: str) -> tuple[str, list[Any]]:
    q = normalize_search_q(q)
    if not q:
        return "", []
    like = f"%{q}%"
    fields = (
        "p.name", "p.brand", "p.sku", "p.barcode", "p.supplier_name",
        "p.model", "p.color", "p.memory", "p.specs_extra", "p.ram",
    )
    text = " OR ".join(f"LOWER({f}) LIKE LOWER(?)" for f in fields)
    params: list[Any] = [like] * len(fields)
    d = only_digits(q)
    if len(d) >= 5 and re.fullmatch(r"[\d\s-]+", q):
        suffix = d[-min(len(d), 15):]
        unit_sql = """
            OR EXISTS (
                SELECT 1 FROM product_units u
                WHERE u.product_id = p.id AND u.status = 'in_stock'
                  AND (u.imei LIKE ? OR u.serial LIKE ? OR u.imei LIKE ? OR u.serial LIKE ?)
            )
        """
        text = f"({text}{unit_sql})"
        params.extend([f"%{suffix}", f"%{suffix}", f"%{d}%", f"%{d}%"])
    else:
        unit_sql = """
            OR EXISTS (
                SELECT 1 FROM product_units u
                WHERE u.product_id = p.id AND u.status = 'in_stock'
                  AND (u.imei LIKE ? OR u.serial LIKE ?)
            )
        """
        text = f"({text}{unit_sql})"
        params.extend([like, like])
    return f" AND {text}", params


def sale_search_sql(q: str) -> tuple[str, list[Any]]:
    q = normalize_search_q(q)
    if not q:
        return "", []
    params: list[Any] = []
    parts: list[str] = []
    id_q = q.lstrip("#").strip()
    if id_q.isdigit():
        parts.append("s.id = ?")
        params.append(int(id_q))
    like = f"%{q}%"
    parts.append(
        """
        EXISTS (
            SELECT 1 FROM sale_items si2
            WHERE si2.sale_id = s.id
              AND (
                LOWER(si2.product_name) LIKE LOWER(?)
                OR LOWER(COALESCE(si2.supplier_name, '')) LIKE LOWER(?)
              )
        )
        """
    )
    params.extend([like, like])
    d = only_digits(q)
    if len(d) >= 5:
        suffix = d[-min(len(d), 15):]
        parts.append(
            """
            EXISTS (
                SELECT 1 FROM sale_items si3
                JOIN sale_item_units siu ON siu.sale_item_id = si3.id
                WHERE si3.sale_id = s.id
                  AND (
                    siu.imei LIKE ? OR siu.serial LIKE ?
                    OR siu.imei LIKE ? OR siu.serial LIKE ?
                  )
            )
            """
        )
        params.extend([f"%{suffix}", f"%{suffix}", f"%{d}%", f"%{d}%"])
    parts.append(
        """
        EXISTS (
            SELECT 1 FROM receivables r
            WHERE r.sale_id = s.id
              AND (
                LOWER(r.customer_name) LIKE LOWER(?)
                OR COALESCE(r.customer_phone, '') LIKE ?
              )
        )
        """
    )
    params.extend([like, like])
    parts.append(
        "(LOWER(COALESCE(s.user_name, '')) LIKE LOWER(?) OR LOWER(COALESCE(s.notes, '')) LIKE LOWER(?))"
    )
    params.extend([like, like])
    return f"({' OR '.join(parts)})", params


def imei_digits_expr(col: str) -> str:
    """SQL expression: strip common separators from IMEI/serial for digit matching."""
    return (
        f"REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(COALESCE({col}, ''), ' ', ''), '-', ''), "
        f"'.', ''), '_', ''), '/', '')"
    )


def unit_search_sql(q: str, imei_col: str = "u.imei", serial_col: str = "u.serial") -> tuple[str, list[Any]]:
    q = normalize_search_q(q)
    if not q:
        return "", []
    d = only_digits(q)
    if len(d) >= 5 and re.fullmatch(r"[\d\s-]+", q):
        suffix = d[-min(len(d), 15):]
        dig_imei = imei_digits_expr(imei_col)
        dig_serial = imei_digits_expr(serial_col)
        return (
            f" AND ("
            f"{imei_col} LIKE ? OR {serial_col} LIKE ? OR {imei_col} LIKE ? OR {serial_col} LIKE ?"
            f" OR {dig_imei} LIKE ? OR {dig_serial} LIKE ?"
            f" OR {dig_imei} LIKE ? OR {dig_serial} LIKE ?"
            f")",
            [
                f"%{suffix}", f"%{suffix}", f"%{d}%", f"%{d}%",
                f"%{suffix}", f"%{suffix}", f"%{d}%", f"%{d}%",
            ],
        )
    like = f"%{q}%"
    return f" AND ({imei_col} LIKE ? OR {serial_col} LIKE ?)", [like, like]


def transfer_product_units(
    conn: sqlite3.Connection,
    product_id: int,
    from_warehouse_id: int,
    to_warehouse_id: int,
    quantity: int,
) -> None:
    product = conn.execute("SELECT track_units FROM products WHERE id = ?", (product_id,)).fetchone()
    if not product or not int(product["track_units"] or 0):
        return
    rows = conn.execute(
        """
        SELECT id FROM product_units
        WHERE product_id = ? AND warehouse_id = ? AND status = 'in_stock'
        ORDER BY id LIMIT ?
        """,
        (product_id, from_warehouse_id, quantity),
    ).fetchall()
    if len(rows) < quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Недостаточно устройств с IMEI: нужно {quantity}, доступно {len(rows)}",
        )
    for row in rows:
        conn.execute(
            "UPDATE product_units SET warehouse_id = ? WHERE id = ?",
            (to_warehouse_id, row["id"]),
        )


def units_for_product_at_warehouse(
    conn: sqlite3.Connection, product_id: int, warehouse_id: int
) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT u.id, u.imei, u.serial, u.status, u.notes, u.created_at,
               u.battery_capacity, u.client_name, u.region, u.arrival_date,
               u.purchase_price AS unit_purchase_price,
               p.color AS product_color, p.name AS product_name, p.model, p.memory,
               p.purchase_price AS product_purchase_price, p.supplier_name, p.sale_price, p.ownership_type
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        WHERE u.product_id = ? AND u.warehouse_id = ? AND u.status = 'in_stock'
        ORDER BY p.color, u.imei, u.serial
        """,
        (product_id, warehouse_id),
    ).fetchall()
    out = []
    for r in rows:
        d = row_to_dict(r) or {}
        d["purchase_price"] = unit_purchase_price(
            {"purchase_price": d.get("unit_purchase_price")},
            {"purchase_price": d.get("product_purchase_price")},
        )
        out.append(d)
    return out


@contextmanager
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
    return bool(
        conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?", (table,)
        ).fetchone()
    )


def _add_column(conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    if not _table_exists(conn, table):
        return
    cols = {r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def migrate_db(conn: sqlite3.Connection) -> None:
    _add_column(conn, "products", "ownership_type", "TEXT NOT NULL DEFAULT 'own'")
    _add_column(conn, "products", "supplier_name", "TEXT DEFAULT ''")
    _add_column(conn, "products", "model", "TEXT DEFAULT ''")
    _add_column(conn, "products", "color", "TEXT DEFAULT ''")
    _add_column(conn, "products", "size", "TEXT DEFAULT ''")
    _add_column(conn, "products", "memory", "TEXT DEFAULT ''")
    _add_column(conn, "products", "ram", "TEXT DEFAULT ''")
    _add_column(conn, "products", "customs_cleared", "INTEGER NOT NULL DEFAULT 0")
    _add_column(conn, "products", "customs_price", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "products", "specs_extra", "TEXT DEFAULT ''")
    _add_column(conn, "products", "condition", "TEXT NOT NULL DEFAULT 'new'")
    _add_column(conn, "sale_items", "ownership_type", "TEXT NOT NULL DEFAULT 'own'")
    _add_column(conn, "sale_items", "supplier_name", "TEXT DEFAULT ''")
    _add_column(conn, "sale_items", "supplier_due", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "sale_items", "shop_profit", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "supplier_payments", "payment_method_code", "TEXT DEFAULT 'cash'")
    _add_column(conn, "sales", "warehouse_id", "INTEGER")
    _add_column(conn, "sales", "cash_amount", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "sales", "card_amount", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "sales", "trade_in_value", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "sales", "shift_id", "INTEGER")
    _add_column(conn, "sales", "user_id", "INTEGER")
    _add_column(conn, "sales", "user_name", "TEXT DEFAULT ''")
    _add_column(conn, "shifts", "expected_payments_json", "TEXT DEFAULT ''")
    _add_column(conn, "shifts", "actual_payments_json", "TEXT DEFAULT ''")
    _add_column(conn, "shifts", "opening_wallets_json", "TEXT DEFAULT ''")
    if _table_exists(conn, "shifts"):
        conn.execute(
            """
            UPDATE sales SET
                user_id = (SELECT user_id FROM shifts WHERE shifts.id = sales.shift_id),
                user_name = COALESCE((SELECT user_name FROM shifts WHERE shifts.id = sales.shift_id), '')
            WHERE user_id IS NULL AND shift_id IS NOT NULL
            """
        )
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS stocktake_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            warehouse_id INTEGER NOT NULL,
            user_id INTEGER,
            user_name TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'open',
            notes TEXT DEFAULT '',
            started_at TEXT NOT NULL,
            completed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS stocktake_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            unit_id INTEGER,
            quantity INTEGER NOT NULL DEFAULT 1,
            imei TEXT DEFAULT '',
            serial TEXT DEFAULT '',
            color TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY (session_id) REFERENCES stocktake_sessions(id) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_stocktake_wh ON stocktake_sessions(warehouse_id, status);
        CREATE TABLE IF NOT EXISTS mutual_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            person_name TEXT NOT NULL,
            person_phone TEXT DEFAULT '',
            direction TEXT NOT NULL DEFAULT 'owe_us',
            amount REAL NOT NULL,
            paid_amount REAL NOT NULL DEFAULT 0,
            amount_due REAL NOT NULL,
            currency_code TEXT NOT NULL DEFAULT 'TJS',
            payment_method_code TEXT DEFAULT '',
            product_note TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'open',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            closed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS mutual_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_id INTEGER NOT NULL,
            amount REAL NOT NULL,
            payment_method_code TEXT DEFAULT 'cash',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY (entry_id) REFERENCES mutual_entries(id) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_mutual_status ON mutual_entries(status, created_at);
        """
    )
    _add_column(conn, "mutual_payments", "pay_amount", "REAL")
    _add_column(conn, "mutual_payments", "pay_currency_code", "TEXT DEFAULT ''")
    _add_column(conn, "mutual_payments", "fx_rate", "REAL")
    _add_column(conn, "receivable_payments", "pay_amount", "REAL")
    _add_column(conn, "receivable_payments", "pay_currency_code", "TEXT DEFAULT ''")
    _add_column(conn, "receivable_payments", "fx_rate", "REAL")
    _add_column(conn, "cash_inflows", "mutual_entry_id", "INTEGER")
    _add_column(conn, "payment_methods", "currency_code", "TEXT NOT NULL DEFAULT ''")
    _backfill_payment_method_currencies(conn)
    if _table_exists(conn, "product_units"):
        conn.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_units_imei ON product_units(imei);
            CREATE INDEX IF NOT EXISTS idx_units_serial ON product_units(serial);
            """
        )
    _add_column(conn, "products", "track_units", "INTEGER NOT NULL DEFAULT 0")
    _add_column(conn, "products", "image_url", "TEXT DEFAULT ''")
    _add_column(conn, "product_units", "customs_cleared", "INTEGER NOT NULL DEFAULT 0")
    _add_column(conn, "product_units", "customs_price", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "sale_item_units", "customs_cleared", "INTEGER NOT NULL DEFAULT 0")
    _add_column(conn, "sale_item_units", "customs_price", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "sale_item_units", "imei_pending", "INTEGER NOT NULL DEFAULT 0")
    _add_column(conn, "product_units", "box_image_url", "TEXT DEFAULT ''")
    _add_column(conn, "product_units", "customs_status", "TEXT NOT NULL DEFAULT 'none'")
    _add_column(conn, "product_units", "battery_capacity", "INTEGER")
    _add_column(conn, "product_units", "client_name", "TEXT DEFAULT ''")
    _add_column(conn, "product_units", "region", "TEXT DEFAULT ''")
    _add_column(conn, "product_units", "arrival_date", "TEXT DEFAULT ''")
    _add_column(conn, "product_units", "purchase_price", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "warehouses", "warehouse_type", "TEXT NOT NULL DEFAULT 'new'")
    _add_column(conn, "warehouses", "currency_code", "TEXT NOT NULL DEFAULT ''")
    if _table_exists(conn, "product_units"):
        conn.execute(
            """
            UPDATE product_units SET customs_status = 'cleared'
            WHERE customs_cleared = 1 AND customs_status = 'none'
            """
        )
        # Backfill unit cost from product card when unit has no own cost yet
        conn.execute(
            """
            UPDATE product_units
            SET purchase_price = (
                SELECT p.purchase_price FROM products p WHERE p.id = product_units.product_id
            )
            WHERE COALESCE(purchase_price, 0) = 0
              AND EXISTS (
                  SELECT 1 FROM products p
                  WHERE p.id = product_units.product_id AND COALESCE(p.purchase_price, 0) > 0
              )
            """
        )
    if _table_exists(conn, "warehouses"):
        conn.execute(
            """
            UPDATE warehouses SET warehouse_type = 'used'
            WHERE name LIKE '%БУ%' OR name LIKE '%бу%' OR name LIKE '%Б/У%' OR name LIKE '%б/у%'
            """
        )
        _merge_duplicate_bu_warehouses(conn)
        _merge_duplicate_accessories_warehouses(conn)
        try:
            resolve_partnership_warehouse_id(conn)
        except Exception as exc:
            logger.warning("Partnership warehouse setup skipped: %s", exc)
        _set_warehouse_currencies(conn)
        _ensure_three_core_warehouses(conn)
        _fix_z_register_warehouse_split(conn)
        _sync_product_purchase_from_sales(conn)
        _sync_receivables_for_voided_sales(conn)
        _reclassify_phones_from_accessories(conn)
        try:
            _dedupe_accessory_products(conn)
        except Exception as exc:
            logger.warning("Accessory dedupe migration skipped: %s", exc)
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS unit_reservations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            warehouse_id INTEGER NOT NULL,
            client_name TEXT NOT NULL,
            client_phone TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            reserved_until TEXT NOT NULL,
            user_id INTEGER,
            user_name TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'active',
            created_at TEXT NOT NULL,
            FOREIGN KEY (unit_id) REFERENCES product_units(id)
        );
        CREATE INDEX IF NOT EXISTS idx_reservations_unit ON unit_reservations(unit_id, status);
        CREATE INDEX IF NOT EXISTS idx_reservations_until ON unit_reservations(reserved_until, status);
        """
    )
    conn.execute("UPDATE products SET track_units = 1 WHERE category = 'phone' AND track_units = 0")
    conn.execute(
        """
        UPDATE sale_items
        SET shop_profit = subtotal - purchase_price * quantity
        WHERE shop_profit = 0 AND subtotal != 0
          AND ABS(subtotal - purchase_price * quantity) > 0.001
        """
    )
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS receivables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER,
            customer_name TEXT NOT NULL,
            customer_phone TEXT DEFAULT '',
            total_amount REAL NOT NULL,
            paid_amount REAL NOT NULL DEFAULT 0,
            amount_due REAL NOT NULL,
            status TEXT NOT NULL DEFAULT 'open',
            warehouse_id INTEGER,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            closed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS receivable_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receivable_id INTEGER NOT NULL,
            amount REAL NOT NULL,
            payment_method_code TEXT DEFAULT 'cash',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY (receivable_id) REFERENCES receivables(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS expense_warehouse_split (
            warehouse_id INTEGER PRIMARY KEY,
            pct REAL NOT NULL DEFAULT 0,
            FOREIGN KEY (warehouse_id) REFERENCES warehouses(id) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_receivables_status ON receivables(status, created_at);
        """
    )
    _add_column(conn, "sales", "amount_paid", "REAL")
    _add_column(conn, "sales", "amount_due", "REAL NOT NULL DEFAULT 0")
    _add_column(conn, "sales", "currency_code", "TEXT NOT NULL DEFAULT 'TJS'")
    _add_column(conn, "expenses", "department", "TEXT NOT NULL DEFAULT 'main'")
    _add_column(conn, "expenses", "created_by", "TEXT NOT NULL DEFAULT ''")
    _add_column(conn, "expenses", "created_by_user_id", "INTEGER")
    _add_column(conn, "expenses", "warehouse_id", "INTEGER")
    _add_column(conn, "expenses", "payee", "TEXT NOT NULL DEFAULT ''")
    _add_column(conn, "expenses", "kind", "TEXT NOT NULL DEFAULT 'expense'")
    _add_column(conn, "sales", "affects_cash", "INTEGER NOT NULL DEFAULT 1")
    _add_column(conn, "expenses", "affects_cash", "INTEGER NOT NULL DEFAULT 1")
    _add_column(conn, "sale_payments", "pay_currency_code", "TEXT NOT NULL DEFAULT ''")
    _add_column(conn, "sale_payments", "pay_amount", "REAL")
    # Mark Excel-imported rows for display when created_by is empty
    if _table_exists(conn, "expenses"):
        conn.execute(
            """
            UPDATE expenses
            SET created_by = 'Excel импорт'
            WHERE COALESCE(TRIM(created_by), '') = ''
              AND description = 'Excel импорт ОПиУ'
            """
        )
    # Excel ОПиУ expenses stay out of cash (they were paper totals, not till moves)
    if _table_exists(conn, "expenses"):
        conn.execute(
            """
            UPDATE expenses SET affects_cash = 0
            WHERE COALESCE(affects_cash, 1) = 1
              AND (
                description = 'Excel импорт ОПиУ'
                OR created_by = 'Excel импорт'
              )
            """
        )
    # Backfill till currency on sale payments (what actually entered the drawer)
    if _table_exists(conn, "sale_payments"):
        conn.execute(
            """
            UPDATE sale_payments
            SET pay_currency_code = COALESCE(
                    (SELECT NULLIF(TRIM(s.currency_code), '') FROM sales s WHERE s.id = sale_payments.sale_id),
                    'TJS'
                ),
                pay_amount = COALESCE(NULLIF(pay_amount, 0), amount)
            WHERE COALESCE(TRIM(pay_currency_code), '') = ''
            """
        )

    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS exchange_rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            currency_code TEXT NOT NULL,
            rate REAL NOT NULL,
            effective_at TEXT NOT NULL,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS payment_methods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            method_type TEXT NOT NULL DEFAULT 'card',
            is_active INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sale_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            method_code TEXT NOT NULL,
            amount REAL NOT NULL,
            FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            amount REAL NOT NULL,
            description TEXT DEFAULT '',
            payment_method_code TEXT DEFAULT 'cash',
            expense_date TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS cash_inflows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount REAL NOT NULL,
            currency_code TEXT NOT NULL DEFAULT 'TJS',
            amount_base REAL NOT NULL,
            payment_method_code TEXT NOT NULL DEFAULT 'cash',
            source_type TEXT NOT NULL DEFAULT 'counterparty',
            counterparty_name TEXT DEFAULT '',
            receivable_id INTEGER,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            created_by TEXT DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_cash_inflows_created ON cash_inflows(created_at);
        CREATE INDEX IF NOT EXISTS idx_exchange_rates_cur ON exchange_rates(currency_code, effective_at);
        CREATE INDEX IF NOT EXISTS idx_sale_payments_sale ON sale_payments(sale_id);
        """
    )
    _seed_finance_defaults(conn)
    _backfill_sale_payments(conn)
    # Z-import phone sales → касса (смн / $). Re-run if re-import cleared the flag.
    if _table_exists(conn, "sales") and get_setting(conn, "z_sales_in_cash_v2", "") != "1":
        conn.execute(
            """
            UPDATE sales SET affects_cash = 1
            WHERE COALESCE(affects_cash, 1) = 0
              AND (
                user_name = 'Z-импорт'
                OR user_name LIKE 'Z-импорт%'
              )
            """
        )
        set_setting(conn, "z_sales_in_cash_v2", "1")
        set_setting(conn, "z_sales_in_cash_v1", "1")

    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS warehouses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            address TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            is_default INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS warehouse_stock (
            warehouse_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (warehouse_id, product_id),
            FOREIGN KEY (warehouse_id) REFERENCES warehouses(id) ON DELETE CASCADE,
            FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS stock_movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            warehouse_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            movement_type TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            target_warehouse_id INTEGER,
            reference_id INTEGER,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY (warehouse_id) REFERENCES warehouses(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        );

        CREATE TABLE IF NOT EXISTS trade_ins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            given_product_id INTEGER NOT NULL,
            given_warehouse_id INTEGER NOT NULL,
            received_name TEXT NOT NULL,
            received_brand TEXT DEFAULT '',
            received_model TEXT DEFAULT '',
            received_color TEXT DEFAULT '',
            received_size TEXT DEFAULT '',
            received_memory TEXT DEFAULT '',
            received_ram TEXT DEFAULT '',
            received_specs_extra TEXT DEFAULT '',
            received_condition TEXT NOT NULL DEFAULT 'used',
            received_purchase_price REAL NOT NULL DEFAULT 0,
            received_sale_price REAL NOT NULL DEFAULT 0,
            received_value REAL NOT NULL DEFAULT 0,
            cash_amount REAL NOT NULL DEFAULT 0,
            card_amount REAL NOT NULL DEFAULT 0,
            received_product_id INTEGER,
            received_warehouse_id INTEGER,
            sale_id INTEGER,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY (given_product_id) REFERENCES products(id),
            FOREIGN KEY (given_warehouse_id) REFERENCES warehouses(id),
            FOREIGN KEY (received_product_id) REFERENCES products(id),
            FOREIGN KEY (received_warehouse_id) REFERENCES warehouses(id),
            FOREIGN KEY (sale_id) REFERENCES sales(id)
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            pin TEXT NOT NULL UNIQUE,
            role TEXT NOT NULL DEFAULT 'cashier',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS product_units (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            warehouse_id INTEGER NOT NULL,
            imei TEXT DEFAULT '',
            serial TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'in_stock',
            sale_id INTEGER,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (warehouse_id) REFERENCES warehouses(id)
        );

        CREATE TABLE IF NOT EXISTS shifts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            user_name TEXT DEFAULT '',
            opened_at TEXT NOT NULL,
            closed_at TEXT,
            opening_cash REAL NOT NULL DEFAULT 0,
            expected_cash REAL NOT NULL DEFAULT 0,
            expected_card REAL NOT NULL DEFAULT 0,
            actual_cash REAL,
            actual_card REAL,
            sales_count INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'open',
            notes TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS sale_item_units (
            sale_item_id INTEGER NOT NULL,
            unit_id INTEGER NOT NULL,
            imei TEXT DEFAULT '',
            serial TEXT DEFAULT '',
            PRIMARY KEY (sale_item_id, unit_id),
            FOREIGN KEY (sale_item_id) REFERENCES sale_items(id) ON DELETE CASCADE,
            FOREIGN KEY (unit_id) REFERENCES product_units(id)
        );
        """
    )

    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_units_product ON product_units(product_id);
        CREATE INDEX IF NOT EXISTS idx_units_warehouse ON product_units(warehouse_id);
        CREATE INDEX IF NOT EXISTS idx_units_status ON product_units(status);
        CREATE INDEX IF NOT EXISTS idx_shifts_status ON shifts(status);
        """
    )

    if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        now = utc_now()
        owner_pin = settings.store_pin or "1234"
        conn.executemany(
            "INSERT INTO users (name, pin, role, is_active, created_at) VALUES (?, ?, ?, 1, ?)",
            [
                ("Владелец", owner_pin, "owner", now),
                ("Кассир", "1111", "cashier", now),
                ("Кладовщик", "2222", "warehouse", now),
            ],
        )

    default_wh = conn.execute(
        "SELECT id FROM warehouses WHERE is_default = 1 LIMIT 1"
    ).fetchone()
    if not default_wh:
        cur = conn.execute(
            """
            INSERT INTO warehouses (name, address, notes, is_default, created_at)
            VALUES (?, '', '', 1, ?)
            """,
            (DEFAULT_WAREHOUSE_NAME, utc_now()),
        )
        default_wh_id = cur.lastrowid
    else:
        default_wh_id = default_wh["id"]

    migrated = conn.execute(
        "SELECT COUNT(*) FROM warehouse_stock"
    ).fetchone()[0]
    if migrated == 0:
        products = conn.execute(
            "SELECT id, stock FROM products WHERE stock > 0"
        ).fetchall()
        now = utc_now()
        for p in products:
            conn.execute(
                """
                INSERT INTO warehouse_stock (warehouse_id, product_id, quantity)
                VALUES (?, ?, ?)
                """,
                (default_wh_id, p["id"], p["stock"]),
            )
            conn.execute(
                """
                INSERT INTO stock_movements
                (warehouse_id, product_id, movement_type, quantity, notes, created_at)
                VALUES (?, ?, 'inbound', ?, 'Миграция начального остатка', ?)
                """,
                (default_wh_id, p["id"], p["stock"], now),
            )

    conn.execute(
        "UPDATE sales SET warehouse_id = ? WHERE warehouse_id IS NULL",
        (default_wh_id,),
    )


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT NOT NULL DEFAULT 'accessory',
                brand TEXT DEFAULT '',
                sku TEXT DEFAULT '',
                barcode TEXT DEFAULT '',
                purchase_price REAL NOT NULL DEFAULT 0,
                sale_price REAL NOT NULL DEFAULT 0,
                stock INTEGER NOT NULL DEFAULT 0,
                min_stock INTEGER NOT NULL DEFAULT 2,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                total REAL NOT NULL,
                discount REAL NOT NULL DEFAULT 0,
                payment_method TEXT NOT NULL DEFAULT 'cash',
                status TEXT NOT NULL DEFAULT 'completed',
                notes TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sale_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL,
                product_id INTEGER,
                product_name TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                unit_price REAL NOT NULL,
                purchase_price REAL NOT NULL DEFAULT 0,
                subtotal REAL NOT NULL,
                FOREIGN KEY (sale_id) REFERENCES sales(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS supplier_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_name TEXT NOT NULL,
                amount REAL NOT NULL,
                notes TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            """
        )
        migrate_db(conn)
        conn.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_products_name ON products(name);
            CREATE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode);
            CREATE INDEX IF NOT EXISTS idx_products_ownership ON products(ownership_type);
            CREATE INDEX IF NOT EXISTS idx_products_model ON products(model);
            CREATE INDEX IF NOT EXISTS idx_sales_created ON sales(created_at);
            CREATE INDEX IF NOT EXISTS idx_sales_warehouse ON sales(warehouse_id);
            CREATE INDEX IF NOT EXISTS idx_warehouse_stock_product ON warehouse_stock(product_id);
            CREATE INDEX IF NOT EXISTS idx_stock_movements_created ON stock_movements(created_at);
            CREATE INDEX IF NOT EXISTS idx_trade_ins_created ON trade_ins(created_at);
            """
        )

        # Каталог пустой — товары добавляются через склад или импорт Excel


def _seed_finance_defaults(conn: sqlite3.Connection) -> None:
    defaults = {
        "base_currency": "TJS",
        "currency_symbol": "смн",
        "currency_name": "Сомони",
    }
    for key, val in defaults.items():
        if not conn.execute("SELECT 1 FROM app_settings WHERE key = ?", (key,)).fetchone():
            conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?)", (key, val))

    if conn.execute("SELECT COUNT(*) FROM payment_methods").fetchone()[0] == 0:
        now = utc_now()
        for code, name, mtype, order in [
            ("cash", "Наличные", "cash", 0),
            ("card", "Банковская карта", "card", 1),
            ("ds", "ДС", "mobile", 2),
            ("alif", "Alif", "mobile", 3),
            ("eskhata", "Эсхата", "mobile", 4),
            ("transfer", "Перевод", "transfer", 5),
        ]:
            conn.execute(
                """
                INSERT INTO payment_methods (code, name, method_type, is_active, sort_order, created_at)
                VALUES (?, ?, ?, 1, ?, ?)
                """,
                (code, name, mtype, order, now),
            )

    if conn.execute("SELECT COUNT(*) FROM exchange_rates").fetchone()[0] == 0:
        now = utc_now()
        conn.execute(
            """
            INSERT INTO exchange_rates (currency_code, rate, effective_at, notes, created_at)
            VALUES ('TJS', 1, ?, 'Базовая валюта', ?)
            """,
            (now, now),
        )

    if not conn.execute("SELECT 1 FROM users WHERE role = 'accessories'").fetchone():
        conn.execute(
            "INSERT INTO users (name, pin, role, is_active, created_at) VALUES (?, ?, ?, 1, ?)",
            ("Ответственный за аксессуары", "3333", "accessories", utc_now()),
        )


def _backfill_sale_payments(conn: sqlite3.Connection) -> None:
    if conn.execute("SELECT COUNT(*) FROM sale_payments").fetchone()[0] > 0:
        return
    sales = conn.execute(
        "SELECT id, payment_method, total, cash_amount, card_amount FROM sales WHERE status = 'completed'"
    ).fetchall()
    for s in sales:
        if float(s["cash_amount"] or 0) > 0:
            conn.execute(
                "INSERT INTO sale_payments (sale_id, method_code, amount) VALUES (?, 'cash', ?)",
                (s["id"], s["cash_amount"]),
            )
        if float(s["card_amount"] or 0) > 0:
            method = s["payment_method"] if s["payment_method"] in ("card", "transfer", "ds", "alif", "eskhata") else "card"
            conn.execute(
                "INSERT INTO sale_payments (sale_id, method_code, amount) VALUES (?, ?, ?)",
                (s["id"], method, s["card_amount"]),
            )
        elif s["payment_method"] not in ("cash", "trade_in") and float(s["total"] or 0) > 0:
            conn.execute(
                "INSERT INTO sale_payments (sale_id, method_code, amount) VALUES (?, ?, ?)",
                (s["id"], s["payment_method"], s["total"]),
            )
        elif s["payment_method"] == "cash" and float(s["cash_amount"] or 0) == 0 and float(s["total"] or 0) > 0:
            conn.execute(
                "INSERT INTO sale_payments (sale_id, method_code, amount) VALUES (?, 'cash', ?)",
                (s["id"], s["total"]),
            )


def get_setting(conn: sqlite3.Connection, key: str, default: str = "") -> str:
    row = conn.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        "INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )


IMPORT_USED_HEADERS = [
    "модель", "цвет", "память", "аккумулятор_%", "imei", "себестоимость", "клиент", "комментарий",
]

IMPORT_NEW_HEADERS = [
    "дата_прихода", "модель", "цвет", "память", "регион", "imei", "себестоимость", "поставщик", "тип", "комментарий",
]

IMPORT_PRODUCT_HEADERS = IMPORT_NEW_HEADERS

IMPORT_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "название": ("название", "name", "product_name", "товар", "модель", "model"),
    "модель": ("модель", "model", "название", "name"),
    "тип": ("тип", "ownership_type", "ownership"),
    "поставщик": ("поставщик", "supplier_name", "supplier"),
    "цвет": ("цвет", "color"),
    "память": ("память", "memory"),
    "состояние": ("состояние", "condition"),
    "закупка": ("закупка", "purchase_price", "закупочная", "себестоимость"),
    "себестоимость": ("себестоимость", "закупка", "purchase_price"),
    "клиент": ("клиент", "client", "client_name"),
    "регион": ("регион", "region"),
    "дата_прихода": ("дата_прихода", "arrival_date", "дата", "date"),
    "аккумулятор_%": ("аккумулятор_%", "батарея_%", "battery_capacity", "батарея", "battery"),
    "категория": ("категория", "category"),
    "цена": ("цена", "sale_price", "продажа"),
    "количество": ("количество", "quantity", "qty", "кол-во"),
    "мин_остаток": ("мин_остаток", "min_stock"),
    "imei": ("imei", "imei1"),
    "серийник": ("серийник", "serial", "serial_number"),
    "батарея_%": ("батарея_%", "battery_capacity", "батарея", "battery"),
    "комментарий": ("комментарий", "notes", "примечание"),
}

IMPORT_SALE_HEADERS = [
    "дата", "чек", "название", "imei", "количество", "цена", "скидка", "оплата", "кассир", "комментарий",
]

IMPORT_SALE_ALIASES: dict[str, tuple[str, ...]] = {
    "дата": ("дата", "date", "datetime", "время"),
    "чек": ("чек", "номер_чека", "receipt", "номер"),
    "название": ("название", "name", "product_name", "товар"),
    "imei": ("imei", "imei1"),
    "количество": ("количество", "quantity", "qty", "кол-во"),
    "цена": ("цена", "sale_price", "продажа", "сумма"),
    "скидка": ("скидка", "discount"),
    "оплата": ("оплата", "payment", "payment_method", "способ_оплаты"),
    "кассир": ("кассир", "cashier", "user", "сотрудник"),
    "комментарий": ("комментарий", "notes", "примечание"),
}


def _norm_category(raw: str) -> str:
    v = raw.strip().lower()
    if v in ("phone", "телефон", "телефоны", "phone"):
        return "phone"
    return "accessory"


def _norm_ownership(raw: str) -> str:
    v = raw.strip().lower()
    if v in ("consignment", "реализация", "комиссия"):
        return "consignment"
    return "own"


def _norm_condition(raw: str) -> str:
    v = raw.strip().lower()
    if v in ("used", "б/у", "бу", "б у"):
        return "used"
    if v in ("refurbished", "восстановленный", "ref"):
        return "refurbished"
    return "new"


def _import_row_cell(
    row: dict[str, Any],
    fields: dict[str, str],
    canonical: str,
    aliases: dict[str, tuple[str, ...]] | None = None,
) -> str:
    alias_map = aliases or IMPORT_HEADER_ALIASES
    for alias in alias_map.get(canonical, (canonical,)):
        key = fields.get(alias.lower())
        if key and row.get(key) not in (None, ""):
            return str(row[key]).strip()
    return ""


def _parse_import_file(raw: bytes, filename: str, sheet: str | None = None) -> list[dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        if sheet and sheet in wb.sheetnames:
            ws = wb[sheet]
        elif sheet == "Продажи" and "Продажи" in wb.sheetnames:
            ws = wb["Продажи"]
        elif "Товары" in wb.sheetnames:
            ws = wb["Товары"]
        elif "Продажи" in wb.sheetnames:
            ws = wb["Продажи"]
        else:
            ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise HTTPException(status_code=400, detail="Пустой файл Excel")
        headers = [str(h or "").strip() for h in header_row]
        fields = {h.lower(): h for h in headers if h}
        out: list[dict[str, Any]] = []
        for cells in rows_iter:
            if not any(c is not None and str(c).strip() for c in cells):
                continue
            row_dict: dict[str, Any] = {}
            for i, h in enumerate(headers):
                if h and i < len(cells) and cells[i] is not None:
                    row_dict[h] = cells[i]
            out.append(row_dict)
        wb.close()
        return out

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="Файл без заголовков")
    return [dict(r) for r in reader]


def build_products_import_xlsx(kind: str = "new") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    headers = IMPORT_USED_HEADERS if kind == "used" else IMPORT_NEW_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    header_fill = PatternFill("solid", fgColor="E8E4FF")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    examples = (
        [["iPhone 13 Pro", "Graphite", "256GB", 87, "123456789012345", 65000, "Иван", "пример — удалите"]]
        if kind == "used"
        else [["2025-06-01", "iPhone 15 Pro", "Black", "256GB", "EU", "123456789012345", 85000, "Поставщик А", "собственный", ""]]
    )
    for r, ex in enumerate(examples, start=2):
        for c, val in enumerate(ex, start=1):
            ws.cell(row=r, column=c, value=val)
    for col in range(1, len(headers) + 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 14
    help_ws = wb.create_sheet("Инструкция")
    lines = (
        ["Шаблон Б/У", "• модель, imei, аккумулятор_%, себестоимость, клиент"]
        if kind == "used"
        else ["Шаблон новых товаров", "• дата_прихода, модель, регион, imei, себестоимость, поставщик", "• продажа — при продаже со склада"]
    )
    for i, line in enumerate(lines, start=1):
        help_ws.cell(row=i, column=1, value=line)
    help_ws.column_dimensions["A"].width = 72
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_sales_import_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"
    header_fill = PatternFill("solid", fgColor="E8F4FF")
    for col, title in enumerate(IMPORT_SALE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    examples = [
        ["2025-06-01 14:30", "1", "iPhone 15 Pro 256GB Black", "123456789012345", 1, 99990, 0, "наличные", "Али", ""],
        ["2025-06-01 14:30", "1", "Чехол силикон iPhone 15", "", 2, 590, 0, "наличные", "Али", ""],
        ["2025-06-02 11:00", "2", "Samsung Galaxy A54 128", "987654321098765", 1, 45000, 500, "карта", "", "пример — удалите"],
    ]
    for r, ex in enumerate(examples, start=2):
        for c, val in enumerate(ex, start=1):
            ws.cell(row=r, column=c, value=val)
    for col in range(1, len(IMPORT_SALE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    help_ws = wb.create_sheet("Инструкция")
    lines = [
        "Как заполнять шаблон продаж TeleStore",
        "",
        "• Сначала импортируйте товары (Склады → шаблон товаров).",
        "• Строки с одинаковыми «дата» + «чек» = одна продажа (несколько позиций).",
        "• название — должно совпадать с товаром в каталоге.",
        "• imei — для телефонов (необязательно при загрузке истории).",
        "• цена — пусто = цена из каталога; скидка — на весь чек (укажите в первой строке чека).",
        "• оплата: наличные, карта, ds, alif, eskhata, перевод.",
        "• Склад не списывается — данные для отчётов и аналитики.",
    ]
    for i, line in enumerate(lines, start=1):
        help_ws.cell(row=i, column=1, value=line)
    help_ws.column_dimensions["A"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def wipe_catalog_data(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DELETE FROM sale_item_units;
        DELETE FROM sale_payments;
        DELETE FROM sale_items;
        DELETE FROM sales;
        DELETE FROM unit_reservations;
        DELETE FROM product_units;
        DELETE FROM stocktake_lines;
        DELETE FROM stocktake_sessions;
        DELETE FROM stock_movements;
        DELETE FROM warehouse_stock;
        DELETE FROM trade_ins;
        DELETE FROM supplier_payments;
        DELETE FROM products;
        DELETE FROM shifts;
        """
    )


def _find_product_for_import(
    conn: sqlite3.Connection, name: str, color: str = "", memory: str = ""
) -> sqlite3.Row | None:
    if not name:
        return None
    if color or memory:
        row = conn.execute(
            """
            SELECT * FROM products
            WHERE name = ? AND COALESCE(color,'') = ? AND COALESCE(memory,'') = ?
            LIMIT 1
            """,
            (name, color, memory),
        ).fetchone()
        if row:
            return row
    return conn.execute("SELECT * FROM products WHERE name = ? LIMIT 1", (name,)).fetchone()


def _find_product_by_name(conn: sqlite3.Connection, name: str) -> sqlite3.Row | None:
    if not name:
        return None
    row = conn.execute("SELECT * FROM products WHERE name = ? LIMIT 1", (name,)).fetchone()
    if row:
        return row
    return conn.execute("SELECT * FROM products WHERE name LIKE ? LIMIT 1", (f"%{name}%",)).fetchone()


def _norm_payment(raw: str) -> str:
    v = raw.strip().lower()
    mapping = {
        "наличные": "cash", "нал": "cash", "cash": "cash", "кэш": "cash",
        "карта": "card", "card": "card", "банковская": "card", "банк": "card",
        "ds": "ds", "дс": "ds",
        "alif": "alif", "алиф": "alif",
        "eskhata": "eskhata", "эсхата": "eskhata",
        "перевод": "transfer", "transfer": "transfer",
    }
    return mapping.get(v, "cash")


def _parse_import_datetime(val: Any) -> str:
    from datetime import datetime as dt

    if isinstance(val, dt):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    s = str(val or "").strip()
    if not s:
        return utc_now()
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y",
    ):
        try:
            return dt.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return utc_now()


def _import_products_rows(
    conn: sqlite3.Connection, rows: list[dict[str, Any]], default_wh_id: int
) -> dict[str, Any]:
    if not rows:
        raise HTTPException(status_code=400, detail="Нет строк для импорта")
    kind = get_warehouse_receipt_kind(conn, default_wh_id)
    first = rows[0]
    fields = {str(k).strip().lower(): k for k in first.keys()}
    created_products = 0
    updated_stock = 0
    created_units = 0
    errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        name = _import_row_cell(row, fields, "модель") or _import_row_cell(row, fields, "название")
        if not name:
            continue
        category = "phone"
        ownership = _norm_ownership(_import_row_cell(row, fields, "тип") or "own")
        supplier = _import_row_cell(row, fields, "поставщик")
        color = _import_row_cell(row, fields, "цвет")
        memory = _import_row_cell(row, fields, "память")
        condition = "used" if kind == "used" else _norm_condition(_import_row_cell(row, fields, "состояние") or "new")
        purchase_raw = _import_row_cell(row, fields, "закупка") or _import_row_cell(row, fields, "себестоимость")
        sale_raw = _import_row_cell(row, fields, "цена")
        qty_raw = "1"
        min_stock_raw = "0"
        imei = _import_row_cell(row, fields, "imei")
        serial = _import_row_cell(row, fields, "серийник")
        notes = _import_row_cell(row, fields, "комментарий")
        battery_raw = _import_row_cell(row, fields, "аккумулятор_%") or _import_row_cell(row, fields, "батарея_%")
        client_name = _import_row_cell(row, fields, "клиент")
        region = _import_row_cell(row, fields, "регион")
        arrival_date = _import_row_cell(row, fields, "дата_прихода")

        try:
            purchase = float(purchase_raw.replace(",", ".") if purchase_raw else 0)
            sale = float(sale_raw.replace(",", ".") if sale_raw else 0)
            qty = 1
            min_stock = 0
        except ValueError:
            errors.append(f"Строка {i}: неверные числа")
            continue

        battery: int | None = None
        if battery_raw:
            try:
                battery = int(float(battery_raw.replace(",", ".")))
            except ValueError:
                errors.append(f"Строка {i}: неверная ёмкость батареи")
                continue

        if kind == "used" and battery is None:
            errors.append(f"Строка {i}: для Б/у укажите аккумулятор_%")
            continue

        if ownership == "consignment" and not supplier:
            errors.append(f"Строка {i}: укажите поставщика для реализации")
            continue

        if not imei and not serial:
            errors.append(f"Строка {i}: укажите IMEI")
            continue

        if sale <= 0:
            sale = max(purchase, 1)

        existing = _find_product_for_import(conn, name, color, memory)
        if existing:
            product_id = int(existing["id"])
            if purchase > 0 or sale > 0:
                conn.execute(
                    "UPDATE products SET purchase_price = ?, sale_price = ? WHERE id = ?",
                    (
                        purchase if purchase > 0 else existing["purchase_price"],
                        sale if sale > 0 else existing["sale_price"],
                        product_id,
                    ),
                )
            product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        else:
            track_val = 1
            cur = conn.execute(
                """
                INSERT INTO products
                (name, category, ownership_type, supplier_name, brand, sku, barcode,
                 purchase_price, sale_price, stock, min_stock, created_at,
                 model, color, size, memory, ram, customs_cleared, customs_price, specs_extra, condition, track_units, image_url)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?,
                        ?, ?, '', ?, ?, 0, 0, '', ?, ?, '')
                """,
                (
                    name, category, ownership, supplier, "", "", "",
                    purchase, sale, min_stock, utc_now(),
                    "", color, memory, "", condition, track_val,
                ),
            )
            product_id = int(cur.lastrowid)
            product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
            created_products += 1

        if imei or serial:
            if imei:
                dup = conn.execute(
                    "SELECT id FROM product_units WHERE imei = ? AND status != 'sold'", (imei,)
                ).fetchone()
                if dup:
                    errors.append(f"Строка {i}: IMEI {imei} уже есть")
                    continue
            conn.execute(
                """
                INSERT INTO product_units
                (product_id, warehouse_id, imei, serial, status, notes, created_at,
                 customs_status, customs_cleared, battery_capacity, client_name, region, arrival_date,
                 purchase_price)
                VALUES (?, ?, ?, ?, 'in_stock', ?, ?, 'none', 0, ?, ?, ?, ?, ?)
                """,
                (
                    product_id, default_wh_id, imei, serial, notes or "Импорт Excel", utc_now(), battery,
                    client_name, region, arrival_date or utc_now()[:10],
                    purchase if purchase > 0 else float(product["purchase_price"] or 0),
                ),
            )
            adjust_warehouse_stock(conn, default_wh_id, product_id, 1, "inbound", notes=f"Импорт: {imei or serial}")
            conn.execute("UPDATE products SET track_units = 1 WHERE id = ?", (product_id,))
            created_units += 1
            updated_stock += 1

    return {
        "created_products": created_products,
        "created_units": created_units,
        "stock_added": updated_stock,
        "errors": errors,
        "total_rows": len(rows),
    }


def _import_sales_rows(
    conn: sqlite3.Connection, rows: list[dict[str, Any]], warehouse_id: int
) -> dict[str, Any]:
    from collections import defaultdict

    if not rows:
        raise HTTPException(status_code=400, detail="Нет строк для импорта")
    first = rows[0]
    fields = {str(k).strip().lower(): k for k in first.keys()}
    parsed: list[tuple[int, dict[str, Any], str, str, str]] = []
    errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        name = _import_row_cell(row, fields, "название", IMPORT_SALE_ALIASES)
        if not name:
            continue
        notes = _import_row_cell(row, fields, "комментарий", IMPORT_SALE_ALIASES)
        if "удалите" in notes.lower() or "пример" in notes.lower():
            continue
        date_raw = row.get(fields.get("дата", "")) if fields.get("дата") else None
        if date_raw is None:
            for alias in IMPORT_SALE_ALIASES["дата"]:
                key = fields.get(alias)
                if key and row.get(key) not in (None, ""):
                    date_raw = row[key]
                    break
        date_key = _parse_import_datetime(date_raw)
        check = _import_row_cell(row, fields, "чек", IMPORT_SALE_ALIASES) or str(i)
        parsed.append((i, row, date_key, check, name))

    if not parsed:
        raise HTTPException(status_code=400, detail="Нет строк с названием товара")

    groups: dict[tuple[str, str], list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for i, row, date_key, check, _name in parsed:
        groups[(date_key, check)].append((i, row))

    created_sales = 0
    created_lines = 0

    for (date_key, _check), group_rows in groups.items():
        discount = 0.0
        payment_method = "cash"
        user_name = "Импорт"
        notes = ""
        lines: list[dict[str, Any]] = []

        for idx, (i, row) in enumerate(group_rows):
            name = _import_row_cell(row, fields, "название", IMPORT_SALE_ALIASES)
            imei = _import_row_cell(row, fields, "imei", IMPORT_SALE_ALIASES)
            qty_raw = _import_row_cell(row, fields, "количество", IMPORT_SALE_ALIASES) or "1"
            price_raw = _import_row_cell(row, fields, "цена", IMPORT_SALE_ALIASES)
            disc_raw = _import_row_cell(row, fields, "скидка", IMPORT_SALE_ALIASES)
            pay_raw = _import_row_cell(row, fields, "оплата", IMPORT_SALE_ALIASES)
            cashier = _import_row_cell(row, fields, "кассир", IMPORT_SALE_ALIASES)
            row_notes = _import_row_cell(row, fields, "комментарий", IMPORT_SALE_ALIASES)

            if idx == 0:
                if pay_raw:
                    payment_method = _norm_payment(pay_raw)
                if cashier:
                    user_name = cashier
                if row_notes:
                    notes = row_notes
            try:
                qty = max(1, int(float(str(qty_raw).replace(",", "."))))
            except ValueError:
                errors.append(f"Строка {i}: неверное количество")
                continue
            unit_price: float | None = None
            if price_raw:
                try:
                    unit_price = float(str(price_raw).replace(",", "."))
                except ValueError:
                    errors.append(f"Строка {i}: неверная цена")
                    continue
            if disc_raw and idx == 0:
                try:
                    discount = max(0.0, float(str(disc_raw).replace(",", ".")))
                except ValueError:
                    errors.append(f"Строка {i}: неверная скидка")
                    continue

            product = _find_product_by_name(conn, name)
            if not product:
                errors.append(f"Строка {i}: товар «{name}» не найден — сначала импортируйте товары")
                continue
            calc = calc_line(product, qty, unit_price)
            lines.append({"product": product, "qty": qty, "imei": imei, **calc})

        if not lines:
            continue

        subtotal = sum(float(l["subtotal"]) for l in lines)
        total = max(0.0, subtotal - discount)
        pay_payload = [{"method_code": payment_method, "amount": total}]
        try:
            cash_amount, card_amount, payment_method, pay_payload = validate_sale_payments(
                conn, pay_payload, total
            )
        except HTTPException as exc:
            errors.append(f"Чек {date_key}: {exc.detail}")
            continue

        sale_notes = notes or "Импорт продаж"
        cur = conn.execute(
            """
            INSERT INTO sales
            (total, discount, payment_method, status, notes, created_at,
             warehouse_id, cash_amount, card_amount, trade_in_value, shift_id, user_id, user_name)
            VALUES (?, ?, ?, 'completed', ?, ?, ?, ?, ?, 0, NULL, NULL, ?)
            """,
            (
                total, discount, payment_method, sale_notes, date_key,
                warehouse_id, cash_amount, card_amount, user_name,
            ),
        )
        sale_id = int(cur.lastrowid)
        insert_sale_payments(conn, sale_id, pay_payload)

        for line in lines:
            p = line["product"]
            cur_item = conn.execute(
                """
                INSERT INTO sale_items
                (sale_id, product_id, product_name, ownership_type, supplier_name, quantity,
                 unit_price, purchase_price, supplier_due, shop_profit, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sale_id, p["id"], p["name"], line["ownership_type"], line["supplier_name"],
                    line["qty"], line["unit_price"], line["purchase_price"],
                    line["supplier_due"], line["shop_profit"], line["subtotal"],
                ),
            )
            sale_item_id = int(cur_item.lastrowid)
            imei = line.get("imei") or ""
            if imei:
                unit = conn.execute(
                    "SELECT id FROM product_units WHERE imei = ? LIMIT 1", (imei,)
                ).fetchone()
                if unit:
                    conn.execute(
                        """
                        INSERT INTO sale_item_units
                        (sale_item_id, unit_id, imei, serial, customs_cleared, customs_price, imei_pending)
                        VALUES (?, ?, ?, '', 0, 0, 0)
                        """,
                        (sale_item_id, unit["id"], imei),
                    )
            created_lines += 1
        created_sales += 1

    return {
        "created_sales": created_sales,
        "created_lines": created_lines,
        "errors": errors,
        "total_rows": len(parsed),
    }


def _z_cell_float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace(",", ".")
    for tok in ("$", "смн", "смн.", "TJS", "tjs", "USD", "usd"):
        s = s.replace(tok, "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _z_cell_date(val: Any) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


def _z_battery_int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        v = int(val)
        return v if 0 <= v <= 100 else None
    s = str(val).strip().lower()
    if s in ("new", "usto", "id", "none"):
        return None
    try:
        v = int(float(s))
        return v if 0 <= v <= 100 else None
    except ValueError:
        return None


def _bu_warehouse_clause(alias: str = "") -> str:
    col = f"{alias}name" if alias else "name"
    return (
        f"({col} LIKE '%БУ%' OR {col} LIKE '%бу%' OR {col} LIKE '%Б/У%' OR {col} LIKE '%б/у%' "
        f"OR LOWER({col}) LIKE '%bu%')"
    )


def _accessories_warehouse_clause(col: str = "name") -> str:
    return (
        f"({col} LIKE '%аксесс%' OR {col} LIKE '%Аксесс%' OR {col} LIKE '%АКСЕСС%' "
        f"OR {col} LIKE '%accessory%' OR {col} LIKE '%Accessory%' OR {col} LIKE '%accessories%' "
        f"OR warehouse_type = 'accessories')"
    )


def _merge_duplicate_accessories_warehouses(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        f"SELECT id FROM warehouses WHERE {_accessories_warehouse_clause()} ORDER BY id"
    ).fetchall()
    if not rows:
        conn.execute(
            """
            INSERT INTO warehouses (name, address, notes, is_default, warehouse_type, created_at)
            VALUES ('Аксессуары', '', 'Склад аксессуаров', 0, 'accessories', ?)
            """,
            (utc_now(),),
        )
        rows = conn.execute(
            "SELECT id FROM warehouses WHERE warehouse_type = 'accessories' ORDER BY id"
        ).fetchall()
    if not rows:
        return
    if len(rows) < 2:
        primary_id = int(rows[0]["id"])
        conn.execute(
            """
            UPDATE warehouses SET warehouse_type = 'accessories', currency_code = 'USD',
                   notes = COALESCE(NULLIF(notes, ''), 'Склад аксессуаров')
            WHERE id = ?
            """,
            (primary_id,),
        )
        return
    primary_id = int(rows[0]["id"])
    for row in rows[1:]:
        dup_id = int(row["id"])
        conn.execute(
            "UPDATE product_units SET warehouse_id = ? WHERE warehouse_id = ?",
            (primary_id, dup_id),
        )
        conn.execute(
            "UPDATE sales SET warehouse_id = ? WHERE warehouse_id = ?",
            (primary_id, dup_id),
        )
        for ws in conn.execute(
            "SELECT product_id, quantity FROM warehouse_stock WHERE warehouse_id = ?", (dup_id,)
        ).fetchall():
            adjust_warehouse_stock(
                conn, primary_id, int(ws["product_id"]), int(ws["quantity"]), "transfer",
                notes="Слияние складов аксессуаров",
            )
        conn.execute("DELETE FROM warehouse_stock WHERE warehouse_id = ?", (dup_id,))
        conn.execute("DELETE FROM stock_movements WHERE warehouse_id = ?", (dup_id,))
        conn.execute("DELETE FROM warehouses WHERE id = ?", (dup_id,))
    conn.execute(
        """
        UPDATE warehouses SET warehouse_type = 'accessories', currency_code = 'USD',
               notes = COALESCE(NULLIF(notes, ''), 'Склад аксессуаров')
        WHERE id = ?
        """,
        (primary_id,),
    )


PHONE_NAME_HINTS = (
    "iphone", "samsung", "xiaomi", "huawei", "pixel", "oneplus", "redmi",
    "honor", "oppo", "vivo", "realme", "poco", "nokia", "sony", "google", "tecno", "infinix", "meizu",
    "ipad", "macbook", "apple watch",
)

ACCESSORY_NAME_HINTS = (
    "dyson", "whoop", "pencil", "pancil", "чехол", "case", "cover", "стекл", "glass",
    "наушник", "headphone", "earphone", "air pod", "airpod", "pods max", "pods pro", "pods ",
    "powerbank", "power bank", "заряд", "charger", "кабель", "cable",
    "magic keyboard", "клавиатур",
    "аксессуар", "accessory", "holder", "подставк", "stand", "band", "ремеш",
)


def _z_row_is_phone(row: dict[str, Any]) -> bool:
    name = (row.get("name") or "").lower().strip()
    nm = f" {name} "
    if any(h in name or h in nm for h in ACCESSORY_NAME_HINTS):
        return False
    if name.startswith("aw ") or name.startswith("apple watch") or "ipad" in name or "macbook" in name:
        return True
    memory = (row.get("memory") or "").strip().lower().replace("gb", "").replace("tb", "").strip()
    if memory and memory.isdigit():
        return True
    return any(h in name for h in PHONE_NAME_HINTS)


def _merge_duplicate_bu_warehouses(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        f"SELECT id FROM warehouses WHERE {_bu_warehouse_clause()} ORDER BY id"
    ).fetchall()
    if len(rows) < 2:
        return
    primary_id = int(rows[0]["id"])
    for row in rows[1:]:
        dup_id = int(row["id"])
        conn.execute(
            "UPDATE product_units SET warehouse_id = ? WHERE warehouse_id = ?",
            (primary_id, dup_id),
        )
        conn.execute(
            "UPDATE sales SET warehouse_id = ? WHERE warehouse_id = ?",
            (primary_id, dup_id),
        )
        conn.execute(
            """
            INSERT INTO warehouse_stock (warehouse_id, product_id, quantity)
            SELECT ?, product_id, quantity FROM warehouse_stock WHERE warehouse_id = ?
            ON CONFLICT(warehouse_id, product_id) DO UPDATE SET
                quantity = warehouse_stock.quantity + excluded.quantity
            """,
            (primary_id, dup_id),
        )
        conn.execute("DELETE FROM warehouse_stock WHERE warehouse_id = ?", (dup_id,))
        conn.execute("DELETE FROM stock_movements WHERE warehouse_id = ?", (dup_id,))
        conn.execute("DELETE FROM warehouses WHERE id = ?", (dup_id,))
    conn.execute(
        "UPDATE warehouses SET warehouse_type = 'used', notes = COALESCE(NULLIF(notes, ''), 'Б/у устройства') WHERE id = ?",
        (primary_id,),
    )


def resolve_bu_warehouse_id(conn: sqlite3.Connection) -> int:
    row = conn.execute(
        f"""
        SELECT id FROM warehouses
        WHERE {_bu_warehouse_clause()}
        ORDER BY id LIMIT 1
        """
    ).fetchone()
    if row:
        return int(row["id"])
    cur = conn.execute(
        """
        INSERT INTO warehouses (name, address, notes, is_default, created_at, warehouse_type)
        VALUES ('БУ', '', 'Б/у устройства', 0, ?, 'used')
        """,
        (utc_now(),),
    )
    return int(cur.lastrowid)


def _partnership_warehouse_clause(col: str = "name") -> str:
    return (
        f"({col} LIKE '%артнер%' OR {col} LIKE '%ARTNER%' OR {col} LIKE '%artner%' "
        f"OR warehouse_type = 'partnership')"
    )


def resolve_partnership_warehouse_id(conn: sqlite3.Connection) -> int:
    row = conn.execute(
        f"""
        SELECT id FROM warehouses
        WHERE {_partnership_warehouse_clause()}
        ORDER BY id LIMIT 1
        """
    ).fetchone()
    if row:
        wh_id = int(row["id"])
        conn.execute(
            """
            UPDATE warehouses SET warehouse_type = 'partnership', currency_code = 'USD',
                   notes = COALESCE(NULLIF(notes, ''), 'Склад партнерство')
            WHERE id = ?
            """,
            (wh_id,),
        )
        return wh_id
    cur = conn.execute(
        """
        INSERT INTO warehouses (name, address, notes, is_default, warehouse_type, currency_code, created_at)
        VALUES ('Партнерство', '', 'Склад партнерство', 0, 'partnership', 'USD', ?)
        """,
        (utc_now(),),
    )
    return int(cur.lastrowid)


def _ensure_three_core_warehouses(conn: sqlite3.Connection) -> None:
    """Б/У (TJS), Основной (USD), Партнёрство (USD)."""
    if not _table_exists(conn, "warehouses"):
        return
    bu_id = resolve_bu_warehouse_id(conn)
    conn.execute(
        """
        UPDATE warehouses SET name = 'Б/У', warehouse_type = 'used', currency_code = 'TJS',
               notes = COALESCE(NULLIF(notes, ''), 'Б/у устройства, сомони')
        WHERE id = ?
        """,
        (bu_id,),
    )
    part_id = resolve_partnership_warehouse_id(conn)
    main = conn.execute(
        f"""
        SELECT id FROM warehouses
        WHERE id NOT IN (?, ?) AND warehouse_type != 'accessories'
        AND NOT ({_accessories_warehouse_clause()})
        ORDER BY is_default DESC, id LIMIT 1
        """,
        (bu_id, part_id),
    ).fetchone()
    if main:
        main_id = int(main["id"])
        conn.execute("UPDATE warehouses SET is_default = 0")
        conn.execute(
            """
            UPDATE warehouses SET is_default = 1, warehouse_type = 'new', currency_code = 'USD',
                   name = CASE
                     WHEN name IN (?, 'БУ', 'Основной склад') OR LOWER(name) LIKE '%снов%' THEN 'Основной'
                     ELSE name END,
                   notes = COALESCE(NULLIF(notes, ''), 'Новые устройства, USD')
            WHERE id = ?
            """,
            (DEFAULT_WAREHOUSE_NAME, main_id),
        )
    else:
        conn.execute("UPDATE warehouses SET is_default = 0")
        conn.execute(
            """
            INSERT INTO warehouses (name, address, notes, is_default, warehouse_type, currency_code, created_at)
            VALUES ('Основной', '', 'Новые устройства', 1, 'new', 'USD', ?)
            """,
            (utc_now(),),
        )


def _parse_z_register_sheet(ws: Any, sheet_kind: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        if sheet_kind == "used":
            arrival = ws.cell(r, 1).value
            name = ws.cell(r, 2).value
            battery = ws.cell(r, 3).value
            imei = ws.cell(r, 4).value
            memory = ws.cell(r, 5).value
            color = ws.cell(r, 6).value
            purchase = ws.cell(r, 7).value
            extra = ws.cell(r, 8).value
            sale_price = ws.cell(r, 9).value
            sale_date = ws.cell(r, 10).value
            profit = ws.cell(r, 11).value
            comments = ws.cell(r, 12).value
            region = ""
            condition = "used"
        else:
            arrival = ws.cell(r, 1).value
            name = ws.cell(r, 2).value
            battery = ws.cell(r, 3).value
            region = ws.cell(r, 4).value
            imei = ws.cell(r, 5).value
            memory = ws.cell(r, 6).value
            color = ws.cell(r, 7).value
            purchase = ws.cell(r, 8).value
            extra = ws.cell(r, 9).value
            sale_price = ws.cell(r, 10).value
            sale_date = ws.cell(r, 11).value
            profit = ws.cell(r, 12).value
            comments = ws.cell(r, 13).value
            condition = "partnership" if sheet_kind == "partnership" else "new"
        if not name or str(name).strip().lower() in ("", "none"):
            continue
        rows.append({
            "row_num": r,
            "arrival_date": _z_cell_date(arrival),
            "name": str(name).strip(),
            "battery": _z_battery_int(battery),
            "region": str(region or "").strip(),
            "imei": str(imei or "").strip(),
            "memory": str(memory or "").strip(),
            "color": str(color or "").strip(),
            "purchase_price": _z_cell_float(purchase) or 0.0,
            "extra_cost": _z_cell_float(extra) or 0.0,
            "sale_price": _z_cell_float(sale_price),
            "sale_date": _z_cell_date(sale_date),
            "profit": _z_cell_float(profit),
            "comments": str(comments or "").strip(),
            "condition": condition,
        })
    return rows


def _rebuild_all_warehouse_stock(conn: sqlite3.Connection) -> None:
    conn.execute("DELETE FROM warehouse_stock")
    for row in conn.execute(
        """
        SELECT warehouse_id, product_id, COUNT(*) AS qty
        FROM product_units WHERE status = 'in_stock'
        GROUP BY warehouse_id, product_id
        """
    ).fetchall():
        qty = int(row["qty"])
        if qty > 0:
            conn.execute(
                "INSERT INTO warehouse_stock (warehouse_id, product_id, quantity) VALUES (?, ?, ?)",
                (int(row["warehouse_id"]), int(row["product_id"]), qty),
            )
    for pid_row in conn.execute("SELECT id FROM products WHERE IFNULL(track_units, 0) = 1").fetchall():
        sync_product_stock(conn, int(pid_row["id"]))


def reset_z_register_import(conn: sqlite3.Connection) -> dict[str, int]:
    unit_ids: set[int] = set()
    try:
        bu_id = resolve_bu_warehouse_id(conn)
        main_id = get_default_warehouse_id(conn)
        partnership_id = resolve_partnership_warehouse_id(conn)
    except HTTPException:
        return {"deleted_sales": 0, "deleted_units": 0}
    wh_ids = [main_id, bu_id, partnership_id]
    for wh_id in wh_ids:
        for r in conn.execute("SELECT id FROM product_units WHERE warehouse_id = ?", (wh_id,)).fetchall():
            unit_ids.add(int(r["id"]))
    sale_ids = [int(r["id"]) for r in conn.execute(
        "SELECT id FROM sales WHERE user_name = 'Z-импорт'"
    ).fetchall()]
    for sid in sale_ids:
        conn.execute("DELETE FROM receivables WHERE sale_id = ?", (sid,))
        conn.execute(
            "DELETE FROM sale_item_units WHERE sale_item_id IN (SELECT id FROM sale_items WHERE sale_id = ?)",
            (sid,),
        )
        conn.execute("DELETE FROM sale_items WHERE sale_id = ?", (sid,))
        conn.execute("DELETE FROM sale_payments WHERE sale_id = ?", (sid,))
        conn.execute("DELETE FROM sales WHERE id = ?", (sid,))

    deleted_units = 0
    for uid in unit_ids:
        conn.execute("DELETE FROM sale_item_units WHERE unit_id = ?", (uid,))
        conn.execute("DELETE FROM unit_reservations WHERE unit_id = ?", (uid,))
        if conn.execute("DELETE FROM product_units WHERE id = ?", (uid,)).rowcount:
            deleted_units += 1
    conn.execute(
        "DELETE FROM product_units WHERE sale_id IS NOT NULL AND sale_id NOT IN (SELECT id FROM sales)"
    )
    _rebuild_all_warehouse_stock(conn)
    _fix_z_register_warehouse_split(conn)
    conn.execute("DELETE FROM expenses WHERE description = 'Excel импорт ОПиУ'")
    return {"deleted_sales": len(sale_ids), "deleted_units": deleted_units}


def _import_z_register_rows(
    conn: sqlite3.Connection,
    rows: list[dict[str, Any]],
    warehouse_id: int,
    *,
    sheet_kind: str = "used",
    affect_cash: bool = True,
) -> dict[str, Any]:
    created_units = sold = skipped = 0
    errors: list[str] = []
    cash_flag = 1 if affect_cash else 0
    expected_condition = (
        "used" if sheet_kind == "used"
        else "partnership" if sheet_kind == "partnership"
        else "new"
    )
    currency_code = get_warehouse_currency(conn, warehouse_id)["code"]

    def resolve_sheet_product(
        *,
        name: str,
        memory: str,
        color: str,
        purchase: float,
        sale_price: float | None,
    ) -> int:
        """Карточка товара строго под лист Excel (used / new / partnership)."""
        existing_prod = conn.execute(
            """
            SELECT id FROM products
            WHERE category = 'phone' AND LOWER(TRIM(name)) = LOWER(TRIM(?))
              AND LOWER(TRIM(COALESCE(model, ''))) = LOWER(TRIM(?))
              AND LOWER(TRIM(COALESCE(memory, ''))) = LOWER(TRIM(?))
              AND LOWER(TRIM(COALESCE(color, ''))) = LOWER(TRIM(?))
              AND condition = ?
            LIMIT 1
            """,
            (name, name, memory, color, expected_condition),
        ).fetchone()
        if existing_prod:
            product_id = int(existing_prod["id"])
            conn.execute(
                """
                UPDATE products
                SET purchase_price = ?, sale_price = COALESCE(?, sale_price), condition = ?
                WHERE id = ?
                """,
                (purchase, sale_price, expected_condition, product_id),
            )
            return product_id
        sp = sale_price if sale_price and sale_price > 0 else round(purchase * 1.15, 2)
        cur = conn.execute(
            """
            INSERT INTO products
            (name, category, ownership_type, supplier_name, brand, sku, barcode,
             purchase_price, sale_price, stock, min_stock, created_at,
             model, color, size, memory, ram, customs_cleared, customs_price, specs_extra,
             condition, track_units, image_url)
            VALUES (?, 'phone', 'own', '', '', '', '', ?, ?, 0, 0, ?,
                    ?, ?, '', ?, '', 0, 0, '', ?, 1, '')
            """,
            (name, purchase, sp, utc_now(), name, color, memory, expected_condition),
        )
        return int(cur.lastrowid)

    def bind_unit_to_sheet(
        unit_id: int,
        *,
        product_id: int,
        purchase: float,
        extra: float,
        row: dict[str, Any],
        imei: str,
        notes: str,
    ) -> None:
        """Единица всегда на складе текущего листа; себестоимость — из строки Excel."""
        old = conn.execute(
            "SELECT product_id, warehouse_id, status FROM product_units WHERE id = ?",
            (unit_id,),
        ).fetchone()
        old_wh = int(old["warehouse_id"] or warehouse_id) if old else warehouse_id
        old_product = int(old["product_id"]) if old else product_id
        old_status = (old["status"] or "").strip() if old else ""
        conn.execute(
            """
            UPDATE product_units
            SET status = 'in_stock', sale_id = NULL, product_id = ?, warehouse_id = ?,
                purchase_price = ?, customs_price = ?, battery_capacity = ?,
                client_name = ?, region = ?, arrival_date = ?, notes = ?,
                serial = COALESCE(NULLIF(serial, ''), ?)
            WHERE id = ?
            """,
            (
                product_id,
                warehouse_id,
                purchase,
                extra,
                row["battery"],
                row["comments"] or "",
                row["region"] or "",
                (row["arrival_date"] or "")[:10],
                notes,
                imei,
                unit_id,
            ),
        )
        if old_status == "in_stock" and old_wh != warehouse_id:
            # Перенос между складами: −1 со старого (если есть), +1 на лист Excel.
            try:
                adjust_warehouse_stock(
                    conn,
                    old_wh,
                    old_product,
                    -1,
                    "transfer",
                    notes=f"Z-импорт перенос IMEI {imei} → склад листа",
                )
            except HTTPException:
                pass
            adjust_warehouse_stock(
                conn,
                warehouse_id,
                product_id,
                1,
                "transfer",
                notes=f"Z-импорт перенос IMEI {imei} на склад листа",
            )
            sync_product_stock(conn, old_product)
            sync_product_stock(conn, product_id)
        elif old_status != "in_stock":
            adjust_warehouse_stock(
                conn,
                warehouse_id,
                product_id,
                1,
                "inbound",
                notes=f"Z-импорт на склад листа IMEI {imei}: {row['name']}",
            )
            if old and (old_wh != warehouse_id or old_product != product_id):
                sync_product_stock(conn, old_product)
            sync_product_stock(conn, product_id)
        elif old_product != product_id:
            sync_product_stock(conn, old_product)
            sync_product_stock(conn, product_id)

    def record_sale(
        *,
        row: dict[str, Any],
        product_id: int,
        unit_id: int,
        name: str,
        purchase: float,
        extra: float,
        imei: str,
        serial: str,
        sale_price: float,
        sale_date: str,
        profit: float | None,
    ) -> None:
        nonlocal sold
        shop_profit = profit if profit is not None else (sale_price - purchase - extra)
        total = float(sale_price)
        cur_s = conn.execute(
            """
            INSERT INTO sales
            (total, discount, payment_method, status, notes, created_at,
             warehouse_id, cash_amount, card_amount, trade_in_value, shift_id, user_id, user_name,
             amount_paid, amount_due, currency_code, affects_cash)
            VALUES (?, 0, 'cash', 'completed', ?, ?, ?, ?, 0, 0, NULL, NULL, 'Z-импорт', ?, 0, ?, ?)
            """,
            (total, row["comments"] or "Z-импорт Excel", sale_date, warehouse_id, total, total, currency_code, cash_flag),
        )
        sale_id = int(cur_s.lastrowid)
        insert_sale_payments(conn, sale_id, [{
            "method_code": "cash",
            "amount": total,
            "pay_currency_code": currency_code,
            "pay_amount": total,
        }])
        cur_item = conn.execute(
            """
            INSERT INTO sale_items
            (sale_id, product_id, product_name, ownership_type, supplier_name, quantity,
             unit_price, purchase_price, supplier_due, shop_profit, subtotal)
            VALUES (?, ?, ?, 'own', '', 1, ?, ?, 0, ?, ?)
            """,
            (sale_id, product_id, name, total, purchase, shop_profit, total),
        )
        sale_item_id = int(cur_item.lastrowid)
        conn.execute(
            """
            INSERT INTO sale_item_units
            (sale_item_id, unit_id, imei, serial, customs_cleared, customs_price, imei_pending)
            VALUES (?, ?, ?, ?, 0, ?, 0)
            """,
            (sale_item_id, unit_id, imei, serial, extra),
        )
        mark_unit_sold_full(conn, unit_id, sale_id, imei, 0, extra)
        adjust_warehouse_stock(conn, warehouse_id, product_id, -1, "sale", reference_id=sale_id)
        sold += 1

    for row in rows:
        name = row["name"]
        imei = row["imei"]
        if imei.lower() in ("id", "none", "n/a", ""):
            imei = ""
        purchase = float(row["purchase_price"])
        extra = float(row["extra_cost"])
        sale_price = row["sale_price"]
        sale_date = row["sale_date"]
        profit = row["profit"]
        has_sale = bool(sale_price and sale_price > 0 and sale_date)
        if sale_price and sale_price > 0 and not sale_date:
            errors.append(f"Строка {row['row_num']}: {name} — нет даты продажи")
        condition = expected_condition
        memory = row["memory"]
        color = row["color"]
        serial = imei if imei else f"Z{row['row_num']}-{warehouse_id}"
        reuse_unit: sqlite3.Row | None = None

        if imei:
            in_stock = conn.execute(
                """
                SELECT id, product_id, warehouse_id FROM product_units
                WHERE imei = ? AND imei != '' AND status = 'in_stock' LIMIT 1
                """,
                (imei,),
            ).fetchone()
            if in_stock:
                product_id = resolve_sheet_product(
                    name=name,
                    memory=memory,
                    color=color,
                    purchase=purchase,
                    sale_price=sale_price,
                )
                unit_notes = (row["comments"] or "").strip()
                unit_notes = f"Z-импорт: {unit_notes}" if unit_notes else "Z-импорт Excel"
                bind_unit_to_sheet(
                    int(in_stock["id"]),
                    product_id=product_id,
                    purchase=purchase,
                    extra=extra,
                    row=row,
                    imei=imei,
                    notes=unit_notes,
                )
                if has_sale:
                    reuse_unit = conn.execute(
                        "SELECT id, product_id FROM product_units WHERE id = ?",
                        (int(in_stock["id"]),),
                    ).fetchone()
                else:
                    skipped += 1
                    continue
            else:
                existing_any = conn.execute(
                    """
                    SELECT id, product_id, warehouse_id, status, sale_id
                    FROM product_units
                    WHERE imei = ? AND imei != ''
                    ORDER BY id DESC LIMIT 1
                    """,
                    (imei,),
                ).fetchone()
                if existing_any and not has_sale:
                    # Return / re-accept: same IMEI was sold, Excel shows it back in stock.
                    # Склад и condition — строго по листу Excel (партнёрство остаётся партнёрством).
                    unit_id = int(existing_any["id"])
                    product_id = resolve_sheet_product(
                        name=name,
                        memory=memory,
                        color=color,
                        purchase=purchase,
                        sale_price=None,
                    )
                    unit_notes = (row["comments"] or "").strip()
                    unit_notes = (
                        f"Z-импорт возврат: {unit_notes}" if unit_notes else "Z-импорт возврат Excel"
                    )
                    bind_unit_to_sheet(
                        unit_id,
                        product_id=product_id,
                        purchase=purchase,
                        extra=extra,
                        row=row,
                        imei=imei,
                        notes=unit_notes,
                    )
                    created_units += 1
                    continue
                if existing_any and has_sale:
                    # Same IMEI sold again in Excel = return + resale (or duplicate of same sale).
                    # Склад продажи = лист Excel, без принудительного Б/У.
                    already = conn.execute(
                        """
                        SELECT s.id
                        FROM sales s
                        JOIN sale_items si ON si.sale_id = s.id
                        JOIN sale_item_units siu ON siu.sale_item_id = si.id
                        WHERE siu.imei = ?
                          AND s.warehouse_id = ?
                          AND substr(s.created_at, 1, 10) = ?
                          AND ABS(COALESCE(s.total, 0) - ?) < 0.02
                        LIMIT 1
                        """,
                        (imei, warehouse_id, str(sale_date)[:10], float(sale_price)),
                    ).fetchone()
                    if already:
                        skipped += 1
                        continue
                    unit_id = int(existing_any["id"])
                    product_id = resolve_sheet_product(
                        name=name,
                        memory=memory,
                        color=color,
                        purchase=purchase,
                        sale_price=sale_price,
                    )
                    unit_notes = (row["comments"] or "").strip()
                    unit_notes = (
                        f"Z-импорт возврат+продажа: {unit_notes}"
                        if unit_notes
                        else "Z-импорт возврат+продажа Excel"
                    )
                    bind_unit_to_sheet(
                        unit_id,
                        product_id=product_id,
                        purchase=purchase,
                        extra=extra,
                        row=row,
                        imei=imei,
                        notes=unit_notes,
                    )
                    reuse_unit = conn.execute(
                        "SELECT id, product_id FROM product_units WHERE id = ?",
                        (unit_id,),
                    ).fetchone()
        else:
            # No IMEI: stable serial Z{row}-{wh}. Skip re-creating the same Excel row.
            by_serial = conn.execute(
                """
                SELECT id, product_id FROM product_units
                WHERE serial = ? AND status = 'in_stock' LIMIT 1
                """,
                (serial,),
            ).fetchone()
            if by_serial:
                product_id = resolve_sheet_product(
                    name=name,
                    memory=memory,
                    color=color,
                    purchase=purchase,
                    sale_price=sale_price,
                )
                bind_unit_to_sheet(
                    int(by_serial["id"]),
                    product_id=product_id,
                    purchase=purchase,
                    extra=extra,
                    row=row,
                    imei=serial,
                    notes=(row["comments"] or "").strip() or "Z-импорт Excel",
                )
                if has_sale:
                    reuse_unit = conn.execute(
                        "SELECT id, product_id FROM product_units WHERE id = ?",
                        (int(by_serial["id"]),),
                    ).fetchone()
                else:
                    skipped += 1
                    continue

        if reuse_unit:
            unit_id = int(reuse_unit["id"])
            product_id = int(reuse_unit["product_id"])
            conn.execute(
                "UPDATE products SET purchase_price = ?, sale_price = COALESCE(?, sale_price) WHERE id = ?",
                (purchase, sale_price, product_id),
            )
            conn.execute(
                """
                UPDATE product_units
                SET purchase_price = ?, customs_price = ?, warehouse_id = ?, product_id = ?
                WHERE id = ?
                """,
                (purchase, extra, warehouse_id, product_id, unit_id),
            )
        else:
            product_id = resolve_sheet_product(
                name=name,
                memory=memory,
                color=color,
                purchase=purchase,
                sale_price=sale_price,
            )

            unit_notes = (row["comments"] or "").strip()
            unit_notes = f"Z-импорт: {unit_notes}" if unit_notes else "Z-импорт Excel"
            cur_u = conn.execute(
                """
                INSERT INTO product_units
                (product_id, warehouse_id, imei, serial, status, notes, created_at,
                 customs_status, customs_cleared, customs_price, battery_capacity,
                 client_name, region, arrival_date, purchase_price)
                VALUES (?, ?, ?, ?, 'in_stock', ?, ?, 'none', 0, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product_id, warehouse_id, imei, serial, unit_notes,
                    utc_now(), extra, row["battery"], row["comments"] or "", row["region"] or "",
                    (row["arrival_date"] or "")[:10], purchase,
                ),
            )
            unit_id = int(cur_u.lastrowid)
            adjust_warehouse_stock(conn, warehouse_id, product_id, 1, "inbound", notes=f"Z-импорт: {name}")
            created_units += 1

        if has_sale:
            record_sale(
                row=row,
                product_id=product_id,
                unit_id=unit_id,
                name=name,
                purchase=purchase,
                extra=extra,
                imei=imei,
                serial=serial,
                sale_price=float(sale_price),
                sale_date=sale_date,
                profit=profit,
            )

    return {
        "created_units": created_units,
        "sold_units": sold,
        "skipped_duplicates": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


def _import_accessory_z_rows(
    conn: sqlite3.Connection,
    rows: list[dict[str, Any]],
    warehouse_id: int,
) -> dict[str, Any]:
    created = sold = skipped = 0
    errors: list[str] = []
    currency_code = get_warehouse_currency(conn, warehouse_id)["code"]
    for row in rows:
        name = row["name"]
        purchase = float(row["purchase_price"])
        extra = float(row["extra_cost"])
        sale_price = row["sale_price"]
        sale_date = row["sale_date"]
        profit = row["profit"]
        model = (row.get("comments") or "").strip()

        existing = conn.execute(
            """
            SELECT id FROM products
            WHERE category = 'accessory' AND LOWER(TRIM(name)) = LOWER(TRIM(?))
            LIMIT 1
            """,
            (name,),
        ).fetchone()
        if existing:
            product_id = int(existing["id"])
            conn.execute(
                "UPDATE products SET purchase_price = ?, sale_price = COALESCE(?, sale_price) WHERE id = ?",
                (purchase, sale_price, product_id),
            )
        else:
            sp = sale_price if sale_price and sale_price > 0 else round(purchase * 1.3, 2)
            cur = conn.execute(
                """
                INSERT INTO products
                (name, category, ownership_type, supplier_name, brand, sku, barcode,
                 purchase_price, sale_price, stock, min_stock, created_at,
                 model, color, size, memory, ram, customs_cleared, customs_price, specs_extra,
                 condition, track_units, image_url)
                VALUES (?, 'accessory', 'own', '', '', '', '', ?, ?, 0, 0, ?,
                        ?, '', '', '', '', 0, 0, '', 'new', 0, '')
                """,
                (name, purchase, sp, utc_now(), model),
            )
            product_id = int(cur.lastrowid)

        if sale_price and sale_price > 0 and sale_date:
            shop_profit = profit if profit is not None else (sale_price - purchase - extra)
            total = float(sale_price)
            cur_s = conn.execute(
                """
                INSERT INTO sales
                (total, discount, payment_method, status, notes, created_at,
                 warehouse_id, cash_amount, card_amount, trade_in_value, shift_id, user_id, user_name,
                 amount_paid, amount_due, currency_code, affects_cash)
                VALUES (?, 0, 'cash', 'completed', ?, ?, ?, ?, 0, 0, NULL, NULL, 'Z-импорт акс', ?, 0, ?, 0)
                """,
                (total, row["comments"] or "Z-импорт аксессуаров", sale_date, warehouse_id, total, total, currency_code),
            )
            sale_id = int(cur_s.lastrowid)
            insert_sale_payments(conn, sale_id, [{
            "method_code": "cash",
            "amount": total,
            "pay_currency_code": currency_code,
            "pay_amount": total,
        }])
            conn.execute(
                """
                INSERT INTO sale_items
                (sale_id, product_id, product_name, ownership_type, supplier_name, quantity,
                 unit_price, purchase_price, supplier_due, shop_profit, subtotal)
                VALUES (?, ?, ?, 'own', '', 1, ?, ?, 0, ?, ?)
                """,
                (sale_id, product_id, name, total, purchase, shop_profit, total),
            )
            sold += 1
        else:
            adjust_warehouse_stock(
                conn, warehouse_id, product_id, 1, "inbound",
                notes=f"Z-импорт акс: {name}",
            )
            created += 1
    for pid_row in conn.execute(
        "SELECT id FROM products WHERE category = 'accessory'"
    ).fetchall():
        sync_product_stock(conn, int(pid_row["id"]))
    return {
        "created_units": created,
        "sold_units": sold,
        "skipped_duplicates": skipped,
        "errors": errors,
        "total_rows": len(rows),
    }


def reset_accessories_z_import(conn: sqlite3.Connection) -> dict[str, int]:
    try:
        acc_wh = resolve_accessories_warehouse_id(conn)
    except HTTPException:
        return {"deleted_sales": 0, "warehouse_id": 0}
    sale_ids = [
        int(r["id"]) for r in conn.execute(
            "SELECT id FROM sales WHERE user_name = 'Z-импорт акс' AND warehouse_id = ?",
            (acc_wh,),
        ).fetchall()
    ]
    for sid in sale_ids:
        conn.execute("DELETE FROM receivables WHERE sale_id = ?", (sid,))
        conn.execute(
            "DELETE FROM sale_item_units WHERE sale_item_id IN (SELECT id FROM sale_items WHERE sale_id = ?)",
            (sid,),
        )
        conn.execute("DELETE FROM sale_items WHERE sale_id = ?", (sid,))
        conn.execute("DELETE FROM sale_payments WHERE sale_id = ?", (sid,))
        conn.execute("DELETE FROM sales WHERE id = ?", (sid,))
    conn.execute("DELETE FROM warehouse_stock WHERE warehouse_id = ?", (acc_wh,))
    conn.execute(
        "DELETE FROM product_units WHERE warehouse_id = ? AND notes LIKE 'Z-импорт%'",
        (acc_wh,),
    )
    _rebuild_all_warehouse_stock(conn)
    return {"deleted_sales": len(sale_ids), "warehouse_id": acc_wh}


def import_accessories_excel(
    conn: sqlite3.Connection, raw: bytes, *, replace: bool = False
) -> dict[str, Any]:
    from openpyxl import load_workbook

    results: dict[str, Any] = {}
    if replace:
        results["reset"] = reset_accessories_z_import(conn)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    all_rows: list[dict[str, Any]] = []
    for sn in wb.sheetnames:
        low = sn.lower()
        if low in ("бу", "bu", "б/у"):
            continue
        rows = _parse_z_register_sheet(wb[sn], "new")
        all_rows.extend([r for r in rows if not _z_row_is_phone(r)])
    acc_wh = resolve_accessories_warehouse_id(conn)
    results["import"] = _import_accessory_z_rows(conn, all_rows, acc_wh) | {"warehouse_id": acc_wh}
    return results


def _parse_opu_expenses_sheet(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    in_expenses = False
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c1s = str(c1 or "").strip().lower()
        c2s = str(c2 or "").strip()
        if c1s in ("расходы", "расход", "expenses"):
            in_expenses = True
            amount = _z_cell_float(c3)
            if c2s and amount and amount > 0 and not c2s.lower().startswith("итого"):
                rows.append({"category": c2s, "amount": float(amount)})
            continue
        if not in_expenses:
            continue
        if c2s.lower().startswith("итого"):
            break
        amount = _z_cell_float(c3)
        if c2s and amount and amount > 0:
            rows.append({"category": c2s, "amount": float(amount)})
    return rows


def _expense_date_from_filename(filename: str) -> str:
    low = (filename or "").lower()
    year = "2026"
    for y in range(2024, 2031):
        if str(y) in low:
            year = str(y)
            break
    month_map = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
        "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12",
        "янв": "01", "фев": "02", "мар": "03", "апр": "04", "май": "05", "мая": "05",
        "июн": "06", "июл": "07", "авг": "08", "сен": "09", "окт": "10", "ноя": "11", "дек": "12",
    }
    for key, mm in month_map.items():
        if key in low:
            import calendar
            last = calendar.monthrange(int(year), int(mm))[1]
            return f"{year}-{mm}-{last:02d}"
    return utc_now()[:10]


def import_opu_expenses(
    conn: sqlite3.Connection,
    rows: list[dict[str, Any]],
    expense_date: str,
    *,
    replace: bool = False,
    affect_cash: bool = False,
) -> dict[str, Any]:
    if replace:
        conn.execute("DELETE FROM expenses WHERE description = 'Excel импорт ОПиУ'")
    created = 0
    total = 0.0
    cash_flag = 1 if affect_cash else 0
    for row in rows:
        cat = row["category"]
        amount = float(row["amount"])
        dept = "accessories" if "аксесс" in cat.lower() else "main"
        conn.execute(
            """
            INSERT INTO expenses (
                category, amount, description, payment_method_code, expense_date, created_at,
                department, created_by, created_by_user_id, affects_cash
            )
            VALUES (?, ?, 'Excel импорт ОПиУ', 'cash', ?, ?, ?, 'Excel импорт', NULL, ?)
            """,
            (cat, amount, expense_date[:10], utc_now(), dept, cash_flag),
        )
        created += 1
        total += amount
    return {"created": created, "total": round(total, 2), "expense_date": expense_date[:10], "affects_cash": cash_flag}


def import_z_register_excel(
    conn: sqlite3.Connection,
    raw: bytes,
    filename: str,
    sheet: str = "",
    *,
    replace: bool = False,
    affect_cash: bool = False,
) -> dict[str, Any]:
    from openpyxl import load_workbook

    results: dict[str, Any] = {"sheets": {}}
    if replace:
        results["reset"] = reset_z_register_import(conn)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    targets: list[tuple[str, str, int]] = []
    if sheet and sheet in wb.sheetnames:
        low = sheet.lower()
        if low in ("бу", "bu", "б/у"):
            kind, wh = "used", resolve_bu_warehouse_id(conn)
        elif "партнер" in low or low == "partnership":
            kind, wh = "partnership", resolve_partnership_warehouse_id(conn)
        else:
            kind, wh = "new", get_default_warehouse_id(conn)
        targets = [(sheet, kind, wh)]
    else:
        for sn in wb.sheetnames:
            low = sn.lower()
            if low in ("бу", "bu", "б/у"):
                targets.append((sn, "used", resolve_bu_warehouse_id(conn)))
            elif "партнер" in low or low == "partnership":
                targets.append((sn, "partnership", resolve_partnership_warehouse_id(conn)))
            elif "доллар" in low or low == "new":
                targets.append((sn, "new", get_default_warehouse_id(conn)))
    if not targets:
        sn = wb.sheetnames[0]
        targets = [(sn, "used", resolve_bu_warehouse_id(conn))]
    for sn, kind, wh_id in targets:
        ws = wb[sn]
        rows = _parse_z_register_sheet(ws, kind)
        results["sheets"][sn] = _import_z_register_rows(
            conn, rows, wh_id, sheet_kind=kind, affect_cash=affect_cash
        ) | {"warehouse_id": wh_id, "kind": kind}
    for sn in wb.sheetnames:
        low = sn.lower().replace(" ", "")
        if low in ("опу", "opu", "опиу", "opiu", "p&l", "pnl"):
            expense_rows = _parse_opu_expenses_sheet(wb[sn])
            if expense_rows:
                exp_date = _expense_date_from_filename(filename)
                results["expenses"] = import_opu_expenses(
                    conn, expense_rows, exp_date, replace=replace, affect_cash=affect_cash
                )
            break
    results["affects_cash"] = bool(affect_cash)
    return results


def get_warehouse_receipt_kind(conn: sqlite3.Connection, warehouse_id: int) -> str:
    row = conn.execute("SELECT name, warehouse_type FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
    if not row:
        return "new"
    wt = (row["warehouse_type"] or "").strip().lower()
    name = (row["name"] or "").lower()
    if wt == "accessories" or "аксесс" in name:
        return "accessories"
    if wt == "used":
        return "used"
    if wt == "partnership" or "артнер" in name:
        return "partnership"
    if wt == "new":
        return "new"
    if "бу" in name or "б/у" in name or "б у" in name:
        return "used"
    return "new"


def _normalize_accessory_name(name: str) -> str:
    return " ".join((name or "").lower().split())


def _reassign_product_id(conn: sqlite3.Connection, old_id: int, new_id: int) -> None:
    if old_id == new_id:
        return
    conn.execute("UPDATE sale_items SET product_id = ? WHERE product_id = ?", (new_id, old_id))
    if _table_exists(conn, "stock_movements"):
        conn.execute("UPDATE stock_movements SET product_id = ? WHERE product_id = ?", (new_id, old_id))
    if _table_exists(conn, "product_units"):
        conn.execute("UPDATE product_units SET product_id = ? WHERE product_id = ?", (new_id, old_id))
    if _table_exists(conn, "trade_ins"):
        conn.execute("UPDATE trade_ins SET given_product_id = ? WHERE given_product_id = ?", (new_id, old_id))
        conn.execute(
            "UPDATE trade_ins SET received_product_id = ? WHERE received_product_id = ?",
            (new_id, old_id),
        )
    if _table_exists(conn, "stocktake_lines"):
        conn.execute("UPDATE stocktake_lines SET product_id = ? WHERE product_id = ?", (new_id, old_id))
    if _table_exists(conn, "unit_reservations"):
        conn.execute("UPDATE unit_reservations SET product_id = ? WHERE product_id = ?", (new_id, old_id))
    conn.execute("DELETE FROM warehouse_stock WHERE product_id = ?", (old_id,))
    conn.execute("DELETE FROM products WHERE id = ?", (old_id,))


def _dedupe_accessory_products(conn: sqlite3.Connection) -> None:
    try:
        acc_wh = resolve_accessories_warehouse_id(conn)
    except HTTPException:
        return
    groups: dict[str, list[int]] = {}
    for r in conn.execute("SELECT id, name FROM products WHERE category = 'accessory'").fetchall():
        key = _normalize_accessory_name(r["name"])
        if not key:
            continue
        groups.setdefault(key, []).append(int(r["id"]))
    changed = False
    for ids in groups.values():
        if len(ids) < 2:
            continue
        primary_id = min(ids)
        for dup_id in ids:
            if dup_id == primary_id:
                continue
            qty = get_warehouse_stock(conn, acc_wh, dup_id)
            if qty > 0:
                adjust_warehouse_stock(
                    conn, acc_wh, dup_id, -qty, "transfer", notes="Слияние дубликата аксессуара",
                )
                adjust_warehouse_stock(
                    conn, acc_wh, primary_id, qty, "transfer", notes="Слияние дубликата аксессуара",
                )
            try:
                _reassign_product_id(conn, dup_id, primary_id)
            except sqlite3.IntegrityError as exc:
                logger.warning("Accessory dedupe: could not merge product %s into %s: %s", dup_id, primary_id, exc)
                continue
            changed = True
    if changed:
        for pid_row in conn.execute("SELECT id FROM products WHERE category = 'accessory'").fetchall():
            sync_product_stock(conn, int(pid_row["id"]))


def _reclassify_phones_from_accessories(conn: sqlite3.Connection) -> None:
    if not _table_exists(conn, "warehouse_stock"):
        return
    try:
        acc_wh = resolve_accessories_warehouse_id(conn)
        main_id = get_default_warehouse_id(conn)
    except HTTPException:
        return
    moved = False
    rows = conn.execute(
        "SELECT id, name, memory FROM products WHERE category = 'accessory'"
    ).fetchall()
    for r in rows:
        if not _z_row_is_phone({"name": r["name"], "memory": r["memory"] or ""}):
            continue
        pid = int(r["id"])
        name = r["name"] or ""
        conn.execute(
            "UPDATE products SET category = 'phone', track_units = 1, condition = 'new' WHERE id = ?",
            (pid,),
        )
        qty = get_warehouse_stock(conn, acc_wh, pid)
        if qty > 0:
            adjust_warehouse_stock(
                conn, acc_wh, pid, -qty, "transfer", notes=f"Перенос на основной склад: {name}",
            )
            for i in range(qty):
                conn.execute(
                    """
                    INSERT INTO product_units
                    (product_id, warehouse_id, imei, serial, status, notes, created_at,
                     customs_status, customs_cleared, customs_price, battery_capacity,
                     client_name, region, arrival_date, purchase_price)
                    VALUES (?, ?, '', ?, 'in_stock', 'Перенос с аксессуаров', ?, 'none', 0, 0, NULL, '', '', '', ?)
                    """,
                    (pid, main_id, f"ACC{pid}-{i + 1}", utc_now(), float(r["purchase_price"] or 0)),
                )
                adjust_warehouse_stock(
                    conn, main_id, pid, 1, "transfer", notes=f"Перенос с аксессуаров: {name}",
                )
            moved = True
        conn.execute(
            """
            UPDATE sales SET warehouse_id = ?
            WHERE warehouse_id = ? AND id IN (SELECT sale_id FROM sale_items WHERE product_id = ?)
            """,
            (main_id, acc_wh, pid),
        )
    if moved:
        _rebuild_all_warehouse_stock(conn)


def _fix_misclassified_accessories(conn: sqlite3.Connection) -> None:
    if not _table_exists(conn, "warehouse_stock"):
        return
    try:
        acc_wh = resolve_accessories_warehouse_id(conn)
        main_id = get_default_warehouse_id(conn)
        bu_id = resolve_bu_warehouse_id(conn)
        partnership_id = resolve_partnership_warehouse_id(conn)
    except HTTPException:
        return
    phone_wh_ids = (main_id, bu_id, partnership_id)
    moved = False

    def move_qty_to_acc(product_id: int, from_wh: int, qty: int, note: str) -> None:
        nonlocal moved
        if qty <= 0:
            return
        adjust_warehouse_stock(conn, from_wh, product_id, -qty, "transfer", notes=note)
        adjust_warehouse_stock(conn, acc_wh, product_id, qty, "transfer", notes=note)
        moved = True

    if _table_exists(conn, "product_units"):
        rows = conn.execute(
            """
            SELECT u.id, u.product_id, p.name, p.memory, p.category
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            WHERE u.warehouse_id IN (?, ?) AND u.status = 'in_stock'
            """,
            phone_wh_ids,
        ).fetchall()
        for r in rows:
            is_acc = r["category"] == "accessory" or (
                r["category"] == "phone"
                and not _z_row_is_phone({"name": r["name"], "memory": r["memory"] or ""})
            )
            if not is_acc:
                continue
            pid = int(r["product_id"])
            uid = int(r["id"])
            wh_id = int(conn.execute(
                "SELECT warehouse_id FROM product_units WHERE id = ?", (uid,)
            ).fetchone()["warehouse_id"])
            conn.execute("UPDATE products SET category = 'accessory', track_units = 0 WHERE id = ?", (pid,))
            conn.execute("DELETE FROM product_units WHERE id = ?", (uid,))
            move_qty_to_acc(pid, wh_id, 1, "Перенос аксессуара со склада телефонов")

    ws_rows = conn.execute(
        """
        SELECT ws.product_id, ws.warehouse_id, ws.quantity, p.name, p.memory, p.category
        FROM warehouse_stock ws
        JOIN products p ON p.id = ws.product_id
        WHERE ws.warehouse_id IN (?, ?) AND ws.quantity > 0
        """,
        phone_wh_ids,
    ).fetchall()
    for r in ws_rows:
        is_acc = r["category"] == "accessory" or (
            r["category"] == "phone"
            and not _z_row_is_phone({"name": r["name"], "memory": r["memory"] or ""})
        )
        if not is_acc:
            continue
        pid = int(r["product_id"])
        wh_id = int(r["warehouse_id"])
        qty = int(r["quantity"])
        conn.execute("UPDATE products SET category = 'accessory', track_units = 0 WHERE id = ?", (pid,))
        move_qty_to_acc(pid, wh_id, qty, "Перенос аксессуара со склада телефонов")

    conn.execute(
        """
        UPDATE sales SET currency_code = 'USD'
        WHERE warehouse_id = ? AND COALESCE(NULLIF(TRIM(currency_code), ''), 'TJS') = 'TJS'
        """,
        (acc_wh,),
    )

    if moved:
        _rebuild_all_warehouse_stock(conn)
    for pid_row in conn.execute("SELECT id FROM products WHERE category = 'accessory'").fetchall():
        sync_product_stock(conn, int(pid_row["id"]))


CURRENCY_META: dict[str, dict[str, str]] = {
    "USD": {"code": "USD", "symbol": "$", "name": "Доллар США"},
    "TJS": {"code": "TJS", "symbol": "смн", "name": "Сомони"},
}


def currency_meta(code: str) -> dict[str, str]:
    c = (code or "TJS").upper()
    return CURRENCY_META.get(c, {"code": c, "symbol": c, "name": c})


def get_warehouse_currency(conn: sqlite3.Connection, warehouse_id: int) -> dict[str, str]:
    row = conn.execute(
        "SELECT currency_code, warehouse_type, name FROM warehouses WHERE id = ?", (warehouse_id,)
    ).fetchone()
    code = (row["currency_code"] or "").strip().upper() if row else ""
    if not code:
        code = "TJS" if get_warehouse_receipt_kind(conn, warehouse_id) == "used" else "USD"
    return currency_meta(code)


def _set_warehouse_currencies(conn: sqlite3.Connection) -> None:
    if not _table_exists(conn, "warehouses"):
        return
    for row in conn.execute("SELECT id, warehouse_type, currency_code, is_default FROM warehouses").fetchall():
        wh_id = int(row["id"])
        kind = get_warehouse_receipt_kind(conn, wh_id)
        if kind == "used":
            expected = "TJS"
        elif kind == "accessories":
            expected = "USD"
        elif kind == "partnership":
            expected = "USD"
        elif int(row["is_default"] or 0):
            expected = "USD"
        else:
            continue
        current = (row["currency_code"] or "").strip().upper()
        if current != expected:
            conn.execute("UPDATE warehouses SET currency_code = ? WHERE id = ?", (expected, wh_id))


def _transfer_unit_between_warehouses(
    conn: sqlite3.Connection,
    unit_id: int,
    product_id: int,
    from_wh: int,
    to_wh: int,
) -> None:
    if from_wh == to_wh:
        return
    unit = conn.execute(
        "SELECT status, sale_id FROM product_units WHERE id = ?", (unit_id,)
    ).fetchone()
    conn.execute("UPDATE product_units SET warehouse_id = ? WHERE id = ?", (to_wh, unit_id))
    if unit and unit["sale_id"]:
        conn.execute("UPDATE sales SET warehouse_id = ? WHERE id = ?", (to_wh, unit["sale_id"]))
    if unit and unit["status"] == "in_stock":
        current = get_warehouse_stock(conn, from_wh, product_id)
        if current > 0:
            adjust_warehouse_stock(
                conn, from_wh, product_id, -1, "transfer",
                target_warehouse_id=to_wh, notes="Перенос по типу NEW/БУ",
            )
            adjust_warehouse_stock(
                conn, to_wh, product_id, 1, "transfer",
                target_warehouse_id=from_wh, notes="Перенос по типу NEW/БУ",
            )


def _fix_z_register_warehouse_split(conn: sqlite3.Connection) -> None:
    if not _table_exists(conn, "product_units") or not _table_exists(conn, "warehouses"):
        return
    try:
        bu_id = resolve_bu_warehouse_id(conn)
        main_id = get_default_warehouse_id(conn)
    except HTTPException:
        return
    if bu_id == main_id:
        return
    try:
        for u in conn.execute(
            """
            SELECT u.id, u.product_id FROM product_units u
            JOIN products p ON p.id = u.product_id
            WHERE u.warehouse_id = ? AND COALESCE(p.condition, 'new') != 'new'
            """,
            (main_id,),
        ).fetchall():
            _transfer_unit_between_warehouses(conn, int(u["id"]), int(u["product_id"]), main_id, bu_id)
        for u in conn.execute(
            """
            SELECT u.id, u.product_id FROM product_units u
            JOIN products p ON p.id = u.product_id
            WHERE u.warehouse_id = ? AND COALESCE(p.condition, 'new') = 'new'
            """,
            (bu_id,),
        ).fetchall():
            _transfer_unit_between_warehouses(conn, int(u["id"]), int(u["product_id"]), bu_id, main_id)
    except Exception as exc:
        logger.warning("Z-register warehouse split fix skipped: %s", exc)


def _unit_line_profit(
    *,
    shop_profit: float | None,
    has_sale: bool,
    sale_price: float | None,
    item_purchase_price: float | None,
    product_purchase_price: float,
    extra_cost: float,
) -> float | None:
    """Прибыль строки Z-отчёта. shop_profit=0 — валидное значение (продажа в ноль), не пересчитывать."""
    if not has_sale:
        return None
    if shop_profit is not None:
        return float(shop_profit)
    if sale_price:
        purchase = float(
            item_purchase_price if item_purchase_price is not None else product_purchase_price
        )
        return float(sale_price) - purchase - float(extra_cost or 0)
    return None


def _sync_product_purchase_from_sales(conn: sqlite3.Connection) -> None:
    """Подтянуть себестоимость в карточку товара из последней продажи (если в карточке 0)."""
    try:
        conn.execute(
            """
            UPDATE products
            SET purchase_price = (
                SELECT si.purchase_price
                FROM product_units u
                JOIN sale_item_units siu ON siu.unit_id = u.id
                JOIN sale_items si ON si.id = siu.sale_item_id
                JOIN sales s ON s.id = si.sale_id AND s.status = 'completed'
                WHERE u.product_id = products.id AND si.purchase_price > 0
                ORDER BY s.created_at DESC
                LIMIT 1
            )
            WHERE purchase_price = 0
              AND category = 'phone'
              AND id IN (
                  SELECT DISTINCT u.product_id
                  FROM product_units u
                  JOIN sale_item_units siu ON siu.unit_id = u.id
                  JOIN sale_items si ON si.id = siu.sale_item_id
                  JOIN sales s ON s.id = si.sale_id AND s.status = 'completed'
                  WHERE si.purchase_price > 0
              )
            """
        )
    except Exception as exc:
        logger.warning("Product purchase sync from sales skipped: %s", exc)


def _z_register_condition_clause(conn: sqlite3.Connection, warehouse_id: int) -> str:
    kind = get_warehouse_receipt_kind(conn, warehouse_id)
    if kind == "used":
        return " AND COALESCE(p.condition, 'used') IN ('used', 'refurbished')"
    if kind == "partnership":
        return " AND COALESCE(p.condition, 'new') = 'partnership'"
    return " AND COALESCE(p.condition, 'new') = 'new'"


def _z_register_lines(
    conn: sqlite3.Connection,
    warehouse_id: int,
    year: int | None = None,
    month: int | None = None,
) -> dict[str, Any]:
    """Строки Z-отчёта склада.

    Продажи — строго по sales.warehouse_id (лист Excel / склад продажи).
    Один IMEI может быть продан с партнёрства и с Б/У в разные циклы:
    каждая продажа остаётся на своём складе; чужой склад не трогаем.
    Остатки — по текущему warehouse_id единицы и её себестоимости из Excel
    (не из старой продажи), иначе после возврата остаток «плывёт».
    Фильтр condition не применяем: склады раздельные, а карточка товара
    после смены листа может ещё быть partnership/used — иначе теряем строки.
    """
    stock_rows = conn.execute(
        """
        SELECT u.id AS unit_id, u.arrival_date, u.imei, u.serial, u.battery_capacity,
               u.client_name, u.region, u.customs_price AS extra_cost, u.status, u.notes,
               u.purchase_price AS unit_purchase_price,
               p.name AS product_name, p.model, p.memory, p.color,
               p.purchase_price AS product_purchase_price, p.condition
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        WHERE u.warehouse_id = ? AND u.status = 'in_stock'
        ORDER BY COALESCE(u.arrival_date, u.created_at), p.name, u.id
        """,
        (warehouse_id,),
    ).fetchall()

    sold_rows = conn.execute(
        """
        SELECT u.id AS unit_id, u.arrival_date, u.imei, u.serial, u.battery_capacity,
               u.client_name, u.region,
               COALESCE(siu.customs_price, u.customs_price, 0) AS extra_cost,
               'sold' AS status, COALESCE(s.notes, u.notes) AS notes,
               COALESCE(si.purchase_price, u.purchase_price) AS unit_purchase_price,
               p.name AS product_name, p.model, p.memory, p.color,
               p.purchase_price AS product_purchase_price, p.condition,
               si.subtotal AS sale_price, si.shop_profit, si.purchase_price AS item_purchase_price,
               s.created_at AS sale_date, s.id AS sale_id
        FROM sales s
        JOIN sale_items si ON si.sale_id = s.id
        JOIN sale_item_units siu ON siu.sale_item_id = si.id
        JOIN product_units u ON u.id = siu.unit_id
        JOIN products p ON p.id = COALESCE(si.product_id, u.product_id)
        WHERE s.status = 'completed' AND s.warehouse_id = ?
        ORDER BY s.created_at, p.name, u.id
        """,
        (warehouse_id,),
    ).fetchall()

    def _build_line(d: dict[str, Any], *, has_sale: bool) -> dict[str, Any]:
        sd = d.get("sale_date") or ""
        product_purchase = unit_purchase_price(
            {"purchase_price": d.get("unit_purchase_price")},
            {"purchase_price": d.get("product_purchase_price")},
        )
        item_purchase_raw = d.get("item_purchase_price")
        item_purchase = float(item_purchase_raw) if item_purchase_raw is not None else None
        extra_cost = float(d.get("extra_cost") or 0)
        sale_price = float(d.get("sale_price") or 0) if d.get("sale_price") else None
        shop_profit_raw = d.get("shop_profit")
        shop_profit = float(shop_profit_raw) if shop_profit_raw is not None else None
        display_purchase = item_purchase if has_sale and item_purchase is not None else product_purchase
        return {
            "arrival_date": (d.get("arrival_date") or "")[:10] or "—",
            "product_name": d.get("product_name") or "",
            "condition": d.get("condition") or "",
            "region": d.get("region") or "",
            "imei": d.get("imei") or d.get("serial") or "—",
            "memory": d.get("memory") or "",
            "color": d.get("color") or "",
            "purchase_price": display_purchase,
            "extra_cost": extra_cost,
            "sale_price": sale_price,
            "sale_date": sd[:10] if sd else "",
            "profit": _unit_line_profit(
                shop_profit=shop_profit,
                has_sale=has_sale,
                sale_price=sale_price,
                item_purchase_price=item_purchase,
                product_purchase_price=product_purchase,
                extra_cost=extra_cost,
            ),
            "comments": d.get("notes") or d.get("client_name") or "",
            "battery": d.get("battery_capacity"),
            "status": d.get("status") or "",
        }

    stock_lines: list[dict[str, Any]] = []
    for r in stock_rows:
        d = row_to_dict(r) or {}
        stock_lines.append(_build_line(d, has_sale=False))

    sold_period: list[dict[str, Any]] = []
    period_profit = period_revenue = 0.0
    period_count = 0
    all_sold_lines: list[dict[str, Any]] = []
    for r in sold_rows:
        d = row_to_dict(r) or {}
        sd = d.get("sale_date") or ""
        line = _build_line(d, has_sale=True)
        all_sold_lines.append(line)
        in_period = False
        if year and month and sd:
            try:
                dt = datetime.strptime(sd[:10], "%Y-%m-%d")
                in_period = dt.year == year and dt.month == month
            except ValueError:
                pass
        elif not year and not month:
            in_period = bool(sd)
        if in_period and sd:
            sold_period.append(line)
            period_profit += float(line["profit"] or 0)
            period_revenue += float(line["sale_price"] or 0)
            period_count += 1

    all_lines = list(stock_lines) + list(all_sold_lines)
    month_names = ["", "январь", "февраль", "март", "апрель", "май", "июнь",
                   "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
    period_label = f"{month_names[month]} {year}" if year and month else "Все данные"
    return {
        "period_label": period_label,
        "year": year,
        "month": month,
        "sold_count": period_count,
        "revenue": period_revenue,
        "profit": period_profit,
        "stock_count": len(stock_lines),
        "stock_value": sum(l["purchase_price"] + l["extra_cost"] for l in stock_lines),
        "sold_lines": sold_period,
        "stock_lines": stock_lines,
        "all_lines": all_lines,
    }


def get_currency_config(conn: sqlite3.Connection) -> dict[str, str]:
    return {
        "code": get_setting(conn, "base_currency", "TJS"),
        "symbol": get_setting(conn, "currency_symbol", "смн"),
        "name": get_setting(conn, "currency_name", "Сомони"),
    }


def latest_exchange_rates(conn: sqlite3.Connection) -> dict[str, float]:
    """Latest rate per currency: 1 unit of currency = rate units of base currency."""
    base = get_setting(conn, "base_currency", "TJS").upper()
    rates: dict[str, float] = {base: 1.0}
    rows = conn.execute(
        """
        SELECT e.currency_code, e.rate
        FROM exchange_rates e
        INNER JOIN (
            SELECT UPPER(currency_code) AS code, MAX(effective_at) AS mx
            FROM exchange_rates
            GROUP BY UPPER(currency_code)
        ) t ON UPPER(e.currency_code) = t.code AND e.effective_at = t.mx
        """
    ).fetchall()
    for r in rows:
        code = str(r["currency_code"] or "").upper()
        if code:
            rates[code] = float(r["rate"] or 1.0)
    rates[base] = 1.0
    return rates


def convert_amount(
    conn: sqlite3.Connection,
    amount: float,
    from_code: str,
    to_code: str,
    *,
    at: str | None = None,
) -> float:
    src = (from_code or "").upper()
    dst = (to_code or "").upper()
    if not src or not dst or src == dst:
        return float(amount)
    base = get_setting(conn, "base_currency", "TJS").upper()
    for code in (src, dst):
        if code == base:
            continue
        exists = conn.execute(
            "SELECT 1 FROM exchange_rates WHERE UPPER(currency_code) = ? LIMIT 1",
            (code,),
        ).fetchone()
        if not exists:
            raise HTTPException(
                status_code=400,
                detail=f"Нет курса {code} в настройках — добавьте курс к базовой валюте",
            )
    from_rate = get_exchange_rate_at(conn, src, at)
    to_rate = get_exchange_rate_at(conn, dst, at)
    if from_rate <= 0 or to_rate <= 0:
        raise HTTPException(status_code=400, detail=f"Некорректный курс {src}/{dst}")
    return float(amount) * from_rate / to_rate


def _infer_payment_method_currency(name: str = "", code: str = "", explicit: str = "") -> str:
    """USD / TJS / '' (any). Prefer explicit; else detect from name/code."""
    ex = (explicit or "").strip().upper()
    if ex in ("USD", "TJS"):
        return ex
    blob = f"{name or ''} {code or ''}".lower()
    if any(x in blob for x in ("доллар", "dollar", "usd", "$")):
        return "USD"
    if any(x in blob for x in ("сомон", "смн", "наличн", "tjs")) or (code or "").lower() == "cash":
        return "TJS"
    return ""


def _backfill_payment_method_currencies(conn: sqlite3.Connection) -> None:
    if not _table_exists(conn, "payment_methods"):
        return
    if "currency_code" not in {r[1] for r in conn.execute("PRAGMA table_info(payment_methods)").fetchall()}:
        return
    for row in conn.execute("SELECT id, code, name, currency_code FROM payment_methods").fetchall():
        current = (row["currency_code"] or "").strip().upper()
        inferred = _infer_payment_method_currency(row["name"], row["code"], current)
        if inferred and inferred != current:
            conn.execute(
                "UPDATE payment_methods SET currency_code = ? WHERE id = ?",
                (inferred, row["id"]),
            )


def payment_method_currency(pm: dict[str, Any] | sqlite3.Row | None) -> str:
    if not pm:
        return ""
    return _infer_payment_method_currency(
        pm["name"] if "name" in pm.keys() else "",
        pm["code"] if "code" in pm.keys() else "",
        pm["currency_code"] if "currency_code" in pm.keys() else "",
    )


def cash_wallet_code_for_currency(conn: sqlite3.Connection, currency: str) -> str | None:
    """Find cash-type payment method bound to USD or TJS."""
    want = (currency or "").strip().upper()
    if want not in ("USD", "TJS"):
        return None
    methods = list_payment_methods(conn, active_only=True)
    exact = [
        m for m in methods
        if (m.get("method_type") or "") == "cash" and payment_method_currency(m) == want
    ]
    if exact:
        # Prefer code containing usd/dollar for USD
        if want == "USD":
            prefer = [m for m in exact if "usd" in m["code"].lower() or "dollar" in m["code"].lower()
                      or "доллар" in (m.get("name") or "").lower()]
            return (prefer or exact)[0]["code"]
        prefer = [m for m in exact if m["code"] == "cash"]
        return (prefer or exact)[0]["code"]
    return None


def resolve_wallet_method_code(
    conn: sqlite3.Connection,
    method_code: str,
    money_currency: str,
) -> str:
    """Route cash payments into the matching currency cash wallet (USD vs сомони)."""
    code = (method_code or "cash").strip() or "cash"
    pm = get_payment_method(conn, code)
    if not pm or (pm["method_type"] or "") != "cash":
        return code
    money_cur = (money_currency or "TJS").strip().upper() or "TJS"
    target = cash_wallet_code_for_currency(conn, money_cur)
    return target or code


def list_payment_methods(conn: sqlite3.Connection, active_only: bool = True) -> list[dict[str, Any]]:
    sql = "SELECT * FROM payment_methods"
    if active_only:
        sql += " WHERE is_active = 1"
    sql += " ORDER BY sort_order, name"
    rows = [row_to_dict(r) for r in conn.execute(sql).fetchall()]
    for r in rows:
        r["currency_code"] = payment_method_currency(r)
    return rows


def get_payment_method(conn: sqlite3.Connection, code: str) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM payment_methods WHERE code = ? AND is_active = 1", (code,)
    ).fetchone()


def get_exchange_rate_at(conn: sqlite3.Connection, currency_code: str, at: str | None = None) -> float:
    code = currency_code.upper()
    base = get_setting(conn, "base_currency", "TJS").upper()
    if code == base:
        return 1.0
    when = at or utc_now()
    row = conn.execute(
        """
        SELECT rate FROM exchange_rates
        WHERE currency_code = ? AND effective_at <= ?
        ORDER BY effective_at DESC LIMIT 1
        """,
        (code, when),
    ).fetchone()
    return float(row["rate"]) if row else 1.0


def sale_payments_for(conn: sqlite3.Connection, sale_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT sp.*, pm.name AS method_name, pm.method_type
        FROM sale_payments sp
        LEFT JOIN payment_methods pm ON pm.code = sp.method_code
        WHERE sp.sale_id = ?
        ORDER BY sp.id
        """,
        (sale_id,),
    ).fetchall()
    return [row_to_dict(r) for r in rows]


def validate_sale_payments(
    conn: sqlite3.Connection,
    payments: list[dict[str, float | str]],
    total: float,
) -> tuple[float, float, str, list[dict[str, float | str]]]:
    cash_amount, card_amount, payment_method, normalized, _, _ = process_sale_payments(
        conn, payments, total, require_full=True
    )
    return cash_amount, card_amount, payment_method, normalized


def process_sale_payments(
    conn: sqlite3.Connection,
    payments: list[dict[str, float | str]],
    total: float,
    *,
    debtor_name: str = "",
    require_full: bool = False,
) -> tuple[float, float, str, list[dict[str, float | str]], float, float]:
    paid = sum(float(p["amount"]) for p in payments if float(p["amount"]) > 0)
    due = max(0.0, total - paid)
    if paid > total + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Сумма оплат ({paid:.2f}) больше итога ({total:.2f})",
        )
    if require_full and due > 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Сумма оплат ({paid:.2f}) не совпадает с итогом ({total:.2f})",
        )
    if due > 0.01 and not debtor_name.strip():
        raise HTTPException(
            status_code=400,
            detail="Сумма оплаты меньше цены продажи — укажите имя и телефон должника",
        )
    if paid <= 0 and due <= 0.01 and total > 0 and not payments:
        raise HTTPException(status_code=400, detail="Укажите способ оплаты")
    cash_amount = 0.0
    card_amount = 0.0
    codes: list[str] = []
    normalized: list[dict[str, float | str]] = []
    for p in payments:
        code = str(p["method_code"]).strip()
        amount = float(p["amount"])
        if amount <= 0:
            continue
        pm = get_payment_method(conn, code)
        if not pm:
            raise HTTPException(status_code=400, detail=f"Способ оплаты «{code}» не найден")
        normalized.append({
            "method_code": code,
            "amount": amount,
            "pay_currency_code": str(p.get("pay_currency_code") or p.get("currency_code") or "").strip().upper(),
            "pay_amount": float(p["pay_amount"]) if p.get("pay_amount") not in (None, "") else amount,
        })
        codes.append(code)
        if pm["method_type"] == "cash":
            cash_amount += amount
        else:
            card_amount += amount
    if not normalized and due <= 0.01 and total > 0:
        raise HTTPException(status_code=400, detail="Укажите способ оплаты")
    payment_method = codes[0] if len(codes) == 1 else ("split" if codes else "credit")
    return cash_amount, card_amount, payment_method, normalized, paid, due


def debt_reduction_from_payment(
    *,
    debt_currency: str,
    pay_amount: float,
    pay_currency: str,
    fx_rate: float | None,
) -> float:
    """Reduce debt by pay_amount in pay_currency. fx_rate = 1 USD in TJS."""
    debt = (debt_currency or "TJS").strip().upper() or "TJS"
    pay = (pay_currency or debt).strip().upper() or debt
    amt = float(pay_amount or 0)
    if amt <= 0:
        raise HTTPException(status_code=400, detail="Сумма оплаты должна быть больше 0")
    if pay == debt:
        return round(amt, 2)
    rate = float(fx_rate or 0)
    if rate <= 0:
        raise HTTPException(
            status_code=400,
            detail="Укажите курс: сколько сомони за 1 доллар",
        )
    if debt == "USD" and pay == "TJS":
        return round(amt / rate, 2)
    if debt == "TJS" and pay == "USD":
        return round(amt * rate, 2)
    raise HTTPException(status_code=400, detail=f"Неподдерживаемая пара валют {pay} → {debt}")


def create_receivable(
    conn: sqlite3.Connection,
    *,
    sale_id: int,
    customer_name: str,
    customer_phone: str,
    total_amount: float,
    paid_amount: float,
    warehouse_id: int | None,
    notes: str = "",
) -> int:
    due = max(0.0, total_amount - paid_amount)
    cur = conn.execute(
        """
        INSERT INTO receivables
        (sale_id, customer_name, customer_phone, total_amount, paid_amount, amount_due,
         status, warehouse_id, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, 'open', ?, ?, ?)
        """,
        (
            sale_id, customer_name.strip(), customer_phone.strip(),
            total_amount, paid_amount, due, warehouse_id, notes, utc_now(),
        ),
    )
    return int(cur.lastrowid)


def _close_receivables_for_return(conn: sqlite3.Connection, sale_id: int, now: str) -> None:
    """Закрыть долг клиента при возврате (полный или частичный)."""
    conn.execute(
        """
        UPDATE receivables
        SET status = 'closed', amount_due = 0,
            closed_at = ?,
            notes = TRIM(COALESCE(notes, '') || ?)
        WHERE sale_id = ?
          AND (status = 'open' OR amount_due > 0.001)
        """,
        (now, f" Закрыто: возврат продажи #{sale_id}", sale_id),
    )


def _sync_receivables_for_voided_sales(conn: sqlite3.Connection) -> None:
    """Убрать из дебиторки долги по уже отменённым продажам."""
    now = utc_now()
    conn.execute(
        """
        UPDATE receivables
        SET status = 'closed', amount_due = 0,
            closed_at = COALESCE(closed_at, ?),
            notes = TRIM(COALESCE(notes, '') || ' [Авто: продажа отменена]')
        WHERE sale_id IN (SELECT id FROM sales WHERE status = 'voided')
          AND (status = 'open' OR amount_due > 0.001)
        """,
        (now,),
    )


def get_expense_warehouse_split(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT ews.warehouse_id, ews.pct, w.name AS warehouse_name
        FROM expense_warehouse_split ews
        JOIN warehouses w ON w.id = ews.warehouse_id
        WHERE ews.pct > 0
        ORDER BY w.name
        """
    ).fetchall()
    if rows:
        return [
            {"warehouse_id": r["warehouse_id"], "warehouse_name": r["warehouse_name"], "pct": float(r["pct"])}
            for r in rows
        ]
    whs = conn.execute(
        """
        SELECT id, name FROM warehouses
        WHERE COALESCE(warehouse_type, '') != 'accessories'
          AND LOWER(name) NOT LIKE '%аксесс%'
        ORDER BY id
        """
    ).fetchall()
    if not whs:
        return []
    pct = round(100.0 / len(whs), 4)
    return [
        {"warehouse_id": w["id"], "warehouse_name": w["name"], "pct": pct}
        for w in whs
    ]


def allocate_expense_amount(amount: float, split: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not split:
        return []
    total_pct = sum(float(s["pct"]) for s in split) or 100.0
    out = []
    for s in split:
        share = amount * float(s["pct"]) / total_pct
        out.append({**s, "amount": round(share, 2)})
    return out


def expenses_allocated_by_warehouse(
    conn: sqlite3.Connection,
    exp_clause: str,
    exp_params: list[Any],
) -> tuple[float, float, list[dict[str, Any]], list[dict[str, Any]]]:
    """Split expenses: warehouse_id set → 100% that WH; NULL → % from settings.

    Returns (total, main_total, expenses_by_warehouse, split_rules).
    """
    rows = conn.execute(
        f"""
        SELECT warehouse_id,
               COALESCE(department, 'main') AS department,
               COALESCE(SUM(amount), 0) AS total
        FROM expenses
        WHERE 1=1 {exp_clause}
        GROUP BY warehouse_id, COALESCE(department, 'main')
        """,
        exp_params,
    ).fetchall()
    shared = 0.0
    shared_main = 0.0
    direct: dict[int, float] = {}
    direct_main: dict[int, float] = {}
    total = 0.0
    main_total = 0.0
    for r in rows:
        amt = float(r["total"] or 0)
        total += amt
        is_main = (r["department"] or "main") != "accessories"
        if is_main:
            main_total += amt
        wid = r["warehouse_id"]
        if wid is None:
            shared += amt
            if is_main:
                shared_main += amt
        else:
            wid_i = int(wid)
            direct[wid_i] = direct.get(wid_i, 0.0) + amt
            if is_main:
                direct_main[wid_i] = direct_main.get(wid_i, 0.0) + amt

    split = get_expense_warehouse_split(conn)
    shared_alloc = allocate_expense_amount(shared, split)
    wh_names = {
        int(w["id"]): w["name"]
        for w in conn.execute("SELECT id, name FROM warehouses").fetchall()
    }
    by_wh: dict[int, dict[str, Any]] = {}
    for s in shared_alloc:
        wid = int(s["warehouse_id"])
        shared_amt = float(s.get("amount") or 0)
        by_wh[wid] = {
            "warehouse_id": wid,
            "warehouse_name": s.get("warehouse_name") or wh_names.get(wid, f"#{wid}"),
            "pct_rule": float(s.get("pct") or 0),
            "shared_amount": shared_amt,
            "direct_amount": 0.0,
            "amount": shared_amt,
        }
    for wid, amt in direct.items():
        if wid not in by_wh:
            by_wh[wid] = {
                "warehouse_id": wid,
                "warehouse_name": wh_names.get(wid, f"#{wid}"),
                "pct_rule": 0.0,
                "shared_amount": 0.0,
                "direct_amount": 0.0,
                "amount": 0.0,
            }
        by_wh[wid]["direct_amount"] = round(amt, 2)
        by_wh[wid]["amount"] = round(float(by_wh[wid]["amount"]) + amt, 2)

    out = sorted(by_wh.values(), key=lambda x: x["warehouse_name"])
    for row in out:
        row["pct"] = round(100.0 * float(row["amount"]) / total, 2) if total else 0.0
        row["amount"] = round(float(row["amount"]), 2)
        row["shared_amount"] = round(float(row["shared_amount"]), 2)
    return round(total, 2), round(main_total, 2), out, split


def expense_amount_for_warehouse(
    expenses_by_warehouse: list[dict[str, Any]],
    warehouse_id: int,
) -> float:
    row = next(
        (x for x in expenses_by_warehouse if int(x.get("warehouse_id") or 0) == int(warehouse_id)),
        None,
    )
    return float(row["amount"]) if row else 0.0


def _cash_net_before(conn: sqlite3.Connection, cutoff: str) -> float:
    sale_in = conn.execute(
        """
        SELECT COALESCE(SUM(sp.amount), 0)
        FROM sale_payments sp
        JOIN sales s ON s.id = sp.sale_id
        WHERE s.status = 'completed' AND COALESCE(s.affects_cash, 1) = 1 AND s.created_at < ?
        """,
        (cutoff,),
    ).fetchone()[0]
    recv_in = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) FROM receivable_payments WHERE created_at < ?",
        (cutoff,),
    ).fetchone()[0]
    manual_in = 0.0
    if _table_exists(conn, "cash_inflows"):
        manual_in = float(conn.execute(
            "SELECT COALESCE(SUM(amount_base), 0) FROM cash_inflows WHERE created_at < ?",
            (cutoff,),
        ).fetchone()[0])
    sup_out = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) FROM supplier_payments WHERE created_at < ?",
        (cutoff,),
    ).fetchone()[0]
    exp_out = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE COALESCE(affects_cash, 1) = 1 AND expense_date < ?",
        (cutoff[:10],),
    ).fetchone()[0]
    return float(sale_in) + float(recv_in) + float(manual_in) - float(sup_out) - float(exp_out)


def _period_cutoff_start(period: str, date_from: str, date_to: str) -> str:
    if date_from:
        return date_from if len(date_from) > 10 else f"{date_from} 00:00:00"
    if period == "all":
        return "9999-12-31"
    return period_start(period)


def insert_sale_payments(conn: sqlite3.Connection, sale_id: int, payments: list[dict[str, float | str]]) -> None:
    for p in payments:
        amount = float(p["amount"])
        pay_cur = str(p.get("pay_currency_code") or p.get("currency_code") or "").strip().upper()
        pay_amt = p.get("pay_amount")
        if pay_amt is None or pay_amt == "":
            pay_amt = amount
        else:
            pay_amt = float(pay_amt)
        conn.execute(
            """
            INSERT INTO sale_payments (sale_id, method_code, amount, pay_currency_code, pay_amount)
            VALUES (?, ?, ?, ?, ?)
            """,
            (sale_id, p["method_code"], amount, pay_cur, pay_amt),
        )


def persist_sale_payments(
    conn: sqlite3.Connection,
    sale_id: int,
    payments: list[dict[str, float | str]],
    total: float,
) -> tuple[float, float, str]:
    cash_amount, card_amount, payment_method, normalized = validate_sale_payments(conn, payments, total)
    insert_sale_payments(conn, sale_id, normalized)
    return cash_amount, card_amount, payment_method


def enrich_sale(conn: sqlite3.Connection, sale: sqlite3.Row) -> dict[str, Any]:
    data = row_to_dict(sale) or {}
    items = conn.execute("SELECT * FROM sale_items WHERE sale_id = ?", (sale["id"],)).fetchall()
    data["items"] = enrich_sale_items(conn, items)
    data["payments"] = sale_payments_for(conn, sale["id"])
    rec = conn.execute(
        "SELECT * FROM receivables WHERE sale_id = ? ORDER BY id DESC LIMIT 1",
        (sale["id"],),
    ).fetchone()
    data["receivable"] = row_to_dict(rec) if rec else None
    data["supplier_due_total"] = sum(
        float(i["supplier_due"] or 0)
        for i in items
        if (i["ownership_type"] or "") == "consignment"
    )
    suppliers = sorted({
        (i["supplier_name"] or "").strip()
        for i in items
        if (i["ownership_type"] or "") == "consignment" and (i["supplier_name"] or "").strip()
    })
    data["supplier_names"] = suppliers
    return data


def payments_for_sales(conn: sqlite3.Connection, sale_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not sale_ids:
        return {}
    placeholders = ",".join("?" * len(sale_ids))
    rows = conn.execute(
        f"""
        SELECT sp.sale_id, sp.method_code, sp.amount, pm.name
        FROM sale_payments sp
        LEFT JOIN payment_methods pm ON pm.code = sp.method_code
        WHERE sp.sale_id IN ({placeholders})
        ORDER BY sp.sale_id, pm.sort_order, sp.method_code
        """,
        sale_ids,
    ).fetchall()
    out: dict[int, list[dict[str, Any]]] = {}
    for r in rows:
        out.setdefault(r["sale_id"], []).append({
            "method_code": r["method_code"],
            "name": r["name"] or r["method_code"],
            "amount": float(r["amount"]),
        })
    return out


def _report_period_clause(
    period: str, date_from: str, date_to: str, column: str = "s.created_at"
) -> tuple[str, list[Any], str]:
    if date_from or date_to:
        clause, params = date_filter_sql(date_from, date_to, column)
        label = f"{date_from or '…'} — {date_to or '…'}"
        return clause, params, label
    if period == "all":
        return "", [], "Всё время"
    labels = {"day": "Сегодня", "week": "Неделя", "month": "Месяц", "quarter": "Квартал", "year": "Год"}
    return f" AND {column} >= ?", [period_start(period)], labels.get( period, period)


def resolve_user(conn: sqlite3.Connection, pin: str | None) -> dict[str, Any] | None:
    if not pin:
        user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if user_count == 0 and not settings.store_pin:
            return {"id": 0, "name": "Система", "role": "owner", "pin": ""}
        return None
    row = conn.execute(
        "SELECT * FROM users WHERE pin = ? AND is_active = 1", (pin,)
    ).fetchone()
    if row:
        return row_to_dict(row)
    if settings.store_pin and pin == settings.store_pin:
        return {"id": 0, "name": "Владелец", "role": "owner", "pin": pin}
    return None


def check_pin(pin: str | None, *, min_role: str | None = None) -> dict[str, Any]:
    with db() as conn:
        user = resolve_user(conn, pin)
        if not user:
            raise HTTPException(status_code=401, detail="Неверный PIN")
        if min_role and ROLE_LEVEL.get(user["role"], 0) < ROLE_LEVEL.get(min_role, 99):
            raise HTTPException(status_code=403, detail="Недостаточно прав")
        return user


def get_open_shift(conn: sqlite3.Connection) -> sqlite3.Row | None:
    return conn.execute(
        "SELECT * FROM shifts WHERE status = 'open' ORDER BY id DESC LIMIT 1"
    ).fetchone()


def shift_payment_breakdown(conn: sqlite3.Connection, shift_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT sp.method_code, pm.name, pm.method_type, COALESCE(SUM(sp.amount), 0) AS amount
        FROM sale_payments sp
        JOIN sales s ON s.id = sp.sale_id
        LEFT JOIN payment_methods pm ON pm.code = sp.method_code
        WHERE s.shift_id = ? AND s.status = 'completed'
        GROUP BY sp.method_code
        ORDER BY pm.sort_order, sp.method_code
        """,
        (shift_id,),
    ).fetchall()
    return [
        {
            "method_code": r["method_code"],
            "name": r["name"] or r["method_code"],
            "method_type": r["method_type"] or "card",
            "amount": float(r["amount"]),
        }
        for r in rows
    ]


def shift_sales_totals(conn: sqlite3.Connection, shift_id: int) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT COUNT(*) AS cnt, COALESCE(SUM(total), 0) AS total
        FROM sales
        WHERE shift_id = ? AND status = 'completed'
        """,
        (shift_id,),
    ).fetchone()
    by_payment = shift_payment_breakdown(conn, shift_id)
    expected_cash = sum(p["amount"] for p in by_payment if p["method_type"] == "cash")
    expected_card = sum(p["amount"] for p in by_payment if p["method_type"] != "cash")
    return {
        "sales_count": row["cnt"],
        "expected_cash": expected_cash,
        "expected_card": expected_card,
        "total_revenue": float(row["total"]),
        "by_payment": by_payment,
    }


def expire_reservations(conn: sqlite3.Connection) -> None:
    now = utc_now()
    rows = conn.execute(
        """
        SELECT id, unit_id FROM unit_reservations
        WHERE status = 'active' AND reserved_until < ?
        """,
        (now,),
    ).fetchall()
    for row in rows:
        conn.execute("UPDATE unit_reservations SET status = 'expired' WHERE id = ?", (row["id"],))
        conn.execute(
            "UPDATE product_units SET status = 'in_stock' WHERE id = ? AND status = 'reserved'",
            (row["unit_id"],),
        )


def fulfill_unit_reservation(conn: sqlite3.Connection, unit_id: int) -> None:
    conn.execute(
        "UPDATE unit_reservations SET status = 'fulfilled' WHERE unit_id = ? AND status = 'active'",
        (unit_id,),
    )


def pick_units(
    conn: sqlite3.Connection,
    product_id: int,
    warehouse_id: int,
    quantity: int,
    unit_ids: list[int] | None,
) -> list[sqlite3.Row]:
    expire_reservations(conn)
    if unit_ids:
        if len(unit_ids) != quantity:
            raise HTTPException(status_code=400, detail="Количество IMEI должно совпадать с количеством товара")
        placeholders = ",".join("?" * len(unit_ids))
        rows = conn.execute(
            f"""
            SELECT * FROM product_units
            WHERE id IN ({placeholders}) AND product_id = ? AND warehouse_id = ?
              AND status IN ('in_stock', 'reserved')
            """,
            (*unit_ids, product_id, warehouse_id),
        ).fetchall()
        if len(rows) != quantity:
            raise HTTPException(status_code=400, detail="Указаны недоступные IMEI/серийники")
        return rows
    rows = conn.execute(
        """
        SELECT * FROM product_units
        WHERE product_id = ? AND warehouse_id = ? AND status = 'in_stock'
        ORDER BY id LIMIT ?
        """,
        (product_id, warehouse_id, quantity),
    ).fetchall()
    if len(rows) < quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Не хватает IMEI на складе: нужно {quantity}, доступно {len(rows)}",
        )
    return rows


def mark_units_sold(conn: sqlite3.Connection, units: list[sqlite3.Row], sale_id: int) -> None:
    for u in units:
        conn.execute(
            "UPDATE product_units SET status = 'sold', sale_id = ? WHERE id = ?",
            (sale_id, u["id"]),
        )




def unit_has_imei(unit: sqlite3.Row | dict[str, Any]) -> bool:
    return bool(str(unit["imei"] if isinstance(unit, dict) else unit["imei"] or "").strip())


def enrich_unit_row(row: sqlite3.Row) -> dict[str, Any]:
    d = row_to_dict(row) or {}
    d["has_imei"] = unit_has_imei(row)
    status = str(d.get("status") or "")
    cs = str(d.get("customs_status") or "none")
    if status == "reserved":
        d["display_status"] = "reserved"
    elif status == "in_stock" and not d["has_imei"]:
        d["display_status"] = "no_imei"
    elif cs == "pending":
        d["display_status"] = "pending_customs"
    else:
        d["display_status"] = "ready"
    return d


def resolve_unit_sale(
    conn: sqlite3.Connection,
    unit: sqlite3.Row,
    checkout: Any | None,
    product: sqlite3.Row,
) -> dict[str, Any]:
    serial = unit["serial"] or ""
    existing = (unit["imei"] or "").strip()
    default_customs = float(product["customs_price"] or 0)

    if existing:
        cleared = int(unit["customs_cleared"] or 0)
        cprice = float(unit["customs_price"] or 0)
        if checkout and int(checkout.customs_cleared or 0) and not cleared:
            cleared = 1
            cprice = float(checkout.customs_price or 0) or default_customs
        return {
            "imei": existing,
            "serial": serial,
            "customs_cleared": cleared,
            "customs_price": cprice,
            "imei_pending": 0,
        }

    if not checkout:
        label = serial or f"#{unit['id']}"
        raise HTTPException(
            status_code=400,
            detail=f"Устройство {label}: укажите IMEI или «активировать позже»",
        )

    if int(checkout.activate_later or 0):
        if (checkout.imei or "").strip():
            raise HTTPException(status_code=400, detail="При «активировать позже» поле IMEI должно быть пустым")
        cleared = int(checkout.customs_cleared or 0)
        cprice = float(checkout.customs_price or 0) or (default_customs if cleared else 0.0)
        return {
            "imei": "",
            "serial": serial,
            "customs_cleared": cleared,
            "customs_price": cprice,
            "imei_pending": 1,
        }

    imei = (checkout.imei or "").strip()
    if not imei:
        raise HTTPException(
            status_code=400,
            detail="Укажите IMEI или отметьте «активировать позже»",
        )
    dup = conn.execute(
        "SELECT id FROM product_units WHERE imei = ? AND id != ?",
        (imei, unit["id"]),
    ).fetchone()
    if dup:
        raise HTTPException(status_code=400, detail=f"IMEI «{imei}» уже в системе")

    cleared = int(checkout.customs_cleared or 0)
    cprice = float(checkout.customs_price or 0) or (default_customs if cleared else 0.0)
    return {
        "imei": imei,
        "serial": serial,
        "customs_cleared": cleared,
        "customs_price": cprice,
        "imei_pending": 0,
    }


def mark_unit_sold_full(
    conn: sqlite3.Connection,
    unit_id: int,
    sale_id: int,
    imei: str,
    customs_cleared: int,
    customs_price: float,
) -> None:
    customs_status = "cleared" if customs_cleared else "pending"
    conn.execute(
        """
        UPDATE product_units
        SET status = 'sold', sale_id = ?, imei = ?, customs_cleared = ?, customs_price = ?,
            customs_status = ?
        WHERE id = ?
        """,
        (sale_id, imei, customs_cleared, customs_price, customs_status, unit_id),
    )
    fulfill_unit_reservation(conn, unit_id)


def is_partnership_warehouse(conn: sqlite3.Connection, warehouse_id: int | None) -> bool:
    if not warehouse_id:
        return False
    row = conn.execute(
        f"SELECT 1 FROM warehouses WHERE id = ? AND {_partnership_warehouse_clause()} LIMIT 1",
        (int(warehouse_id),),
    ).fetchone()
    return bool(row)


def sale_return_target_warehouse(
    conn: sqlite3.Connection, sale_id: int, sale_warehouse_id: int | None
) -> tuple[int, bool]:
    """Куда вернуть товар при отмене продажи.

    Всегда на исходный склад продажи (партнёрство остаётся партнёрством и т.д.).
    """
    sale_wh = int(sale_warehouse_id or get_default_warehouse_id(conn))
    return sale_wh, False


def find_or_create_used_product_for_return(
    conn: sqlite3.Connection,
    product: sqlite3.Row,
    purchase_price: float,
) -> int:
    """Карточка Б/У для возврата партнёрского телефона (тот же name/memory/color)."""
    name = (product["name"] or "").strip()
    model = (product["model"] or name).strip()
    memory = (product["memory"] or "").strip()
    color = (product["color"] or "").strip()
    category = (product["category"] or "phone").strip() or "phone"
    existing = conn.execute(
        """
        SELECT id FROM products
        WHERE category = ?
          AND LOWER(TRIM(name)) = LOWER(TRIM(?))
          AND LOWER(TRIM(COALESCE(model, ''))) = LOWER(TRIM(?))
          AND LOWER(TRIM(COALESCE(memory, ''))) = LOWER(TRIM(?))
          AND LOWER(TRIM(COALESCE(color, ''))) = LOWER(TRIM(?))
          AND condition = 'used'
        LIMIT 1
        """,
        (category, name, model, memory, color),
    ).fetchone()
    if existing:
        pid = int(existing["id"])
        if purchase_price > 0:
            conn.execute(
                "UPDATE products SET purchase_price = ? WHERE id = ?",
                (purchase_price, pid),
            )
        return pid
    sale_price = float(product["sale_price"] or 0) or round(purchase_price * 1.15, 2)
    cur = conn.execute(
        """
        INSERT INTO products
        (name, category, ownership_type, supplier_name, brand, sku, barcode,
         purchase_price, sale_price, stock, min_stock, created_at,
         model, color, size, memory, ram, customs_cleared, customs_price, specs_extra,
         condition, track_units, image_url)
        VALUES (?, ?, ?, '', ?, '', '', ?, ?, 0, 0, ?,
                ?, ?, ?, ?, ?, 0, 0, ?, 'used', 1, '')
        """,
        (
            name,
            category,
            product["ownership_type"] or "own",
            product["brand"] or "",
            purchase_price,
            sale_price,
            utc_now(),
            model,
            color,
            product["size"] or "",
            memory,
            product["ram"] or "",
            product["specs_extra"] or "",
        ),
    )
    return int(cur.lastrowid)


def restore_units_for_sale(
    conn: sqlite3.Connection,
    sale_id: int,
    *,
    return_warehouse_id: int | None = None,
    to_bu: bool = False,
) -> list[dict[str, Any]]:
    """Вернуть единицы в наличие. При to_bu — на склад Б/У и в карточку used (партнёрский возврат)."""
    units = conn.execute(
        "SELECT * FROM product_units WHERE sale_id = ?",
        (sale_id,),
    ).fetchall()
    restored: list[dict[str, Any]] = []
    for unit in units:
        old_product_id = int(unit["product_id"])
        old_wh = int(unit["warehouse_id"])
        target_wh = int(return_warehouse_id or old_wh)
        product = conn.execute(
            "SELECT * FROM products WHERE id = ?", (old_product_id,)
        ).fetchone()
        product_id = old_product_id
        purchase = float(unit["purchase_price"] or 0) if "purchase_price" in unit.keys() else 0.0
        if purchase <= 0 and product:
            purchase = float(product["purchase_price"] or 0)
        customs = float(unit["customs_price"] or 0) if "customs_price" in unit.keys() else 0.0

        if to_bu and target_wh != old_wh:
            from_cur = get_warehouse_currency(conn, old_wh)["code"]
            to_cur = get_warehouse_currency(conn, target_wh)["code"]
            if from_cur != to_cur:
                if purchase > 0:
                    converted = _safe_convert(conn, purchase, from_cur, to_cur)
                    if converted > 0:
                        purchase = converted
                if customs > 0:
                    converted_c = _safe_convert(conn, customs, from_cur, to_cur)
                    if converted_c > 0:
                        customs = converted_c
            if product and (
                (product["condition"] or "") == "partnership"
                or is_partnership_warehouse(conn, old_wh)
            ):
                product_id = find_or_create_used_product_for_return(conn, product, purchase)

        note_extra = ""
        if to_bu:
            note_extra = "Возврат на склад продажи"
        old_notes = (unit["notes"] or "").strip() if "notes" in unit.keys() else ""
        if note_extra and note_extra not in old_notes:
            new_notes = f"{old_notes}; {note_extra}".strip("; ").strip()
        else:
            new_notes = old_notes

        conn.execute(
            """
            UPDATE product_units
            SET status = 'in_stock', sale_id = NULL,
                product_id = ?, warehouse_id = ?,
                purchase_price = CASE WHEN ? > 0 THEN ? ELSE purchase_price END,
                customs_price = CASE WHEN ? > 0 THEN ? ELSE customs_price END,
                notes = ?
            WHERE id = ?
            """,
            (
                product_id,
                target_wh,
                purchase,
                purchase,
                customs,
                customs,
                new_notes,
                int(unit["id"]),
            ),
        )
        if old_product_id != product_id:
            sync_product_stock(conn, old_product_id)
        restored.append(
            {
                "unit_id": int(unit["id"]),
                "product_id": product_id,
                "warehouse_id": target_wh,
                "old_product_id": old_product_id,
                "old_warehouse_id": old_wh,
            }
        )
    return restored


def get_default_warehouse_id(conn: sqlite3.Connection) -> int:
    row = conn.execute(
        "SELECT id FROM warehouses WHERE is_default = 1 LIMIT 1"
    ).fetchone()
    if not row:
        raise HTTPException(status_code=500, detail="Склад по умолчанию не найден")
    return int(row["id"])


def resolve_warehouse_id(conn: sqlite3.Connection, warehouse_id: int | None) -> int:
    if warehouse_id is not None:
        row = conn.execute("SELECT id FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Склад не найден")
        return warehouse_id
    return get_default_warehouse_id(conn)


def resolve_accessories_warehouse_id(conn: sqlite3.Connection) -> int:
    row = conn.execute(
        f"""
        SELECT id FROM warehouses
        WHERE warehouse_type = 'accessories' OR {_accessories_warehouse_clause()}
        ORDER BY id LIMIT 1
        """
    ).fetchone()
    if row:
        return int(row["id"])
    cur = conn.execute(
        """
        INSERT INTO warehouses (name, address, notes, is_default, warehouse_type, currency_code, created_at)
        VALUES ('Аксессуары', '', 'Склад аксессуаров', 0, 'accessories', 'TJS', ?)
        """,
        (utc_now(),),
    )
    return int(cur.lastrowid)


def _accessories_category_clause() -> str:
    return " AND EXISTS (SELECT 1 FROM products p WHERE p.id = si.product_id AND p.category = 'accessory')"


def _accessories_finance_report(
    conn: sqlite3.Connection, period: str, date_from: str, date_to: str
) -> dict[str, Any]:
    if date_from or date_to:
        since_clause, params = date_filter_sql(date_from, date_to)
        period_label = f"{date_from or '…'} — {date_to or '…'}"
    elif period == "all":
        since_clause, params = "", []
        period_label = "Всё время"
    else:
        since_clause = " AND s.created_at >= ?"
        params = [period_start(period)]
        period_label = {"day": "Сегодня", "week": "Неделя", "month": "Месяц", "quarter": "Квартал", "year": "Год"}.get(period, period)

    cat = _accessories_category_clause()
    agg = conn.execute(
        f"""
        SELECT COUNT(DISTINCT s.id) AS sales_count,
               COALESCE(SUM(si.subtotal), 0) AS gross_revenue,
               COALESCE(SUM(si.purchase_price * si.quantity), 0) AS cogs,
               COALESCE(SUM(si.shop_profit), 0) AS shop_profit,
               COALESCE(SUM(si.quantity), 0) AS items_sold
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        WHERE s.status = 'completed' {since_clause} {cat}
        """,
        params,
    ).fetchone()

    if period != "all" and not date_from and not date_to:
        exp_clause = " AND expense_date >= ?"
        exp_params = [period_start(period)[:10]]
    elif date_from or date_to:
        exp_clause, exp_params = date_filter_sql(date_from, date_to, "expense_date")
        exp_params = [p[:10] if isinstance(p, str) and len(p) > 10 else p for p in exp_params]
    else:
        exp_clause, exp_params = "", []
    expenses_total = conn.execute(
        f"SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE department = 'accessories' {exp_clause}",
        exp_params,
    ).fetchone()[0]
    exp_by_cat = conn.execute(
        f"""
        SELECT category, COALESCE(SUM(amount), 0) AS total
        FROM expenses WHERE department = 'accessories' {exp_clause}
        GROUP BY category ORDER BY total DESC
        """,
        exp_params,
    ).fetchall()

    revenue = float(agg["gross_revenue"] or 0)
    profit = float(agg["shop_profit"] or 0)
    cogs = float(agg["cogs"] or 0)
    exp = float(expenses_total)
    return {
        "period_label": period_label,
        "sales_count": int(agg["sales_count"] or 0),
        "items_sold": int(agg["items_sold"] or 0),
        "revenue": revenue,
        "cogs": cogs,
        "gross_profit": revenue - cogs,
        "shop_profit": profit,
        "expenses": exp,
        "net_profit": profit - exp,
        "expenses_by_category": [{"category": r["category"], "amount": float(r["total"])} for r in exp_by_cat],
    }


def sync_product_stock(conn: sqlite3.Connection, product_id: int) -> int:
    total = conn.execute(
        """
        SELECT COALESCE(SUM(quantity), 0)
        FROM warehouse_stock WHERE product_id = ?
        """,
        (product_id,),
    ).fetchone()[0]
    conn.execute("UPDATE products SET stock = ? WHERE id = ?", (total, product_id))
    return int(total)


def get_warehouse_stock(conn: sqlite3.Connection, warehouse_id: int, product_id: int) -> int:
    row = conn.execute(
        """
        SELECT quantity FROM warehouse_stock
        WHERE warehouse_id = ? AND product_id = ?
        """,
        (warehouse_id, product_id),
    ).fetchone()
    return int(row["quantity"]) if row else 0


def record_stock_movement(
    conn: sqlite3.Connection,
    warehouse_id: int,
    product_id: int,
    movement_type: str,
    quantity: int,
    *,
    target_warehouse_id: int | None = None,
    reference_id: int | None = None,
    notes: str = "",
) -> int:
    cur = conn.execute(
        """
        INSERT INTO stock_movements
        (warehouse_id, product_id, movement_type, quantity, target_warehouse_id, reference_id, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (warehouse_id, product_id, movement_type, quantity, target_warehouse_id, reference_id, notes, utc_now()),
    )
    return int(cur.lastrowid)


def adjust_warehouse_stock(
    conn: sqlite3.Connection,
    warehouse_id: int,
    product_id: int,
    delta: int,
    movement_type: str,
    *,
    target_warehouse_id: int | None = None,
    reference_id: int | None = None,
    notes: str = "",
) -> tuple[int, int | None]:
    if delta == 0:
        return get_warehouse_stock(conn, warehouse_id, product_id), None
    current = get_warehouse_stock(conn, warehouse_id, product_id)
    new_qty = current + delta
    if new_qty < 0:
        product = conn.execute("SELECT name FROM products WHERE id = ?", (product_id,)).fetchone()
        name = product["name"] if product else f"#{product_id}"
        raise HTTPException(
            status_code=400,
            detail=f"Недостаточно «{name}» на складе: доступно {current}, нужно {abs(delta)}",
        )
    has_row = conn.execute(
        "SELECT 1 FROM warehouse_stock WHERE warehouse_id = ? AND product_id = ?",
        (warehouse_id, product_id),
    ).fetchone()
    if has_row:
        conn.execute(
            """
            UPDATE warehouse_stock SET quantity = ?
            WHERE warehouse_id = ? AND product_id = ?
            """,
            (new_qty, warehouse_id, product_id),
        )
    elif delta > 0:
        conn.execute(
            """
            INSERT INTO warehouse_stock (warehouse_id, product_id, quantity)
            VALUES (?, ?, ?)
            """,
            (warehouse_id, product_id, delta),
        )
    movement_id = record_stock_movement(
        conn, warehouse_id, product_id, movement_type, abs(delta),
        target_warehouse_id=target_warehouse_id,
        reference_id=reference_id,
        notes=notes,
    )
    sync_product_stock(conn, product_id)
    return new_qty, movement_id


def enrich_product(conn: sqlite3.Connection, product: sqlite3.Row) -> dict[str, Any]:
    data = row_to_dict(product) or {}
    rows = conn.execute(
        """
        SELECT ws.warehouse_id, ws.quantity, w.name AS warehouse_name
        FROM warehouse_stock ws
        JOIN warehouses w ON w.id = ws.warehouse_id
        WHERE ws.product_id = ?
        ORDER BY w.is_default DESC, w.name
        """,
        (product["id"],),
    ).fetchall()
    stock_by_warehouse = {str(r["warehouse_id"]): r["quantity"] for r in rows}
    data["stock_by_warehouse"] = stock_by_warehouse
    data["stock"] = int(sum(stock_by_warehouse.values()))
    if int(product["track_units"] or 0):
        unit_rows = conn.execute(
            """
            SELECT warehouse_id, COUNT(*) AS cnt FROM product_units
            WHERE product_id = ? AND status = 'in_stock'
            GROUP BY warehouse_id
            """,
            (product["id"],),
        ).fetchall()
        units_by_warehouse = {str(r["warehouse_id"]): int(r["cnt"]) for r in unit_rows}
        data["units_by_warehouse"] = units_by_warehouse
        data["units_available"] = int(sum(units_by_warehouse.values()))
    return data


def enrich_sale_items(conn: sqlite3.Connection, items: list[sqlite3.Row]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for item in items:
        d = row_to_dict(item) or {}
        units = conn.execute(
            "SELECT * FROM sale_item_units WHERE sale_item_id = ?", (item["id"],)
        ).fetchall()
        d["units"] = [row_to_dict(u) for u in units]
        result.append(d)
    return result


def period_start(period: str) -> str:
    now = datetime.now()
    if period == "week":
        start = now - timedelta(days=now.weekday())
    elif period == "month":
        start = now.replace(day=1)
    elif period == "quarter":
        q_month = ((now.month - 1) // 3) * 3 + 1
        start = now.replace(month=q_month, day=1)
    elif period == "year":
        start = now.replace(month=1, day=1)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    return start.strftime("%Y-%m-%d %H:%M:%S")


def date_filter_sql(date_from: str, date_to: str, column: str = "s.created_at") -> tuple[str, list[Any]]:
    clause = ""
    params: list[Any] = []
    if date_from:
        clause += f" AND {column} >= ?"
        params.append(date_from)
    if date_to:
        clause += f" AND {column} <= ?"
        params.append(date_to + " 23:59:59")
    return clause, params


def ownership_filter_sql(scope: str, column: str = "si.ownership_type") -> tuple[str, list[Any]]:
    if scope == "own":
        return f" AND {column} = 'own'", []
    if scope == "consignment":
        return f" AND {column} = 'consignment'", []
    return "", []


def unit_purchase_price(
    unit: sqlite3.Row | dict[str, Any] | None,
    product: sqlite3.Row | dict[str, Any] | None = None,
) -> float:
    """Cost of one device: unit.purchase_price if set, else product card."""
    if unit is not None:
        try:
            up = float(unit["purchase_price"] or 0)
        except (KeyError, IndexError, TypeError):
            up = 0.0
        if up > 0:
            return up
    if product is not None:
        try:
            return float(product["purchase_price"] or 0)
        except (KeyError, IndexError, TypeError):
            return 0.0
    return 0.0


def unit_extra_cost(unit: sqlite3.Row | dict[str, Any] | None) -> float:
    """Per-IMEI extra (repair / customs column «Расходы»)."""
    if unit is None:
        return 0.0
    try:
        return max(0.0, float(unit["customs_price"] or 0))
    except (KeyError, IndexError, TypeError):
        return 0.0


def unit_total_cost(
    unit: sqlite3.Row | dict[str, Any] | None,
    product: sqlite3.Row | dict[str, Any] | None = None,
) -> float:
    """Landed cost for profit: себестоимость + расходы на IMEI."""
    return unit_purchase_price(unit, product) + unit_extra_cost(unit)


def calc_line(
    product: sqlite3.Row,
    qty: int,
    unit_price: float | None = None,
    unit_costs: list[float] | None = None,
) -> dict[str, float | str]:
    price = unit_price if unit_price is not None else float(product["sale_price"])
    subtotal = price * qty
    ownership = product["ownership_type"] or "own"
    if unit_costs is not None and len(unit_costs) == qty and qty > 0:
        total_cost = float(sum(unit_costs))
        cost = total_cost / qty
    else:
        cost = float(product["purchase_price"] or 0)
        total_cost = cost * qty
    if ownership == "consignment":
        supplier_due = total_cost
        shop_profit = subtotal - supplier_due
    else:
        supplier_due = 0.0
        shop_profit = subtotal - total_cost
    return {
        "ownership_type": ownership,
        "supplier_name": product["supplier_name"] or "",
        "unit_price": price,
        "purchase_price": cost,
        "subtotal": subtotal,
        "supplier_due": supplier_due,
        "shop_profit": shop_profit,
    }


def inventory_stock_value_by_currency(
    conn: sqlite3.Connection, scope: str = "all", warehouse_id: int | None = None
) -> list[dict[str, Any]]:
    """Stock value: per-IMEI cost for tracked phones + product cost × qty for accessories."""
    scope_sql = ""
    params: list[Any] = []
    if scope != "all":
        scope_sql = " AND p.ownership_type = ?"
        params.append(scope)
    wh_sql = ""
    if warehouse_id:
        wh_sql = " AND w.id = ?"
        params.append(warehouse_id)
    totals: dict[str, float] = {}
    for row in conn.execute(
        f"""
        SELECT COALESCE(w.currency_code, 'TJS') AS currency_code,
               COALESCE(SUM(
                   (CASE WHEN COALESCE(u.purchase_price, 0) > 0 THEN u.purchase_price
                         ELSE p.purchase_price END)
                   + COALESCE(u.customs_price, 0)
               ), 0) AS val
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        JOIN warehouses w ON w.id = u.warehouse_id
        WHERE u.status = 'in_stock'{scope_sql}{wh_sql}
        GROUP BY COALESCE(w.currency_code, 'TJS')
        """,
        params,
    ).fetchall():
        code = (row["currency_code"] or "TJS").strip().upper() or "TJS"
        totals[code] = totals.get(code, 0.0) + float(row["val"])
    qty_params = list(params)
    for row in conn.execute(
        f"""
        SELECT COALESCE(w.currency_code, 'TJS') AS currency_code,
               COALESCE(SUM(p.purchase_price * ws.quantity), 0) AS val
        FROM warehouse_stock ws
        JOIN products p ON p.id = ws.product_id
        JOIN warehouses w ON w.id = ws.warehouse_id
        WHERE ws.quantity > 0 AND IFNULL(p.track_units, 0) = 0{scope_sql}{wh_sql}
        GROUP BY COALESCE(w.currency_code, 'TJS')
        """,
        qty_params,
    ).fetchall():
        code = (row["currency_code"] or "TJS").strip().upper() or "TJS"
        totals[code] = totals.get(code, 0.0) + float(row["val"])
    return [{**currency_meta(code), "value": val} for code, val in sorted(totals.items())]


def _safe_convert(
    conn: sqlite3.Connection, amount: float, from_code: str, to_code: str
) -> float:
    try:
        return round(convert_amount(conn, amount, from_code, to_code), 2)
    except HTTPException:
        return 0.0


def warehouse_stock_money(conn: sqlite3.Connection, warehouse_id: int) -> dict[str, Any]:
    """Stock cost on one warehouse: native currency + USD + TJS equivalents."""
    wh = conn.execute(
        "SELECT id, name, currency_code FROM warehouses WHERE id = ?", (warehouse_id,)
    ).fetchone()
    if not wh:
        raise HTTPException(status_code=404, detail="Склад не найден")
    native_code = (wh["currency_code"] or "TJS").strip().upper() or "TJS"
    by_cur = inventory_stock_value_by_currency(conn, "all", warehouse_id)
    native_val = round(sum(float(x["value"]) for x in by_cur), 2)
    units_count = int(conn.execute(
        "SELECT COUNT(*) FROM product_units WHERE warehouse_id = ? AND status = 'in_stock'",
        (warehouse_id,),
    ).fetchone()[0])
    qty_items = int(conn.execute(
        """
        SELECT COALESCE(SUM(ws.quantity), 0)
        FROM warehouse_stock ws
        JOIN products p ON p.id = ws.product_id
        WHERE ws.warehouse_id = ? AND ws.quantity > 0 AND IFNULL(p.track_units, 0) = 0
        """,
        (warehouse_id,),
    ).fetchone()[0])
    return {
        "warehouse_id": warehouse_id,
        "warehouse_name": wh["name"],
        "currency_code": native_code,
        "currency": currency_meta(native_code),
        "value": native_val,
        "value_usd": _safe_convert(conn, native_val, native_code, "USD"),
        "value_tjs": _safe_convert(conn, native_val, native_code, "TJS"),
        "units_count": units_count,
        "qty_items": qty_items,
        "by_currency": by_cur,
    }


def _pay_bucket(method_type: str | None) -> str:
    return "cash" if (method_type or "") == "cash" else "wallet"


def _money_sides_by_currency(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Bag signed cash lines into per-currency plus/minus/net. Never mix USD+TJS."""
    bags: dict[str, dict[str, float]] = {}
    for line in lines:
        code = str(line.get("currency_code") or "TJS").strip().upper() or "TJS"
        bag = bags.setdefault(code, {"plus": 0.0, "minus": 0.0})
        amt = float(line.get("amount") or 0)
        if line.get("side") == "+":
            bag["plus"] += amt
        else:
            bag["minus"] += amt
    out: list[dict[str, Any]] = []
    for code in sorted(bags.keys(), key=lambda c: (0 if c == "USD" else 1, c)):
        plus = round(bags[code]["plus"], 2)
        minus = round(bags[code]["minus"], 2)
        out.append({
            **currency_meta(code),
            "plus": plus,
            "minus": minus,
            "net": round(plus - minus, 2),
            "inflow": plus,
            "outflow": minus,
        })
    return out


def _period_cash_by_currency(
    conn: sqlite3.Connection,
    period: str,
    *,
    method_code: str = "",
    date_from: str = "",
    date_to: str = "",
) -> dict[str, Any]:
    """Period cash in/out bags by currency (+ optional per payment method bags)."""
    sale_clause, sale_params, label = _report_period_clause(period, date_from, date_to, "s.created_at")
    method = (method_code or "").strip()
    bags: dict[str, dict[str, float]] = {}
    method_bags: dict[str, dict[str, dict[str, float]]] = {}

    def add(cur: str, side: str, amt: float, mcode: str = "") -> None:
        code = (cur or "TJS").strip().upper() or "TJS"
        bag = bags.setdefault(code, {"inflow": 0.0, "outflow": 0.0})
        val = float(amt or 0)
        if side == "+":
            bag["inflow"] += val
        else:
            bag["outflow"] += val
        if mcode:
            mb = method_bags.setdefault(mcode, {})
            mb2 = mb.setdefault(code, {"inflow": 0.0, "outflow": 0.0})
            if side == "+":
                mb2["inflow"] += val
            else:
                mb2["outflow"] += val

    def wallet_code(raw_method: str, money_cur: str) -> str:
        resolved = resolve_wallet_method_code(conn, raw_method or "cash", money_cur)
        if method and resolved != method:
            return ""
        return resolved

    pay_sql = f"""
        SELECT UPPER(COALESCE(
                   NULLIF(TRIM(sp.pay_currency_code), ''),
                   NULLIF(TRIM(s.currency_code), ''),
                   'TJS'
               )) AS cur,
               sp.method_code,
               SUM(COALESCE(NULLIF(sp.pay_amount, 0), sp.amount)) AS amt
        FROM sale_payments sp
        JOIN sales s ON s.id = sp.sale_id
        WHERE s.status = 'completed' AND COALESCE(s.affects_cash, 1) = 1 {sale_clause}
        GROUP BY 1, 2
    """
    for r in conn.execute(pay_sql, list(sale_params)).fetchall():
        cur = r["cur"] or "TJS"
        wcode = wallet_code(r["method_code"] or "cash", cur)
        if not wcode:
            continue
        add(cur, "+", r["amt"], wcode)

    recv_clause, recv_params, _ = _report_period_clause(period, date_from, date_to, "rp.created_at")
    recv_sql = f"""
        SELECT UPPER(COALESCE(
                   NULLIF(TRIM(rp.pay_currency_code), ''),
                   NULLIF(TRIM(s.currency_code), ''),
                   'TJS'
               )) AS cur,
               rp.payment_method_code AS method_code,
               SUM(COALESCE(NULLIF(rp.pay_amount, 0), rp.amount)) AS amt
        FROM receivable_payments rp
        JOIN receivables r ON r.id = rp.receivable_id
        LEFT JOIN sales s ON s.id = r.sale_id
        WHERE COALESCE(s.affects_cash, 1) = 1 {recv_clause}
        GROUP BY 1, 2
    """
    if _table_exists(conn, "receivable_payments"):
        for r in conn.execute(recv_sql, list(recv_params)).fetchall():
            cur = r["cur"] or "TJS"
            wcode = wallet_code(r["method_code"] or "cash", cur)
            if not wcode:
                continue
            add(cur, "+", r["amt"], wcode)

    if _table_exists(conn, "mutual_payments"):
        m_clause, m_params, _ = _report_period_clause(period, date_from, date_to, "mp.created_at")
        m_sql = f"""
            SELECT me.direction AS direction,
                   UPPER(COALESCE(
                       NULLIF(TRIM(mp.pay_currency_code), ''),
                       NULLIF(TRIM(me.currency_code), ''),
                       'TJS'
                   )) AS cur,
                   mp.payment_method_code AS method_code,
                   SUM(COALESCE(NULLIF(mp.pay_amount, 0), mp.amount)) AS amt
            FROM mutual_payments mp
            JOIN mutual_entries me ON me.id = mp.entry_id
            WHERE me.direction IN ('owe_us', 'we_owe') {m_clause}
            GROUP BY 1, 2, 3
        """
        for r in conn.execute(m_sql, list(m_params)).fetchall():
            cur = r["cur"] or "TJS"
            wcode = wallet_code(r["method_code"] or "cash", cur)
            if not wcode:
                continue
            # owe_us: должник принёс деньги → приход; we_owe: мы отдали → расход
            add(cur, "+" if r["direction"] == "owe_us" else "−", r["amt"], wcode)

    if _table_exists(conn, "cash_inflows"):
        c_clause, c_params, _ = _report_period_clause(period, date_from, date_to, "created_at")
        c_sql = f"""
            SELECT UPPER(COALESCE(NULLIF(TRIM(currency_code), ''), 'TJS')) AS cur,
                   payment_method_code AS method_code,
                   SUM(amount) AS amt
            FROM cash_inflows
            WHERE source_type IN ('counterparty', 'shift_close', 'shift_opening') {c_clause}
            GROUP BY 1, 2
        """
        for r in conn.execute(c_sql, list(c_params)).fetchall():
            cur = r["cur"] or "TJS"
            wcode = wallet_code(r["method_code"] or "cash", cur)
            if not wcode:
                continue
            add(cur, "+", r["amt"], wcode)

    # expense_date хранится как YYYY-MM-DD — нельзя сравнивать с datetime opened_at,
    # иначе все расходы дня отсекаются (минусы «не считаются»).
    exp_from = (date_from or "")[:10]
    exp_to = (date_to or "")[:10]
    exp_clause, exp_params, _ = _report_period_clause(period, exp_from, exp_to, "expense_date")
    if period != "all" and not exp_from and not exp_to and not exp_params:
        exp_clause = " AND expense_date >= ?"
        exp_params = [period_start(period)[:10]]
    exp_sql = f"""
        SELECT payment_method_code AS method_code, SUM(amount) AS amt
        FROM expenses WHERE COALESCE(affects_cash, 1) = 1 {exp_clause}
        GROUP BY 1
    """
    for r in conn.execute(exp_sql, list(exp_params)).fetchall():
        wcode = wallet_code(r["method_code"] or "cash", "TJS")
        if not wcode:
            continue
        add("TJS", "−", r["amt"], wcode)

    if _table_exists(conn, "supplier_payments"):
        s_clause, s_params, _ = _report_period_clause(period, date_from, date_to, "created_at")
        s_sql = f"""
            SELECT COALESCE(payment_method_code, 'cash') AS method_code, SUM(amount) AS amt
            FROM supplier_payments WHERE 1=1 {s_clause}
            GROUP BY 1
        """
        for r in conn.execute(s_sql, list(s_params)).fetchall():
            wcode = wallet_code(r["method_code"] or "cash", "TJS")
            if not wcode:
                continue
            add("TJS", "−", r["amt"], wcode)

    by_currency = []
    for code in sorted(bags.keys(), key=lambda c: (0 if c == "USD" else 1, c)):
        inn = round(bags[code]["inflow"], 2)
        out = round(bags[code]["outflow"], 2)
        by_currency.append({
            **currency_meta(code),
            "inflow": inn,
            "outflow": out,
            "net": round(inn - out, 2),
            "plus": inn,
            "minus": out,
        })

    balances_by_method: list[dict[str, Any]] = []
    methods = list_payment_methods(conn, active_only=True)
    for pm in methods:
        code = pm["code"]
        cur_map = method_bags.get(code) or {}
        if method and code != method:
            continue
        pm_cur = payment_method_currency(pm)
        mbags = []
        for ccode in sorted(cur_map.keys(), key=lambda c: (0 if c == "USD" else 1, c)):
            # Bound wallet: ignore foreign-currency leakage
            if pm_cur and ccode != pm_cur:
                continue
            inn = round(cur_map[ccode]["inflow"], 2)
            out = round(cur_map[ccode]["outflow"], 2)
            mbags.append({
                **currency_meta(ccode),
                "inflow": inn,
                "outflow": out,
                "net": round(inn - out, 2),
            })
        if not mbags:
            show = pm_cur or "TJS"
            mbags = [{**currency_meta(show), "inflow": 0.0, "outflow": 0.0, "net": 0.0}]
            if not pm_cur:
                mbags.insert(0, {**currency_meta("USD"), "inflow": 0.0, "outflow": 0.0, "net": 0.0})
        balances_by_method.append({
            "code": code,
            "name": pm["name"],
            "method_type": pm["method_type"],
            "currency_code": pm_cur,
            "by_currency": mbags,
            "inflow": None,
            "outflow": None,
            "net": None,
        })

    return {
        "period_label": label,
        "by_currency": by_currency,
        "balances": balances_by_method,
    }


def till_summary_from_balances(balances: list[dict[str, Any]]) -> dict[str, Any]:
    """Сводка кассы: нал смн / нал $ / карта (безнал) — то, что реально в ящике."""
    cash_tjs = 0.0
    cash_usd = 0.0
    card_by_cur: dict[str, float] = {}
    for b in balances or []:
        mtype = (b.get("method_type") or "").strip().lower()
        bound = (b.get("currency_code") or "").strip().upper()
        for row in b.get("by_currency") or []:
            code = str(row.get("code") or "TJS").upper()
            net = float(row.get("net") or 0)
            if mtype == "cash":
                use = bound or code
                if use == "USD":
                    cash_usd += net
                else:
                    cash_tjs += net
            else:
                card_by_cur[code] = card_by_cur.get(code, 0.0) + net
    card_bags = [
        {**currency_meta(c), "amount": round(v, 2), "net": round(v, 2)}
        for c, v in sorted(card_by_cur.items(), key=lambda x: (0 if x[0] == "USD" else 1, x[0]))
        if abs(v) > 0.0001
    ]
    if not card_bags:
        card_bags = [{**currency_meta("TJS"), "amount": 0.0, "net": 0.0}]
    return {
        "cash_tjs": round(cash_tjs, 2),
        "cash_usd": round(cash_usd, 2),
        "card": card_bags,
        "cash_tjs_meta": currency_meta("TJS"),
        "cash_usd_meta": currency_meta("USD"),
    }


def funds_snapshot(conn: sqlite3.Connection) -> dict[str, Any]:
    """Current money: stock per warehouse + cash/wallet by USD/TJS."""
    rates = latest_exchange_rates(conn)
    warehouses = []
    stock_usd = 0.0
    stock_tjs = 0.0
    for w in conn.execute(
        "SELECT id FROM warehouses ORDER BY is_default DESC, name"
    ).fetchall():
        row = warehouse_stock_money(conn, w["id"])
        warehouses.append(row)
        stock_usd += float(row["value_usd"] or 0)
        stock_tjs += float(row["value_tjs"] or 0)

    bals: dict[str, dict[str, float]] = {}

    def add(cur: str, kind: str, amt: float) -> None:
        code = (cur or "TJS").strip().upper() or "TJS"
        bals.setdefault(code, {"cash": 0.0, "wallet": 0.0})
        bals[code][kind] = bals[code].get(kind, 0.0) + float(amt or 0)

    for r in conn.execute(
        """
        SELECT UPPER(COALESCE(
                   NULLIF(TRIM(sp.pay_currency_code), ''),
                   NULLIF(TRIM(s.currency_code), ''),
                   'TJS'
               )) AS cur,
               COALESCE(pm.method_type, 'card') AS method_type,
               SUM(COALESCE(NULLIF(sp.pay_amount, 0), sp.amount)) AS amt
        FROM sale_payments sp
        JOIN sales s ON s.id = sp.sale_id
        LEFT JOIN payment_methods pm ON pm.code = sp.method_code
        WHERE s.status = 'completed' AND COALESCE(s.affects_cash, 1) = 1
        GROUP BY 1, 2
        """
    ).fetchall():
        add(r["cur"], _pay_bucket(r["method_type"]), r["amt"])

    if _table_exists(conn, "receivable_payments"):
        for r in conn.execute(
            """
            SELECT UPPER(COALESCE(
                       NULLIF(TRIM(rp.pay_currency_code), ''),
                       NULLIF(TRIM(s.currency_code), ''),
                       'TJS'
                   )) AS cur,
                   COALESCE(pm.method_type, 'cash') AS method_type,
                   SUM(COALESCE(NULLIF(rp.pay_amount, 0), rp.amount)) AS amt
            FROM receivable_payments rp
            JOIN receivables r ON r.id = rp.receivable_id
            JOIN sales s ON s.id = r.sale_id
            LEFT JOIN payment_methods pm ON pm.code = rp.payment_method_code
            WHERE s.status = 'completed' AND COALESCE(s.affects_cash, 1) = 1
            GROUP BY 1, 2
            """
        ).fetchall():
            add(r["cur"], _pay_bucket(r["method_type"]), r["amt"])

    if _table_exists(conn, "cash_inflows"):
        for r in conn.execute(
            """
            SELECT UPPER(COALESCE(NULLIF(TRIM(currency_code), ''), 'TJS')) AS cur,
                   payment_method_code,
                   SUM(amount) AS amt
            FROM cash_inflows
            WHERE source_type IN ('counterparty', 'shift_close', 'shift_opening')
            GROUP BY 1, 2
            """
        ).fetchall():
            pm = get_payment_method(conn, r["payment_method_code"] or "cash")
            mtype = pm["method_type"] if pm else "cash"
            add(r["cur"], _pay_bucket(mtype), r["amt"])

    if _table_exists(conn, "mutual_payments"):
        for r in conn.execute(
            """
            SELECT me.direction AS direction,
                   UPPER(COALESCE(
                       NULLIF(TRIM(mp.pay_currency_code), ''),
                       NULLIF(TRIM(me.currency_code), ''),
                       'TJS'
                   )) AS cur,
                   mp.payment_method_code AS method_code,
                   SUM(COALESCE(NULLIF(mp.pay_amount, 0), mp.amount)) AS amt
            FROM mutual_payments mp
            JOIN mutual_entries me ON me.id = mp.entry_id
            WHERE me.direction IN ('owe_us', 'we_owe')
            GROUP BY 1, 2, 3
            """
        ).fetchall():
            pm = get_payment_method(conn, r["method_code"] or "cash")
            mtype = pm["method_type"] if pm else "cash"
            signed = float(r["amt"] or 0) if r["direction"] == "owe_us" else -float(r["amt"] or 0)
            add(r["cur"], _pay_bucket(mtype), signed)

    for r in conn.execute(
        """
        SELECT payment_method_code, SUM(amount) AS amt
        FROM expenses
        WHERE COALESCE(affects_cash, 1) = 1
        GROUP BY payment_method_code
        """
    ).fetchall():
        pm = get_payment_method(conn, r["payment_method_code"] or "cash")
        mtype = pm["method_type"] if pm else "cash"
        add("TJS", _pay_bucket(mtype), -float(r["amt"] or 0))

    if _table_exists(conn, "supplier_payments"):
        for r in conn.execute(
            """
            SELECT COALESCE(payment_method_code, 'cash') AS payment_method_code,
                   SUM(amount) AS amt
            FROM supplier_payments
            GROUP BY 1
            """
        ).fetchall():
            pm = get_payment_method(conn, r["payment_method_code"] or "cash")
            mtype = pm["method_type"] if pm else "cash"
            add("TJS", _pay_bucket(mtype), -float(r["amt"] or 0))

    cash_by_currency = []
    for code in sorted(bals.keys()):
        cash = round(bals[code].get("cash", 0.0), 2)
        wallet = round(bals[code].get("wallet", 0.0), 2)
        cash_by_currency.append({
            **currency_meta(code),
            "cash": cash,
            "wallet": wallet,
            "total": round(cash + wallet, 2),
        })

    def cur_amount(code: str, field: str) -> float:
        row = next((x for x in cash_by_currency if x["code"] == code), None)
        return float(row[field]) if row else 0.0

    return {
        "as_of": utc_now(),
        "exchange_rates": rates,
        "warehouses": warehouses,
        "cash_by_currency": cash_by_currency,
        "totals": {
            "stock_usd": round(stock_usd, 2),
            "stock_tjs": round(stock_tjs, 2),
            "cash_usd": cur_amount("USD", "cash"),
            "cash_tjs": cur_amount("TJS", "cash"),
            "wallet_usd": cur_amount("USD", "wallet"),
            "wallet_tjs": cur_amount("TJS", "wallet"),
            "money_usd": cur_amount("USD", "total"),
            "money_tjs": cur_amount("TJS", "total"),
        },
    }


def warehouse_sales_filter(warehouse_id: int | None, alias: str = "s") -> tuple[str, list[Any]]:
    if warehouse_id:
        return f" AND {alias}.warehouse_id = ?", [int(warehouse_id)]
    return "", []


class ProductIn(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    category: str = "accessory"
    ownership_type: OwnershipType = "own"
    supplier_name: str = ""
    brand: str = ""
    sku: str = ""
    barcode: str = ""
    purchase_price: float = Field(ge=0)
    sale_price: float = Field(ge=0)
    stock: int = Field(ge=0)
    min_stock: int = Field(ge=0, default=2)
    model: str = ""
    color: str = ""
    size: str = ""
    memory: str = ""
    ram: str = ""
    customs_cleared: int = Field(ge=0, le=1, default=0)
    customs_price: float = Field(ge=0, default=0)
    specs_extra: str = ""
    condition: ProductCondition = "new"
    warehouse_id: int | None = None
    track_units: int | None = None
    image_url: str = ""


class ProductUpdate(BaseModel):
    name: str | None = None
    category: str | None = None
    ownership_type: OwnershipType | None = None
    supplier_name: str | None = None
    brand: str | None = None
    sku: str | None = None
    barcode: str | None = None
    purchase_price: float | None = Field(default=None, ge=0)
    sale_price: float | None = Field(default=None, ge=0)
    stock: int | None = Field(default=None, ge=0)
    min_stock: int | None = Field(default=None, ge=0)
    model: str | None = None
    color: str | None = None
    size: str | None = None
    memory: str | None = None
    ram: str | None = None
    customs_cleared: int | None = Field(default=None, ge=0, le=1)
    customs_price: float | None = Field(default=None, ge=0)
    specs_extra: str | None = None
    condition: ProductCondition | None = None
    warehouse_id: int | None = None
    track_units: int | None = Field(default=None, ge=0, le=1)
    image_url: str | None = None


class UnitCheckoutIn(BaseModel):
    unit_id: int
    imei: str = ""
    activate_later: int = Field(default=0, ge=0, le=1)
    customs_cleared: int = Field(default=0, ge=0, le=1)
    customs_price: float = Field(default=0, ge=0)


class CartItem(BaseModel):
    product_id: int
    quantity: int = Field(ge=1)
    unit_ids: list[int] = Field(default_factory=list)
    units: list[UnitCheckoutIn] = Field(default_factory=list)
    unit_price: float | None = Field(default=None, gt=0)


class PaymentPart(BaseModel):
    method_code: str = Field(min_length=1)
    amount: float = Field(ge=0)
    currency_code: str = ""


class SaleIn(BaseModel):
    items: list[CartItem]
    discount: float = Field(ge=0, default=0)
    payment_method: str = "cash"
    payments: list[PaymentPart] = Field(default_factory=list)
    notes: str = ""
    warehouse_id: int | None = None
    shift_id: int | None = None
    debtor_name: str = ""
    debtor_phone: str = ""
    pay_currency: str = ""  # alternate currency for FX payments
    fx_rate: float | None = Field(default=None, gt=0)  # 1 sale_currency = fx_rate pay_currency


class CurrencySettingsIn(BaseModel):
    base_currency: str = Field(min_length=3, max_length=3)
    currency_symbol: str = Field(min_length=1, max_length=8)
    currency_name: str = ""


class ExchangeRateIn(BaseModel):
    currency_code: str = Field(min_length=3, max_length=3)
    rate: float = Field(gt=0)
    effective_at: str = ""
    notes: str = ""


def normalize_exchange_effective_at(raw: str | None) -> str:
    """datetime-local / ручной ввод → 'YYYY-MM-DD HH:MM:SS'. Пустое/битое → сейчас."""
    s = (raw or "").strip().replace("T", " ")
    if not s or s in (":00", "00:00", "00:00:00"):
        return utc_now()
    # already full
    if len(s) >= 19 and s[4] == "-" and s[7] == "-":
        return s[:19]
    # YYYY-MM-DD HH:MM
    if len(s) >= 16 and s[4] == "-" and s[7] == "-":
        return s[:16] + ":00"
    # YYYY-MM-DD
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10] + " 00:00:00"
    raise HTTPException(
        status_code=400,
        detail="Некорректная дата курса. Укажите «Действует с» или оставьте пустым (сейчас).",
    )


class PaymentMethodIn(BaseModel):
    code: str = Field(min_length=1, max_length=32)
    name: str = Field(min_length=1, max_length=100)
    method_type: str = "card"
    currency_code: str = Field(default="", pattern="^(|TJS|USD)$")
    sort_order: int = 0


class PaymentMethodUpdate(BaseModel):
    name: str | None = None
    method_type: str | None = None
    currency_code: str | None = Field(default=None, pattern="^(|TJS|USD)$")
    is_active: int | None = Field(default=None, ge=0, le=1)
    sort_order: int | None = None


class ExpenseIn(BaseModel):
    category: str = Field(default="", max_length=100)
    amount: float = Field(gt=0)
    description: str = ""
    payment_method_code: str = "cash"
    expense_date: str = ""
    department: str = "main"
    warehouse_id: int | None = None
    payee: str = ""
    kind: Literal["expense", "payout"] = "expense"


class CashInflowIn(BaseModel):
    amount: float = Field(gt=0)
    currency_code: str = "TJS"
    payment_method_code: str = "cash"
    source_type: Literal["counterparty", "debtor"] = "counterparty"
    counterparty_name: str = ""
    receivable_id: int | None = None
    mutual_entry_id: int | None = None
    fx_rate: float | None = Field(default=None, gt=0)  # 1 USD = fx_rate TJS
    notes: str = ""


class AccessoryInboundIn(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    model: str = ""
    quantity: int = Field(ge=1, le=10000)
    supplier_name: str = ""
    purchase_price: float = Field(gt=0)
    sale_price: float | None = Field(default=None, gt=0)


class ExpenseAllocationRule(BaseModel):
    warehouse_id: int
    pct: float = Field(ge=0, le=100)


class ExpenseAllocationIn(BaseModel):
    rules: list[ExpenseAllocationRule]


class ReceivablePaymentIn(BaseModel):
    amount: float = Field(gt=0)  # сумма в валюте оплаты
    payment_method_code: str = "cash"
    notes: str = ""
    currency_code: str = ""  # валюта оплаты; пусто = валюта долга
    fx_rate: float | None = Field(default=None, gt=0)  # 1 USD = fx_rate TJS


class MutualEntryIn(BaseModel):
    person_name: str = Field(min_length=1, max_length=200)
    person_phone: str = ""
    direction: Literal["owe_us", "we_owe"] = "owe_us"
    amount: float = Field(gt=0)
    currency_code: str = "TJS"
    payment_method_code: str = ""
    product_note: str = ""
    notes: str = ""


class MutualEntryUpdate(BaseModel):
    person_name: str | None = Field(default=None, min_length=1, max_length=200)
    person_phone: str | None = None
    direction: Literal["owe_us", "we_owe"] | None = None
    amount: float | None = Field(default=None, gt=0)
    currency_code: str | None = None
    product_note: str | None = None
    notes: str | None = None


class MutualPaymentIn(BaseModel):
    amount: float = Field(gt=0)  # сумма в валюте оплаты
    payment_method_code: str = "cash"
    notes: str = ""
    currency_code: str = ""
    fx_rate: float | None = Field(default=None, gt=0)  # 1 USD = fx_rate TJS


class UnitIn(BaseModel):
    product_id: int
    warehouse_id: int
    imei: str = ""
    serial: str = ""
    notes: str = ""
    customs_status: Literal["none", "pending", "cleared"] = "none"
    purchase_price: float | None = Field(default=None, ge=0)


class BulkUnitsIn(BaseModel):
    product_id: int
    warehouse_id: int
    quantity: int = Field(ge=1, le=500)
    serial_prefix: str = Field(default="", max_length=40)
    notes: str = ""
    mark_pending_customs: int = Field(default=0, ge=0, le=1)
    purchase_price: float | None = Field(default=None, ge=0)


class CompleteImeiIn(BaseModel):
    imei: str = Field(min_length=5, max_length=20)


class CustomsStatusIn(BaseModel):
    customs_status: Literal["none", "pending", "cleared"]


class ReservationIn(BaseModel):
    unit_id: int
    client_name: str = Field(min_length=1, max_length=120)
    client_phone: str = ""
    notes: str = ""
    reserved_until: str = Field(min_length=10, max_length=30)


class UserIn(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    pin: str = Field(min_length=4, max_length=12)
    role: UserRole = "cashier"


class UserUpdate(BaseModel):
    name: str | None = None
    pin: str | None = Field(default=None, min_length=4, max_length=12)
    role: UserRole | None = None
    is_active: int | None = Field(default=None, ge=0, le=1)


class ShiftOpeningWallet(BaseModel):
    method_code: str = Field(min_length=1)
    currency_code: str = Field(default="TJS", min_length=3, max_length=3)
    amount: float = Field(ge=0, default=0)


class ShiftOpenIn(BaseModel):
    opening_cash: float = Field(ge=0, default=0)
    opening_wallets: list[ShiftOpeningWallet] = Field(default_factory=list)


class ShiftPaymentActual(BaseModel):
    method_code: str
    amount: float = Field(ge=0)
    currency_code: str = Field(default="TJS", min_length=3, max_length=3)


class ShiftCloseIn(BaseModel):
    actual_cash: float = Field(ge=0, default=0)
    actual_card: float = Field(ge=0, default=0)
    actual_payments: list[ShiftPaymentActual] = Field(default_factory=list)
    actual_wallets: list[ShiftOpeningWallet] = Field(default_factory=list)
    notes: str = ""


class SupplierPaymentIn(BaseModel):
    supplier_name: str = Field(min_length=1)
    amount: float = Field(gt=0)
    notes: str = ""
    payment_method_code: str = "cash"


class WarehouseIn(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    address: str = ""
    notes: str = ""
    is_default: bool = False
    warehouse_type: str = Field(default="new", pattern="^(new|used|partnership|accessories)$")
    currency_code: str = Field(default="", pattern="^(|TJS|USD)$")


class WarehouseUpdate(BaseModel):
    name: str | None = None
    address: str | None = None
    notes: str | None = None
    is_default: bool | None = None
    warehouse_type: str | None = Field(default=None, pattern="^(new|used|partnership|accessories)$")
    currency_code: str | None = Field(default=None, pattern="^(|TJS|USD)$")


class StockMovementIn(BaseModel):
    warehouse_id: int
    product_id: int
    quantity: int = Field(ge=1)
    notes: str = ""


class InboundProductNew(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    category: str = "phone"
    ownership_type: OwnershipType = "own"
    supplier_name: str = ""
    brand: str = ""
    sku: str = ""
    barcode: str = ""
    purchase_price: float = Field(ge=0)
    sale_price: float = Field(ge=0)
    min_stock: int = Field(ge=0, default=2)
    model: str = ""
    color: str = ""
    size: str = ""
    memory: str = ""
    ram: str = ""
    customs_cleared: int = Field(ge=0, le=1, default=0)
    customs_price: float = Field(ge=0, default=0)
    specs_extra: str = ""
    condition: ProductCondition = "new"


class InboundReceiptIn(BaseModel):
    warehouse_id: int
    mode: Literal["new", "existing"]
    product_id: int | None = None
    quantity: int = Field(default=1, ge=1)
    imei: str = ""
    serial: str = ""
    battery_capacity: int | None = Field(default=None, ge=0, le=100)
    notes: str = ""
    client_name: str = ""
    region: str = ""
    arrival_date: str = ""
    unit_purchase_price: float | None = Field(default=None, ge=0)
    unit_extra_cost: float = Field(default=0, ge=0)
    product: InboundProductNew | None = None


class WarehouseQuickSellIn(BaseModel):
    unit_id: int
    sale_price: float = Field(gt=0)
    payment_method: str = "cash"
    paid_amount: float | None = Field(default=None, ge=0)
    payments: list[PaymentPart] = Field(default_factory=list)
    discount: float = Field(ge=0, default=0)
    notes: str = ""
    debtor_name: str = ""
    debtor_phone: str = ""


class StockTransferIn(BaseModel):
    product_id: int
    from_warehouse_id: int
    to_warehouse_id: int
    quantity: int = Field(ge=1)
    notes: str = ""


class TradeInIn(BaseModel):
    given_product_id: int
    given_warehouse_id: int
    received_name: str = Field(min_length=1, max_length=200)
    received_brand: str = ""
    received_model: str = ""
    received_color: str = ""
    received_size: str = ""
    received_memory: str = ""
    received_ram: str = ""
    received_specs_extra: str = ""
    received_condition: ProductCondition = "used"
    received_purchase_price: float = Field(ge=0, default=0)
    received_sale_price: float = Field(ge=0, default=0)
    received_value: float = Field(ge=0)
    cash_amount: float = Field(ge=0, default=0)
    card_amount: float = Field(ge=0, default=0)
    received_warehouse_id: int | None = None
    given_unit_id: int | None = None
    received_imei: str = ""
    received_serial: str = ""
    notes: str = ""


@asynccontextmanager
async def lifespan(_: FastAPI):
    init_db()
    logger.info("DB: %s", DB_PATH)
    yield


app = FastAPI(title=settings.store_name, lifespan=lifespan)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
app.mount("/uploads", StaticFiles(directory=UPLOADS_DIR), name="uploads")


@app.get("/")
async def index():
    return FileResponse(
        STATIC_DIR / "index.html",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
        },
    )


@app.get("/sw.js")
async def service_worker():
    return FileResponse(
        STATIC_DIR / "sw.js",
        media_type="application/javascript",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Service-Worker-Allowed": "/",
        },
    )


@app.get("/manifest.webmanifest")
async def web_manifest():
    return FileResponse(STATIC_DIR / "manifest.webmanifest", media_type="application/manifest+json")


@app.get("/api/health")
async def health():
    return {
        "status": "ok",
        "build": "we-owe-wallet-minus-v1",
        "db": str(DB_PATH),
        "db_exists": DB_PATH.exists(),
    }


@app.get("/api/config")
async def config():
    with db() as conn:
        user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        currency = get_currency_config(conn)
        payment_methods = list_payment_methods(conn, active_only=True)
        rates = latest_exchange_rates(conn)
    return {
        "auth_required": user_count > 0 or bool(settings.store_pin),
        "store_name": settings.store_name,
        "simple_ui": settings.simple_ui,
        "role_pages": ROLE_PAGES,
        "simple_role_pages": SIMPLE_ROLE_PAGES,
        "currency": currency,
        "payment_methods": payment_methods,
        "exchange_rates": rates,
    }


@app.post("/api/auth/check")
async def auth_check(body: dict):
    pin = body.get("pin", "")
    with db() as conn:
        user = resolve_user(conn, pin)
        if not user:
            raise HTTPException(status_code=401, detail="Неверный PIN")
        shift = get_open_shift(conn)
        return {
            "ok": True,
            "user": {"id": user["id"], "name": user["name"], "role": user["role"]},
            "pages": pages_for_role(user["role"], simple=settings.simple_ui),
            "simple_ui": settings.simple_ui,
            "open_shift": row_to_dict(shift) if shift else None,
        }



# --- Settings (owner) ---


@app.get("/api/settings")
async def get_settings(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        return {
            "currency": get_currency_config(conn),
            "exchange_rates": [
                row_to_dict(r)
                for r in conn.execute(
                    "SELECT * FROM exchange_rates ORDER BY effective_at DESC, id DESC LIMIT 200"
                ).fetchall()
            ],
            "payment_methods": list_payment_methods(conn, active_only=False),
        }


@app.put("/api/settings/currency")
async def update_currency(body: CurrencySettingsIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    code = body.base_currency.upper()
    with db() as conn:
        set_setting(conn, "base_currency", code)
        set_setting(conn, "currency_symbol", body.currency_symbol)
        set_setting(conn, "currency_name", body.currency_name or code)
        if not conn.execute(
            "SELECT 1 FROM exchange_rates WHERE currency_code = ? AND rate = 1", (code,)
        ).fetchone():
            conn.execute(
                "INSERT INTO exchange_rates (currency_code, rate, effective_at, notes, created_at) VALUES (?, 1, ?, 'Базовая валюта', ?)",
                (code, utc_now(), utc_now()),
            )
    return {"ok": True}


@app.post("/api/settings/exchange-rates")
async def add_exchange_rate(body: ExchangeRateIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    when = normalize_exchange_effective_at(body.effective_at)
    code = (body.currency_code or "").strip().upper()
    if len(code) != 3:
        raise HTTPException(status_code=400, detail="Код валюты: 3 буквы, например USD")
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO exchange_rates (currency_code, rate, effective_at, notes, created_at) VALUES (?, ?, ?, ?, ?)",
            (code, float(body.rate), when, (body.notes or "").strip(), utc_now()),
        )
        row = conn.execute("SELECT * FROM exchange_rates WHERE id = ?", (cur.lastrowid,)).fetchone()
    return row_to_dict(row)


@app.delete("/api/settings/exchange-rates/{rate_id}")
async def delete_exchange_rate(rate_id: str, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    try:
        rid = int(str(rate_id).strip())
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Некорректный id курса — обновите страницу (Ctrl+F5)")
    if rid <= 0:
        raise HTTPException(status_code=400, detail="Некорректный id курса")
    with db() as conn:
        row = conn.execute("SELECT * FROM exchange_rates WHERE id = ?", (rid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Курс не найден")
        conn.execute("DELETE FROM exchange_rates WHERE id = ?", (rid,))
    return {"ok": True, "deleted": row_to_dict(row)}


@app.post("/api/settings/payment-methods")
async def add_payment_method(body: PaymentMethodIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    code = body.code.strip().lower()
    with db() as conn:
        if conn.execute("SELECT 1 FROM payment_methods WHERE code = ?", (code,)).fetchone():
            raise HTTPException(status_code=400, detail="Код уже существует")
        cur_code = _infer_payment_method_currency(body.name, code, body.currency_code)
        cur = conn.execute(
            """
            INSERT INTO payment_methods (code, name, method_type, is_active, sort_order, created_at, currency_code)
            VALUES (?, ?, ?, 1, ?, ?, ?)
            """,
            (code, body.name.strip(), body.method_type, body.sort_order, utc_now(), cur_code),
        )
        row = conn.execute("SELECT * FROM payment_methods WHERE id = ?", (cur.lastrowid,)).fetchone()
    return row_to_dict(row)


@app.put("/api/settings/payment-methods/{method_id}")
async def update_payment_method(
    method_id: int, body: PaymentMethodUpdate, x_pin: str | None = Header(default=None, alias="X-Pin")
):
    check_pin(x_pin, min_role="owner")
    fields = {k: v for k, v in body.model_dump().items() if v is not None}
    if not fields:
        raise HTTPException(status_code=400, detail="Нет данных")
    sets = ", ".join(f"{k} = ?" for k in fields)
    with db() as conn:
        conn.execute(f"UPDATE payment_methods SET {sets} WHERE id = ?", (*fields.values(), method_id))
        row = conn.execute("SELECT * FROM payment_methods WHERE id = ?", (method_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Не найдено")
    return row_to_dict(row)


@app.get("/api/expenses")
async def list_expenses(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    category: str = "",
    department: str = "",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    clause, params, _ = _report_period_clause(period, date_from, date_to, "e.expense_date")
    with db() as conn:
        sql = f"""
            SELECT e.*, COALESCE(w.name, '') AS warehouse_name
            FROM expenses e
            LEFT JOIN warehouses w ON w.id = e.warehouse_id
            WHERE 1=1 {clause}
        """
        if category.strip():
            sql += " AND e.category = ?"
            params.append(category.strip())
        if department.strip():
            sql += " AND COALESCE(e.department, 'main') = ?"
            params.append(department.strip())
        sql += " ORDER BY e.expense_date DESC, e.id DESC"
        rows = conn.execute(sql, params).fetchall()
        pay_map = {
            r["code"]: r["name"]
            for r in conn.execute("SELECT code, name FROM payment_methods").fetchall()
        }
    out = []
    for r in rows:
        d = row_to_dict(r)
        d["payment_method_name"] = pay_map.get(d.get("payment_method_code") or "", d.get("payment_method_code") or "—")
        if not (d.get("created_by") or "").strip():
            if (d.get("description") or "") == "Excel импорт ОПиУ":
                d["created_by"] = "Excel импорт"
            else:
                d["created_by"] = "—"
        out.append(d)
    return out


@app.post("/api/expenses")
async def create_expense(body: ExpenseIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    user = check_pin(x_pin)
    dept = (body.department or "main").strip() or "main"
    kind = (body.kind or "expense").strip().lower()
    if kind not in ("expense", "payout"):
        kind = "expense"
    payee = (body.payee or "").strip()
    category = (body.category or "").strip()
    if kind == "payout":
        if not body.warehouse_id:
            raise HTTPException(status_code=400, detail="Для выплаты выберите склад")
        if not payee:
            raise HTTPException(status_code=400, detail="Укажите, кому выплатили")
        if not category:
            category = "Выплата"
    elif not category:
        raise HTTPException(status_code=400, detail="Укажите категорию")
    if dept != "accessories" and ROLE_LEVEL.get(user.get("role") or "", 0) < ROLE_LEVEL["cashier"]:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    when = body.expense_date or utc_now()[:10]
    with db() as conn:
        if user.get("role") == "accessories" and dept != "accessories":
            raise HTTPException(status_code=403, detail="Доступны только расходы аксессуаров")
        warehouse_id = body.warehouse_id
        if warehouse_id is not None:
            wh = conn.execute(
                """
                SELECT id, name, COALESCE(warehouse_type, '') AS warehouse_type
                FROM warehouses WHERE id = ?
                """,
                (warehouse_id,),
            ).fetchone()
            if not wh:
                raise HTTPException(status_code=400, detail="Склад не найден")
            if (wh["warehouse_type"] or "") == "accessories" or "аксесс" in (wh["name"] or "").lower():
                if dept != "accessories":
                    raise HTTPException(status_code=400, detail="Выберите склад телефонов")
        elif kind == "payout":
            raise HTTPException(status_code=400, detail="Для выплаты выберите склад")
        else:
            warehouse_id = None
        desc = body.description or ""
        if kind == "payout" and payee and payee.lower() not in desc.lower():
            desc = f"{payee}" + (f" — {desc}" if desc.strip() else "")
        cur = conn.execute(
            """
            INSERT INTO expenses (
                category, amount, description, payment_method_code, expense_date,
                created_at, department, created_by, created_by_user_id,
                warehouse_id, payee, kind
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                category,
                body.amount,
                desc,
                body.payment_method_code,
                when,
                utc_now(),
                dept,
                (user.get("name") or "").strip() or "—",
                user.get("id") or None,
                warehouse_id,
                payee,
                kind,
            ),
        )
        row = conn.execute(
            """
            SELECT e.*, COALESCE(w.name, '') AS warehouse_name
            FROM expenses e
            LEFT JOIN warehouses w ON w.id = e.warehouse_id
            WHERE e.id = ?
            """,
            (cur.lastrowid,),
        ).fetchone()
    return row_to_dict(row)


@app.delete("/api/expenses/{expense_id}")
async def delete_expense(expense_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="cashier")
    with db() as conn:
        row = conn.execute("SELECT * FROM expenses WHERE id = ?", (expense_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Не найдено")
        conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    return {"ok": True, "deleted": row_to_dict(row)}


@app.delete("/api/pos/cash-inflows/{inflow_id}")
async def delete_cash_inflow(inflow_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="cashier")
    with db() as conn:
        if not _table_exists(conn, "cash_inflows"):
            raise HTTPException(status_code=404, detail="Не найдено")
        row = conn.execute("SELECT * FROM cash_inflows WHERE id = ?", (inflow_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Не найдено")
        conn.execute("DELETE FROM cash_inflows WHERE id = ?", (inflow_id,))
    return {"ok": True, "deleted": row_to_dict(row)}


def _pos_lines_window(period: str) -> tuple[str, str, str, str]:
    """Return (period_key, date_from, date_to, label) for POS line lists."""
    today = utc_now()[:10]
    if period == "yesterday":
        y = (datetime.strptime(today, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        return "all", y, y, f"Вчера · {y}"
    if period == "day":
        return "all", today, today, f"Сегодня · {today}"
    if period == "week":
        start = period_start("week")[:10]
        return "all", start, today, f"Неделя · {start} — {today}"
    if period == "month":
        start = period_start("month")[:10]
        return "all", start, today, f"Месяц · {start} — {today}"
    return "all", "", "", "Все записи"


@app.get("/api/pos/cash-lines")
async def pos_cash_lines(
    period: str = Query(default="day", pattern="^(day|yesterday|week|month|all)$"),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    """Expenses + inflows for a selectable period (for delete / review of past days)."""
    check_pin(x_pin)
    period_key, date_from, date_to, label = _pos_lines_window(period)
    with db() as conn:
        exp_clause, exp_params, _ = _report_period_clause(period_key, date_from, date_to, "e.expense_date")
        exp_rows = conn.execute(
            f"""
            SELECT e.*, COALESCE(w.name, '') AS warehouse_name
            FROM expenses e
            LEFT JOIN warehouses w ON w.id = e.warehouse_id
            WHERE COALESCE(e.affects_cash, 1) = 1 {exp_clause}
            ORDER BY e.expense_date DESC, e.id DESC
            LIMIT 300
            """,
            list(exp_params),
        ).fetchall()
        inflows: list[dict[str, Any]] = []
        if _table_exists(conn, "cash_inflows"):
            iclause, iparams, _ = _report_period_clause(period_key, date_from, date_to, "created_at")
            inflows = [
                row_to_dict(r)
                for r in conn.execute(
                    f"""
                    SELECT * FROM cash_inflows
                    WHERE 1=1 {iclause}
                    ORDER BY created_at DESC
                    LIMIT 300
                    """,
                    list(iparams),
                ).fetchall()
            ]
    return {
        "period": period,
        "period_label": label,
        "date_from": date_from,
        "date_to": date_to,
        "expenses": [row_to_dict(r) for r in exp_rows],
        "cash_inflows": inflows,
    }


@app.get("/api/pos/cash-inflows")
async def list_cash_inflows(
    period: str = Query(default="day", pattern="^(day|week|month|all)$"),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        if not _table_exists(conn, "cash_inflows"):
            return []
        clause, params, _ = _report_period_clause(period, "", "", "created_at")
        rows = conn.execute(
            f"""
            SELECT * FROM cash_inflows
            WHERE 1=1 {clause}
            ORDER BY created_at DESC LIMIT 40
            """,
            params,
        ).fetchall()
    return [row_to_dict(r) for r in rows]


def _usd_to_tjs_rate(conn: sqlite3.Connection, at: str | None = None, override: float | None = None) -> float | None:
    if override is not None and float(override) > 0:
        return float(override)
    base = get_setting(conn, "base_currency", "TJS").upper()
    if base == "TJS":
        rate = get_exchange_rate_at(conn, "USD", at)
        return float(rate) if rate and rate > 0 else None
    # If base is USD, invert TJS rate when present
    tjs = get_exchange_rate_at(conn, "TJS", at)
    if tjs and tjs > 0 and base == "USD":
        return round(1.0 / float(tjs), 6)
    return None


@app.post("/api/pos/cash-inflow")
async def create_cash_inflow(body: CashInflowIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    user = check_pin(x_pin)
    source = (body.source_type or "counterparty").strip().lower()
    if source not in ("counterparty", "debtor"):
        raise HTTPException(status_code=400, detail="Тип прихода: counterparty или debtor")
    with db() as conn:
        pm = get_payment_method(conn, body.payment_method_code)
        if not pm:
            raise HTTPException(status_code=400, detail="Способ оплаты не найден")
        now = utc_now()
        currency = (body.currency_code or "TJS").upper().strip()
        base = get_setting(conn, "base_currency", "TJS").upper()
        amount = float(body.amount)
        if source == "debtor":
            if body.mutual_entry_id and body.receivable_id:
                raise HTTPException(status_code=400, detail="Укажите одного должника")
            if not body.mutual_entry_id and not body.receivable_id:
                raise HTTPException(status_code=400, detail="Выберите должника")

            if body.mutual_entry_id:
                row = conn.execute(
                    "SELECT * FROM mutual_entries WHERE id = ?", (body.mutual_entry_id,)
                ).fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail="Долг не найден")
                if row["status"] == "closed":
                    raise HTTPException(status_code=400, detail="Долг уже закрыт")
                if row["direction"] != "owe_us":
                    raise HTTPException(
                        status_code=400,
                        detail="Это «мы должны» — погашение через Взаиморасчёты → Оплата",
                    )
                debt_cur = (row["currency_code"] or "TJS").strip().upper() or "TJS"
                pay_cur = currency or debt_cur
                fx = None
                if pay_cur != debt_cur:
                    fx = _usd_to_tjs_rate(conn, now, body.fx_rate)
                    if not fx:
                        raise HTTPException(
                            status_code=400,
                            detail="Укажите курс: сколько сомони за 1 доллар",
                        )
                reduction = debt_reduction_from_payment(
                    debt_currency=debt_cur,
                    pay_amount=amount,
                    pay_currency=pay_cur,
                    fx_rate=fx,
                )
                due = float(row["amount_due"])
                if reduction > due + 0.01:
                    raise HTTPException(
                        status_code=400,
                        detail=f"К долгу спишется {reduction:.2f} {debt_cur}, а остаток {due:.2f}",
                    )
                note = body.notes or f"Приход с кассы ({pay_cur} {amount:g})"
                conn.execute(
                    """
                    INSERT INTO mutual_payments
                    (entry_id, amount, payment_method_code, notes, created_at,
                     pay_amount, pay_currency_code, fx_rate)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        body.mutual_entry_id, reduction, body.payment_method_code, note, now,
                        amount, pay_cur, fx if pay_cur != debt_cur else None,
                    ),
                )
                new_paid = float(row["paid_amount"]) + reduction
                new_due = max(0.0, float(row["amount"]) - new_paid)
                new_status = "closed" if new_due <= 0.01 else "open"
                conn.execute(
                    """
                    UPDATE mutual_entries
                    SET paid_amount = ?, amount_due = ?, status = ?,
                        closed_at = CASE WHEN ? = 'closed' THEN ? ELSE closed_at END
                    WHERE id = ?
                    """,
                    (new_paid, new_due, new_status, new_status, now, body.mutual_entry_id),
                )
                amount_base = (
                    reduction if debt_cur == base
                    else (convert_amount(conn, reduction, debt_cur, base, at=now) or reduction)
                )
                cur = conn.execute(
                    """
                    INSERT INTO cash_inflows
                    (amount, currency_code, amount_base, payment_method_code, source_type,
                     counterparty_name, receivable_id, mutual_entry_id, notes, created_at, created_by)
                    VALUES (?, ?, ?, ?, 'debtor', ?, NULL, ?, ?, ?, ?)
                    """,
                    (
                        amount, pay_cur, round(float(amount_base), 2), body.payment_method_code,
                        row["person_name"] or "", body.mutual_entry_id,
                        note, now, (user or {}).get("name", ""),
                    ),
                )
                saved = conn.execute("SELECT * FROM cash_inflows WHERE id = ?", (cur.lastrowid,)).fetchone()
                return {"ok": True, "type": "debtor", "inflow": row_to_dict(saved), "amount_due": new_due}

            # Sale receivable
            row = conn.execute("SELECT * FROM receivables WHERE id = ?", (body.receivable_id,)).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Долг не найден")
            if row["status"] == "closed":
                raise HTTPException(status_code=400, detail="Долг уже закрыт")
            if row["sale_id"]:
                sale_row = conn.execute(
                    "SELECT status, currency_code FROM sales WHERE id = ?", (row["sale_id"],)
                ).fetchone()
                if sale_row and sale_row["status"] == "voided":
                    raise HTTPException(status_code=400, detail="Продажа отменена — долг снят")
                debt_cur = (sale_row["currency_code"] if sale_row else None) or "TJS"
            else:
                debt_cur = "TJS"
            debt_cur = str(debt_cur).strip().upper() or "TJS"
            pay_cur = currency or debt_cur
            fx = None
            if pay_cur != debt_cur:
                fx = _usd_to_tjs_rate(conn, now, body.fx_rate)
                if not fx:
                    raise HTTPException(
                        status_code=400,
                        detail="Укажите курс: сколько сомони за 1 доллар",
                    )
            reduction = debt_reduction_from_payment(
                debt_currency=debt_cur,
                pay_amount=amount,
                pay_currency=pay_cur,
                fx_rate=fx,
            )
            due = float(row["amount_due"])
            if reduction > due + 0.01:
                raise HTTPException(
                    status_code=400,
                    detail=f"К долгу спишется {reduction:.2f} {debt_cur}, а остаток {due:.2f}",
                )
            note = body.notes or f"Приход с кассы ({pay_cur} {amount:g})"
            conn.execute(
                """
                INSERT INTO receivable_payments
                (receivable_id, amount, payment_method_code, notes, created_at,
                 pay_amount, pay_currency_code, fx_rate)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    body.receivable_id, reduction, body.payment_method_code, note, now,
                    amount, pay_cur, fx if pay_cur != debt_cur else None,
                ),
            )
            new_paid = float(row["paid_amount"]) + reduction
            new_due = max(0.0, float(row["total_amount"]) - new_paid)
            new_status = "closed" if new_due <= 0.01 else "open"
            conn.execute(
                """
                UPDATE receivables
                SET paid_amount = ?, amount_due = ?, status = ?,
                    closed_at = CASE WHEN ? = 'closed' THEN ? ELSE closed_at END
                WHERE id = ?
                """,
                (new_paid, new_due, new_status, new_status, now, body.receivable_id),
            )
            if row["sale_id"]:
                conn.execute(
                    "UPDATE sales SET amount_paid = COALESCE(amount_paid, 0) + ?, amount_due = ? WHERE id = ?",
                    (reduction, new_due, row["sale_id"]),
                )
            amount_base = (
                reduction if debt_cur == base
                else (convert_amount(conn, reduction, debt_cur, base, at=now) or reduction)
            )
            cur = conn.execute(
                """
                INSERT INTO cash_inflows
                (amount, currency_code, amount_base, payment_method_code, source_type,
                 counterparty_name, receivable_id, mutual_entry_id, notes, created_at, created_by)
                VALUES (?, ?, ?, ?, 'debtor', ?, ?, NULL, ?, ?, ?)
                """,
                (
                    amount, pay_cur, round(float(amount_base), 2), body.payment_method_code,
                    row["customer_name"] or "", body.receivable_id,
                    note, now, (user or {}).get("name", ""),
                ),
            )
            saved = conn.execute("SELECT * FROM cash_inflows WHERE id = ?", (cur.lastrowid,)).fetchone()
            return {"ok": True, "type": "debtor", "inflow": row_to_dict(saved), "amount_due": new_due}

        name = body.counterparty_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Укажите контрагента / источник")
        amount_base = amount if currency == base else convert_amount(conn, amount, currency, base, at=now)
        amount_base = round(float(amount_base), 2)
        note = body.notes.strip()
        if currency != base:
            note = (f"{currency} {amount:g} → {base} {amount_base:g}" + (f" · {note}" if note else "")).strip()
        cur = conn.execute(
            """
            INSERT INTO cash_inflows
            (amount, currency_code, amount_base, payment_method_code, source_type,
             counterparty_name, receivable_id, mutual_entry_id, notes, created_at, created_by)
            VALUES (?, ?, ?, ?, 'counterparty', ?, NULL, NULL, ?, ?, ?)
            """,
            (
                amount, currency, amount_base, body.payment_method_code,
                name, note, now, (user or {}).get("name", ""),
            ),
        )
        saved = conn.execute("SELECT * FROM cash_inflows WHERE id = ?", (cur.lastrowid,)).fetchone()
    return {"ok": True, "type": "counterparty", "inflow": row_to_dict(saved)}


@app.get("/api/settings/expense-allocation")
async def get_expense_allocation(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        rules = get_expense_warehouse_split(conn)
    return {"rules": rules}


@app.put("/api/settings/expense-allocation")
async def save_expense_allocation(body: ExpenseAllocationIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    total = sum(r.pct for r in body.rules)
    if body.rules and abs(total - 100) > 0.5:
        raise HTTPException(status_code=400, detail=f"Сумма процентов должна быть 100% (сейчас {total:.1f}%)")
    with db() as conn:
        allowed = {
            int(r["id"]) for r in conn.execute(
                """
                SELECT id FROM warehouses
                WHERE COALESCE(warehouse_type, '') != 'accessories'
                  AND LOWER(name) NOT LIKE '%аксесс%'
                """
            ).fetchall()
        }
        conn.execute("DELETE FROM expense_warehouse_split")
        for r in body.rules:
            if r.warehouse_id not in allowed:
                continue
            if r.pct <= 0:
                continue
            conn.execute(
                "INSERT INTO expense_warehouse_split (warehouse_id, pct) VALUES (?, ?)",
                (r.warehouse_id, r.pct),
            )
        rules = get_expense_warehouse_split(conn)
    return {"rules": rules}


@app.get("/api/receivables")
async def list_receivables(
    status: str = Query(default="", pattern="^(|open|closed)$"),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        sql = """
            SELECT r.*, COALESCE(w.name, '—') AS warehouse_name,
                   UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), NULLIF(TRIM(w.currency_code), ''), 'TJS')) AS currency_code,
                   (SELECT GROUP_CONCAT(si.product_name, ', ')
                    FROM sale_items si WHERE si.sale_id = r.sale_id) AS products
            FROM receivables r
            LEFT JOIN sales s ON s.id = r.sale_id
            LEFT JOIN warehouses w ON w.id = COALESCE(r.warehouse_id, s.warehouse_id)
            WHERE (r.sale_id IS NULL OR s.status != 'voided')
        """
        params: list[Any] = []
        if status:
            sql += " AND r.status = ?"
            params.append(status)
        sql += " ORDER BY r.status ASC, r.created_at DESC"
        rows = conn.execute(sql, params).fetchall()
    return [row_to_dict(r) for r in rows]


@app.post("/api/receivables/{receivable_id}/pay")
async def pay_receivable(
    receivable_id: int,
    body: ReceivablePaymentIn,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        row = conn.execute("SELECT * FROM receivables WHERE id = ?", (receivable_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Долг не найден")
        if row["status"] == "closed":
            raise HTTPException(status_code=400, detail="Долг уже закрыт")
        if row["sale_id"]:
            sale_row = conn.execute(
                "SELECT status, currency_code FROM sales WHERE id = ?", (row["sale_id"],)
            ).fetchone()
            if sale_row and sale_row["status"] == "voided":
                raise HTTPException(status_code=400, detail="Продажа отменена — долг снят")
            debt_cur = (sale_row["currency_code"] if sale_row else None) or "TJS"
        else:
            debt_cur = "TJS"
        debt_cur = str(debt_cur).strip().upper() or "TJS"
        pay_cur = (body.currency_code or debt_cur).strip().upper() or debt_cur
        reduction = debt_reduction_from_payment(
            debt_currency=debt_cur,
            pay_amount=body.amount,
            pay_currency=pay_cur,
            fx_rate=body.fx_rate,
        )
        due = float(row["amount_due"])
        if reduction > due + 0.01:
            raise HTTPException(
                status_code=400,
                detail=f"К долгу спишется {reduction:.2f} {debt_cur}, а остаток {due:.2f}",
            )
        pm = get_payment_method(conn, body.payment_method_code)
        if not pm:
            raise HTTPException(status_code=400, detail="Способ оплаты не найден")
        now = utc_now()
        conn.execute(
            """
            INSERT INTO receivable_payments
            (receivable_id, amount, payment_method_code, notes, created_at,
             pay_amount, pay_currency_code, fx_rate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                receivable_id, reduction, body.payment_method_code, body.notes, now,
                body.amount, pay_cur, body.fx_rate if pay_cur != debt_cur else None,
            ),
        )
        new_paid = float(row["paid_amount"]) + reduction
        new_due = max(0.0, float(row["total_amount"]) - new_paid)
        new_status = "closed" if new_due <= 0.01 else "open"
        conn.execute(
            """
            UPDATE receivables
            SET paid_amount = ?, amount_due = ?, status = ?,
                closed_at = CASE WHEN ? = 'closed' THEN ? ELSE closed_at END
            WHERE id = ?
            """,
            (new_paid, new_due, new_status, new_status, now, receivable_id),
        )
        if row["sale_id"]:
            conn.execute(
                "UPDATE sales SET amount_paid = COALESCE(amount_paid, 0) + ?, amount_due = ? WHERE id = ?",
                (reduction, new_due, row["sale_id"]),
            )
        updated = conn.execute("SELECT * FROM receivables WHERE id = ?", (receivable_id,)).fetchone()
    return row_to_dict(updated)


@app.get("/api/receivables/{receivable_id}/payments")
async def list_receivable_payments(
    receivable_id: int,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        if not conn.execute("SELECT id FROM receivables WHERE id = ?", (receivable_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Долг не найден")
        rows = conn.execute(
            """
            SELECT * FROM receivable_payments
            WHERE receivable_id = ?
            ORDER BY created_at DESC
            """,
            (receivable_id,),
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/mutual-entries")
async def list_mutual_entries(
    status: str = Query(default="", pattern="^(|open|closed)$"),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        sql = "SELECT * FROM mutual_entries WHERE 1=1"
        params: list[Any] = []
        if status:
            sql += " AND status = ?"
            params.append(status)
        sql += " ORDER BY status ASC, created_at DESC"
        rows = conn.execute(sql, params).fetchall()
    return [row_to_dict(r) for r in rows]


@app.post("/api/mutual-entries")
async def create_mutual_entry(
    body: MutualEntryIn,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    now = utc_now()
    cur_code = (body.currency_code or "TJS").strip().upper()
    with db() as conn:
        cur = conn.execute(
            """
            INSERT INTO mutual_entries
            (person_name, person_phone, direction, amount, paid_amount, amount_due,
             currency_code, payment_method_code, product_note, status, notes, created_at)
            VALUES (?, ?, ?, ?, 0, ?, ?, ?, ?, 'open', ?, ?)
            """,
            (
                body.person_name.strip(),
                body.person_phone.strip(),
                body.direction,
                body.amount,
                body.amount,
                cur_code,
                body.payment_method_code,
                body.product_note,
                body.notes,
                now,
            ),
        )
        row = conn.execute("SELECT * FROM mutual_entries WHERE id = ?", (cur.lastrowid,)).fetchone()
    return row_to_dict(row)


@app.get("/api/mutual-entries/{entry_id}/payments")
async def list_mutual_payments(
    entry_id: int,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        if not conn.execute("SELECT id FROM mutual_entries WHERE id = ?", (entry_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Запись не найдена")
        rows = conn.execute(
            "SELECT * FROM mutual_payments WHERE entry_id = ? ORDER BY created_at DESC",
            (entry_id,),
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.post("/api/mutual-entries/{entry_id}/pay")
async def pay_mutual_entry(
    entry_id: int,
    body: MutualPaymentIn,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        row = conn.execute("SELECT * FROM mutual_entries WHERE id = ?", (entry_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Запись не найдена")
        if row["status"] == "closed":
            raise HTTPException(status_code=400, detail="Запись уже закрыта")
        debt_cur = (row["currency_code"] or "TJS").strip().upper() or "TJS"
        pay_cur = (body.currency_code or debt_cur).strip().upper() or debt_cur
        reduction = debt_reduction_from_payment(
            debt_currency=debt_cur,
            pay_amount=body.amount,
            pay_currency=pay_cur,
            fx_rate=body.fx_rate,
        )
        due = float(row["amount_due"])
        if reduction > due + 0.01:
            raise HTTPException(
                status_code=400,
                detail=f"К долгу спишется {reduction:.2f} {debt_cur}, а остаток {due:.2f}",
            )
        pm = get_payment_method(conn, body.payment_method_code)
        if not pm:
            raise HTTPException(status_code=400, detail="Способ оплаты не найден")
        now = utc_now()
        conn.execute(
            """
            INSERT INTO mutual_payments
            (entry_id, amount, payment_method_code, notes, created_at,
             pay_amount, pay_currency_code, fx_rate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                entry_id, reduction, body.payment_method_code, body.notes, now,
                body.amount, pay_cur, body.fx_rate if pay_cur != debt_cur else None,
            ),
        )
        new_paid = float(row["paid_amount"]) + reduction
        new_due = max(0.0, float(row["amount"]) - new_paid)
        new_status = "closed" if new_due <= 0.01 else "open"
        conn.execute(
            """
            UPDATE mutual_entries
            SET paid_amount = ?, amount_due = ?, status = ?,
                closed_at = CASE WHEN ? = 'closed' THEN ? ELSE closed_at END
            WHERE id = ?
            """,
            (new_paid, new_due, new_status, new_status, now, entry_id),
        )
        updated = conn.execute("SELECT * FROM mutual_entries WHERE id = ?", (entry_id,)).fetchone()
    return row_to_dict(updated)


@app.put("/api/mutual-entries/{entry_id}")
async def update_mutual_entry(
    entry_id: int,
    body: MutualEntryUpdate,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        row = conn.execute("SELECT * FROM mutual_entries WHERE id = ?", (entry_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Запись не найдена")
        paid = float(row["paid_amount"] or 0)
        name = body.person_name.strip() if body.person_name is not None else row["person_name"]
        phone = body.person_phone if body.person_phone is not None else row["person_phone"]
        direction = body.direction if body.direction is not None else row["direction"]
        amount = float(body.amount) if body.amount is not None else float(row["amount"])
        currency = (body.currency_code or row["currency_code"] or "TJS").strip().upper()
        product = body.product_note if body.product_note is not None else row["product_note"]
        notes = body.notes if body.notes is not None else row["notes"]
        if amount + 0.01 < paid:
            raise HTTPException(
                status_code=400,
                detail=f"Сумма не может быть меньше уже оплаченного ({paid:.2f})",
            )
        due = max(0.0, amount - paid)
        status = "closed" if due <= 0.01 else "open"
        conn.execute(
            """
            UPDATE mutual_entries
            SET person_name = ?, person_phone = ?, direction = ?, amount = ?,
                amount_due = ?, currency_code = ?, product_note = ?, notes = ?, status = ?,
                closed_at = CASE WHEN ? = 'closed' THEN COALESCE(closed_at, ?) ELSE NULL END
            WHERE id = ?
            """,
            (
                name, phone or "", direction, amount, due, currency,
                product or "", notes or "", status, status, utc_now(), entry_id,
            ),
        )
        updated = conn.execute("SELECT * FROM mutual_entries WHERE id = ?", (entry_id,)).fetchone()
    return row_to_dict(updated)


@app.delete("/api/mutual-entries/{entry_id}")
async def delete_mutual_entry(entry_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        row = conn.execute("SELECT * FROM mutual_entries WHERE id = ?", (entry_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Запись не найдена")
        conn.execute("DELETE FROM mutual_payments WHERE entry_id = ?", (entry_id,))
        conn.execute("DELETE FROM mutual_entries WHERE id = ?", (entry_id,))
    return {"ok": True, "id": entry_id}


# --- Users (owner) ---


@app.get("/api/users")
async def list_users(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        rows = conn.execute(
            "SELECT id, name, role, is_active, created_at FROM users ORDER BY role, name"
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.post("/api/users")
async def create_user(body: UserIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO users (name, pin, role, is_active, created_at) VALUES (?, ?, ?, 1, ?)",
                (body.name.strip(), body.pin, body.role, utc_now()),
            )
            row = conn.execute(
                "SELECT id, name, role, is_active, created_at FROM users WHERE id = ?",
                (cur.lastrowid,),
            ).fetchone()
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="PIN уже используется")
    return row_to_dict(row)


@app.put("/api/users/{user_id}")
async def update_user(user_id: int, body: UserUpdate, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    fields = {k: v for k, v in body.model_dump().items() if v is not None}
    if not fields:
        raise HTTPException(status_code=400, detail="Нет данных")
    sets = ", ".join(f"{k} = ?" for k in fields)
    with db() as conn:
        try:
            conn.execute(f"UPDATE users SET {sets} WHERE id = ?", (*fields.values(), user_id))
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="PIN уже используется")
        row = conn.execute(
            "SELECT id, name, role, is_active, created_at FROM users WHERE id = ?", (user_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
    return row_to_dict(row)


# --- Shifts ---


def _parse_shift_wallets(raw: str | None) -> list[dict[str, Any]]:
    try:
        data = json.loads(raw or "[]")
    except json.JSONDecodeError:
        data = []
    return data if isinstance(data, list) else []


def shift_expected_wallets(conn: sqlite3.Connection, shift: sqlite3.Row) -> list[dict[str, Any]]:
    """Ожидаемые остатки: старт + приход − расход за смену (не net!)."""
    opening = _parse_shift_wallets(shift["opening_wallets_json"] if "opening_wallets_json" in shift.keys() else "")
    bags: dict[tuple[str, str], dict[str, Any]] = {}
    for w in opening:
        code = (w.get("method_code") or "").strip()
        cur = (w.get("currency_code") or "TJS").strip().upper() or "TJS"
        if not code:
            continue
        key = (code, cur)
        bags[key] = {
            "method_code": code,
            "method_name": w.get("method_name") or code,
            "method_type": w.get("method_type") or "",
            "currency_code": cur,
            "opening": float(w.get("amount") or 0),
            "inflow": 0.0,
            "outflow": 0.0,
            "movement": 0.0,
            "expected": float(w.get("amount") or 0),
        }
    opened_at = (shift["opened_at"] or "")[:19]
    day = (opened_at or utc_now())[:10]
    # Движение за календарный день смены (как KPI кассы): иначе expense_date
    # не попадает в фильтр opened_at с временем.
    period = _period_cash_by_currency(
        conn, "all", date_from=day, date_to=day,
    )
    for bal in period.get("balances") or []:
        code = bal.get("code") or ""
        mtype = bal.get("method_type") or ""
        name = bal.get("name") or code
        for row in bal.get("by_currency") or []:
            cur = (row.get("code") or "TJS").strip().upper() or "TJS"
            inn = float(row.get("inflow") or 0)
            out = float(row.get("outflow") or 0)
            net = round(inn - out, 2)
            key = (code, cur)
            if key not in bags:
                bags[key] = {
                    "method_code": code,
                    "method_name": name,
                    "method_type": mtype,
                    "currency_code": cur,
                    "opening": 0.0,
                    "inflow": 0.0,
                    "outflow": 0.0,
                    "movement": 0.0,
                    "expected": 0.0,
                }
            bags[key]["inflow"] = round(inn, 2)
            bags[key]["outflow"] = round(out, 2)
            bags[key]["movement"] = net
            bags[key]["expected"] = round(float(bags[key]["opening"]) + net, 2)
            if not bags[key].get("method_type"):
                bags[key]["method_type"] = mtype
            if not bags[key].get("method_name"):
                bags[key]["method_name"] = name
    # Ensure all active methods appear (even zeros)
    for pm in list_payment_methods(conn, active_only=True):
        bound = payment_method_currency(pm)
        currencies = [bound] if bound in ("USD", "TJS") else ["TJS", "USD"]
        for cur in currencies:
            key = (pm["code"], cur)
            if key not in bags:
                bags[key] = {
                    "method_code": pm["code"],
                    "method_name": pm["name"],
                    "method_type": pm["method_type"],
                    "currency_code": cur,
                    "opening": 0.0,
                    "inflow": 0.0,
                    "outflow": 0.0,
                    "movement": 0.0,
                    "expected": 0.0,
                }
    order = {m["code"]: i for i, m in enumerate(list_payment_methods(conn, active_only=True))}
    return sorted(
        bags.values(),
        key=lambda x: (order.get(x["method_code"], 999), 0 if x["currency_code"] == "TJS" else 1),
    )


def balances_from_shift_expected(
    conn: sqlite3.Connection,
    expected: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Кошельки кассы: остаток = старт + приход − расход. Старт НЕ входит в «+»."""
    grouped: dict[str, dict[str, Any]] = {}
    for row in expected:
        code = (row.get("method_code") or "").strip()
        if not code:
            continue
        cur = (row.get("currency_code") or "TJS").strip().upper() or "TJS"
        opening = float(row.get("opening") or 0)
        inn = float(row.get("inflow") or 0)
        out = float(row.get("outflow") or 0)
        # backward compat if only net movement present
        if inn == 0 and out == 0 and row.get("movement") is not None:
            mov = float(row.get("movement") or 0)
            if mov >= 0:
                inn = mov
            else:
                out = abs(mov)
        expected_amt = float(row.get("expected") if row.get("expected") is not None else (opening + inn - out))
        bag = grouped.get(code)
        if not bag:
            pm = get_payment_method(conn, code)
            bag = {
                "code": code,
                "name": row.get("method_name") or (pm["name"] if pm else code),
                "method_type": row.get("method_type") or (pm["method_type"] if pm else ""),
                "currency_code": payment_method_currency(pm) if pm else (row.get("currency_code") or ""),
                "by_currency": [],
                "inflow": None,
                "outflow": None,
                "net": None,
            }
            grouped[code] = bag
        bag["by_currency"].append({
            **currency_meta(cur),
            "inflow": round(inn, 2),
            "outflow": round(out, 2),
            "net": round(expected_amt, 2),
            "opening": round(opening, 2),
            "movement": round(inn - out, 2),
        })
    order = {m["code"]: i for i, m in enumerate(list_payment_methods(conn, active_only=True))}
    result = list(grouped.values())
    for bag in result:
        bag["by_currency"].sort(key=lambda r: (0 if r.get("code") == "TJS" else 1, r.get("code") or ""))
    return sorted(result, key=lambda b: order.get(b["code"], 999))


def apply_shift_close_to_cash(
    conn: sqlite3.Connection,
    *,
    shift_id: int,
    targets: list[dict[str, Any]],
    user_name: str = "",
) -> list[dict[str, Any]]:
    """При закрытии смены: подогнать остатки кассы под фактический пересчёт.

    В ledger (продажи/приходы/расходы) нет стартовых сумм открытия смены.
    Здесь пишем разницу actual − текущий остаток кошелька как cash_inflows
    source_type=shift_close (сумма может быть отрицательной при недостаче).
    Идемпотентно по маркеру shift_close:{id} в notes.
    """
    if not _table_exists(conn, "cash_inflows"):
        return []
    marker = f"shift_close:{int(shift_id)}"
    already = conn.execute(
        "SELECT 1 FROM cash_inflows WHERE notes LIKE ? LIMIT 1",
        (f"%{marker}%",),
    ).fetchone()
    if already:
        return []

    ledger: dict[tuple[str, str], float] = {}
    for b in (_period_cash_by_currency(conn, "all").get("balances") or []):
        code = (b.get("code") or "").strip()
        for row in b.get("by_currency") or []:
            cur = str(row.get("code") or "TJS").upper()
            ledger[(code, cur)] = float(row.get("net") or 0)

    base = get_setting(conn, "base_currency", "TJS").upper()
    now = utc_now()
    posted: list[dict[str, Any]] = []
    for aw in targets:
        code = (aw.get("method_code") or "").strip()
        cur = (aw.get("currency_code") or "TJS").strip().upper() or "TJS"
        if not code:
            continue
        actual = round(float(aw.get("amount") or 0), 2)
        current = round(float(ledger.get((code, cur), 0.0)), 2)
        delta = round(actual - current, 2)
        if abs(delta) < 0.01:
            continue
        amount_base = delta if cur == base else (convert_amount(conn, delta, cur, base, at=now) or delta)
        amount_base = round(float(amount_base), 2)
        name = aw.get("method_name") or code
        note = (
            f"{marker} · пересчёт смены #{shift_id}: "
            f"{name} {cur} факт {actual:g} − учёт {current:g} = {delta:+g}"
        )
        cur_ins = conn.execute(
            """
            INSERT INTO cash_inflows
            (amount, currency_code, amount_base, payment_method_code, source_type,
             counterparty_name, receivable_id, mutual_entry_id, notes, created_at, created_by)
            VALUES (?, ?, ?, ?, 'shift_close', ?, NULL, NULL, ?, ?, ?)
            """,
            (
                delta,
                cur,
                amount_base,
                code,
                f"Смена #{shift_id}",
                note,
                now,
                (user_name or "").strip() or "Смена",
            ),
        )
        posted.append({
            "id": int(cur_ins.lastrowid),
            "method_code": code,
            "currency_code": cur,
            "amount": delta,
            "actual": actual,
            "ledger_before": current,
        })
        ledger[(code, cur)] = actual
    return posted


@app.get("/api/shifts/current")
async def current_shift(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        shift = get_open_shift(conn)
        if not shift:
            return {"shift": None, "summary": None, "expected_wallets": [], "opening_wallets": []}
        summary = shift_sales_totals(conn, shift["id"])
        data = row_to_dict(shift) or {}
        opening = _parse_shift_wallets(data.get("opening_wallets_json"))
        data["opening_wallets"] = opening
        expected = shift_expected_wallets(conn, shift)
        return {
            "shift": data,
            "summary": summary,
            "opening_wallets": opening,
            "expected_wallets": expected,
        }


@app.post("/api/shifts/open")
async def open_shift(body: ShiftOpenIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    user = check_pin(x_pin)
    with db() as conn:
        if get_open_shift(conn):
            raise HTTPException(status_code=400, detail="Смена уже открыта")
        wallets: list[dict[str, Any]] = []
        opening_cash = float(body.opening_cash or 0)
        for w in body.opening_wallets:
            code = (w.method_code or "").strip()
            cur = (w.currency_code or "TJS").strip().upper() or "TJS"
            amt = round(float(w.amount or 0), 2)
            if not code or amt < 0:
                continue
            pm = get_payment_method(conn, code)
            if not pm:
                raise HTTPException(status_code=400, detail=f"Кошелёк «{code}» не найден")
            wallets.append({
                "method_code": code,
                "method_name": pm["name"],
                "method_type": pm["method_type"],
                "currency_code": cur,
                "amount": amt,
            })
        if wallets:
            # Нал смн — для совместимости со старым closing_cash
            opening_cash = round(
                sum(
                    float(x["amount"])
                    for x in wallets
                    if (x.get("method_type") or "") == "cash"
                    and (x.get("currency_code") or "") == "TJS"
                ),
                2,
            )
        elif opening_cash > 0:
            cash_code = cash_wallet_code_for_currency(conn, "TJS") or "cash"
            pm = get_payment_method(conn, cash_code)
            wallets = [{
                "method_code": cash_code,
                "method_name": pm["name"] if pm else "Наличные",
                "method_type": "cash",
                "currency_code": "TJS",
                "amount": opening_cash,
            }]
        cur = conn.execute(
            """
            INSERT INTO shifts (user_id, user_name, opened_at, opening_cash, status, opening_wallets_json)
            VALUES (?, ?, ?, ?, 'open', ?)
            """,
            (
                user.get("id") or None,
                user.get("name", ""),
                utc_now(),
                opening_cash,
                json.dumps(wallets, ensure_ascii=False),
            ),
        )
        row = conn.execute("SELECT * FROM shifts WHERE id = ?", (cur.lastrowid,)).fetchone()
    data = row_to_dict(row) or {}
    try:
        data["opening_wallets"] = json.loads(data.get("opening_wallets_json") or "[]")
    except json.JSONDecodeError:
        data["opening_wallets"] = []
    return data


@app.post("/api/shifts/{shift_id}/close")
async def close_shift(shift_id: int, body: ShiftCloseIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        shift = conn.execute(
            "SELECT * FROM shifts WHERE id = ? AND status = 'open'", (shift_id,)
        ).fetchone()
        if not shift:
            raise HTTPException(status_code=404, detail="Открытая смена не найдена")
        totals = shift_sales_totals(conn, shift_id)
        expected_cash_in_drawer = float(shift["opening_cash"]) + totals["expected_cash"]
        expected_wallets = shift_expected_wallets(conn, shift)

        actual_wallets: list[dict[str, Any]] = []
        for w in body.actual_wallets:
            code = (w.method_code or "").strip()
            cur = (w.currency_code or "TJS").strip().upper() or "TJS"
            amt = round(float(w.amount or 0), 2)
            if not code:
                continue
            pm = get_payment_method(conn, code)
            if not pm:
                raise HTTPException(status_code=400, detail=f"Кошелёк «{code}» не найден")
            actual_wallets.append({
                "method_code": code,
                "method_name": pm["name"],
                "method_type": pm["method_type"],
                "currency_code": cur,
                "amount": amt,
            })
        if not actual_wallets and body.actual_payments:
            for p in body.actual_payments:
                pm = get_payment_method(conn, p.method_code)
                actual_wallets.append({
                    "method_code": p.method_code,
                    "method_name": pm["name"] if pm else p.method_code,
                    "method_type": pm["method_type"] if pm else "card",
                    "currency_code": (p.currency_code or "TJS").upper(),
                    "amount": float(p.amount or 0),
                })

        if actual_wallets:
            actual_cash = round(
                sum(
                    float(x["amount"])
                    for x in actual_wallets
                    if (x.get("method_type") or "") == "cash"
                    and (x.get("currency_code") or "") == "TJS"
                ),
                2,
            )
            actual_card = round(
                sum(
                    float(x["amount"])
                    for x in actual_wallets
                    if (x.get("method_type") or "") != "cash"
                ),
                2,
            )
        else:
            actual_cash = float(body.actual_cash or 0)
            actual_card = float(body.actual_card or 0)
            if actual_cash > 0:
                cash_code = cash_wallet_code_for_currency(conn, "TJS") or "cash"
                pm = get_payment_method(conn, cash_code)
                actual_wallets = [{
                    "method_code": cash_code,
                    "method_name": pm["name"] if pm else "Наличные",
                    "method_type": "cash",
                    "currency_code": "TJS",
                    "amount": actual_cash,
                }]

        exp_map = {
            (e["method_code"], e["currency_code"]): e
            for e in expected_wallets
        }
        wallet_diffs = []
        for aw in actual_wallets:
            key = (aw["method_code"], aw["currency_code"])
            exp = exp_map.get(key) or {}
            expected_amt = float(exp.get("expected") or 0)
            wallet_diffs.append({
                **aw,
                "expected": expected_amt,
                "difference": round(float(aw["amount"]) - expected_amt, 2),
            })
        # Missing expected wallets with no actual entered
        seen = {(d["method_code"], d["currency_code"]) for d in wallet_diffs}
        for exp in expected_wallets:
            key = (exp["method_code"], exp["currency_code"])
            if key in seen:
                continue
            if abs(float(exp.get("expected") or 0)) < 0.001:
                continue
            wallet_diffs.append({
                "method_code": exp["method_code"],
                "method_name": exp["method_name"],
                "method_type": exp["method_type"],
                "currency_code": exp["currency_code"],
                "amount": 0.0,
                "expected": float(exp["expected"]),
                "difference": round(0.0 - float(exp["expected"]), 2),
            })

        # Цель пересчёта: факт с формы, иначе ожидаемое (старт + движение)
        reconcile_targets = actual_wallets or [
            {
                "method_code": e["method_code"],
                "method_name": e["method_name"],
                "method_type": e["method_type"],
                "currency_code": e["currency_code"],
                "amount": float(e.get("expected") or 0),
            }
            for e in expected_wallets
            if abs(float(e.get("expected") or 0)) >= 0.01
            or abs(float(e.get("opening") or 0)) >= 0.01
        ]
        cash_posts = apply_shift_close_to_cash(
            conn,
            shift_id=shift_id,
            targets=reconcile_targets,
            user_name=(shift["user_name"] or "") if "user_name" in shift.keys() else "",
        )

        conn.execute(
            """
            UPDATE shifts SET
                closed_at = ?, status = 'closed',
                expected_cash = ?, expected_card = ?,
                actual_cash = ?, actual_card = ?,
                sales_count = ?, notes = ?,
                expected_payments_json = ?, actual_payments_json = ?
            WHERE id = ?
            """,
            (
                utc_now(),
                totals["expected_cash"],
                totals["expected_card"],
                actual_cash,
                actual_card,
                totals["sales_count"],
                body.notes,
                json.dumps(expected_wallets, ensure_ascii=False),
                json.dumps(actual_wallets, ensure_ascii=False),
                shift_id,
            ),
        )
        row = conn.execute("SELECT * FROM shifts WHERE id = ?", (shift_id,)).fetchone()
    result = row_to_dict(row) or {}
    result["expected_cash_in_drawer"] = expected_cash_in_drawer
    result["cash_difference"] = actual_cash - expected_cash_in_drawer
    result["card_difference"] = actual_card - totals["expected_card"]
    result["wallet_differences"] = wallet_diffs
    result["expected_wallets"] = expected_wallets
    result["actual_wallets"] = actual_wallets
    result["cash_ledger_posts"] = cash_posts
    result["by_payment"] = totals["by_payment"]
    return result


@app.get("/api/shifts")
async def list_shifts(limit: int = Query(default=30, ge=1, le=200), x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM shifts ORDER BY opened_at DESC LIMIT ?", (limit,)
        ).fetchall()
    return [row_to_dict(r) for r in rows]


# --- Product units (IMEI) ---


@app.get("/api/units")
async def list_units(
    product_id: int | None = None,
    warehouse_id: int | None = None,
    status: str = "in_stock",
    q: str = "",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        expire_reservations(conn)
        sql = """
        SELECT u.*, p.name AS product_name, p.model, p.color AS product_color, w.name AS warehouse_name
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        JOIN warehouses w ON w.id = u.warehouse_id
        WHERE 1=1
    """
        params: list[Any] = []
        if product_id:
            sql += " AND u.product_id = ?"
            params.append(product_id)
        if warehouse_id:
            sql += " AND u.warehouse_id = ?"
            params.append(warehouse_id)
        if status:
            sql += " AND u.status = ?"
            params.append(status)
        if q:
            uclause, uparams = unit_search_sql(q)
            sql += uclause
            params.extend(uparams)
            like = f"%{q.strip()}%"
            sql += " AND (p.name LIKE ? OR p.model LIKE ? OR p.color LIKE ?)"
            params.extend([like, like, like])
        sql += " ORDER BY u.created_at DESC LIMIT 500"
        rows = conn.execute(sql, params).fetchall()
    return [enrich_unit_row(r) for r in rows]




@app.post("/api/units/bulk")
async def bulk_create_units(body: BulkUnitsIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    prefix = body.serial_prefix.strip() or f"P{body.product_id}"
    created = []
    with db() as conn:
        product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="Товар не найден")
        wh = resolve_warehouse_id(conn, body.warehouse_id)
        existing = conn.execute(
            "SELECT COUNT(*) FROM product_units WHERE product_id = ? AND warehouse_id = ?",
            (body.product_id, wh),
        ).fetchone()[0]
        now = utc_now()
        customs_status = "pending" if body.mark_pending_customs else "none"
        for i in range(body.quantity):
            serial = f"{prefix}-{(existing + i + 1):03d}"
            cur = conn.execute(
                """
                INSERT INTO product_units
                (product_id, warehouse_id, imei, serial, status, notes, created_at, customs_status,
                 purchase_price)
                VALUES (?, ?, '', ?, 'in_stock', ?, ?, ?, ?)
                """,
                (
                    body.product_id, wh, serial, body.notes or "Партия без IMEI", now, customs_status,
                    float(body.purchase_price) if body.purchase_price is not None and body.purchase_price > 0
                    else float(product["purchase_price"] or 0),
                ),
            )
            adjust_warehouse_stock(
                conn, wh, body.product_id, 1, "inbound",
                notes=f"Партия: {serial}",
            )
            row = conn.execute("SELECT * FROM product_units WHERE id = ?", (cur.lastrowid,)).fetchone()
            created.append(enrich_unit_row(row))
        conn.execute("UPDATE products SET track_units = 1 WHERE id = ?", (body.product_id,))
    return {"created": len(created), "units": created}


@app.get("/api/units/pending-imei")
async def units_pending_imei(
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        sql = """
            SELECT u.*, p.name AS product_name, p.model, p.color AS product_color,
                   w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE u.status = 'in_stock' AND TRIM(COALESCE(u.imei, '')) = ''
        """
        params: list[Any] = []
        if warehouse_id:
            sql += " AND u.warehouse_id = ?"
            params.append(warehouse_id)
        sql += " ORDER BY u.created_at DESC LIMIT 500"
        rows = conn.execute(sql, params).fetchall()
    return [enrich_unit_row(r) for r in rows]


@app.get("/api/reports/imei-pending")
async def report_imei_pending(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        rows = conn.execute(
            """
            SELECT siu.*, si.product_name, s.id AS sale_id, s.created_at AS sale_date,
                   s.user_name AS cashier_name
            FROM sale_item_units siu
            JOIN sale_items si ON si.id = siu.sale_item_id
            JOIN sales s ON s.id = si.sale_id
            WHERE siu.imei_pending = 1 AND s.status = 'completed'
            ORDER BY s.created_at DESC
            LIMIT 200
            """
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/units/lookup")
async def lookup_unit(
    q: str = Query(min_length=1),
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        expire_reservations(conn)
        sql = """
            SELECT u.*, p.name AS product_name, p.model, p.color AS product_color,
                   p.barcode, p.sale_price, p.track_units, p.category,
                   w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE u.status IN ('in_stock', 'reserved')
        """
        params: list[Any] = []
        uclause, uparams = unit_search_sql(q)
        sql += uclause
        params.extend(uparams)
        # IMEI / digit fragment: search all warehouses (last 5 digits etc.)
        # Text/name search: keep selected warehouse filter.
        if warehouse_id and not is_code_query(q):
            sql += " AND u.warehouse_id = ?"
            params.append(warehouse_id)
        order = " ORDER BY CASE WHEN u.warehouse_id = ? THEN 0 ELSE 1 END, u.created_at DESC LIMIT 20"
        order_params = [warehouse_id or 0]
        units = conn.execute(sql + order, params + order_params).fetchall()
        if units:
            matches = []
            for u in units:
                d = row_to_dict(u)
                d["is_reserved"] = u["status"] == "reserved"
                matches.append(d)
            return {"match_type": "unit", "matches": matches}
        exact = conn.execute("SELECT * FROM products WHERE barcode = ?", (q.strip(),)).fetchone()
        if exact:
            return {"match_type": "product", "matches": [enrich_product(conn, exact)]}
        clause, sparams = product_search_sql(q)
        prows = conn.execute(f"SELECT p.* FROM products p WHERE 1=1 {clause} LIMIT 20", sparams).fetchall()
        if prows:
            return {"match_type": "product", "matches": [enrich_product(conn, r) for r in prows]}
    return {"match_type": "none", "matches": []}


def _unit_ops_count(conn: sqlite3.Connection, unit_id: int) -> int:
    sales = conn.execute(
        "SELECT COUNT(*) FROM sale_item_units WHERE unit_id = ?", (unit_id,)
    ).fetchone()[0]
    reserves = conn.execute(
        "SELECT COUNT(*) FROM unit_reservations WHERE unit_id = ?", (unit_id,)
    ).fetchone()[0]
    # inbound/create always counts as 1 operation
    return 1 + int(sales or 0) + int(reserves or 0)


def _unit_history_events(conn: sqlite3.Connection, unit_id: int) -> list[dict[str, Any]]:
    unit = conn.execute(
        """
        SELECT u.*, p.name AS product_name, p.model, p.color AS product_color,
               p.memory, p.sale_price, w.name AS warehouse_name
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        JOIN warehouses w ON w.id = u.warehouse_id
        WHERE u.id = ?
        """,
        (unit_id,),
    ).fetchone()
    if not unit:
        return []
    events: list[dict[str, Any]] = []
    inbound_at = (unit["arrival_date"] or "").strip() or (unit["created_at"] or "")
    if inbound_at and len(inbound_at) == 10:
        inbound_at = f"{inbound_at} 00:00:00"
    events.append({
        "kind": "inbound",
        "title": "Приход на склад",
        "at": inbound_at,
        "user_name": "",
        "amount": float(unit["purchase_price"] or 0),
        "currency_code": get_warehouse_currency(conn, unit["warehouse_id"])["code"],
        "notes": unit["notes"] or "",
        "warehouse_name": unit["warehouse_name"],
        "sale_id": None,
        "status": "ok",
    })
    for r in conn.execute(
        """
        SELECT r.*, w.name AS warehouse_name
        FROM unit_reservations r
        LEFT JOIN warehouses w ON w.id = r.warehouse_id
        WHERE r.unit_id = ?
        ORDER BY r.created_at
        """,
        (unit_id,),
    ).fetchall():
        events.append({
            "kind": "reservation",
            "title": f"Резерв ({r['status']})",
            "at": r["created_at"] or "",
            "user_name": r["user_name"] or r["client_name"] or "",
            "amount": 0.0,
            "currency_code": "",
            "notes": (r["notes"] or "") + (f" · клиент {r['client_name']}" if r["client_name"] else ""),
            "warehouse_name": r["warehouse_name"] or "",
            "sale_id": None,
            "status": r["status"],
        })
    for r in conn.execute(
        """
        SELECT s.id AS sale_id, s.created_at, s.user_name, s.status AS sale_status,
               s.currency_code, s.total AS sale_total, s.notes AS sale_notes,
               si.unit_price, si.subtotal, si.shop_profit, si.product_name,
               si.quantity, w.name AS warehouse_name
        FROM sale_item_units siu
        JOIN sale_items si ON si.id = siu.sale_item_id
        JOIN sales s ON s.id = si.sale_id
        LEFT JOIN warehouses w ON w.id = s.warehouse_id
        WHERE siu.unit_id = ?
        ORDER BY s.created_at
        """,
        (unit_id,),
    ).fetchall():
        qty = max(1, int(r["quantity"] or 1))
        amount = float(r["unit_price"] or 0) if qty == 1 else float(r["subtotal"] or 0) / qty
        title = "Продажа" if r["sale_status"] == "completed" else f"Продажа ({r['sale_status']})"
        events.append({
            "kind": "sale",
            "title": title,
            "at": r["created_at"] or "",
            "user_name": r["user_name"] or "",
            "amount": round(amount, 2),
            "currency_code": (r["currency_code"] or "TJS").upper(),
            "notes": r["sale_notes"] or "",
            "warehouse_name": r["warehouse_name"] or "",
            "sale_id": r["sale_id"],
            "status": r["sale_status"],
            "profit": float(r["shop_profit"] or 0) / qty if qty else float(r["shop_profit"] or 0),
            "product_name": r["product_name"] or "",
        })
    events.sort(key=lambda e: e.get("at") or "")
    return events


@app.get("/api/units/imei-history")
async def search_units_imei_history(
    q: str = Query(min_length=1),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    """Smart IMEI search for dashboard: any status, with ops count and sold flag."""
    check_pin(x_pin)
    q = normalize_search_q(q)
    if len(only_digits(q)) < 5 and not re.search(r"[A-Za-zА-Яа-я]", q):
        return {"matches": []}
    with db() as conn:
        expire_reservations(conn)
        sql = """
            SELECT u.*, p.name AS product_name, p.model, p.color AS product_color,
                   p.memory, p.sale_price, w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE 1=1
        """
        uclause, uparams = unit_search_sql(q)
        if not uclause:
            # also allow short text model search via digits-only fail — require code or text
            like = f"%{q}%"
            sql += " AND (u.imei LIKE ? OR u.serial LIKE ? OR LOWER(p.name) LIKE LOWER(?) OR LOWER(p.model) LIKE LOWER(?))"
            uparams = [like, like, like, like]
        else:
            sql += uclause
        rows = conn.execute(
            sql + " ORDER BY CASE u.status WHEN 'in_stock' THEN 0 WHEN 'reserved' THEN 1 ELSE 2 END, u.created_at DESC LIMIT 30",
            uparams,
        ).fetchall()
        matches = []
        for u in rows:
            d = enrich_unit_row(u)
            ops = _unit_ops_count(conn, u["id"])
            sales_n = conn.execute(
                "SELECT COUNT(*) FROM sale_item_units siu JOIN sale_items si ON si.id = siu.sale_item_id JOIN sales s ON s.id = si.sale_id WHERE siu.unit_id = ? AND s.status = 'completed'",
                (u["id"],),
            ).fetchone()[0]
            last_sale = conn.execute(
                """
                SELECT s.id, s.created_at, s.user_name, si.unit_price, s.currency_code
                FROM sale_item_units siu
                JOIN sale_items si ON si.id = siu.sale_item_id
                JOIN sales s ON s.id = si.sale_id
                WHERE siu.unit_id = ? AND s.status = 'completed'
                ORDER BY s.created_at DESC LIMIT 1
                """,
                (u["id"],),
            ).fetchone()
            d["ops_count"] = ops
            d["sales_count"] = int(sales_n or 0)
            d["is_sold"] = u["status"] == "sold" or int(sales_n or 0) > 0
            d["last_sale"] = row_to_dict(last_sale) if last_sale else None
            matches.append(d)
    return {"matches": matches}


@app.get("/api/units/{unit_id}/history")
async def get_unit_history(unit_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        expire_reservations(conn)
        row = conn.execute(
            """
            SELECT u.*, p.name AS product_name, p.model, p.color AS product_color,
                   p.memory, p.sale_price, w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE u.id = ?
            """,
            (unit_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
        unit = enrich_unit_row(row)
        events = _unit_history_events(conn, unit_id)
        unit["ops_count"] = _unit_ops_count(conn, unit_id)
        unit["is_sold"] = unit.get("status") == "sold" or any(
            e["kind"] == "sale" and e.get("status") == "completed" for e in events
        )
    return {"unit": unit, "events": events}


@app.post("/api/units/{unit_id}/complete-imei")
async def complete_unit_imei(
    unit_id: int, body: CompleteImeiIn, x_pin: str | None = Header(default=None, alias="X-Pin")
):
    check_pin(x_pin, min_role="warehouse")
    imei = body.imei.strip()
    with db() as conn:
        unit = conn.execute("SELECT * FROM product_units WHERE id = ?", (unit_id,)).fetchone()
        if not unit:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
        pending = conn.execute(
            """
            SELECT siu.* FROM sale_item_units siu
            WHERE siu.unit_id = ? AND siu.imei_pending = 1
            ORDER BY siu.sale_item_id DESC LIMIT 1
            """,
            (unit_id,),
        ).fetchone()
        if not pending and unit_has_imei(unit):
            raise HTTPException(status_code=400, detail="IMEI уже указан")
        dup = conn.execute(
            "SELECT id FROM product_units WHERE imei = ? AND id != ?", (imei, unit_id)
        ).fetchone()
        if dup:
            raise HTTPException(status_code=400, detail="IMEI уже в системе")
        conn.execute("UPDATE product_units SET imei = ? WHERE id = ?", (imei, unit_id))
        if pending:
            conn.execute(
                """
                UPDATE sale_item_units SET imei = ?, imei_pending = 0
                WHERE sale_item_id = ? AND unit_id = ?
                """,
                (imei, pending["sale_item_id"], unit_id),
            )
        row = conn.execute("SELECT * FROM product_units WHERE id = ?", (unit_id,)).fetchone()
    return enrich_unit_row(row)


@app.get("/api/reports/customs")
async def report_customs(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    clause, params, label = _report_period_clause(period, date_from, date_to, "s.created_at")
    wh_clause, wh_params = warehouse_sales_filter(warehouse_id)
    with db() as conn:
        total = conn.execute(
            f"""
            SELECT COALESCE(SUM(siu.customs_price), 0) AS total, COUNT(*) AS cnt
            FROM sale_item_units siu
            JOIN sale_items si ON si.id = siu.sale_item_id
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status = 'completed' AND siu.customs_cleared = 1 {clause}{wh_clause}
            """,
            params + wh_params,
        ).fetchone()
        rows = conn.execute(
            f"""
            SELECT s.id AS sale_id, s.created_at, si.product_name, siu.imei, siu.serial,
                   siu.customs_price, s.user_name
            FROM sale_item_units siu
            JOIN sale_items si ON si.id = siu.sale_item_id
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status = 'completed' AND siu.customs_cleared = 1 {clause}{wh_clause}
            ORDER BY s.created_at DESC LIMIT 100
            """,
            params + wh_params,
        ).fetchall()
        wh_name = ""
        if warehouse_id:
            row = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
            wh_name = row["name"] if row else f"#{warehouse_id}"
            label = f"{label} · {wh_name}"
    return {
        "period_label": label,
        "warehouse_id": warehouse_id,
        "warehouse_name": wh_name,
        "total_customs": float(total["total"]),
        "units_count": int(total["cnt"]),
        "items": [row_to_dict(r) for r in rows],
    }

@app.post("/api/units")
async def create_unit(body: UnitIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    if not body.imei.strip() and not body.serial.strip():
        raise HTTPException(status_code=400, detail="Укажите IMEI или серийный номер")
    with db() as conn:
        product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="Товар не найден")
        wh = resolve_warehouse_id(conn, body.warehouse_id)
        imei = body.imei.strip()
        serial = body.serial.strip()
        cs = body.customs_status
        if cs == "cleared":
            customs_cleared, customs_price = 1, 0.0
        else:
            customs_cleared, customs_price = 0, 0.0
        if imei:
            dup = conn.execute(
                "SELECT id FROM product_units WHERE imei = ? AND status != 'sold'", (imei,)
            ).fetchone()
            if dup:
                raise HTTPException(status_code=400, detail="IMEI уже в системе")
        target_pid = body.product_id
        unit_note = body.notes
        cur = conn.execute(
            """
            INSERT INTO product_units
            (product_id, warehouse_id, imei, serial, status, notes, created_at,
             customs_status, customs_cleared, customs_price, purchase_price)
            VALUES (?, ?, ?, ?, 'in_stock', ?, ?, ?, ?, ?, ?)
            """,
            (
                target_pid, wh, imei, serial, unit_note, utc_now(), cs, customs_cleared, customs_price,
                float(body.purchase_price) if body.purchase_price is not None and body.purchase_price > 0
                else float(product["purchase_price"] or 0),
            ),
        )
        adjust_warehouse_stock(
            conn, wh, target_pid, 1, "inbound",
            notes=f"IMEI: {imei or serial}",
        )
        conn.execute("UPDATE products SET track_units = 1 WHERE id = ?", (target_pid,))
        row = conn.execute(
            """
            SELECT u.*, p.name AS product_name, w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE u.id = ?
            """,
            (cur.lastrowid,),
        ).fetchone()
    return row_to_dict(row)


@app.delete("/api/units/{unit_id}")
async def delete_unit(unit_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        unit = conn.execute(
            "SELECT * FROM product_units WHERE id = ? AND status = 'in_stock'", (unit_id,)
        ).fetchone()
        if not unit:
            raise HTTPException(status_code=404, detail="Единица не найдена или уже продана")
        conn.execute("DELETE FROM product_units WHERE id = ?", (unit_id,))
        adjust_warehouse_stock(
            conn, unit["warehouse_id"], unit["product_id"], -1,
            "outbound", notes=f"Удаление IMEI #{unit_id}",
        )
    return {"ok": True}


def _remove_unit_photos(unit_id: int) -> None:
    for path in UPLOADS_DIR.glob(f"unit_{unit_id}.*"):
        path.unlink(missing_ok=True)


@app.get("/api/units/{unit_id}")
async def get_unit(unit_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        expire_reservations(conn)
        row = conn.execute(
            """
            SELECT u.*, p.name AS product_name, p.model, p.color AS product_color,
                   w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE u.id = ?
            """,
            (unit_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
        res = conn.execute(
            """
            SELECT * FROM unit_reservations
            WHERE unit_id = ? AND status = 'active'
            ORDER BY id DESC LIMIT 1
            """,
            (unit_id,),
        ).fetchone()
    data = enrich_unit_row(row)
    if res:
        data["reservation"] = row_to_dict(res)
    return data


@app.get("/api/units/{unit_id}/label")
async def unit_label(unit_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        row = conn.execute(
            """
            SELECT u.*, p.name AS product_name, p.model, p.color AS product_color
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            WHERE u.id = ?
            """,
            (unit_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
    qr_data = f"UNIT:{unit_id}|{row['serial'] or ''}|{row['imei'] or ''}"
    return {
        "unit_id": unit_id,
        "serial": row["serial"] or "",
        "imei": row["imei"] or "",
        "product_name": row["product_name"],
        "color": row["product_color"] or "",
        "model": row["model"] or "",
        "qr_data": qr_data,
    }


@app.patch("/api/units/{unit_id}/customs-status")
async def update_unit_customs_status(
    unit_id: int, body: CustomsStatusIn, x_pin: str | None = Header(default=None, alias="X-Pin")
):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        unit = conn.execute("SELECT * FROM product_units WHERE id = ?", (unit_id,)).fetchone()
        if not unit:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
        if unit["status"] == "sold":
            raise HTTPException(status_code=400, detail="Устройство уже продано")
        cleared = 1 if body.customs_status == "cleared" else 0
        conn.execute(
            "UPDATE product_units SET customs_status = ?, customs_cleared = ? WHERE id = ?",
            (body.customs_status, cleared, unit_id),
        )
        row = conn.execute(
            """
            SELECT u.*, p.name AS product_name, p.color AS product_color, w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE u.id = ?
            """,
            (unit_id,),
        ).fetchone()
    return enrich_unit_row(row)


class UnitPurchaseIn(BaseModel):
    purchase_price: float = Field(ge=0)


class UnitExtraCostIn(BaseModel):
    extra_cost: float = Field(ge=0)


def _unit_detail_row(conn: sqlite3.Connection, unit_id: int) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT u.*, p.name AS product_name, p.color AS product_color,
               p.purchase_price AS product_purchase_price, w.name AS warehouse_name
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        JOIN warehouses w ON w.id = u.warehouse_id
        WHERE u.id = ?
        """,
        (unit_id,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Устройство не найдено")
    d = enrich_unit_row(row)
    d["purchase_price"] = unit_purchase_price(row, {"purchase_price": row["product_purchase_price"]})
    d["extra_cost"] = unit_extra_cost(row)
    d["customs_price"] = d["extra_cost"]
    return d


@app.patch("/api/units/{unit_id}/purchase-price")
async def update_unit_purchase_price(
    unit_id: int, body: UnitPurchaseIn, x_pin: str | None = Header(default=None, alias="X-Pin")
):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        unit = conn.execute("SELECT * FROM product_units WHERE id = ?", (unit_id,)).fetchone()
        if not unit:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
        if unit["status"] == "sold":
            raise HTTPException(status_code=400, detail="Устройство уже продано — себестоимость в чеке не меняется")
        conn.execute(
            "UPDATE product_units SET purchase_price = ? WHERE id = ?",
            (body.purchase_price, unit_id),
        )
        return _unit_detail_row(conn, unit_id)


@app.patch("/api/units/{unit_id}/extra-cost")
async def update_unit_extra_cost(
    unit_id: int, body: UnitExtraCostIn, x_pin: str | None = Header(default=None, alias="X-Pin")
):
    """Расходы на Б/У (ремонт и т.п.) — в карточке устройства, колонка «Расходы»."""
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        unit = conn.execute("SELECT * FROM product_units WHERE id = ?", (unit_id,)).fetchone()
        if not unit:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
        if unit["status"] == "sold":
            raise HTTPException(status_code=400, detail="Устройство уже продано — расходы не меняются")
        conn.execute(
            "UPDATE product_units SET customs_price = ? WHERE id = ?",
            (float(body.extra_cost or 0), unit_id),
        )
        return _unit_detail_row(conn, unit_id)


@app.post("/api/units/{unit_id}/photo")
async def upload_unit_photo(
    unit_id: int,
    file: UploadFile = File(...),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        raise HTTPException(status_code=400, detail="Допустимы JPG, PNG, WEBP, GIF")
    data = await file.read()
    if len(data) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Максимальный размер файла — 5 МБ")
    with db() as conn:
        unit = conn.execute("SELECT id FROM product_units WHERE id = ?", (unit_id,)).fetchone()
        if not unit:
            raise HTTPException(status_code=404, detail="Устройство не найдено")
    _remove_unit_photos(unit_id)
    filename = f"unit_{unit_id}{ext}"
    (UPLOADS_DIR / filename).write_bytes(data)
    image_url = f"/uploads/{filename}"
    with db() as conn:
        conn.execute("UPDATE product_units SET box_image_url = ? WHERE id = ?", (image_url, unit_id))
        row = conn.execute(
            """
            SELECT u.*, p.name AS product_name, p.color AS product_color, w.name AS warehouse_name
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            JOIN warehouses w ON w.id = u.warehouse_id
            WHERE u.id = ?
            """,
            (unit_id,),
        ).fetchone()
    return enrich_unit_row(row)


@app.post("/api/units/import-csv")
async def import_units_csv(
    file: UploadFile = File(...),
    warehouse_id: int | None = Query(default=None),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Пустой файл")
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV без заголовков")
    fields = {f.strip().lower(): f for f in reader.fieldnames}
    created = []
    errors: list[str] = []
    with db() as conn:
        default_wh = resolve_warehouse_id(conn, warehouse_id)
        for i, row in enumerate(reader, start=2):
            def cell(*names: str) -> str:
                for n in names:
                    key = fields.get(n.lower())
                    if key and row.get(key):
                        return str(row[key]).strip()
                return ""

            product_id_raw = cell("product_id", "id")
            product_name = cell("product_name", "name", "product", "товар")
            imei = cell("imei", "imei1")
            serial = cell("serial", "serial_number", "серийник")
            wh_raw = cell("warehouse_id", "warehouse", "склад")
            cs_raw = cell("customs_status", "таможня").lower()
            customs_status = "pending" if cs_raw in ("pending", "1", "да", "yes", "на растamожке", "на растаможке") else "none"

            if not imei and not serial:
                errors.append(f"Строка {i}: нужен IMEI или серийник")
                continue

            pid: int | None = None
            if product_id_raw.isdigit():
                pid = int(product_id_raw)
            elif product_name:
                prod = conn.execute(
                    "SELECT id FROM products WHERE name = ? OR sku = ? LIMIT 1",
                    (product_name, product_name),
                ).fetchone()
                if prod:
                    pid = int(prod["id"])
            if not pid:
                errors.append(f"Строка {i}: товар не найден")
                continue

            wh_id = default_wh
            if wh_raw.isdigit():
                wh_id = resolve_warehouse_id(conn, int(wh_raw))

            product = conn.execute("SELECT * FROM products WHERE id = ?", (pid,)).fetchone()
            if not product:
                errors.append(f"Строка {i}: товар #{pid} не найден")
                continue

            if imei:
                dup = conn.execute(
                    "SELECT id FROM product_units WHERE imei = ? AND status != 'sold'", (imei,)
                ).fetchone()
                if dup:
                    errors.append(f"Строка {i}: IMEI {imei} уже есть")
                    continue

            cleared = 1 if customs_status == "cleared" else 0
            cur = conn.execute(
                """
                INSERT INTO product_units
                (product_id, warehouse_id, imei, serial, status, notes, created_at,
                 customs_status, customs_cleared, purchase_price)
                VALUES (?, ?, ?, ?, 'in_stock', ?, ?, ?, ?, ?)
                """,
                (
                    pid, wh_id, imei, serial, "CSV импорт", utc_now(), customs_status, cleared,
                    float(product["purchase_price"] or 0),
                ),
            )
            adjust_warehouse_stock(conn, wh_id, pid, 1, "inbound", notes=f"CSV: {imei or serial}")
            conn.execute("UPDATE products SET track_units = 1 WHERE id = ?", (pid,))
            unit_row = conn.execute("SELECT * FROM product_units WHERE id = ?", (cur.lastrowid,)).fetchone()
            created.append(enrich_unit_row(unit_row))

    return {"created": len(created), "errors": errors, "units": created[:50]}


@app.get("/api/reservations")
async def list_reservations(
    status: str = "active",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        expire_reservations(conn)
        sql = """
            SELECT r.*, u.imei, u.serial, u.box_image_url, u.customs_status,
                   p.name AS product_name, p.color AS product_color,
                   w.name AS warehouse_name
            FROM unit_reservations r
            JOIN product_units u ON u.id = r.unit_id
            JOIN products p ON p.id = r.product_id
            JOIN warehouses w ON w.id = r.warehouse_id
            WHERE 1=1
        """
        params: list[Any] = []
        if status:
            sql += " AND r.status = ?"
            params.append(status)
        sql += " ORDER BY r.reserved_until ASC LIMIT 200"
        rows = conn.execute(sql, params).fetchall()
    return [row_to_dict(r) for r in rows]


@app.post("/api/reservations")
async def create_reservation(body: ReservationIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    user = None
    with db() as conn:
        expire_reservations(conn)
        user = resolve_user(conn, x_pin)
        unit = conn.execute(
            "SELECT * FROM product_units WHERE id = ? AND status = 'in_stock'", (body.unit_id,)
        ).fetchone()
        if not unit:
            raise HTTPException(status_code=400, detail="Устройство недоступно для резерва")
        active = conn.execute(
            "SELECT id FROM unit_reservations WHERE unit_id = ? AND status = 'active'", (body.unit_id,)
        ).fetchone()
        if active:
            raise HTTPException(status_code=400, detail="Устройство уже зарезервировано")
        until = body.reserved_until.replace("T", " ")
        if len(until) == 10:
            until += " 23:59:59"
        cur = conn.execute(
            """
            INSERT INTO unit_reservations
            (unit_id, product_id, warehouse_id, client_name, client_phone, notes,
             reserved_until, user_id, user_name, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)
            """,
            (
                body.unit_id, unit["product_id"], unit["warehouse_id"],
                body.client_name.strip(), body.client_phone.strip(), body.notes.strip(),
                until, user.get("id") if user else None, user.get("name", "") if user else "",
                utc_now(),
            ),
        )
        conn.execute("UPDATE product_units SET status = 'reserved' WHERE id = ?", (body.unit_id,))
        row = conn.execute(
            """
            SELECT r.*, u.imei, u.serial, p.name AS product_name, p.color AS product_color,
                   w.name AS warehouse_name
            FROM unit_reservations r
            JOIN product_units u ON u.id = r.unit_id
            JOIN products p ON p.id = r.product_id
            JOIN warehouses w ON w.id = r.warehouse_id
            WHERE r.id = ?
            """,
            (cur.lastrowid,),
        ).fetchone()
    return row_to_dict(row)


@app.delete("/api/reservations/{reservation_id}")
async def cancel_reservation(reservation_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        res = conn.execute(
            "SELECT * FROM unit_reservations WHERE id = ? AND status = 'active'", (reservation_id,)
        ).fetchone()
        if not res:
            raise HTTPException(status_code=404, detail="Резерв не найден")
        conn.execute("UPDATE unit_reservations SET status = 'cancelled' WHERE id = ?", (reservation_id,))
        conn.execute(
            "UPDATE product_units SET status = 'in_stock' WHERE id = ? AND status = 'reserved'",
            (res["unit_id"],),
        )
    return {"ok": True}


@app.get("/api/products/{product_id}/units")
async def product_units(
    product_id: int,
    warehouse_id: int | None = None,
    status: str = "in_stock",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    return await list_units(
        product_id=product_id, warehouse_id=warehouse_id, status=status, q="", x_pin=x_pin
    )


# --- Warehouses ---


@app.get("/api/warehouses")
async def list_warehouses(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM warehouses ORDER BY is_default DESC, name"
        ).fetchall()
        return [
            row_to_dict(r) | {"currency": get_warehouse_currency(conn, int(r["id"]))}
            for r in rows
        ]


@app.post("/api/warehouses")
async def create_warehouse(body: WarehouseIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        nm = body.name.strip().lower()
        if "аксесс" in nm or "accessory" in nm:
            if conn.execute(
                f"SELECT 1 FROM warehouses WHERE {_accessories_warehouse_clause()} LIMIT 1"
            ).fetchone():
                raise HTTPException(
                    status_code=400,
                    detail="Склад аксессуаров уже есть — используйте раздел «Аксессуары»",
                )
        if body.is_default:
            conn.execute("UPDATE warehouses SET is_default = 0")
        wh_type = body.warehouse_type or "new"
        cur_code = (body.currency_code or "").strip().upper()
        if not cur_code:
            cur_code = "TJS" if wh_type == "used" else "USD"
        cur = conn.execute(
            """
            INSERT INTO warehouses (name, address, notes, is_default, warehouse_type, currency_code, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (body.name.strip(), body.address, body.notes, int(body.is_default), wh_type, cur_code, utc_now()),
        )
        row = conn.execute("SELECT * FROM warehouses WHERE id = ?", (cur.lastrowid,)).fetchone()
        wh_id = int(row["id"])
        currency = get_warehouse_currency(conn, wh_id)
    return row_to_dict(row) | {"currency": currency}


@app.put("/api/warehouses/{warehouse_id}")
async def update_warehouse(
    warehouse_id: int,
    body: WarehouseUpdate,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    fields = {k: v for k, v in body.model_dump().items() if v is not None}
    if not fields:
        raise HTTPException(status_code=400, detail="Нет данных для обновления")
    if "is_default" in fields:
        fields["is_default"] = int(fields["is_default"])
    with db() as conn:
        if not conn.execute("SELECT id FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Склад не найден")
        if fields.get("is_default"):
            conn.execute("UPDATE warehouses SET is_default = 0")
        sets = ", ".join(f"{k} = ?" for k in fields)
        conn.execute(f"UPDATE warehouses SET {sets} WHERE id = ?", (*fields.values(), warehouse_id))
        row = conn.execute("SELECT * FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        currency = get_warehouse_currency(conn, warehouse_id)
    return row_to_dict(row) | {"currency": currency}


@app.delete("/api/warehouses/{warehouse_id}")
async def delete_warehouse(warehouse_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        wh = conn.execute("SELECT * FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        if not wh:
            raise HTTPException(status_code=404, detail="Склад не найден")
        if wh["is_default"]:
            raise HTTPException(status_code=400, detail="Нельзя удалить склад по умолчанию")
        if get_warehouse_receipt_kind(conn, warehouse_id) == "accessories":
            raise HTTPException(
                status_code=400,
                detail="Склад аксессуаров удаляется только через раздел «Аксессуары»",
            )
        stock_count = conn.execute(
            "SELECT COUNT(*) FROM warehouse_stock WHERE warehouse_id = ? AND quantity > 0",
            (warehouse_id,),
        ).fetchone()[0]
        if stock_count > 0:
            raise HTTPException(status_code=400, detail="На складе есть остатки — сначала переместите товар")
        conn.execute("DELETE FROM warehouses WHERE id = ?", (warehouse_id,))
    return {"ok": True}


@app.get("/api/warehouses/{warehouse_id}/stock")
async def warehouse_stock(warehouse_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        if not conn.execute("SELECT id FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Склад не найден")
        rows = conn.execute(
            """
            SELECT p.*, ws.quantity AS warehouse_quantity
            FROM warehouse_stock ws
            JOIN products p ON p.id = ws.product_id
            WHERE ws.warehouse_id = ? AND ws.quantity > 0
            ORDER BY p.name
            """,
            (warehouse_id,),
        ).fetchall()
        result = []
        for r in rows:
            item = enrich_product(conn, r)
            item["warehouse_quantity"] = r["warehouse_quantity"]
            if int(r["track_units"] or 0):
                item["units"] = units_for_product_at_warehouse(conn, r["id"], warehouse_id)
            else:
                item["units"] = []
            result.append(item)
    return result


@app.get("/api/warehouses/{warehouse_id}/devices")
async def warehouse_devices(warehouse_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        if not conn.execute("SELECT id FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Склад не найден")
        kind = get_warehouse_receipt_kind(conn, warehouse_id)
        rows = conn.execute(
            """
            SELECT u.*, p.name AS model, p.color, p.memory,
                   p.purchase_price AS product_purchase_price,
                   p.supplier_name, p.sale_price, p.ownership_type, p.condition, p.category
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            WHERE u.warehouse_id = ? AND u.status = 'in_stock' AND p.category != 'accessory'
            ORDER BY COALESCE(u.arrival_date, u.created_at) DESC, u.id DESC
            """,
            (warehouse_id,),
        ).fetchall()
        items = []
        for r in rows:
            d = enrich_unit_row(r)
            d["model"] = r["model"] or ""
            d["color"] = r["color"] or ""
            d["memory"] = r["memory"] or ""
            d["purchase_price"] = unit_purchase_price(r, {"purchase_price": r["product_purchase_price"]})
            d["supplier_name"] = r["supplier_name"] or ""
            d["ownership_type"] = r["ownership_type"] or "own"
            items.append(d)
        acc = []
        accessories: list[dict[str, Any]] = []
        if kind == "accessories":
            acc = conn.execute(
                """
                SELECT p.*, ws.quantity AS warehouse_quantity
                FROM warehouse_stock ws
                JOIN products p ON p.id = ws.product_id
                WHERE ws.warehouse_id = ? AND ws.quantity > 0
                  AND p.category = 'accessory' AND IFNULL(p.track_units, 0) = 0
                ORDER BY p.name
                """,
                (warehouse_id,),
            ).fetchall()
            accessories = [
                enrich_product(conn, r) | {"warehouse_quantity": r["warehouse_quantity"]} for r in acc
            ]
        totals = warehouse_stock_money(conn, warehouse_id)
    return {"kind": kind, "items": items, "accessories": accessories, "totals": totals}


@app.get("/api/warehouses/{warehouse_id}/stock-value")
async def warehouse_stock_value(warehouse_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        resolve_warehouse_id(conn, warehouse_id)
        return warehouse_stock_money(conn, warehouse_id)


@app.get("/api/warehouses/{warehouse_id}/z-report")
async def warehouse_z_report(
    warehouse_id: int,
    period: str = Query(default="day", pattern="^(day|week|month|all|custom)$"),
    year: int | None = Query(default=None, ge=2020, le=2035),
    month: int | None = Query(default=None, ge=1, le=12),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        resolve_warehouse_id(conn, warehouse_id)
        if year and month:
            reg = _z_register_lines(conn, warehouse_id, year, month)
            currency = get_warehouse_currency(conn, warehouse_id)
            return {
                "warehouse_id": warehouse_id,
                "period": "custom",
                "period_label": reg["period_label"],
                "year": year,
                "month": month,
                "currency": currency,
                "sales_count": reg["sold_count"],
                "revenue": reg["revenue"],
                "discounts": 0.0,
                "profit": reg["profit"],
                "stock_count": reg["stock_count"],
                "stock_value": reg["stock_value"],
                "by_payment": [],
                "lines": reg["sold_lines"],
                "stock_lines": reg["stock_lines"],
                "all_lines": reg["all_lines"],
            }
        since = period_start(period) if period not in ("all", "custom") else "1970-01-01"
        agg = conn.execute(
            """
            SELECT COUNT(DISTINCT s.id) AS sales_count,
                   COALESCE(SUM(s.total), 0) AS revenue,
                   COALESCE(SUM(s.discount), 0) AS discounts
            FROM sales s
            WHERE s.status = 'completed' AND s.warehouse_id = ? AND s.created_at >= ?
            """,
            (warehouse_id, since),
        ).fetchone()
        profit_row = conn.execute(
            """
            SELECT COALESCE(SUM(si.shop_profit), 0)
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status = 'completed' AND s.warehouse_id = ? AND s.created_at >= ?
            """,
            (warehouse_id, since),
        ).fetchone()
        by_payment = conn.execute(
            """
            SELECT sp.method_code, pm.name, COALESCE(SUM(sp.amount), 0) AS amount
            FROM sale_payments sp
            JOIN sales s ON s.id = sp.sale_id
            LEFT JOIN payment_methods pm ON pm.code = sp.method_code
            WHERE s.status = 'completed' AND s.warehouse_id = ? AND s.created_at >= ?
            GROUP BY sp.method_code
            ORDER BY amount DESC
            """,
            (warehouse_id, since),
        ).fetchall()
        reg = _z_register_lines(conn, warehouse_id, None, None)
        currency = get_warehouse_currency(conn, warehouse_id)
        period_labels = {"day": "Сегодня", "week": "Неделя", "month": "Месяц", "all": "Всё время"}
    return {
        "warehouse_id": warehouse_id,
        "period": period,
        "period_label": period_labels.get(period, period),
        "currency": currency,
        "sales_count": int(agg["sales_count"] or 0),
        "revenue": float(agg["revenue"] or 0),
        "discounts": float(agg["discounts"] or 0),
        "profit": float(profit_row[0] or 0),
        "stock_count": reg["stock_count"],
        "stock_value": reg["stock_value"],
        "by_payment": [row_to_dict(r) for r in by_payment],
        "lines": reg["sold_lines"][:200],
        "stock_lines": reg["stock_lines"],
        "all_lines": reg["all_lines"],
    }


@app.post("/api/import/z-register")
async def import_z_register_file(
    file: UploadFile = File(...),
    sheet: str = Query(default=""),
    replace: bool = Query(default=False),
    affect_cash: bool = Query(default=True),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Пустой файл")
    try:
        with db() as conn:
            result = import_z_register_excel(
                conn, raw, file.filename or "import.xlsx", sheet,
                replace=replace, affect_cash=affect_cash,
            )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Z-register import failed")
        raise HTTPException(status_code=500, detail=f"Ошибка импорта: {exc}") from exc
    return result


@app.post("/api/warehouse/quick-sell")
async def warehouse_quick_sell(body: WarehouseQuickSellIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        unit = conn.execute(
            "SELECT * FROM product_units WHERE id = ? AND status = 'in_stock'", (body.unit_id,)
        ).fetchone()
        if not unit:
            raise HTTPException(status_code=404, detail="Устройство не найдено или уже продано")
        wh_id = int(unit["warehouse_id"])
        product_id = int(unit["product_id"])
    total = body.sale_price - body.discount
    if body.payments:
        pay_list = [{"method_code": p.method_code, "amount": p.amount} for p in body.payments]
    elif body.paid_amount is not None:
        pay_list = [{"method_code": body.payment_method, "amount": body.paid_amount}]
    else:
        pay_list = [{"method_code": body.payment_method, "amount": total}]
    sale_in = SaleIn(
        items=[CartItem(product_id=product_id, quantity=1, unit_ids=[body.unit_id], unit_price=body.sale_price)],
        discount=body.discount,
        payments=[PaymentPart(**p) for p in pay_list],
        notes=body.notes or "Продажа со склада",
        warehouse_id=wh_id,
        debtor_name=body.debtor_name,
        debtor_phone=body.debtor_phone,
    )
    return await create_sale(sale_in, x_pin)


def _amount_bags(rows: list[tuple[str, float]]) -> list[dict[str, Any]]:
    bags: dict[str, float] = {}
    for cur, amt in rows:
        code = (cur or "TJS").strip().upper() or "TJS"
        bags[code] = bags.get(code, 0.0) + float(amt or 0)
    out = []
    for code in sorted(bags.keys(), key=lambda c: (0 if c == "USD" else 1, c)):
        out.append({**currency_meta(code), "amount": round(bags[code], 2)})
    return out


def _shift_report_window(conn: sqlite3.Connection) -> tuple[str, str, str, dict[str, Any] | None]:
    """Return (period, date_from, date_to, shift_dict|None). Open shift → day of opened_at."""
    shift = get_open_shift(conn)
    if shift:
        day = (shift["opened_at"] or utc_now())[:10]
        return "day", day, day, row_to_dict(shift)
    # No open shift: calendar today
    day = utc_now()[:10]
    return "day", day, day, None


@app.get("/api/pos/cash-register")
async def pos_cash_register(
    period: str = Query(default="auto", pattern="^(auto|day|week|month|all)$"),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    """POS cash: period KPIs (shift day or today) + all-time wallet balances."""
    check_pin(x_pin)
    with db() as conn:
        shift_row = get_open_shift(conn)
        shift = row_to_dict(shift_row) if shift_row else None
        if period == "auto":
            _, use_from, use_to, _ = _shift_report_window(conn)
            period_key = "day"
        elif period == "all":
            use_from, use_to, period_key = "", "", "all"
        else:
            use_from, use_to, period_key = "", "", period

        cash = _period_cash_by_currency(
            conn, "all" if use_from else period_key,
            date_from=use_from, date_to=use_to,
        )
        if use_from and use_to and use_from == use_to:
            cash["period_label"] = (f"Смена · {use_from}" if shift else f"Сегодня · {use_from}")

        wallets = _period_cash_by_currency(conn, "all")  # all-time money in wallets
        fin = _finance_report(conn, "all" if use_from else period_key, "all", use_from, use_to)

        # Split: sales / manual inflows / expenses / payouts for the period
        sale_clause, sale_params, _ = _report_period_clause(
            "all" if use_from else period_key, use_from, use_to, "s.created_at"
        )
        sales_rows = conn.execute(
            f"""
            SELECT UPPER(COALESCE(
                       NULLIF(TRIM(sp.pay_currency_code), ''),
                       NULLIF(TRIM(s.currency_code), ''),
                       'TJS'
                   )) AS cur,
                   COALESCE(SUM(COALESCE(NULLIF(sp.pay_amount, 0), sp.amount)), 0) AS amt
            FROM sale_payments sp
            JOIN sales s ON s.id = sp.sale_id
            WHERE s.status = 'completed' AND COALESCE(s.affects_cash, 1) = 1 {sale_clause}
            GROUP BY 1
            """,
            list(sale_params),
        ).fetchall()
        sales_bags = _amount_bags([(r["cur"], r["amt"]) for r in sales_rows])

        inflows_bags: list[dict[str, Any]] = []
        recent_inflows: list[dict[str, Any]] = []
        if _table_exists(conn, "cash_inflows"):
            iclause, iparams, _ = _report_period_clause(
                "all" if use_from else period_key, use_from, use_to, "created_at"
            )
            in_rows = conn.execute(
                f"""
                SELECT UPPER(COALESCE(NULLIF(TRIM(currency_code), ''), 'TJS')) AS cur,
                       COALESCE(SUM(amount), 0) AS amt
                FROM cash_inflows
                WHERE source_type = 'counterparty' {iclause}
                GROUP BY 1
                """,
                list(iparams),
            ).fetchall()
            inflows_bags = _amount_bags([(r["cur"], r["amt"]) for r in in_rows])
            recent_inflows = [
                row_to_dict(r)
                for r in conn.execute(
                    f"SELECT * FROM cash_inflows WHERE 1=1 {iclause} ORDER BY created_at DESC LIMIT 30",
                    list(iparams),
                ).fetchall()
            ]

        exp_clause, exp_params, _ = _report_period_clause(
            "all" if use_from else period_key, use_from, use_to, "expense_date"
        )
        if period_key != "all" and not use_from and not exp_params:
            exp_clause = " AND expense_date >= ?"
            exp_params = [period_start(period_key)[:10]]
        exp_kind_rows = conn.execute(
            f"""
            SELECT COALESCE(kind, 'expense') AS kind,
                   COALESCE(SUM(amount), 0) AS amt
            FROM expenses WHERE COALESCE(affects_cash, 1) = 1 {exp_clause}
            GROUP BY 1
            """,
            list(exp_params),
        ).fetchall()
        exp_amt = sum(float(r["amt"]) for r in exp_kind_rows if (r["kind"] or "expense") != "payout")
        pay_amt = sum(float(r["amt"]) for r in exp_kind_rows if (r["kind"] or "") == "payout")
        expenses_bags = _amount_bags([("TJS", exp_amt)]) if exp_amt else []
        payouts_bags = _amount_bags([("TJS", pay_amt)]) if pay_amt else []

        exp_rows = conn.execute(
            f"""
            SELECT e.*, COALESCE(w.name, '') AS warehouse_name
            FROM expenses e
            LEFT JOIN warehouses w ON w.id = e.warehouse_id
            WHERE COALESCE(e.affects_cash, 1) = 1
              {exp_clause.replace('expense_date', 'e.expense_date') if exp_clause else ''}
            ORDER BY e.expense_date DESC, e.id DESC LIMIT 40
            """,
            list(exp_params),
        ).fetchall()

        by_cur = cash["by_currency"]
        wallets_all = wallets["balances"]
        balances_source = "all_time"
        balances_label = "Сейчас в кошельках (весь период)"
        drawer_balances = wallets_all
        if shift_row:
            expected = shift_expected_wallets(conn, shift_row)
            drawer_balances = balances_from_shift_expected(conn, expected)
            balances_source = "shift"
            balances_label = "Сейчас в кассе (старт + приход − расход)"
        till = till_summary_from_balances(drawer_balances)
        till_period = till_summary_from_balances(cash["balances"])
    return {
        "period": period_key,
        "date_from": use_from,
        "date_to": use_to,
        "period_label": cash["period_label"],
        "shift": shift,
        "by_currency": by_cur,
        "sales": sales_bags,
        "manual_inflows": inflows_bags,
        "expenses_bags": expenses_bags,
        "payouts_bags": payouts_bags,
        "total_inflows": None,
        "total_outflows": None,
        "net_cash": None,
        "revenue": fin["gross_revenue"],
        "profit": fin["shop_profit"],
        "by_currency_finance": fin.get("by_currency") or [],
        "expenses_total": exp_amt + pay_amt,
        "expenses": [row_to_dict(r) for r in exp_rows],
        # Period movement per wallet (shift/today)
        "balances_period": cash["balances"],
        # Drawer now: with open shift = opening + movement; else all-time ops
        "balances": drawer_balances,
        "balances_source": balances_source,
        "balances_label": balances_label,
        # All-time money from sales/inflows/expenses (без стартовых сумм смены)
        "wallets_all_time": wallets_all,
        "till": till,
        "till_period": till_period,
        "cash_inflows": recent_inflows,
    }


@app.get("/api/pos/cash-register/detail")
async def pos_cash_register_detail(
    kind: str = Query(default="inflow", pattern="^(inflow|outflow|net|wallet|profit)$"),
    method_code: str = Query(default=""),
    period: str = Query(default="auto", pattern="^(auto|day|week|month|all)$"),
    date_from: str = "",
    date_to: str = "",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    """Lines explaining cash KPI / wallet balances (where +/− came from)."""
    check_pin(x_pin)
    with db() as conn:
        if period == "auto" or (not date_from and not date_to and period == "day"):
            _, date_from, date_to, _ = _shift_report_window(conn)
            period = "all"  # use explicit dates
        period_label = (
            f"{date_from} — {date_to}" if date_from or date_to
            else ("Всё время" if period == "all" else {"day": "Сегодня", "week": "Неделя", "month": "Месяц"}.get(period, period))
        )
        sale_clause, sale_params, _ = _report_period_clause(period, date_from, date_to, "s.created_at")
        exp_clause, exp_params, _ = _report_period_clause(period, date_from, date_to, "expense_date")
        if period != "all" and not date_from and not date_to and not exp_params:
            exp_clause = " AND expense_date >= ?"
            exp_params = [period_start(period)[:10]]
        in_clause, in_params, _ = _report_period_clause(period, date_from, date_to, "created_at")
        method = (method_code or "").strip()
        lines: list[dict[str, Any]] = []

        def add_line(
            *,
            at: str,
            side: str,
            source: str,
            who: str,
            amount: float,
            currency: str,
            note: str = "",
            method_name: str = "",
            ref: str = "",
        ) -> None:
            lines.append({
                "at": at or "",
                "side": side,
                "source": source,
                "who": who or "—",
                "amount": float(amount or 0),
                "currency_code": (currency or "TJS").upper(),
                "note": note or "",
                "method_name": method_name or "",
                "ref": ref or "",
            })

        want_in = kind in ("inflow", "net", "wallet")
        want_out = kind in ("outflow", "net", "wallet")
        if kind == "profit":
            want_in = want_out = False

        if want_in:
            pay_sql = f"""
                SELECT s.created_at AS at, s.id AS sale_id,
                       COALESCE(NULLIF(sp.pay_amount, 0), sp.amount) AS amount,
                       UPPER(COALESCE(
                           NULLIF(TRIM(sp.pay_currency_code), ''),
                           NULLIF(TRIM(s.currency_code), ''),
                           'TJS'
                       )) AS currency_code,
                       sp.method_code, COALESCE(pm.name, sp.method_code) AS method_name,
                       COALESCE(NULLIF(s.user_name, ''), '—') AS who,
                       COALESCE(w.name, '') AS warehouse_name
                FROM sale_payments sp
                JOIN sales s ON s.id = sp.sale_id
                LEFT JOIN payment_methods pm ON pm.code = sp.method_code
                LEFT JOIN warehouses w ON w.id = s.warehouse_id
                WHERE s.status = 'completed' AND COALESCE(s.affects_cash, 1) = 1 {sale_clause}
                ORDER BY s.created_at DESC LIMIT 400
            """
            for r in conn.execute(pay_sql, list(sale_params)).fetchall():
                cur = r["currency_code"] or "TJS"
                wcode = resolve_wallet_method_code(conn, r["method_code"] or "cash", cur)
                if kind == "wallet" and method and wcode != method:
                    continue
                pm_show = get_payment_method(conn, wcode)
                add_line(
                    at=r["at"], side="+", source="Продажа",
                    who=r["who"], amount=r["amount"], currency=cur,
                    note=r["warehouse_name"] or "",
                    method_name=(pm_show["name"] if pm_show else r["method_name"]),
                    ref=f"#{r['sale_id']}",
                )

            recv_clause, recv_params, _ = _report_period_clause(period, "", "", "rp.created_at")
            recv_sql = f"""
                SELECT rp.created_at AS at,
                       COALESCE(NULLIF(rp.pay_amount, 0), rp.amount) AS amount,
                       rp.payment_method_code,
                       COALESCE(pm.name, rp.payment_method_code) AS method_name,
                       r.customer_name AS who, r.sale_id,
                       UPPER(COALESCE(
                           NULLIF(TRIM(rp.pay_currency_code), ''),
                           NULLIF(TRIM(s.currency_code), ''),
                           NULLIF(TRIM(w.currency_code), ''),
                           'TJS'
                       )) AS currency_code
                FROM receivable_payments rp
                JOIN receivables r ON r.id = rp.receivable_id
                LEFT JOIN sales s ON s.id = r.sale_id
                LEFT JOIN warehouses w ON w.id = COALESCE(r.warehouse_id, s.warehouse_id)
                LEFT JOIN payment_methods pm ON pm.code = rp.payment_method_code
                WHERE 1=1 {recv_clause}
            """
            recv_params = list(recv_params)
            if kind == "wallet" and method:
                recv_sql += " AND rp.payment_method_code = ?"
                recv_params.append(method)
            recv_sql += " ORDER BY rp.created_at DESC LIMIT 200"
            for r in conn.execute(recv_sql, recv_params).fetchall():
                add_line(
                    at=r["at"], side="+", source="Погашение долга",
                    who=r["who"], amount=r["amount"], currency=r["currency_code"],
                    note=f"чек #{r['sale_id']}" if r["sale_id"] else "",
                    method_name=r["method_name"],
                )

            if _table_exists(conn, "mutual_payments"):
                m_clause, m_params, _ = _report_period_clause(period, "", "", "mp.created_at")
                m_sql = f"""
                    SELECT mp.created_at AS at,
                           COALESCE(NULLIF(mp.pay_amount, 0), mp.amount) AS amount,
                           mp.payment_method_code,
                           COALESCE(pm.name, mp.payment_method_code) AS method_name,
                           me.person_name AS who,
                           me.product_note AS note,
                           UPPER(COALESCE(
                               NULLIF(TRIM(mp.pay_currency_code), ''),
                               NULLIF(TRIM(me.currency_code), ''),
                               'TJS'
                           )) AS currency_code
                    FROM mutual_payments mp
                    JOIN mutual_entries me ON me.id = mp.entry_id
                    LEFT JOIN payment_methods pm ON pm.code = mp.payment_method_code
                    WHERE me.direction = 'owe_us' {m_clause}
                """
                mparams = list(m_params)
                if kind == "wallet" and method:
                    m_sql += " AND mp.payment_method_code = ?"
                    mparams.append(method)
                m_sql += " ORDER BY mp.created_at DESC LIMIT 200"
                for r in conn.execute(m_sql, mparams).fetchall():
                    add_line(
                        at=r["at"], side="+", source="Погашение займа",
                        who=r["who"], amount=r["amount"], currency=r["currency_code"],
                        note=r["note"] or "", method_name=r["method_name"],
                    )

            if _table_exists(conn, "cash_inflows"):
                man_sql = f"""
                    SELECT created_at AS at, amount, amount_base, currency_code,
                           payment_method_code, counterparty_name, source_type, notes
                    FROM cash_inflows
                    WHERE source_type IN ('counterparty', 'shift_close', 'shift_opening') {in_clause}
                """
                man_params = list(in_params)
                if kind == "wallet" and method:
                    man_sql += " AND payment_method_code = ?"
                    man_params.append(method)
                man_sql += " ORDER BY created_at DESC LIMIT 200"
                for r in conn.execute(man_sql, man_params).fetchall():
                    pm = get_payment_method(conn, r["payment_method_code"] or "cash")
                    src = (r["source_type"] or "counterparty")
                    src_label = (
                        "Пересчёт смены" if src == "shift_close"
                        else "Старт смены" if src == "shift_opening"
                        else "Приход (контрагент)"
                    )
                    amt = float(r["amount"] or 0)
                    add_line(
                        at=r["at"],
                        side="+" if amt >= 0 else "−",
                        source=src_label,
                        who=r["counterparty_name"] or "—",
                        amount=abs(amt),
                        currency=r["currency_code"] or "TJS",
                        note=r["notes"] or "",
                        method_name=pm["name"] if pm else (r["payment_method_code"] or ""),
                    )

        if want_out:
            out_sql = f"""
                SELECT e.expense_date AS at, e.amount, e.category, e.description,
                       e.payment_method_code, COALESCE(pm.name, e.payment_method_code) AS method_name,
                       COALESCE(NULLIF(e.created_by, ''), '—') AS who
                FROM expenses e
                LEFT JOIN payment_methods pm ON pm.code = e.payment_method_code
                WHERE COALESCE(e.affects_cash, 1) = 1 {exp_clause}
            """
            out_params = list(exp_params)
            if kind == "wallet" and method:
                out_sql += " AND e.payment_method_code = ?"
                out_params.append(method)
            out_sql += " ORDER BY e.expense_date DESC, e.id DESC LIMIT 200"
            for r in conn.execute(out_sql, out_params).fetchall():
                add_line(
                    at=r["at"], side="−", source="Расход",
                    who=r["who"], amount=r["amount"], currency="TJS",
                    note=f"{r['category'] or ''} {r['description'] or ''}".strip(),
                    method_name=r["method_name"] or "",
                )

            if kind in ("outflow", "net") or (kind == "wallet" and method in ("", "cash")):
                # supplier payments (usually cash TJS)
                if _table_exists(conn, "supplier_payments"):
                    sup_clause, sup_params, _ = _report_period_clause(period, "", "", "created_at")
                    sup_sql = f"""
                        SELECT created_at AS at, amount, supplier_name, notes,
                               COALESCE(payment_method_code, 'cash') AS payment_method_code
                        FROM supplier_payments
                        WHERE 1=1 {sup_clause}
                    """
                    sup_params = list(sup_params)
                    if kind == "wallet" and method:
                        sup_sql += " AND COALESCE(payment_method_code, 'cash') = ?"
                        sup_params.append(method)
                    for r in conn.execute(sup_sql + " ORDER BY created_at DESC LIMIT 100", sup_params).fetchall():
                        pm = get_payment_method(conn, r["payment_method_code"] or "cash")
                        add_line(
                            at=r["at"], side="−", source="Оплата поставщику",
                            who=r["supplier_name"] or "—", amount=r["amount"], currency="TJS",
                            note=r["notes"] or "", method_name=pm["name"] if pm else "Наличные",
                        )

            if _table_exists(conn, "mutual_payments"):
                m_clause, m_params, _ = _report_period_clause(period, "", "", "mp.created_at")
                m_sql = f"""
                    SELECT mp.created_at AS at,
                           COALESCE(NULLIF(mp.pay_amount, 0), mp.amount) AS amount,
                           mp.payment_method_code,
                           COALESCE(pm.name, mp.payment_method_code) AS method_name,
                           me.person_name AS who,
                           me.product_note AS note,
                           UPPER(COALESCE(
                               NULLIF(TRIM(mp.pay_currency_code), ''),
                               NULLIF(TRIM(me.currency_code), ''),
                               'TJS'
                           )) AS currency_code
                    FROM mutual_payments mp
                    JOIN mutual_entries me ON me.id = mp.entry_id
                    LEFT JOIN payment_methods pm ON pm.code = mp.payment_method_code
                    WHERE me.direction = 'we_owe' {m_clause}
                """
                mparams = list(m_params)
                if kind == "wallet" and method:
                    m_sql += " AND mp.payment_method_code = ?"
                    mparams.append(method)
                m_sql += " ORDER BY mp.created_at DESC LIMIT 200"
                for r in conn.execute(m_sql, mparams).fetchall():
                    add_line(
                        at=r["at"], side="−", source="Оплата: мы должны",
                        who=r["who"], amount=r["amount"], currency=r["currency_code"],
                        note=r["note"] or "", method_name=r["method_name"],
                    )

        if kind == "profit":
            fin = _finance_report(conn, period, "all", "", "")
            title = f"Прибыль · {period_label}"
            return {
                "kind": kind,
                "title": title,
                "period_label": period_label,
                "summary": {
                    "revenue": fin.get("gross_revenue"),
                    "profit": fin.get("shop_profit"),
                    "sales_count": fin.get("sales_count"),
                },
                "lines": [],
                "by_currency": fin.get("by_currency") or [],
            }

        lines.sort(key=lambda x: x.get("at") or "", reverse=True)
        method_name = ""
        if method:
            pm = get_payment_method(conn, method)
            method_name = pm["name"] if pm else method
        titles = {
            "inflow": "Приход — откуда деньги",
            "outflow": "Расход — куда ушли",
            "net": "Чистыми — плюсы и минусы",
            "wallet": f"Кошелёк: {method_name or method}",
        }
        by_currency = _money_sides_by_currency(lines)
        return {
            "kind": kind,
            "title": f"{titles.get(kind, 'Детали')} · {period_label}",
            "period_label": period_label,
            "method_code": method,
            "method_name": method_name,
            "summary": {
                "by_currency": by_currency,
                "lines_count": len(lines),
                # Never mix USD+TJS into one number
                "plus": None,
                "minus": None,
                "net": None,
            },
            "by_currency": by_currency,
            "lines": lines,
        }


@app.get("/api/stock/total")
async def stock_total(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        rows = conn.execute(
            """
            SELECT p.id, p.name, p.sku, p.barcode, p.category, p.brand,
                   COALESCE(SUM(ws.quantity), 0) AS total_stock
            FROM products p
            LEFT JOIN warehouse_stock ws ON ws.product_id = p.id
            GROUP BY p.id
            HAVING total_stock > 0
            ORDER BY p.name
            """
        ).fetchall()
        enriched = []
        for r in rows:
            product = conn.execute("SELECT * FROM products WHERE id = ?", (r["id"],)).fetchone()
            if product:
                enriched.append(enrich_product(conn, product))
    return enriched


@app.get("/api/import/products/template")
async def import_products_template(
    warehouse_id: int | None = Query(default=None),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        kind = get_warehouse_receipt_kind(conn, warehouse_id) if warehouse_id else "new"
    data = build_products_import_xlsx(kind)
    fname = "telestore_used_import.xlsx" if kind == "used" else "telestore_new_import.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.post("/api/import/products")
async def import_products_file(
    file: UploadFile = File(...),
    warehouse_id: int | None = Query(default=None),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Пустой файл")
    rows = _parse_import_file(raw, file.filename or "import.csv")
    with db() as conn:
        wh_id = resolve_warehouse_id(conn, warehouse_id)
        result = _import_products_rows(conn, rows, wh_id)
    return result


@app.get("/api/import/sales/template")
async def import_sales_template(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="cashier")
    data = build_sales_import_xlsx()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="telestore_sales_import.xlsx"'},
    )


@app.post("/api/import/sales")
async def import_sales_file(
    file: UploadFile = File(...),
    warehouse_id: int | None = Query(default=None),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="cashier")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Пустой файл")
    rows = _parse_import_file(raw, file.filename or "import.csv", sheet="Продажи")
    with db() as conn:
        wh_id = resolve_warehouse_id(conn, warehouse_id)
        result = _import_sales_rows(conn, rows, wh_id)
    return result


@app.post("/api/store/wipe-catalog")
async def wipe_catalog(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        wipe_catalog_data(conn)
    return {"ok": True, "message": "Демо-данные удалены. Можно загружать свои товары."}


@app.post("/api/stock/inbound")
async def stock_inbound(body: StockMovementIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        if not conn.execute("SELECT id FROM products WHERE id = ?", (body.product_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Товар не найден")
        resolve_warehouse_id(conn, body.warehouse_id)
        adjust_warehouse_stock(
            conn, body.warehouse_id, body.product_id, body.quantity,
            "inbound", notes=body.notes or "Приход на склад",
        )
        product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
        return enrich_product(conn, product) if product else {"ok": True}


def _require_battery_for_used(condition: str, battery: int | None) -> None:
    if condition in ("used", "refurbished") and battery is None:
        raise HTTPException(status_code=400, detail="Для Б/у укажите ёмкость батареи (%)")


@app.post("/api/stock/inbound-receipt")
async def stock_inbound_receipt(body: InboundReceiptIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        wh_id = resolve_warehouse_id(conn, body.warehouse_id)
        if body.mode == "new":
            if not body.product:
                raise HTTPException(status_code=400, detail="Заполните данные нового товара")
            p = body.product
            if p.ownership_type == "consignment" and not p.supplier_name.strip():
                raise HTTPException(status_code=400, detail="Укажите поставщика для реализации")
            _require_battery_for_used(p.condition, body.battery_capacity)
            track = 1 if p.category == "phone" else 0
            if track and not body.imei.strip() and not body.serial.strip():
                raise HTTPException(status_code=400, detail="Укажите IMEI или серийный номер")
            cur = conn.execute(
                """
                INSERT INTO products
                (name, category, ownership_type, supplier_name, brand, sku, barcode,
                 purchase_price, sale_price, stock, min_stock, created_at,
                 model, color, size, memory, ram, customs_cleared, customs_price, specs_extra, condition, track_units, image_url)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?,
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '')
                """,
                (
                    p.name, p.category, p.ownership_type, p.supplier_name.strip(),
                    p.brand, p.sku, p.barcode, p.purchase_price, p.sale_price,
                    p.min_stock, utc_now(),
                    p.model, p.color, p.size, p.memory, p.ram,
                    p.customs_cleared, p.customs_price, p.specs_extra, p.condition, track,
                ),
            )
            product_id = int(cur.lastrowid)
            product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        else:
            if not body.product_id:
                raise HTTPException(status_code=400, detail="Выберите товар")
            product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
            if not product:
                raise HTTPException(status_code=404, detail="Товар не найден")
            product_id = int(product["id"])
            _require_battery_for_used(product["condition"], body.battery_capacity)
            if int(product["track_units"] or 0) and not body.imei.strip() and not body.serial.strip():
                raise HTTPException(status_code=400, detail="Укажите IMEI или серийный номер")

        track = int(product["track_units"] or 0)
        qty = 1 if track else body.quantity
        unit_row = None
        if track:
            imei = body.imei.strip()
            serial = body.serial.strip()
            if imei:
                dup = conn.execute(
                    "SELECT id FROM product_units WHERE imei = ? AND status != 'sold'", (imei,)
                ).fetchone()
                if dup:
                    raise HTTPException(status_code=400, detail="IMEI уже в системе")
            cs = "pending" if not int(product["customs_cleared"] or 0) and product["category"] == "phone" else "none"
            unit_cost = (
                float(body.unit_purchase_price)
                if body.unit_purchase_price is not None and body.unit_purchase_price > 0
                else float(product["purchase_price"] or 0)
            )
            if body.mode == "new" and body.product and body.product.purchase_price > 0:
                unit_cost = float(body.product.purchase_price)
            unit_extra = float(body.unit_extra_cost or 0)
            if unit_extra <= 0 and body.mode == "new" and body.product:
                unit_extra = float(body.product.customs_price or 0)
            unit_note = body.notes or "Приход на склад"
            cur_u = conn.execute(
                """
                INSERT INTO product_units
                (product_id, warehouse_id, imei, serial, status, notes, created_at,
                 customs_status, customs_cleared, customs_price, battery_capacity,
                 client_name, region, arrival_date, purchase_price)
                VALUES (?, ?, ?, ?, 'in_stock', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product_id, wh_id, imei, serial,
                    unit_note, utc_now(), cs,
                    int(product["customs_cleared"] or 0), unit_extra, body.battery_capacity,
                    body.client_name.strip(), body.region.strip(),
                    body.arrival_date.strip() or utc_now()[:10],
                    unit_cost,
                ),
            )
            unit_row = conn.execute("SELECT * FROM product_units WHERE id = ?", (cur_u.lastrowid,)).fetchone()
            adjust_warehouse_stock(
                conn, wh_id, product_id, 1, "inbound",
                notes=f"Приход: {imei or serial or f'#{cur_u.lastrowid}'}",
            )
        else:
            adjust_warehouse_stock(
                conn, wh_id, product_id, qty, "inbound",
                notes=body.notes or "Приход на склад",
            )

        product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        result = enrich_product(conn, product)
        if unit_row:
            result["unit"] = enrich_unit_row(unit_row)
        return result


@app.post("/api/stock/outbound")
async def stock_outbound(body: StockMovementIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        if not conn.execute("SELECT id FROM products WHERE id = ?", (body.product_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Товар не найден")
        resolve_warehouse_id(conn, body.warehouse_id)
        adjust_warehouse_stock(
            conn, body.warehouse_id, body.product_id, -body.quantity,
            "outbound", notes=body.notes or "Расход со склада",
        )
        product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
        return enrich_product(conn, product) if product else {"ok": True}


def _build_transfer_document(
    conn: sqlite3.Connection,
    movement_id: int,
    body: StockTransferIn,
    created_at: str,
) -> dict[str, Any]:
    product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
    from_wh = conn.execute("SELECT * FROM warehouses WHERE id = ?", (body.from_warehouse_id,)).fetchone()
    to_wh = conn.execute("SELECT * FROM warehouses WHERE id = ?", (body.to_warehouse_id,)).fetchone()
    return {
        "id": movement_id,
        "created_at": created_at,
        "from_warehouse_id": body.from_warehouse_id,
        "from_warehouse_name": from_wh["name"] if from_wh else "",
        "from_warehouse_address": from_wh["address"] if from_wh else "",
        "to_warehouse_id": body.to_warehouse_id,
        "to_warehouse_name": to_wh["name"] if to_wh else "",
        "to_warehouse_address": to_wh["address"] if to_wh else "",
        "product_id": body.product_id,
        "product_name": product["name"] if product else "",
        "product_model": product["model"] if product else "",
        "product_color": product["color"] if product else "",
        "product_memory": product["memory"] if product else "",
        "product_sku": product["sku"] if product else "",
        "quantity": body.quantity,
        "notes": body.notes or "",
    }


@app.get("/api/stock/movements")
async def list_stock_movements(
    warehouse_id: int | None = None,
    movement_type: str = "",
    limit: int = Query(default=50, ge=1, le=200),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    sql = """
        SELECT m.*,
               p.name AS product_name, p.model AS product_model, p.sku AS product_sku,
               w.name AS warehouse_name,
               tw.name AS target_warehouse_name
        FROM stock_movements m
        JOIN products p ON p.id = m.product_id
        JOIN warehouses w ON w.id = m.warehouse_id
        LEFT JOIN warehouses tw ON tw.id = m.target_warehouse_id
        WHERE 1=1
    """
    params: list[Any] = []
    if warehouse_id is not None:
        sql += " AND (m.warehouse_id = ? OR m.target_warehouse_id = ?)"
        params.extend([warehouse_id, warehouse_id])
    if movement_type:
        sql += " AND m.movement_type = ?"
        params.append(movement_type)
    sql += " ORDER BY m.created_at DESC LIMIT ?"
    params.append(limit)
    with db() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/stock/transfers/{movement_id}/document")
async def get_transfer_document(movement_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        m = conn.execute(
            """
            SELECT m.*, p.name AS product_name, p.model AS product_model,
                   p.color AS product_color, p.memory AS product_memory, p.sku AS product_sku,
                   w.name AS from_warehouse_name, w.address AS from_warehouse_address,
                   tw.name AS to_warehouse_name, tw.address AS to_warehouse_address
            FROM stock_movements m
            JOIN products p ON p.id = m.product_id
            JOIN warehouses w ON w.id = m.warehouse_id
            LEFT JOIN warehouses tw ON tw.id = m.target_warehouse_id
            WHERE m.id = ? AND m.movement_type = 'transfer_out'
            """,
            (movement_id,),
        ).fetchone()
        if not m:
            raise HTTPException(status_code=404, detail="Накладная не найдена")
    return {
        "id": m["id"],
        "created_at": m["created_at"],
        "from_warehouse_id": m["warehouse_id"],
        "from_warehouse_name": m["from_warehouse_name"],
        "from_warehouse_address": m["from_warehouse_address"] or "",
        "to_warehouse_id": m["target_warehouse_id"],
        "to_warehouse_name": m["to_warehouse_name"] or "",
        "to_warehouse_address": m["to_warehouse_address"] or "",
        "product_id": m["product_id"],
        "product_name": m["product_name"],
        "product_model": m["product_model"] or "",
        "product_color": m["product_color"] or "",
        "product_memory": m["product_memory"] or "",
        "product_sku": m["product_sku"] or "",
        "quantity": m["quantity"],
        "notes": m["notes"] or "",
    }


@app.post("/api/stock/transfer")
async def stock_transfer(body: StockTransferIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    if body.from_warehouse_id == body.to_warehouse_id:
        raise HTTPException(status_code=400, detail="Склады отправления и назначения совпадают")
    with db() as conn:
        if not conn.execute("SELECT id FROM products WHERE id = ?", (body.product_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Товар не найден")
        resolve_warehouse_id(conn, body.from_warehouse_id)
        resolve_warehouse_id(conn, body.to_warehouse_id)
        now = utc_now()
        notes = body.notes or "Перемещение между складами"
        transfer_product_units(
            conn, body.product_id, body.from_warehouse_id, body.to_warehouse_id, body.quantity
        )
        _, movement_id = adjust_warehouse_stock(
            conn, body.from_warehouse_id, body.product_id, -body.quantity,
            "transfer_out", target_warehouse_id=body.to_warehouse_id, notes=notes,
        )
        adjust_warehouse_stock(
            conn, body.to_warehouse_id, body.product_id, body.quantity,
            "transfer_in", target_warehouse_id=body.from_warehouse_id, notes=notes,
        )
        document = _build_transfer_document(conn, movement_id or 0, body, now)
        product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
        result = enrich_product(conn, product) if product else {"ok": True}
        if isinstance(result, dict):
            result["transfer_document"] = document
        else:
            result = {"ok": True, "transfer_document": document}
        return result


# --- Products ---


@app.get("/api/products/meta/colors")
async def product_colors(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT TRIM(color) AS color
            FROM products
            WHERE TRIM(color) != ''
            ORDER BY LOWER(color)
            """
        ).fetchall()
    return [r["color"] for r in rows]


@app.get("/api/products")
async def list_products(
    q: str = "",
    category: str = "",
    ownership_type: str = "",
    supplier: str = "",
    color: str = "",
    warehouse_id: int | None = None,
    low_stock: bool = False,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    sql = "SELECT p.* FROM products p WHERE 1=1"
    params: list[Any] = []
    with db() as conn:
        user = resolve_user(conn, x_pin)
        if user and user.get("role") == "accessories":
            category = "accessory"
        if warehouse_id is not None:
            sql += """
                AND EXISTS (
                    SELECT 1 FROM warehouse_stock ws
                    WHERE ws.product_id = p.id AND ws.warehouse_id = ? AND ws.quantity > 0
                )
            """
            params.append(warehouse_id)
            wh_row = conn.execute(
                "SELECT warehouse_type, name FROM warehouses WHERE id = ?", (warehouse_id,)
            ).fetchone()
            if wh_row and get_warehouse_receipt_kind(conn, warehouse_id) != "accessories":
                sql += " AND p.category != 'accessory'"
        if q:
            clause, sparams = product_search_sql(q)
            sql += clause
            params.extend(sparams)
        if category:
            sql += " AND p.category = ?"
            params.append(category)
        if ownership_type:
            sql += " AND p.ownership_type = ?"
            params.append(ownership_type)
        if supplier:
            sql += " AND p.supplier_name LIKE ?"
            params.append(f"%{supplier}%")
        if color:
            sql += " AND LOWER(TRIM(p.color)) LIKE LOWER(?)"
            params.append(f"%{color.strip()}%")
        if low_stock:
            sql += " AND p.stock <= p.min_stock"
        sql += " ORDER BY p.ownership_type, p.name"
        rows = conn.execute(sql, params).fetchall()
        return [enrich_product(conn, r) for r in rows]


@app.get("/api/accessories/warehouse")
async def accessories_warehouse(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        wh_id = resolve_accessories_warehouse_id(conn)
        row = conn.execute("SELECT * FROM warehouses WHERE id = ?", (wh_id,)).fetchone()
        return row_to_dict(row) | {"currency": get_warehouse_currency(conn, wh_id)}


@app.get("/api/accessories/stock")
async def accessories_stock(
    q: str = "",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        wh_id = resolve_accessories_warehouse_id(conn)
        sql = """
            SELECT p.*, COALESCE(ws.quantity, 0) AS warehouse_quantity
            FROM products p
            LEFT JOIN warehouse_stock ws ON ws.product_id = p.id AND ws.warehouse_id = ?
            WHERE p.category = 'accessory'
        """
        params: list[Any] = [wh_id]
        if q.strip():
            clause, sp = product_search_sql(q)
            sql += clause
            params.extend(sp)
        sql += " ORDER BY p.name"
        rows = conn.execute(sql, params).fetchall()
        return [enrich_product(conn, r) | {"warehouse_quantity": r["warehouse_quantity"]} for r in rows]


@app.post("/api/accessories/inbound")
async def accessories_inbound(body: AccessoryInboundIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        wh_id = resolve_accessories_warehouse_id(conn)
        sale_price = body.sale_price if body.sale_price and body.sale_price > 0 else round(body.purchase_price * 1.3, 2)
        existing = conn.execute(
            """
            SELECT * FROM products
            WHERE category = 'accessory' AND LOWER(TRIM(name)) = LOWER(TRIM(?))
              AND LOWER(TRIM(COALESCE(model, ''))) = LOWER(TRIM(?))
              AND LOWER(TRIM(COALESCE(supplier_name, ''))) = LOWER(TRIM(?))
            LIMIT 1
            """,
            (body.name, body.model, body.supplier_name),
        ).fetchone()
        if existing:
            product_id = int(existing["id"])
            conn.execute(
                """
                UPDATE products SET purchase_price = ?, supplier_name = ?, model = ?
                WHERE id = ?
                """,
                (body.purchase_price, body.supplier_name.strip(), body.model.strip(), product_id),
            )
        else:
            cur = conn.execute(
                """
                INSERT INTO products
                (name, category, ownership_type, supplier_name, brand, sku, barcode,
                 purchase_price, sale_price, stock, min_stock, created_at,
                 model, color, size, memory, ram, customs_cleared, customs_price, specs_extra,
                 condition, track_units, image_url)
                VALUES (?, 'accessory', 'own', ?, '', '', '', ?, ?, 0, 0, ?,
                        ?, '', '', '', '', 0, 0, '', 'new', 0, '')
                """,
                (
                    body.name.strip(), body.supplier_name.strip(),
                    body.purchase_price, sale_price, utc_now(), body.model.strip(),
                ),
            )
            product_id = int(cur.lastrowid)
        adjust_warehouse_stock(
            conn, wh_id, product_id, body.quantity, "inbound",
            notes=f"Приход аксессуаров: {body.name} ×{body.quantity}",
        )
        product = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        result = enrich_product(conn, product)
        result["warehouse_quantity"] = get_warehouse_stock(conn, wh_id, product_id)
        return result


@app.post("/api/accessories/import/excel")
async def import_accessories_excel_file(
    file: UploadFile = File(...),
    replace: bool = Query(default=False),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="warehouse")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Пустой файл")
    try:
        with db() as conn:
            result = import_accessories_excel(conn, raw, replace=replace)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Accessories import failed")
        raise HTTPException(status_code=500, detail=f"Ошибка импорта: {exc}") from exc
    return result


@app.put("/api/accessories/products/{product_id}/price")
async def accessories_update_price(
    product_id: int,
    sale_price: float = Query(gt=0),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        row = conn.execute(
            "SELECT * FROM products WHERE id = ? AND category = 'accessory'", (product_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Товар не найден")
        conn.execute("UPDATE products SET sale_price = ? WHERE id = ?", (sale_price, product_id))
        return enrich_product(conn, conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone())


@app.get("/api/accessories/reports/finance")
async def accessories_finance_report(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        fin = _accessories_finance_report(conn, period, date_from, date_to)
        wh_id = resolve_accessories_warehouse_id(conn)
        clause, params, _ = _report_period_clause(period, date_from, date_to, "s.created_at")
        lines = conn.execute(
            f"""
            SELECT s.id, s.created_at, s.total, si.product_name, si.quantity, si.subtotal, si.shop_profit
            FROM sales s
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE s.status = 'completed' AND p.category = 'accessory' {clause}
            ORDER BY s.created_at DESC LIMIT 50
            """,
            params,
        ).fetchall()
        fin["recent_sales"] = [row_to_dict(r) for r in lines]
        fin["warehouse_id"] = wh_id
        return fin


@app.get("/api/accessories/cash-register")
async def accessories_cash_register(
    period: str = Query(default="day", pattern="^(day|week|month|all)$"),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        fin = _accessories_finance_report(conn, period, "", "")
        clause = ""
        params: list[Any] = []
        if period != "all":
            clause = " AND s.created_at >= ?"
            params = [period_start(period)]
        inflows = conn.execute(
            f"""
            SELECT sp.method_code, pm.name, COALESCE(SUM(sp.amount), 0) AS amount
            FROM sale_payments sp
            JOIN sales s ON s.id = sp.sale_id
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            LEFT JOIN payment_methods pm ON pm.code = sp.method_code
            WHERE s.status = 'completed' AND p.category = 'accessory' {clause}
            GROUP BY sp.method_code
            """,
            params,
        ).fetchall()
        total_in = sum(float(r["amount"]) for r in inflows)
        exp_clause = " AND expense_date >= ?" if period != "all" else ""
        exp_params = [period_start(period)[:10]] if period != "all" else []
        exp_out = conn.execute(
            f"SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE department = 'accessories' {exp_clause}",
            exp_params,
        ).fetchone()[0]
        exp_rows = conn.execute(
            f"SELECT * FROM expenses WHERE department = 'accessories' {exp_clause} ORDER BY expense_date DESC LIMIT 20",
            exp_params,
        ).fetchall()
        return {
            "period": period,
            "period_label": fin["period_label"],
            "total_inflows": None,
            "total_outflows": None,
            "net_cash": None,
            "profit": fin["shop_profit"],
            "revenue": fin["revenue"],
            "expenses": fin["expenses"],
            "by_currency": [
                {**currency_meta("USD"), "inflow": round(float(total_in), 2), "outflow": 0.0, "net": round(float(total_in), 2)},
                {**currency_meta("TJS"), "inflow": 0.0, "outflow": round(float(exp_out), 2), "net": round(-float(exp_out), 2)},
            ],
            "inflows": [{"method_code": r["method_code"], "name": r["name"] or r["method_code"], "amount": float(r["amount"]), "currency_code": "USD"} for r in inflows],
            "expense_lines": [row_to_dict(r) for r in exp_rows],
        }


@app.get("/api/suppliers")
async def list_suppliers(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        return _creditors_list(conn)


def _creditors_list(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    names = [
        r[0]
        for r in conn.execute(
            """
            SELECT DISTINCT supplier_name FROM products
            WHERE ownership_type = 'consignment' AND supplier_name != ''
            ORDER BY supplier_name
            """
        ).fetchall()
    ]
    result = []
    for name in names:
        products_count = conn.execute(
            "SELECT COUNT(*) FROM products WHERE ownership_type = 'consignment' AND supplier_name = ?",
            (name,),
        ).fetchone()[0]
        total_stock = conn.execute(
            """
            SELECT COALESCE(SUM(ws.quantity), 0)
            FROM warehouse_stock ws
            JOIN products p ON p.id = ws.product_id
            WHERE p.ownership_type = 'consignment' AND p.supplier_name = ?
            """,
            (name,),
        ).fetchone()[0]
        accrued_rows = conn.execute(
            """
            SELECT UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS cur,
                   COALESCE(SUM(si.supplier_due), 0) AS amt
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE si.ownership_type = 'consignment'
              AND si.supplier_name = ?
              AND s.status = 'completed'
            GROUP BY 1
            """,
            (name,),
        ).fetchall()
        paid_rows = conn.execute(
            """
            SELECT payment_method_code, COALESCE(SUM(amount), 0) AS amt
            FROM supplier_payments
            WHERE supplier_name = ?
            GROUP BY 1
            """,
            (name,),
        ).fetchall()
        bags: dict[str, dict[str, float]] = {}
        for r in accrued_rows:
            code = (r["cur"] or "TJS").upper()
            bags.setdefault(code, {"accrued_due": 0.0, "paid": 0.0})
            bags[code]["accrued_due"] += float(r["amt"] or 0)
        for r in paid_rows:
            pm = get_payment_method(conn, r["payment_method_code"] or "cash")
            code = payment_method_currency(pm) or "TJS"
            bags.setdefault(code, {"accrued_due": 0.0, "paid": 0.0})
            bags[code]["paid"] += float(r["amt"] or 0)
        by_currency = []
        for code in sorted(bags.keys(), key=lambda c: (0 if c == "USD" else 1, c)):
            acc = round(bags[code]["accrued_due"], 2)
            paid = round(bags[code]["paid"], 2)
            by_currency.append({
                **currency_meta(code),
                "accrued_due": acc,
                "paid": paid,
                "balance": round(max(0.0, acc - paid), 2),
            })
        sales_count = conn.execute(
            """
            SELECT COUNT(DISTINCT s.id)
            FROM sales s
            JOIN sale_items si ON si.sale_id = s.id
            WHERE s.status = 'completed' AND si.ownership_type = 'consignment' AND si.supplier_name = ?
            """,
            (name,),
        ).fetchone()[0]
        result.append({
            "supplier_name": name,
            "products_count": products_count,
            "total_stock": total_stock,
            "stock_value": None,
            "accrued_due": None,
            "paid": None,
            "balance": None,
            "by_currency": by_currency,
            "sales_count": sales_count,
        })
    return result


@app.get("/api/creditors")
async def list_creditors(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        creditors = _creditors_list(conn)
        totals: dict[str, float] = {}
        for c in creditors:
            for bag in c.get("by_currency") or []:
                totals[bag["code"]] = totals.get(bag["code"], 0.0) + float(bag.get("balance") or 0)
        by_currency = [
            {**currency_meta(code), "balance": round(amt, 2)}
            for code, amt in sorted(totals.items(), key=lambda x: (0 if x[0] == "USD" else 1, x[0]))
        ]
        payments = conn.execute(
            "SELECT * FROM supplier_payments ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
        pay_out = []
        for r in payments:
            d = row_to_dict(r)
            pm = get_payment_method(conn, d.get("payment_method_code") or "cash")
            d["currency_code"] = payment_method_currency(pm) or "TJS"
            pay_out.append(d)
    return {
        "creditors": creditors,
        "total_balance": None,
        "by_currency": by_currency,
        "recent_payments": pay_out,
    }


@app.get("/api/creditors/sales")
async def list_creditor_sales(
    supplier_name: str = Query(min_length=1),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        rows = conn.execute(
            """
            SELECT s.id AS sale_id, s.created_at, s.warehouse_id,
                   UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
                   si.product_name, si.quantity, si.subtotal, si.supplier_due, si.shop_profit,
                   COALESCE(w.name, '—') AS warehouse_name
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            WHERE si.ownership_type = 'consignment'
              AND si.supplier_name = ?
              AND s.status = 'completed'
              AND si.supplier_due > 0
            ORDER BY s.created_at DESC
            LIMIT 200
            """,
            (supplier_name.strip(),),
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.post("/api/products")
async def create_product(body: ProductIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    if body.ownership_type == "consignment" and not body.supplier_name.strip():
        raise HTTPException(status_code=400, detail="Укажите поставщика для товара под реализацию")
    with db() as conn:
        wh_id = resolve_warehouse_id(conn, body.warehouse_id)
        track = body.track_units if body.track_units is not None else (1 if body.category == "phone" else 0)
        cur = conn.execute(
            """
            INSERT INTO products
            (name, category, ownership_type, supplier_name, brand, sku, barcode,
             purchase_price, sale_price, stock, min_stock, created_at,
             model, color, size, memory, ram, customs_cleared, customs_price, specs_extra, condition, track_units, image_url)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                body.name, body.category, body.ownership_type, body.supplier_name.strip(),
                body.brand, body.sku, body.barcode, body.purchase_price, body.sale_price,
                body.min_stock, utc_now(),
                body.model, body.color, body.size, body.memory, body.ram,
                body.customs_cleared, body.customs_price, body.specs_extra, body.condition, track,
                body.image_url.strip(),
            ),
        )
        product_id = cur.lastrowid
        if body.stock > 0:
            adjust_warehouse_stock(
                conn, wh_id, product_id, body.stock,
                "inbound", notes="Начальный остаток при создании товара",
            )
        else:
            sync_product_stock(conn, product_id)
        row = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        return enrich_product(conn, row)


@app.put("/api/products/{product_id}")
async def update_product(product_id: int, body: ProductUpdate, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    fields = {k: v for k, v in body.model_dump().items() if v is not None and k != "warehouse_id"}
    warehouse_id = body.warehouse_id
    stock_update = fields.pop("stock", None)
    if not fields and stock_update is None:
        raise HTTPException(status_code=400, detail="Нет данных для обновления")
    if fields.get("ownership_type") == "consignment" and not (fields.get("supplier_name") or "").strip():
        with db() as conn:
            existing = conn.execute("SELECT supplier_name FROM products WHERE id = ?", (product_id,)).fetchone()
            if not existing or not existing["supplier_name"]:
                raise HTTPException(status_code=400, detail="Укажите поставщика")
    with db() as conn:
        if not conn.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Товар не найден")
        if fields:
            sets = ", ".join(f"{k} = ?" for k in fields)
            conn.execute(f"UPDATE products SET {sets} WHERE id = ?", (*fields.values(), product_id))
        if stock_update is not None:
            wh_id = resolve_warehouse_id(conn, warehouse_id)
            current = get_warehouse_stock(conn, wh_id, product_id)
            delta = stock_update - current
            if delta != 0:
                adjust_warehouse_stock(
                    conn, wh_id, product_id, delta,
                    "inbound" if delta > 0 else "outbound",
                    notes="Корректировка остатка через карточку товара",
                )
            else:
                sync_product_stock(conn, product_id)
        row = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        return enrich_product(conn, row)


@app.delete("/api/products/{product_id}")
async def delete_product(product_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        row = conn.execute("SELECT image_url FROM products WHERE id = ?", (product_id,)).fetchone()
        cur = conn.execute("DELETE FROM products WHERE id = ?", (product_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Товар не найден")
        if row and row["image_url"] and row["image_url"].startswith("/uploads/"):
            path = UPLOADS_DIR / Path(row["image_url"]).name
            if path.exists():
                path.unlink(missing_ok=True)
    return {"ok": True}


def _remove_product_images(product_id: int) -> None:
    for path in UPLOADS_DIR.glob(f"product_{product_id}.*"):
        path.unlink(missing_ok=True)


@app.post("/api/products/{product_id}/image")
async def upload_product_image(
    product_id: int,
    file: UploadFile = File(...),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        raise HTTPException(status_code=400, detail="Допустимы JPG, PNG, WEBP, GIF")
    data = await file.read()
    if len(data) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Максимальный размер файла — 5 МБ")
    with db() as conn:
        if not conn.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Товар не найден")
    _remove_product_images(product_id)
    filename = f"product_{product_id}{ext}"
    (UPLOADS_DIR / filename).write_bytes(data)
    image_url = f"/uploads/{filename}"
    with db() as conn:
        conn.execute("UPDATE products SET image_url = ? WHERE id = ?", (image_url, product_id))
        row = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        return enrich_product(conn, row)


@app.delete("/api/products/{product_id}/image")
async def delete_product_image(product_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        if not conn.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Товар не найден")
        conn.execute("UPDATE products SET image_url = '' WHERE id = ?", (product_id,))
    _remove_product_images(product_id)
    return {"ok": True}


@app.get("/api/products/{product_id}")
async def get_product(product_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        row = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Товар не найден")
        return enrich_product(conn, row)


# --- Sales ---


@app.post("/api/sales")
async def create_sale(body: SaleIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    if not body.items:
        raise HTTPException(status_code=400, detail="Корзина пуста")

    with db() as conn:
        warehouse_id = resolve_warehouse_id(conn, body.warehouse_id)
        user = resolve_user(conn, x_pin)
        shift = get_open_shift(conn)
        shift_id = body.shift_id or (shift["id"] if shift else None)

        subtotal = 0.0
        lines: list[dict[str, Any]] = []
        for item in body.items:
            product = conn.execute("SELECT * FROM products WHERE id = ?", (item.product_id,)).fetchone()
            if not product:
                raise HTTPException(status_code=404, detail=f"Товар #{item.product_id} не найден")
            if user and user.get("role") == "accessories" and product["category"] != "accessory":
                raise HTTPException(status_code=403, detail="Доступны только аксессуары")
            if int(product["track_units"] or 0):
                avail = conn.execute(
                    """
                    SELECT COUNT(*) FROM product_units
                    WHERE product_id = ? AND warehouse_id = ? AND status = 'in_stock'
                    """,
                    (item.product_id, warehouse_id),
                ).fetchone()[0]
                if avail < item.quantity:
                    raise HTTPException(
                        status_code=400,
                        detail=f"«{product['name']}»: нужны IMEI — доступно {avail} из {item.quantity}",
                    )
                picked_units = pick_units(
                    conn, product["id"], warehouse_id, item.quantity, item.unit_ids or None
                )
                # Landed cost: себестоимость + колонка «Расходы» (repair/customs)
                unit_costs = [unit_total_cost(u, product) for u in picked_units]
            else:
                wh_stock = get_warehouse_stock(conn, warehouse_id, item.product_id)
                if wh_stock < item.quantity:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Недостаточно «{product['name']}» на складе: доступно {wh_stock}",
                    )
                picked_units = None
                unit_costs = None
            calc = calc_line(product, item.quantity, item.unit_price, unit_costs)
            subtotal += calc["subtotal"]
            lines.append({
                "product": product,
                "qty": item.quantity,
                "unit_ids": item.unit_ids,
                "unit_checkouts": {u.unit_id: u for u in item.units},
                "picked_units": picked_units,
                **calc,
            })

        total = max(0.0, subtotal - body.discount)
        now = utc_now()
        sale_currency = get_warehouse_currency(conn, warehouse_id)["code"]
        pay_currency = (body.pay_currency or sale_currency).upper().strip() or sale_currency
        fx_rate = float(body.fx_rate) if body.fx_rate and body.fx_rate > 0 else None
        fx_note = ""
        converted_payments: list[dict[str, Any]] = []
        for p in body.payments:
            cur = (p.currency_code or pay_currency or sale_currency).upper().strip() or sale_currency
            amt_in = float(p.amount)
            if cur == sale_currency:
                amt = amt_in
            elif fx_rate and fx_rate > 0 and cur == pay_currency and pay_currency != sale_currency:
                amt = amt_in / fx_rate
            else:
                amt = convert_amount(conn, amt_in, cur, sale_currency, at=now)
            converted_payments.append({
                "method_code": p.method_code,
                "amount": round(amt, 2),
                "currency_code": cur,
                "amount_input": amt_in,
            })
        # Absorb tiny FX rounding so full payment still matches total
        paid_conv = sum(p["amount"] for p in converted_payments)
        if converted_payments and abs(paid_conv - total) <= 0.05 and abs(paid_conv - total) > 0.001:
            converted_payments[-1]["amount"] = round(
                converted_payments[-1]["amount"] + (total - paid_conv), 2
            )
        if any(p["currency_code"] != sale_currency for p in converted_payments) or (
            pay_currency != sale_currency and body.payments
        ):
            if fx_rate and pay_currency != sale_currency:
                fx_note = f"Курс: 1 {sale_currency} = {fx_rate:.4f} {pay_currency}"
            elif pay_currency != sale_currency:
                rate_pay = get_exchange_rate_at(conn, pay_currency, now)
                rate_sale = get_exchange_rate_at(conn, sale_currency, now)
                cross = rate_pay / rate_sale if rate_sale else rate_pay
                fx_note = f"Конвертация: 1 {pay_currency} = {cross:.4f} {sale_currency}"
            mixed = []
            for p in converted_payments:
                if p["currency_code"] != sale_currency:
                    mixed.append(f"{p['amount_input']:.2f} {p['currency_code']}→{p['amount']:.2f} {sale_currency}")
            if mixed:
                fx_note = (fx_note + " · " if fx_note else "") + "; ".join(mixed)
        pay_payload_src = [{
            "method_code": p["method_code"],
            "amount": p["amount"],
            "pay_currency_code": p["currency_code"],
            "pay_amount": p["amount_input"],
        } for p in converted_payments]

        sale_notes = body.notes.strip()
        if fx_note:
            sale_notes = f"{sale_notes} | {fx_note}".strip(" |") if sale_notes else fx_note

        if pay_payload_src:
            cash_amount, card_amount, payment_method, pay_payload, amount_paid, amount_due = process_sale_payments(
                conn, pay_payload_src, total, debtor_name=body.debtor_name
            )
        else:
            if total <= 0:
                amount_paid = 0.0
                amount_due = 0.0
                cash_amount = card_amount = 0.0
                payment_method = body.payment_method
                pay_payload = []
            elif body.debtor_name.strip():
                amount_paid = 0.0
                amount_due = total
                cash_amount = card_amount = 0.0
                payment_method = "credit"
                pay_payload = []
            else:
                cash_amount = total if body.payment_method == "cash" else 0.0
                card_amount = total if body.payment_method not in ("cash", "trade_in", "credit") else 0.0
                payment_method = body.payment_method
                pay_payload = [{"method_code": body.payment_method, "amount": total}]
                amount_paid = total
                amount_due = 0.0

        cur = conn.execute(
            """
            INSERT INTO sales
            (total, discount, payment_method, status, notes, created_at,
             warehouse_id, cash_amount, card_amount, trade_in_value, shift_id, user_id, user_name,
             amount_paid, amount_due, currency_code, affects_cash)
            VALUES (?, ?, ?, 'completed', ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, 1)
            """,
            (total, body.discount, payment_method, sale_notes, now,
             warehouse_id, cash_amount, card_amount, shift_id,
             user.get("id") if user else None, user.get("name", "") if user else "",
             amount_paid, amount_due, sale_currency),
        )
        sale_id = cur.lastrowid
        if pay_payload:
            insert_sale_payments(conn, sale_id, pay_payload)
        if amount_due > 0.01:
            create_receivable(
                conn,
                sale_id=sale_id,
                customer_name=body.debtor_name,
                customer_phone=body.debtor_phone,
                total_amount=total,
                paid_amount=amount_paid,
                warehouse_id=warehouse_id,
                notes=body.notes,
            )
        for line in lines:
            p = line["product"]
            cur_item = conn.execute(
                """
                INSERT INTO sale_items
                (sale_id, product_id, product_name, ownership_type, supplier_name, quantity,
                 unit_price, purchase_price, supplier_due, shop_profit, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sale_id, p["id"], p["name"], line["ownership_type"], line["supplier_name"],
                    line["qty"], line["unit_price"], line["purchase_price"],
                    line["supplier_due"], line["shop_profit"], line["subtotal"],
                ),
            )
            sale_item_id = cur_item.lastrowid
            if int(p["track_units"] or 0):
                checkout_map = line.get("unit_checkouts") or {}
                units = line.get("picked_units") or pick_units(
                    conn, p["id"], warehouse_id, line["qty"], line["unit_ids"] or None
                )
                for u in units:
                    co = checkout_map.get(u["id"])
                    resolved = resolve_unit_sale(conn, u, co, p)
                    conn.execute(
                        """
                        INSERT INTO sale_item_units
                        (sale_item_id, unit_id, imei, serial, customs_cleared, customs_price, imei_pending)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            sale_item_id, u["id"], resolved["imei"], resolved["serial"],
                            resolved["customs_cleared"], resolved["customs_price"], resolved["imei_pending"],
                        ),
                    )
                    mark_unit_sold_full(
                        conn, u["id"], sale_id, resolved["imei"],
                        resolved["customs_cleared"], resolved["customs_price"],
                    )
            adjust_warehouse_stock(
                conn, warehouse_id, p["id"], -line["qty"],
                "sale", reference_id=sale_id, notes=f"Продажа #{sale_id}",
            )

        sale = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
        return enrich_sale(conn, sale)


@app.get("/api/sales")
async def list_sales(
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    date_from: str = "",
    date_to: str = "",
    ownership_type: str = "",
    warehouse_id: int | None = None,
    user_id: int | None = None,
    q: str = "",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    joins: list[str] = []
    wheres = ["s.status = 'completed'"]
    params: list[Any] = []
    if ownership_type:
        joins.append("JOIN sale_items si ON si.sale_id = s.id")
        wheres.append("si.ownership_type = ?")
        params.append(ownership_type)
    if warehouse_id is not None:
        wheres.append("s.warehouse_id = ?")
        params.append(warehouse_id)
    if user_id is not None:
        wheres.append("s.user_id = ?")
        params.append(user_id)
    search_clause, search_params = sale_search_sql(q)
    if search_clause:
        wheres.append(search_clause)
        params.extend(search_params)
    df, dp = date_filter_sql(date_from, date_to, "s.created_at")
    params.extend(dp)
    join_sql = " " + " ".join(dict.fromkeys(joins)) if joins else ""
    where_sql = " AND ".join(wheres)
    sql = f"SELECT DISTINCT s.* FROM sales s{join_sql} WHERE {where_sql}{df} ORDER BY s.created_at DESC LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    count_sql = f"SELECT COUNT(DISTINCT s.id) FROM sales s{join_sql} WHERE {where_sql}{df}"
    count_params = params[:-2]
    with db() as conn:
        sales = conn.execute(sql, params).fetchall()
        total_count = conn.execute(count_sql, count_params).fetchone()[0]
        wh_map = {
            int(r["id"]): r["name"]
            for r in conn.execute("SELECT id, name FROM warehouses").fetchall()
        }
        sale_ids = [s["id"] for s in sales]
        pay_map = payments_for_sales(conn, sale_ids)
        recv_map: dict[int, dict[str, Any]] = {}
        if sale_ids:
            placeholders = ",".join("?" * len(sale_ids))
            for r in conn.execute(
                f"""
                SELECT sale_id, customer_name, customer_phone, amount_due, status
                FROM receivables WHERE sale_id IN ({placeholders})
                ORDER BY id DESC
                """,
                sale_ids,
            ).fetchall():
                sid = int(r["sale_id"])
                if sid not in recv_map:
                    recv_map[sid] = row_to_dict(r) or {}
    items = []
    for s in sales:
        d = row_to_dict(s)
        d["payments"] = pay_map.get(s["id"], [])
        wid = int(s["warehouse_id"]) if s["warehouse_id"] else None
        d["warehouse_name"] = wh_map.get(wid, "—") if wid else "—"
        rec = recv_map.get(int(s["id"]))
        if rec:
            d["debtor_name"] = rec.get("customer_name") or ""
            d["debtor_phone"] = rec.get("customer_phone") or ""
            d["amount_due"] = float(rec.get("amount_due") or 0)
            d["receivable_status"] = rec.get("status") or ""
        else:
            d["debtor_name"] = ""
            d["debtor_phone"] = ""
            d["amount_due"] = float(s["amount_due"] or 0) if s["amount_due"] else 0.0
            d["receivable_status"] = ""
        items.append(d)
    return {"items": items, "total": int(total_count)}


@app.get("/api/sales/{sale_id}")
async def get_sale(sale_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        sale = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
        if not sale:
            raise HTTPException(status_code=404, detail="Продажа не найдена")
        return enrich_sale(conn, sale)


@app.post("/api/sales/{sale_id}/void")
async def void_sale(sale_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="owner")
    now = utc_now()
    with db() as conn:
        sale = conn.execute("SELECT * FROM sales WHERE id = ? AND status = 'completed'", (sale_id,)).fetchone()
        if not sale:
            raise HTTPException(status_code=404, detail="Продажа не найдена")
        sale_wh = sale["warehouse_id"] or get_default_warehouse_id(conn)
        target_wh, to_bu = sale_return_target_warehouse(conn, sale_id, sale_wh)
        restored = restore_units_for_sale(
            conn,
            sale_id,
            return_warehouse_id=target_wh,
            to_bu=False,
        )
        # Единицы по IMEI — обратно на склад продажи (склад не меняем).
        for ru in restored:
            adjust_warehouse_stock(
                conn,
                int(ru["warehouse_id"]),
                int(ru["product_id"]),
                1,
                "void",
                reference_id=sale_id,
                notes=f"Возврат продажи #{sale_id}",
            )
        # Позиции без серийников (аксессуары и т.п.) — на склад продажи.
        for item in conn.execute("SELECT * FROM sale_items WHERE sale_id = ?", (sale_id,)).fetchall():
            if not item["product_id"]:
                continue
            linked = conn.execute(
                """
                SELECT COUNT(*) FROM sale_item_units siu
                WHERE siu.sale_item_id = ?
                """,
                (item["id"],),
            ).fetchone()[0]
            if int(linked or 0) > 0:
                continue
            adjust_warehouse_stock(
                conn,
                sale_wh,
                item["product_id"],
                item["quantity"],
                "void",
                reference_id=sale_id,
                notes=f"Возврат продажи #{sale_id}",
            )
        _close_receivables_for_return(conn, sale_id, now)
        note_suffix = f" [Возврат {now[:10]}]"
        conn.execute(
            """
            UPDATE sales
            SET status = 'voided', amount_due = 0, amount_paid = 0,
                notes = TRIM(COALESCE(notes, '') || ?)
            WHERE id = ?
            """,
            (note_suffix, sale_id),
        )
    return {"ok": True, "return_to_bu": to_bu, "warehouse_id": target_wh}


# --- Trade-in ---


@app.get("/api/trade-ins")
async def list_trade_ins(
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        rows = conn.execute(
            """
            SELECT t.*,
                   gp.name AS given_product_name,
                   rp.name AS received_product_name,
                   gw.name AS given_warehouse_name,
                   rw.name AS received_warehouse_name
            FROM trade_ins t
            LEFT JOIN products gp ON gp.id = t.given_product_id
            LEFT JOIN products rp ON rp.id = t.received_product_id
            LEFT JOIN warehouses gw ON gw.id = t.given_warehouse_id
            LEFT JOIN warehouses rw ON rw.id = t.received_warehouse_id
            ORDER BY t.created_at DESC
            LIMIT ? OFFSET ?
            """,
            (limit, offset),
        ).fetchall()
        total = conn.execute("SELECT COUNT(*) FROM trade_ins").fetchone()[0]
    return {"items": [row_to_dict(r) for r in rows], "total": total}


@app.post("/api/trade-ins")
async def create_trade_in(body: TradeInIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        given_product = conn.execute(
            "SELECT * FROM products WHERE id = ?", (body.given_product_id,)
        ).fetchone()
        if not given_product:
            raise HTTPException(status_code=404, detail="Выдаваемый товар не найден")
        given_wh = resolve_warehouse_id(conn, body.given_warehouse_id)
        received_wh = resolve_warehouse_id(conn, body.received_warehouse_id)
        wh_stock = get_warehouse_stock(conn, given_wh, body.given_product_id)
        if wh_stock < 1:
            raise HTTPException(
                status_code=400,
                detail=f"Недостаточно «{given_product['name']}» на складе: доступно {wh_stock}",
            )

        calc = calc_line(given_product, 1)
        subtotal = calc["subtotal"]
        total = max(0.0, subtotal)
        paid = body.cash_amount + body.card_amount + body.received_value
        if abs(paid - total) > 0.01:
            raise HTTPException(
                status_code=400,
                detail=f"Сумма оплаты ({paid:.2f}) не совпадает с ценой товара ({total:.2f})",
            )

        shift = get_open_shift(conn)
        shift_id = shift["id"] if shift else None

        now = utc_now()
        cur = conn.execute(
            """
            INSERT INTO sales
            (total, discount, payment_method, status, notes, created_at,
             warehouse_id, cash_amount, card_amount, trade_in_value, shift_id)
            VALUES (?, 0, 'trade_in', 'completed', ?, ?, ?, ?, ?, ?, ?)
            """,
            (total, body.notes or "Обмен (trade-in)", now,
             given_wh, body.cash_amount, body.card_amount, body.received_value, shift_id),
        )
        sale_id = cur.lastrowid

        cur_item = conn.execute(
            """
            INSERT INTO sale_items
            (sale_id, product_id, product_name, ownership_type, supplier_name, quantity,
             unit_price, purchase_price, supplier_due, shop_profit, subtotal)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)
            """,
            (
                sale_id, given_product["id"], given_product["name"],
                calc["ownership_type"], calc["supplier_name"],
                calc["unit_price"], calc["purchase_price"],
                calc["supplier_due"], calc["shop_profit"], calc["subtotal"],
            ),
        )
        sale_item_id = cur_item.lastrowid
        if int(given_product["track_units"] or 0):
            units = pick_units(
                conn, given_product["id"], given_wh, 1,
                [body.given_unit_id] if body.given_unit_id else None,
            )
            for u in units:
                conn.execute(
                    "INSERT INTO sale_item_units (sale_item_id, unit_id, imei, serial) VALUES (?, ?, ?, ?)",
                    (sale_item_id, u["id"], u["imei"] or "", u["serial"] or ""),
                )
            mark_units_sold(conn, units, sale_id)

        adjust_warehouse_stock(
            conn, given_wh, body.given_product_id, -1,
            "trade_in_given", reference_id=sale_id, notes="Trade-in: выдача товара клиенту",
        )

        recv_cur = conn.execute(
            """
            INSERT INTO products
            (name, category, ownership_type, supplier_name, brand, sku, barcode,
             purchase_price, sale_price, stock, min_stock, created_at,
             model, color, size, memory, ram, customs_cleared, customs_price, specs_extra, condition, track_units)
            VALUES (?, 'phone', 'own', '', ?, '', '', ?, ?, 0, 1, ?,
                    ?, ?, ?, ?, ?, 0, 0, ?, ?, 1)
            """,
            (
                body.received_name, body.received_brand, body.received_purchase_price,
                body.received_sale_price, now,
                body.received_model, body.received_color, body.received_size,
                body.received_memory, body.received_ram, body.received_specs_extra,
                body.received_condition,
            ),
        )
        received_product_id = recv_cur.lastrowid

        recv_imei = body.received_imei.strip()
        recv_note = "Trade-in"

        if recv_imei or body.received_serial.strip():
            conn.execute(
                """
                INSERT INTO product_units
                (product_id, warehouse_id, imei, serial, status, notes, created_at, purchase_price)
                VALUES (?, ?, ?, ?, 'in_stock', ?, ?, ?)
                """,
                (
                    received_product_id, received_wh,
                    recv_imei, body.received_serial.strip(), recv_note, now,
                    float(body.received_purchase_price or 0),
                ),
            )
        adjust_warehouse_stock(
            conn, received_wh, received_product_id, 1,
            "trade_in_received", reference_id=sale_id,
            notes="Trade-in: принят устройство от клиента",
        )

        ti_cur = conn.execute(
            """
            INSERT INTO trade_ins
            (given_product_id, given_warehouse_id,
             received_name, received_brand, received_model, received_color,
             received_size, received_memory, received_ram, received_specs_extra,
             received_condition, received_purchase_price, received_sale_price,
             received_value, cash_amount, card_amount,
             received_product_id, received_warehouse_id, sale_id, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                body.given_product_id, given_wh,
                body.received_name, body.received_brand, body.received_model, body.received_color,
                body.received_size, body.received_memory, body.received_ram, body.received_specs_extra,
                body.received_condition, body.received_purchase_price, body.received_sale_price,
                body.received_value, body.cash_amount, body.card_amount,
                received_product_id, received_wh, sale_id, body.notes, now,
            ),
        )

        row = conn.execute(
            """
            SELECT t.*,
                   gp.name AS given_product_name,
                   rp.name AS received_product_name,
                   gw.name AS given_warehouse_name,
                   rw.name AS received_warehouse_name
            FROM trade_ins t
            LEFT JOIN products gp ON gp.id = t.given_product_id
            LEFT JOIN products rp ON rp.id = t.received_product_id
            LEFT JOIN warehouses gw ON gw.id = t.given_warehouse_id
            LEFT JOIN warehouses rw ON rw.id = t.received_warehouse_id
            WHERE t.id = ?
            """,
            (ti_cur.lastrowid,),
        ).fetchone()
    return row_to_dict(row)


# --- Supplier payments ---


@app.get("/api/supplier-payments")
async def list_supplier_payments(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        rows = conn.execute("SELECT * FROM supplier_payments ORDER BY created_at DESC LIMIT 200").fetchall()
    return [row_to_dict(r) for r in rows]


@app.post("/api/supplier-payments")
async def create_supplier_payment(body: SupplierPaymentIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        pm = get_payment_method(conn, body.payment_method_code)
        if not pm:
            raise HTTPException(status_code=400, detail="Способ оплаты не найден")
        balance_row = next(
            (c for c in _creditors_list(conn) if c["supplier_name"] == body.supplier_name.strip()),
            None,
        )
        if balance_row and body.amount > balance_row["balance"] + 0.01:
            raise HTTPException(
                status_code=400,
                detail=f"Сумма больше долга ({balance_row['balance']:.2f})",
            )
        cur = conn.execute(
            """
            INSERT INTO supplier_payments
            (supplier_name, amount, notes, created_at, payment_method_code)
            VALUES (?, ?, ?, ?, ?)
            """,
            (body.supplier_name.strip(), body.amount, body.notes, utc_now(), body.payment_method_code),
        )
        row = conn.execute("SELECT * FROM supplier_payments WHERE id = ?", (cur.lastrowid,)).fetchone()
    return row_to_dict(row)



def _report_opiu(
    conn: sqlite3.Connection,
    period: str,
    date_from: str,
    date_to: str,
    warehouse_id: int | None = None,
) -> dict[str, Any]:
    clause, params, label = _report_period_clause(period, date_from, date_to)
    fin = _finance_report(conn, period, "all", date_from, date_to, warehouse_id)
    exp_clause = clause.replace("s.created_at", "expense_date") if clause else ""
    if period != "all" and not date_from and not date_to:
        exp_clause = " AND expense_date >= ?"
        exp_params = [period_start(period)[:10]]
    elif clause:
        exp_params = [p[:10] if isinstance(p, str) and len(p) > 10 else p for p in params]
    else:
        exp_params = []
    expenses = conn.execute(
        f"""
        SELECT category, department, warehouse_id,
               COALESCE(SUM(amount), 0) AS total
        FROM expenses
        WHERE 1=1 {exp_clause}
        GROUP BY category, department, warehouse_id
        """,
        exp_params,
    ).fetchall()
    total_expenses, main_expenses, expenses_by_warehouse, split = expenses_allocated_by_warehouse(
        conn, exp_clause, exp_params
    )
    category_rows: list[dict[str, Any]] = []
    if warehouse_id:
        wh_amt = expense_amount_for_warehouse(expenses_by_warehouse, warehouse_id)
        wh_row = next(
            (x for x in expenses_by_warehouse if int(x.get("warehouse_id") or 0) == int(warehouse_id)),
            None,
        )
        shared_share = 0.0
        shared_total = sum(float(r["total"]) for r in expenses if r["warehouse_id"] is None)
        if wh_row and shared_total > 0:
            shared_share = float(wh_row.get("shared_amount") or 0) / shared_total
        main_expenses = 0.0
        for r in expenses:
            amt = float(r["total"] or 0)
            if r["warehouse_id"] is None:
                part = round(amt * shared_share, 2)
            elif int(r["warehouse_id"]) == int(warehouse_id):
                part = amt
            else:
                part = 0.0
            if part <= 0.0001:
                continue
            if (r["department"] or "main") != "accessories":
                main_expenses += part
            category_rows.append({
                "category": r["category"],
                "amount": part,
                "department": r["department"] or "main",
            })
        # merge same category+department
        merged: dict[tuple[str, str], float] = {}
        for c in category_rows:
            key = (c["category"], c["department"])
            merged[key] = merged.get(key, 0.0) + float(c["amount"])
        category_rows = [
            {"category": k[0], "amount": round(v, 2), "department": k[1]}
            for k, v in merged.items()
            if v > 0.0001
        ]
        total_expenses = wh_amt
        main_expenses = round(main_expenses, 2)
        expenses_by_warehouse = [wh_row] if wh_row else []
    else:
        merged_cat: dict[tuple[str, str], float] = {}
        for r in expenses:
            key = (r["category"], r["department"] or "main")
            merged_cat[key] = merged_cat.get(key, 0.0) + float(r["total"] or 0)
        category_rows = [
            {"category": k[0], "amount": round(v, 2), "department": k[1]}
            for k, v in merged_cat.items()
            if v > 0.0001
        ]
    by_currency: list[dict[str, Any]] = []
    for c in fin.get("by_currency") or []:
        cur = dict(c)
        if cur.get("code") == "TJS":
            cur["operating_expenses"] = total_expenses
            cur["net_profit"] = float(cur.get("gross_profit", cur.get("shop_profit", 0))) - total_expenses
        else:
            cur["operating_expenses"] = 0.0
            cur["net_profit"] = float(cur.get("gross_profit", cur.get("shop_profit", 0)))
        by_currency.append(cur)
    multi = len(by_currency) > 1
    gross_profit = fin["gross_revenue"] - fin["own_cogs"] - fin["supplier_due"]
    operating_profit = gross_profit - total_expenses if not multi else None
    return {
        "period_label": fin.get("period_label") or label,
        "warehouse_id": warehouse_id,
        "warehouse_name": fin.get("warehouse_name") or "",
        "revenue": fin["gross_revenue"],
        "discounts": fin["discounts"],
        "net_revenue": fin["net_revenue"],
        "cogs_own": fin["own_cogs"],
        "supplier_due": fin["supplier_due"],
        "gross_profit": gross_profit,
        "operating_expenses": total_expenses,
        "main_operating_expenses": main_expenses,
        "expenses_by_category": category_rows,
        "expenses_by_warehouse": expenses_by_warehouse,
        "expense_allocation": split,
        "operating_profit": operating_profit,
        "shop_profit": fin["shop_profit"],
        "net_profit": operating_profit,
        "by_currency": by_currency,
        "multi_currency": multi,
    }


def _report_dds(
    conn: sqlite3.Connection,
    period: str,
    date_from: str,
    date_to: str,
    warehouse_id: int | None = None,
) -> dict[str, Any]:
    clause, params, label = _report_period_clause(period, date_from, date_to, "s.created_at")
    wh_clause, wh_params = warehouse_sales_filter(warehouse_id)
    all_params = params + wh_params
    inflows = conn.execute(
        f"""
        SELECT sp.method_code, pm.name, pm.method_type, COALESCE(SUM(sp.amount), 0) AS amount
        FROM sale_payments sp
        JOIN sales s ON s.id = sp.sale_id
        LEFT JOIN payment_methods pm ON pm.code = sp.method_code
        WHERE s.status = 'completed' {clause}{wh_clause}
        GROUP BY sp.method_code
        """,
        all_params,
    ).fetchall()
    total_in = sum(float(r["amount"]) for r in inflows)
    recv_clause = clause.replace("s.created_at", "rp.created_at")
    if warehouse_id:
        recv_in = conn.execute(
            f"""
            SELECT COALESCE(SUM(rp.amount), 0)
            FROM receivable_payments rp
            JOIN receivables r ON r.id = rp.receivable_id
            JOIN sales s ON s.id = r.sale_id
            WHERE s.warehouse_id = ? {recv_clause.replace('s.', 'rp.') if 's.' in recv_clause else recv_clause}
            """,
            [warehouse_id] + (params if recv_clause else []),
        ).fetchone()[0]
    else:
        recv_in = conn.execute(
            f"""
            SELECT COALESCE(SUM(rp.amount), 0)
            FROM receivable_payments rp
            WHERE 1=1 {recv_clause.replace('s.', '') if 's.' in recv_clause else recv_clause}
            """,
            params if recv_clause else [],
        ).fetchone()[0]
    total_in += float(recv_in)
    manual_inflows: list[dict[str, Any]] = []
    manual_total = 0.0
    # Manual cash inflows are company-wide — include only for "all warehouses"
    if not warehouse_id and _table_exists(conn, "cash_inflows"):
        man_clause = clause.replace("s.created_at", "ci.created_at")
        man_rows = conn.execute(
            f"""
            SELECT ci.payment_method_code AS method_code, pm.name, pm.method_type,
                   COALESCE(SUM(ci.amount_base), 0) AS amount
            FROM cash_inflows ci
            LEFT JOIN payment_methods pm ON pm.code = ci.payment_method_code
            WHERE ci.source_type = 'counterparty' {man_clause}
            GROUP BY ci.payment_method_code
            """,
            params if man_clause else [],
        ).fetchall()
        manual_inflows = [
            {"method_code": r["method_code"], "name": r["name"] or r["method_code"], "amount": float(r["amount"])}
            for r in man_rows
        ]
        manual_total = sum(float(r["amount"]) for r in man_rows)
        total_in += manual_total
    inflow_map: dict[str, dict[str, Any]] = {}
    for r in inflows:
        code = r["method_code"]
        inflow_map[code] = {
            "method_code": code,
            "name": r["name"] or code,
            "amount": float(r["amount"]),
        }
    for r in manual_inflows:
        code = r["method_code"]
        if code in inflow_map:
            inflow_map[code]["amount"] += r["amount"]
        else:
            inflow_map[code] = dict(r)
    operating_inflows = list(inflow_map.values())
    # Supplier payments / expenses: warehouse view = direct payouts + % of shared
    if warehouse_id:
        supplier_out = 0.0
        exp_clause = clause.replace("s.created_at", "expense_date")
        if period != "all" and not date_from and not date_to:
            exp_clause = " AND expense_date >= ?"
            exp_params = [period_start(period)[:10]]
        else:
            exp_params = [p[:10] if isinstance(p, str) and len(p) > 10 else p for p in params] if exp_clause else []
        _, _, allocated, _ = expenses_allocated_by_warehouse(conn, exp_clause, exp_params)
        expenses_out = expense_amount_for_warehouse(allocated, warehouse_id)
    else:
        sup_clause = clause.replace("s.created_at", "created_at")
        supplier_out = conn.execute(
            f"SELECT COALESCE(SUM(amount), 0) FROM supplier_payments WHERE 1=1 {sup_clause.replace('s.', '') if 's.' in sup_clause else sup_clause}",
            params if sup_clause else [],
        ).fetchone()[0]
        exp_clause = clause.replace("s.created_at", "expense_date")
        if period != "all" and not date_from and not date_to:
            exp_clause = " AND expense_date >= ?"
            exp_params = [period_start(period)[:10]]
        else:
            exp_params = [p[:10] if isinstance(p, str) and len(p) > 10 else p for p in params] if exp_clause else []
        expenses_out = conn.execute(
            f"SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE 1=1 {exp_clause}",
            exp_params,
        ).fetchone()[0]
    operating_out = float(supplier_out) + float(expenses_out)
    net_operating_cash = total_in - operating_out
    cutoff = _period_cutoff_start(period, date_from, date_to)
    if warehouse_id:
        opening_balance = 0.0
        if period != "all" or date_from:
            opening_balance = float(conn.execute(
                """
                SELECT COALESCE(SUM(sp.amount), 0)
                FROM sale_payments sp
                JOIN sales s ON s.id = sp.sale_id
                WHERE s.status = 'completed' AND s.warehouse_id = ? AND s.created_at < ?
                """,
                (warehouse_id, cutoff),
            ).fetchone()[0])
    else:
        opening_balance = _cash_net_before(conn, cutoff) if period != "all" or date_from else 0.0
    closing_balance = opening_balance + net_operating_cash
    wh_name = ""
    if warehouse_id:
        row = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        wh_name = row["name"] if row else f"#{warehouse_id}"
        label = f"{label} · {wh_name}"
    # Per-currency bags (no USD+TJS mix). Warehouse filter: sale currency only for that wh.
    cash_bags = _period_cash_by_currency(conn, period if not date_from else "all")
    by_currency = cash_bags["by_currency"]
    if warehouse_id or date_from or date_to:
        # Recompute narrow period/warehouse bags via sale payments only + TJS outs when company-wide
        bags: dict[str, dict[str, float]] = {}
        for r in conn.execute(
            f"""
            SELECT UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS cur,
                   COALESCE(SUM(sp.amount), 0) AS amt
            FROM sale_payments sp
            JOIN sales s ON s.id = sp.sale_id
            WHERE s.status = 'completed' {clause}{wh_clause}
            GROUP BY 1
            """,
            all_params,
        ).fetchall():
            code = (r["cur"] or "TJS").upper()
            bags.setdefault(code, {"inflow": 0.0, "outflow": 0.0})
            bags[code]["inflow"] += float(r["amt"] or 0)
        if not warehouse_id:
            bags.setdefault("TJS", {"inflow": 0.0, "outflow": 0.0})
            bags["TJS"]["outflow"] += float(operating_out)
        by_currency = []
        for code in sorted(bags.keys(), key=lambda c: (0 if c == "USD" else 1, c)):
            inn = round(bags[code]["inflow"], 2)
            out = round(bags[code]["outflow"], 2)
            by_currency.append({
                **currency_meta(code),
                "inflow": inn, "outflow": out, "net": round(inn - out, 2),
                "plus": inn, "minus": out,
            })
    return {
        "period_label": label,
        "warehouse_id": warehouse_id,
        "warehouse_name": wh_name,
        "operating_inflows": operating_inflows,
        "receivable_collections": float(recv_in),
        "manual_inflows": manual_total,
        "total_inflows": None,
        "supplier_payments": float(supplier_out),
        "operating_expenses": float(expenses_out),
        "total_outflows": None,
        "net_operating_cash": None,
        "opening_balance": None,
        "closing_balance": None,
        "by_currency": by_currency,
        # Legacy mixed numbers kept only for debugging; UI must not use them
        "_mixed_total_inflows": total_in,
        "_mixed_total_outflows": operating_out,
        "_mixed_net": net_operating_cash,
        "_mixed_opening": opening_balance,
        "_mixed_closing": closing_balance,
    }


def _report_financial_detail(
    conn: sqlite3.Connection,
    kind: str,
    period: str,
    date_from: str,
    date_to: str,
    warehouse_id: int | None = None,
) -> dict[str, Any]:
    clause, params, label = _report_period_clause(period, date_from, date_to, "s.created_at")
    wh_clause = " AND s.warehouse_id = ?" if warehouse_id else ""
    wh_params = list(params) + ([warehouse_id] if warehouse_id else [])

    sale_lines = conn.execute(
        f"""
        SELECT s.id, s.created_at, s.total, s.amount_paid, s.amount_due,
               COALESCE(w.name, '—') AS warehouse_name, s.warehouse_id,
               COALESCE(SUM(si.shop_profit), 0) AS profit
        FROM sales s
        LEFT JOIN warehouses w ON w.id = s.warehouse_id
        LEFT JOIN sale_items si ON si.sale_id = s.id
        WHERE s.status = 'completed' {clause}{wh_clause}
        GROUP BY s.id
        ORDER BY s.created_at DESC
        LIMIT 200
        """,
        wh_params,
    ).fetchall()

    pay_lines = conn.execute(
        f"""
        SELECT sp.method_code, pm.name, sp.amount, s.created_at, s.id AS sale_id,
               UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
               COALESCE(w.name, '—') AS warehouse_name
        FROM sale_payments sp
        JOIN sales s ON s.id = sp.sale_id
        LEFT JOIN payment_methods pm ON pm.code = sp.method_code
        LEFT JOIN warehouses w ON w.id = s.warehouse_id
        WHERE s.status = 'completed' {clause}{wh_clause}
        ORDER BY s.created_at DESC
        LIMIT 300
        """,
        wh_params,
    ).fetchall()

    exp_clause = clause.replace("s.created_at", "expense_date")
    if period != "all" and not date_from and not date_to:
        exp_clause = " AND expense_date >= ?"
        exp_params: list[Any] = [period_start(period)[:10]]
    else:
        exp_params = [p[:10] if isinstance(p, str) and len(p) > 10 else p for p in params] if exp_clause else []
    expense_rows = conn.execute(
        f"SELECT * FROM expenses WHERE 1=1 {exp_clause} ORDER BY expense_date DESC LIMIT 200",
        exp_params,
    ).fetchall()

    sup_clause = clause.replace("s.created_at", "created_at")
    supplier_rows = conn.execute(
        f"SELECT * FROM supplier_payments WHERE 1=1 {sup_clause.replace('s.', '') if 's.' in sup_clause else sup_clause} ORDER BY created_at DESC LIMIT 100",
        params if sup_clause else [],
    ).fetchall()

    recv_clause = clause.replace("s.created_at", "rp.created_at")
    recv_rows = conn.execute(
        f"""
        SELECT rp.*, r.customer_name, r.customer_phone,
               UPPER(COALESCE(NULLIF(TRIM(rp.pay_currency_code), ''), NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code
        FROM receivable_payments rp
        JOIN receivables r ON r.id = rp.receivable_id
        LEFT JOIN sales s ON s.id = r.sale_id
        WHERE 1=1 {recv_clause.replace('s.', 'rp.') if 's.' in recv_clause else recv_clause}
        ORDER BY rp.created_at DESC LIMIT 100
        """,
        params if recv_clause else [],
    ).fetchall()

    opiu = _report_opiu(conn, period, date_from, date_to)
    split = opiu.get("expense_allocation") or get_expense_warehouse_split(conn)
    expenses_by_wh = opiu.get("expenses_by_warehouse") or []
    wh_breakdown: list[dict[str, Any]] = []
    warehouses = conn.execute("SELECT id, name FROM warehouses ORDER BY name").fetchall()
    for wh in warehouses:
        wid = wh["id"]
        cur_rows = conn.execute(
            f"""
            SELECT UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS cur,
                   COALESCE(SUM(s.total), 0) AS revenue,
                   COALESCE(SUM((SELECT SUM(si.shop_profit) FROM sale_items si WHERE si.sale_id = s.id)), 0) AS profit,
                   COALESCE(SUM((SELECT SUM(sp.amount) FROM sale_payments sp WHERE sp.sale_id = s.id)), 0) AS inflows
            FROM sales s
            WHERE s.status = 'completed' AND s.warehouse_id = ? {clause}
            GROUP BY 1
            """,
            [wid, *params],
        ).fetchall()
        by_currency = []
        rev = profit = inflow = 0.0
        for r in cur_rows:
            code = (r["cur"] or "TJS").upper()
            rv, pf, inf = float(r["revenue"] or 0), float(r["profit"] or 0), float(r["inflows"] or 0)
            rev += rv
            profit += pf
            inflow += inf
            by_currency.append({
                **currency_meta(code),
                "revenue": round(rv, 2),
                "profit": round(pf, 2),
                "inflows": round(inf, 2),
            })
        exp_row = next((x for x in expenses_by_wh if int(x.get("warehouse_id") or 0) == int(wid)), None)
        wh_split = next((s for s in split if int(s["warehouse_id"]) == int(wid)), None)
        pct_rule = float(wh_split["pct"]) if wh_split else 0.0
        exp_alloc = float(exp_row["amount"]) if exp_row else 0.0
        wh_breakdown.append({
            "warehouse_id": wid,
            "warehouse_name": wh["name"],
            "revenue": None,
            "profit": None,
            "inflows": None,
            "by_currency": by_currency,
            "expenses_allocated": round(exp_alloc, 2),  # TJS: shared% + direct payouts
            "pct": float(exp_row["pct"]) if exp_row else pct_rule,
            "pct_rule": pct_rule,
            "direct_amount": float(exp_row.get("direct_amount") or 0) if exp_row else 0.0,
            "shared_amount": float(exp_row.get("shared_amount") or 0) if exp_row else 0.0,
        })

    dds = _report_dds(conn, period, date_from, date_to)
    return {
        "kind": kind,
        "period_label": label,
        "warehouse_id": warehouse_id,
        "summary": opiu if kind == "opiu" else dds,
        "warehouses": wh_breakdown,
        "sales": [row_to_dict(r) for r in sale_lines],
        "inflow_lines": [row_to_dict(r) for r in pay_lines],
        "expense_lines": [row_to_dict(r) for r in expense_rows],
        "expenses_by_category": opiu.get("expenses_by_category", []),
        "supplier_payments": [row_to_dict(r) for r in supplier_rows],
        "receivable_collections": [row_to_dict(r) for r in recv_rows],
        "expense_allocation": split,
    }


def _report_balance(conn: sqlite3.Connection, warehouse_id: int | None = None) -> dict[str, Any]:
    """Balance sheet as currency bags — never mix USD+TJS into one float."""
    cash = _period_cash_by_currency(conn, "all")
    cash_bags = {
        (b["code"] or "TJS").upper(): float(b.get("net") or 0)
        for b in (cash.get("by_currency") or [])
    }
    inv_rows = inventory_stock_value_by_currency(conn, "all", warehouse_id)
    inv_bags = {(r.get("code") or "TJS").upper(): float(r.get("value") or 0) for r in inv_rows}
    wh_clause, wh_params = warehouse_sales_filter(warehouse_id)
    recv_rows = conn.execute(
        f"""
        SELECT UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS cur,
               COALESCE(SUM(r.amount_due), 0) AS amt
        FROM receivables r
        JOIN sales s ON s.id = r.sale_id
        WHERE r.status = 'open' AND s.status = 'completed'{wh_clause}
        GROUP BY 1
        """,
        wh_params,
    ).fetchall()
    recv_bags = {(r["cur"] or "TJS").upper(): float(r["amt"] or 0) for r in recv_rows}
    pay_bags: dict[str, float] = {}
    if warehouse_id:
        for r in conn.execute(
            f"""
            SELECT UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS cur,
                   COALESCE(SUM(si.supplier_due), 0) AS amt
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status = 'completed' AND si.ownership_type = 'consignment'{wh_clause}
            GROUP BY 1
            """,
            wh_params,
        ).fetchall():
            pay_bags[(r["cur"] or "TJS").upper()] = float(r["amt"] or 0)
    else:
        for c in _creditors_list(conn):
            for b in c.get("by_currency") or []:
                code = b["code"]
                pay_bags[code] = pay_bags.get(code, 0.0) + float(b.get("balance") or 0)
    codes = sorted(set(cash_bags) | set(inv_bags) | set(recv_bags) | set(pay_bags) | {"USD", "TJS"},
                   key=lambda c: (0 if c == "USD" else 1, c))
    by_currency = []
    for code in codes:
        cash_v = round(cash_bags.get(code, 0.0), 2)
        inv_v = round(inv_bags.get(code, 0.0), 2)
        recv_v = round(recv_bags.get(code, 0.0), 2)
        pay_v = round(pay_bags.get(code, 0.0), 2)
        assets = round(cash_v + inv_v + recv_v, 2)
        equity = round(assets - pay_v, 2)
        by_currency.append({
            **currency_meta(code),
            "cash": cash_v,
            "inventory": inv_v,
            "receivables": recv_v,
            "supplier_payables": pay_v,
            "assets": assets,
            "equity": equity,
        })
    wh_name = ""
    if warehouse_id:
        row = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        wh_name = row["name"] if row else f"#{warehouse_id}"
    return {
        "warehouse_id": warehouse_id,
        "warehouse_name": wh_name,
        "by_currency": by_currency,
        "assets": {"cash": None, "inventory": None, "receivables": None, "total": None},
        "liabilities": {"supplier_payables": None, "total": None},
        "equity": None,
    }


def _report_by_cashier(
    conn: sqlite3.Connection,
    period: str,
    date_from: str,
    date_to: str,
    warehouse_id: int | None = None,
) -> dict[str, Any]:
    clause, params, label = _report_period_clause(period, date_from, date_to, "s.created_at")
    wh_clause, wh_params = warehouse_sales_filter(warehouse_id)
    all_params = params + wh_params
    rows = conn.execute(
        f"""
        SELECT
            COALESCE(NULLIF(s.user_name, ''), sh.user_name, '—') AS cashier_name,
            COALESCE(s.user_id, sh.user_id, 0) AS user_id,
            UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
            COUNT(*) AS sales_count,
            COALESCE(SUM(s.total), 0) AS revenue
        FROM sales s
        LEFT JOIN shifts sh ON sh.id = s.shift_id
        WHERE s.status = 'completed' {clause}{wh_clause}
        GROUP BY COALESCE(s.user_id, sh.user_id),
                 COALESCE(NULLIF(s.user_name, ''), sh.user_name, '—'),
                 UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS'))
        ORDER BY revenue DESC
        """,
        all_params,
    ).fetchall()
    profit_rows = conn.execute(
        f"""
        SELECT COALESCE(s.user_id, sh.user_id, 0) AS user_id,
               UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
               COALESCE(SUM(si.shop_profit), 0) AS profit
        FROM sales s
        JOIN sale_items si ON si.sale_id = s.id
        LEFT JOIN shifts sh ON sh.id = s.shift_id
        WHERE s.status = 'completed' {clause}{wh_clause}
        GROUP BY COALESCE(s.user_id, sh.user_id, 0),
                 UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS'))
        """,
        all_params,
    ).fetchall()
    profit_map = {(r["user_id"], (r["currency_code"] or "TJS").upper()): float(r["profit"]) for r in profit_rows}
    pay_rows = conn.execute(
        f"""
        SELECT
            COALESCE(s.user_id, sh.user_id, 0) AS user_id,
            UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
            sp.method_code,
            pm.name,
            COALESCE(SUM(sp.amount), 0) AS amount
        FROM sale_payments sp
        JOIN sales s ON s.id = sp.sale_id
        LEFT JOIN shifts sh ON sh.id = s.shift_id
        LEFT JOIN payment_methods pm ON pm.code = sp.method_code
        WHERE s.status = 'completed' {clause}{wh_clause}
        GROUP BY COALESCE(s.user_id, sh.user_id, 0),
                 UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')),
                 sp.method_code
        ORDER BY amount DESC
        """,
        all_params,
    ).fetchall()
    pay_by_user: dict[tuple[Any, str], list[dict[str, Any]]] = {}
    for r in pay_rows:
        key = (r["user_id"], (r["currency_code"] or "TJS").upper())
        pay_by_user.setdefault(key, []).append({
            "method_code": r["method_code"],
            "name": r["name"] or r["method_code"],
            "amount": float(r["amount"]),
            "currency_code": (r["currency_code"] or "TJS").upper(),
        })
    cashiers = []
    for r in rows:
        uid = r["user_id"] or 0
        cur = (r["currency_code"] or "TJS").upper()
        cashiers.append({
            "user_id": uid,
            "cashier_name": r["cashier_name"],
            "currency_code": cur,
            "sales_count": r["sales_count"],
            "revenue": float(r["revenue"]),
            "profit": profit_map.get((uid, cur), 0.0),
            "by_payment": pay_by_user.get((uid, cur), []),
        })
    wh_name = ""
    if warehouse_id:
        row = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        wh_name = row["name"] if row else f"#{warehouse_id}"
        label = f"{label} · {wh_name}"
    return {"period_label": label, "warehouse_id": warehouse_id, "warehouse_name": wh_name, "cashiers": cashiers}


@app.get("/api/reports/opiu")
async def report_opiu(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        return _report_opiu(conn, period, date_from, date_to, warehouse_id)


@app.get("/api/reports/dds")
async def report_dds(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        return _report_dds(conn, period, date_from, date_to, warehouse_id)


@app.get("/api/reports/detail")
async def report_detail(
    kind: str = Query(pattern="^(opiu|dds|finance)$"),
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        return _report_financial_detail(conn, kind, period, date_from, date_to, warehouse_id)


@app.get("/api/reports/balance")
async def report_balance(
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        return _report_balance(conn, warehouse_id)


@app.get("/api/reports/cashiers")
async def report_cashiers(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        return _report_by_cashier(conn, period, date_from, date_to, warehouse_id)


def _finance_by_currency(
    conn: sqlite3.Connection,
    since_clause: str,
    params: list[Any],
    own_clause: str,
    own_params: list[Any],
    warehouse_id: int | None = None,
) -> list[dict[str, Any]]:
    wh_clause, wh_params = warehouse_sales_filter(warehouse_id)
    rows = conn.execute(
        f"""
        SELECT COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS') AS currency_code,
               COUNT(DISTINCT s.id) AS sales_count,
               COALESCE(SUM(si.subtotal), 0) AS gross_revenue,
               COALESCE(SUM(si.shop_profit), 0) AS shop_profit,
               COALESCE(SUM(CASE WHEN si.ownership_type = 'own' THEN si.purchase_price * si.quantity ELSE 0 END), 0) AS own_cogs,
               COALESCE(SUM(CASE WHEN si.ownership_type = 'consignment' THEN si.supplier_due ELSE 0 END), 0) AS supplier_due,
               COALESCE(SUM(si.quantity), 0) AS items_sold
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        WHERE s.status = 'completed' {since_clause} {own_clause}{wh_clause}
        GROUP BY COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')
        ORDER BY currency_code
        """,
        params + own_params + wh_params,
    ).fetchall()
    result: list[dict[str, Any]] = []
    for r in rows:
        code = (r["currency_code"] or "TJS").upper()
        meta = currency_meta(code)
        rev = float(r["gross_revenue"])
        profit = float(r["shop_profit"])
        cogs = float(r["own_cogs"] or 0)
        sup_due = float(r["supplier_due"] or 0)
        result.append({
            **meta,
            "sales_count": int(r["sales_count"]),
            "gross_revenue": rev,
            "shop_profit": profit,
            "own_cogs": cogs,
            "supplier_due": sup_due,
            "gross_profit": rev - cogs - sup_due,
            "items_sold": int(r["items_sold"]),
            "margin_pct": round(profit / rev * 100, 1) if rev else 0,
        })
    return result


def _finance_report(
    conn: sqlite3.Connection,
    period: str,
    scope: str,
    date_from: str,
    date_to: str,
    warehouse_id: int | None = None,
) -> dict[str, Any]:
    if date_from or date_to:
        since_clause, params = date_filter_sql(date_from, date_to)
        period_label = f"{date_from or '…'} — {date_to or '…'}"
    elif period == "all":
        since_clause, params = "", []
        period_label = "Всё время"
    else:
        since_clause = " AND s.created_at >= ?"
        params = [period_start(period)]
        period_label = {"day": "Сегодня", "week": "Неделя", "month": "Месяц", "quarter": "Квартал", "year": "Год"}.get(period, period)

    own_clause, own_params = ownership_filter_sql(scope)
    wh_clause, wh_params = warehouse_sales_filter(warehouse_id)

    base = f"""
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        WHERE s.status = 'completed' {since_clause} {own_clause}{wh_clause}
    """
    all_params = params + own_params + wh_params
    sale_params = params + wh_params

    agg = conn.execute(
        f"""
        SELECT COUNT(DISTINCT s.id) AS sales_count,
               COALESCE(SUM(si.subtotal), 0) AS gross_revenue,
               COALESCE(SUM(CASE WHEN si.ownership_type = 'own' THEN si.purchase_price * si.quantity ELSE 0 END), 0) AS own_cogs,
               COALESCE(SUM(CASE WHEN si.ownership_type = 'consignment' THEN si.supplier_due ELSE 0 END), 0) AS supplier_due,
               COALESCE(SUM(si.shop_profit), 0) AS shop_profit,
               COALESCE(SUM(si.quantity), 0) AS items_sold
        {base}
        """,
        all_params,
    ).fetchone()

    discounts = conn.execute(
        f"""
        SELECT COALESCE(SUM(s.discount), 0)
        FROM sales s
        WHERE s.status = 'completed' {since_clause}{wh_clause}
        """,
        sale_params,
    ).fetchone()[0] if scope == "all" else 0

    payment_rows = conn.execute(
        f"""
        SELECT s.payment_method, COUNT(DISTINCT s.id) AS cnt, COALESCE(SUM(si.subtotal), 0) AS amount
        {base}
        GROUP BY s.payment_method
        """,
        all_params,
    ).fetchall()

    payment_totals = conn.execute(
        f"""
        SELECT
            COALESCE(SUM(
                CASE WHEN s.payment_method = 'cash' THEN s.total
                     ELSE s.cash_amount END
            ), 0) AS total_cash,
            COALESCE(SUM(
                CASE WHEN s.payment_method = 'card' THEN s.total
                     ELSE s.card_amount END
            ), 0) AS total_card,
            COALESCE(SUM(s.trade_in_value), 0) AS total_trade_in
        FROM sales s
        WHERE s.status = 'completed' {since_clause}{wh_clause}
        """,
        sale_params,
    ).fetchone()

    own_revenue = conn.execute(
        f"SELECT COALESCE(SUM(si.subtotal), 0) {base} AND si.ownership_type = 'own'",
        params + (own_params if scope != "consignment" else []) + wh_params if scope != "own" else all_params,
    ).fetchone()[0] if scope == "all" else (float(agg["gross_revenue"]) if scope == "own" else 0)

    cons_revenue = conn.execute(
        f"SELECT COALESCE(SUM(si.subtotal), 0) {base} AND si.ownership_type = 'consignment'",
        all_params if scope != "own" else params + wh_params,
    ).fetchone()[0] if scope in ("all", "consignment") else 0

    revenue = float(agg["gross_revenue"])
    shop_profit = float(agg["shop_profit"])
    own_cogs = float(agg["own_cogs"])
    supplier_due = float(agg["supplier_due"])

    wh_name = ""
    if warehouse_id:
        row = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        wh_name = row["name"] if row else f"#{warehouse_id}"
        period_label = f"{period_label} · {wh_name}"

    return {
        "scope": scope,
        "period": period,
        "period_label": period_label,
        "warehouse_id": warehouse_id,
        "warehouse_name": wh_name,
        "sales_count": agg["sales_count"],
        "items_sold": agg["items_sold"],
        "gross_revenue": revenue,
        "discounts": float(discounts),
        "net_revenue": revenue - float(discounts) if scope == "all" else revenue,
        "own_cogs": own_cogs,
        "supplier_due": supplier_due,
        "shop_profit": shop_profit,
        "margin_pct": round(shop_profit / revenue * 100, 1) if revenue else 0,
        "own_revenue": float(own_revenue) if scope == "all" else (revenue if scope == "own" else 0),
        "consignment_revenue": float(cons_revenue) if scope in ("all", "consignment") else 0,
        "total_cash": float(payment_totals["total_cash"]),
        "total_card": float(payment_totals["total_card"]),
        "total_trade_in": float(payment_totals["total_trade_in"]),
        "by_payment": [
            {"method": r["payment_method"], "count": r["cnt"], "amount": float(r["amount"])}
            for r in payment_rows
        ],
        "by_currency": _finance_by_currency(
            conn, since_clause, params, own_clause, own_params, warehouse_id
        ),
    }


@app.get("/api/reports/trade-ins")
async def trade_ins_report(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    with db() as conn:
        if date_from or date_to:
            since_clause, params = date_filter_sql(date_from, date_to, "t.created_at")
            period_label = f"{date_from or '…'} — {date_to or '…'}"
        elif period == "all":
            since_clause, params = "", []
            period_label = "Всё время"
        else:
            since_clause = " AND t.created_at >= ?"
            params = [period_start(period)]
            period_label = {
                "day": "Сегодня", "week": "Неделя", "month": "Месяц",
                "quarter": "Квартал", "year": "Год",
            }.get(period, period)

        wh_clause = ""
        if warehouse_id:
            wh_clause = " AND t.given_warehouse_id = ?"
            params = list(params) + [warehouse_id]
            row = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
            if row:
                period_label = f"{period_label} · {row['name']}"

        agg = conn.execute(
            f"""
            SELECT COUNT(*) AS deals_count,
                   COALESCE(SUM(t.received_value), 0) AS total_trade_credit,
                   COALESCE(SUM(t.cash_amount), 0) AS total_cash,
                   COALESCE(SUM(t.card_amount), 0) AS total_card,
                   COALESCE(SUM(t.cash_amount + t.card_amount), 0) AS total_money,
                   COALESCE(SUM(t.cash_amount + t.card_amount + t.received_value), 0) AS total_deal_value
            FROM trade_ins t
            WHERE 1=1 {since_clause}{wh_clause}
            """,
            params,
        ).fetchone()

        items = conn.execute(
            f"""
            SELECT t.*,
                   gp.name AS given_product_name, gp.sale_price AS given_sale_price,
                   rp.name AS received_product_name,
                   gw.name AS given_warehouse_name,
                   rw.name AS received_warehouse_name
            FROM trade_ins t
            LEFT JOIN products gp ON gp.id = t.given_product_id
            LEFT JOIN products rp ON rp.id = t.received_product_id
            LEFT JOIN warehouses gw ON gw.id = t.given_warehouse_id
            LEFT JOIN warehouses rw ON rw.id = t.received_warehouse_id
            WHERE 1=1 {since_clause}{wh_clause}
            ORDER BY t.created_at DESC
            LIMIT 500
            """,
            params,
        ).fetchall()

        by_warehouse = conn.execute(
            f"""
            SELECT gw.name AS warehouse_name,
                   COUNT(*) AS deals,
                   COALESCE(SUM(t.cash_amount + t.card_amount), 0) AS money,
                   COALESCE(SUM(t.received_value), 0) AS trade_credit
            FROM trade_ins t
            JOIN warehouses gw ON gw.id = t.given_warehouse_id
            WHERE 1=1 {since_clause}{wh_clause}
            GROUP BY gw.name
            ORDER BY deals DESC
            """,
            params,
        ).fetchall()

    return {
        "period": period,
        "period_label": period_label,
        "deals_count": agg["deals_count"],
        "total_trade_credit": float(agg["total_trade_credit"]),
        "total_cash": float(agg["total_cash"]),
        "total_card": float(agg["total_card"]),
        "total_money": float(agg["total_money"]),
        "total_deal_value": float(agg["total_deal_value"]),
        "items": [row_to_dict(i) for i in items],
        "by_warehouse": [row_to_dict(w) for w in by_warehouse],
    }


@app.get("/api/reports/finance")
async def finance_report(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    scope: ReportScope = "all",
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        report = _finance_report(conn, period, scope, date_from, date_to, warehouse_id)
        if scope in ("all", "consignment"):
            if date_from or date_to:
                sup_since, sup_params = date_filter_sql(date_from, date_to)
            elif period == "all":
                sup_since, sup_params = "", []
            else:
                sup_since = " AND s.created_at >= ?"
                sup_params = [period_start(period)]
            wh_clause, wh_params = warehouse_sales_filter(warehouse_id)
            suppliers = conn.execute(
                f"""
                SELECT si.supplier_name,
                       SUM(si.quantity) AS qty,
                       SUM(si.subtotal) AS revenue,
                       SUM(si.supplier_due) AS due,
                       SUM(si.shop_profit) AS profit
                FROM sale_items si
                JOIN sales s ON s.id = si.sale_id
                WHERE s.status = 'completed' AND si.ownership_type = 'consignment' {sup_since}{wh_clause}
                GROUP BY si.supplier_name
                ORDER BY revenue DESC
                """,
                list(sup_params) + wh_params,
            ).fetchall()
            report["by_supplier"] = [row_to_dict(r) for r in suppliers]
        else:
            report["by_supplier"] = []
        if scope == "all" and not warehouse_id:
            report["funds"] = funds_snapshot(conn)
    return report


@app.get("/api/reports/funds")
async def reports_funds(x_pin: str | None = Header(default=None, alias="X-Pin")):
    """Снимок: остатки по складам + наличка/кошелёк в $ и смн."""
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        return funds_snapshot(conn)


@app.get("/api/reports/combined")
async def combined_report(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    date_from: str = "",
    date_to: str = "",
    warehouse_id: int | None = None,
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        all_rep = _finance_report(conn, period, "all", date_from, date_to, warehouse_id)
        if not warehouse_id:
            all_rep["funds"] = funds_snapshot(conn)
        return {
            "all": all_rep,
            "own": _finance_report(conn, period, "own", date_from, date_to, warehouse_id),
            "consignment": _finance_report(conn, period, "consignment", date_from, date_to, warehouse_id),
        }


@app.get("/api/analytics/summary")
async def analytics_summary(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    scope: ReportScope = "all",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    with db() as conn:
        report = _finance_report(conn, period, scope, "", "")
        opiu = _report_opiu(conn, period, "", "")
        low_sql = "SELECT COUNT(*) FROM products WHERE stock <= min_stock"
        params: list[Any] = []
        if scope != "all":
            low_sql += " AND ownership_type = ?"
            params.append(scope)
        low_stock = conn.execute(low_sql, params).fetchone()[0]
        stock_by_cur = inventory_stock_value_by_currency(conn, scope)
        stock_value = sum(s["value"] for s in stock_by_cur)
        products_count = conn.execute(
            f"SELECT COUNT(*) FROM products {'WHERE ownership_type = ?' if scope != 'all' else ''}",
            params,
        ).fetchone()[0]

    return {
        "period": period,
        "scope": scope,
        "sales_count": report["sales_count"],
        "revenue": report["gross_revenue"],
        "profit": report["shop_profit"],
        "supplier_due": report["supplier_due"],
        "own_cogs": report["own_cogs"],
        "margin_pct": report["margin_pct"],
        "low_stock_count": low_stock,
        "stock_value": float(stock_value),
        "stock_by_currency": stock_by_cur,
        "products_count": products_count,
        "total_cash": report["total_cash"],
        "total_card": report["total_card"],
        "total_trade_in": report["total_trade_in"],
        "by_currency": report["by_currency"],
        "expenses": opiu["operating_expenses"],
        "expenses_by_category": opiu["expenses_by_category"],
        "multi_currency": len(report.get("by_currency") or []) > 1,
    }


@app.get("/api/analytics/top-products")
async def analytics_top(
    period: str = Query(default="month", pattern="^(day|week|month|quarter|year|all)$"),
    scope: ReportScope = "all",
    limit: int = Query(default=10, ge=1, le=50),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    since_clause, params = ("", []) if period == "all" else ("AND s.created_at >= ?", [period_start(period)])
    own_clause, own_params = ownership_filter_sql(scope)
    params = params + own_params + [limit]
    with db() as conn:
        rows = conn.execute(
            f"""
            SELECT si.product_name AS name, si.ownership_type,
                   COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS') AS currency_code,
                   SUM(si.quantity) AS qty,
                   SUM(si.subtotal) AS revenue,
                   SUM(si.shop_profit) AS profit
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status = 'completed' {since_clause} {own_clause}
            GROUP BY si.product_name, si.ownership_type, COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')
            ORDER BY revenue DESC
            LIMIT ?
            """,
            params,
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/analytics/daily")
async def analytics_daily(
    days: int = Query(default=30, ge=1, le=90),
    scope: ReportScope = "all",
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin, min_role="owner")
    since = (datetime.now() - timedelta(days=days - 1)).strftime("%Y-%m-%d")
    own_clause, own_params = ownership_filter_sql(scope)
    with db() as conn:
        rows = conn.execute(
            f"""
            SELECT DATE(s.created_at) AS day,
                   UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
                   COUNT(DISTINCT s.id) AS sales,
                   COALESCE(SUM(si.subtotal), 0) AS revenue,
                   COALESCE(SUM(si.shop_profit), 0) AS profit
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status = 'completed' AND DATE(s.created_at) >= ? {own_clause}
            GROUP BY DATE(s.created_at), UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS'))
            ORDER BY day, currency_code
            """,
            [since, *own_params],
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/dashboard")
async def dashboard(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin)
    with db() as conn:
        today = _finance_report(conn, "day", "all", "", "")
        month = _finance_report(conn, "month", "all", "", "")
        own_month = _finance_report(conn, "month", "own", "", "")
        cons_month = _finance_report(conn, "month", "consignment", "", "")
        balances = [
            {
                "supplier_name": c["supplier_name"],
                "balance": None,
                "by_currency": [b for b in (c.get("by_currency") or []) if float(b.get("balance") or 0) > 0.01],
            }
            for c in _creditors_list(conn)
            if any(float(b.get("balance") or 0) > 0.01 for b in (c.get("by_currency") or []))
        ]
        low_stock = conn.execute("SELECT COUNT(*) FROM products WHERE stock <= min_stock").fetchone()[0]
        expire_reservations(conn)
        stale_cutoff = (datetime.now() - timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S")
        imei_pending_stale = conn.execute(
            """
            SELECT COUNT(*) FROM sale_item_units siu
            JOIN sale_items si ON si.id = siu.sale_item_id
            JOIN sales s ON s.id = si.sale_id
            WHERE siu.imei_pending = 1 AND s.status = 'completed' AND s.created_at < ?
            """,
            (stale_cutoff,),
        ).fetchone()[0]
        no_imei_stock = conn.execute(
            """
            SELECT COUNT(*) FROM product_units
            WHERE status = 'in_stock' AND TRIM(COALESCE(imei, '')) = ''
            """
        ).fetchone()[0]
        pending_customs = conn.execute(
            """
            SELECT COUNT(*) FROM product_units
            WHERE status IN ('in_stock', 'reserved') AND customs_status = 'pending'
            """
        ).fetchone()[0]
        active_reservations = conn.execute(
            "SELECT COUNT(*) FROM unit_reservations WHERE status = 'active'"
        ).fetchone()[0]
        stale_items = conn.execute(
            """
            SELECT s.id AS sale_id, s.created_at, si.product_name, siu.serial, siu.imei,
                   s.user_name AS cashier_name
            FROM sale_item_units siu
            JOIN sale_items si ON si.id = siu.sale_item_id
            JOIN sales s ON s.id = si.sale_id
            WHERE siu.imei_pending = 1 AND s.status = 'completed' AND s.created_at < ?
            ORDER BY s.created_at ASC LIMIT 10
            """,
            (stale_cutoff,),
        ).fetchall()
        warehouses = conn.execute(
            """
            SELECT w.id, w.name, w.is_default,
                   COALESCE(SUM(ws.quantity), 0) AS total_items,
                   COUNT(DISTINCT CASE WHEN ws.quantity > 0 THEN ws.product_id END) AS products_count
            FROM warehouses w
            LEFT JOIN warehouse_stock ws ON ws.warehouse_id = w.id
            GROUP BY w.id
            ORDER BY w.is_default DESC, w.name
            """
        ).fetchall()
    return {
        "today": today,
        "month": month,
        "own_month": own_month,
        "consignment_month": cons_month,
        "supplier_balances": balances,
        "low_stock_count": low_stock,
        "warehouses": [row_to_dict(w) for w in warehouses],
        "alerts": {
            "imei_pending_stale": int(imei_pending_stale),
            "no_imei_stock": int(no_imei_stock),
            "pending_customs": int(pending_customs),
            "active_reservations": int(active_reservations),
            "stale_imei_items": [row_to_dict(r) for r in stale_items],
        },
    }


def _kpi_since_clause(metric: str) -> tuple[str, list[Any], str]:
    titles = {
        "revenue_today": "Выручка сегодня",
        "profit_today": "Прибыль сегодня",
        "revenue_month": "Выручка за месяц",
        "low_stock": "Мало на складе",
    }
    if metric not in titles:
        raise HTTPException(status_code=400, detail="Неизвестный показатель")
    if metric == "low_stock":
        return "", [], titles[metric]
    period = "day" if metric.endswith("_today") else "month"
    since = period_start(period)
    return " AND s.created_at >= ?", [since], titles[metric]


@app.get("/api/dashboard/kpi-detail")
async def dashboard_kpi_detail(
    metric: str = Query(..., pattern="^(revenue_today|profit_today|revenue_month|low_stock)$"),
    warehouse_id: int | None = Query(default=None),
    x_pin: str | None = Header(default=None, alias="X-Pin"),
):
    check_pin(x_pin)
    since_clause, since_params, title = _kpi_since_clause(metric)

    with db() as conn:
        if metric == "low_stock":
            wh_filter = ""
            wh_params: list[Any] = []
            if warehouse_id is not None:
                resolve_warehouse_id(conn, warehouse_id)
                wh_filter = " AND ws.warehouse_id = ?"
                wh_params = [warehouse_id]
            rows = conn.execute(
                f"""
                SELECT p.id AS product_id, p.name AS product_name, p.min_stock,
                       w.id AS warehouse_id, w.name AS warehouse_name,
                       ws.quantity
                FROM warehouse_stock ws
                JOIN products p ON p.id = ws.product_id
                JOIN warehouses w ON w.id = ws.warehouse_id
                WHERE ws.quantity <= p.min_stock {wh_filter}
                ORDER BY w.name, p.name
                """,
                wh_params,
            ).fetchall()
            items = [row_to_dict(r) for r in rows]
            return {
                "metric": metric,
                "title": title,
                "total": len(items),
                "by_warehouse": [],
                "items": items,
            }

        wh_filter = ""
        wh_params: list[Any] = []
        if warehouse_id is not None:
            resolve_warehouse_id(conn, warehouse_id)
            wh_filter = " AND s.warehouse_id = ?"
            wh_params = [warehouse_id]

        by_wh = conn.execute(
            f"""
            SELECT COALESCE(w.name, '—') AS warehouse_name,
                   s.warehouse_id,
                   UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
                   COUNT(DISTINCT s.id) AS sales_count,
                   COALESCE(SUM(si.subtotal), 0) AS revenue,
                   COALESCE(SUM(si.shop_profit), 0) AS profit
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            WHERE s.status = 'completed' {since_clause} {wh_filter}
            GROUP BY s.warehouse_id, UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS'))
            ORDER BY revenue DESC
            """,
            since_params + wh_params,
        ).fetchall()

        lines = conn.execute(
            f"""
            SELECT s.id AS sale_id, s.created_at, s.user_name AS cashier,
                   COALESCE(w.name, '—') AS warehouse_name,
                   UPPER(COALESCE(NULLIF(TRIM(s.currency_code), ''), 'TJS')) AS currency_code,
                   si.product_name, si.quantity, si.subtotal, si.shop_profit,
                   si.ownership_type
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            WHERE s.status = 'completed' {since_clause} {wh_filter}
            ORDER BY s.created_at DESC, s.id DESC
            LIMIT 500
            """,
            since_params + wh_params,
        ).fetchall()

        bags: dict[str, dict[str, float]] = {}
        for r in by_wh:
            code = (r["currency_code"] or "TJS").upper()
            bags.setdefault(code, {"gross_revenue": 0.0, "shop_profit": 0.0})
            bags[code]["gross_revenue"] += float(r["revenue"] or 0)
            bags[code]["shop_profit"] += float(r["profit"] or 0)
        by_currency = [
            {**currency_meta(code), **{k: round(v, 2) for k, v in vals.items()}}
            for code, vals in sorted(bags.items(), key=lambda x: (0 if x[0] == "USD" else 1, x[0]))
        ]

        return {
            "metric": metric,
            "title": title,
            "total": None,
            "total_revenue": None,
            "total_profit": None,
            "by_currency": by_currency,
            "by_warehouse": [row_to_dict(r) for r in by_wh],
            "items": [row_to_dict(r) for r in lines],
        }


class StocktakeStartIn(BaseModel):
    warehouse_id: int
    notes: str = ""


class StocktakeScanIn(BaseModel):
    q: str = ""
    unit_id: int | None = None
    product_id: int | None = None


class StocktakeCountIn(BaseModel):
    product_id: int
    quantity: int = Field(ge=1)


def _stocktake_expected(conn: sqlite3.Connection, warehouse_id: int) -> dict[str, Any]:
    products = conn.execute(
        """
        SELECT p.id, p.name, p.model, p.color, p.category, p.track_units, p.barcode,
               ws.quantity AS qty
        FROM warehouse_stock ws
        JOIN products p ON p.id = ws.product_id
        WHERE ws.warehouse_id = ? AND ws.quantity > 0
        ORDER BY p.name
        """,
        (warehouse_id,),
    ).fetchall()
    units = conn.execute(
        """
        SELECT u.id, u.product_id, u.imei, u.serial, p.color, p.name AS product_name,
               p.model, p.barcode, p.track_units
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        WHERE u.warehouse_id = ? AND u.status = 'in_stock'
        ORDER BY p.name, u.imei, u.serial
        """,
        (warehouse_id,),
    ).fetchall()
    return {
        "products": [row_to_dict(r) for r in products],
        "units": [row_to_dict(r) for r in units],
    }


def _stocktake_summary(conn: sqlite3.Connection, session_id: int) -> dict[str, Any]:
    session = conn.execute("SELECT * FROM stocktake_sessions WHERE id = ?", (session_id,)).fetchone()
    if not session:
        raise HTTPException(status_code=404, detail="Инвентаризация не найдена")
    lines = conn.execute(
        """
        SELECT l.*, p.name AS product_name, p.model, p.track_units
        FROM stocktake_lines l
        JOIN products p ON p.id = l.product_id
        WHERE l.session_id = ?
        ORDER BY l.created_at DESC
        """,
        (session_id,),
    ).fetchall()
    expected = _stocktake_expected(conn, session["warehouse_id"])
    scanned_unit_ids = {r["unit_id"] for r in lines if r["unit_id"]}
    counted_by_product: dict[int, int] = {}
    for line in lines:
        counted_by_product[line["product_id"]] = counted_by_product.get(line["product_id"], 0) + int(line["quantity"])
    variances = []
    for p in expected["products"]:
        pid = p["id"]
        exp = int(p["qty"])
        cnt = counted_by_product.get(pid, 0)
        if exp != cnt:
            variances.append({
                "product_id": pid,
                "product_name": p["name"],
                "color": p["color"],
                "expected": exp,
                "counted": cnt,
                "difference": cnt - exp,
                "track_units": int(p["track_units"] or 0),
            })
    missing_units = [u for u in expected["units"] if u["id"] not in scanned_unit_ids]
    unit_expected = len(expected["units"])
    acc_expected = sum(
        int(p["qty"]) for p in expected["products"] if not int(p.get("track_units") or 0)
    )
    return {
        "session": row_to_dict(session),
        "lines": [row_to_dict(r) for r in lines],
        "expected": expected,
        "variances": variances,
        "missing_units": missing_units,
        "counted_total": sum(counted_by_product.values()),
        "expected_total": unit_expected + acc_expected,
    }


def _resolve_stocktake_scan(conn: sqlite3.Connection, warehouse_id: int, q: str) -> dict[str, Any]:
    q = normalize_search_q(q)
    if not q:
        raise HTTPException(status_code=400, detail="Введите IMEI, серийник, штрихкод или название")
    uclause, uparams = unit_search_sql(q)
    units = conn.execute(
        f"""
        SELECT u.*, p.name AS product_name, p.color
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        WHERE u.warehouse_id = ? AND u.status = 'in_stock'
        {uclause}
        LIMIT 8
        """,
        [warehouse_id, *uparams],
    ).fetchall()
    if not units:
        # Name / model search → units of matching products on this warehouse
        clause, params = product_search_sql(q)
        units = conn.execute(
            f"""
            SELECT u.*, p.name AS product_name, p.color
            FROM product_units u
            JOIN products p ON p.id = u.product_id
            WHERE u.warehouse_id = ? AND u.status = 'in_stock'
            {clause}
            ORDER BY p.name, u.imei
            LIMIT 8
            """,
            [warehouse_id, *params],
        ).fetchall()
    if len(units) > 1:
        raise HTTPException(
            status_code=400,
            detail="Найдено несколько устройств — выберите из списка или введите больше цифр IMEI",
        )
    if len(units) == 1:
        unit = units[0]
        return {
            "product_id": unit["product_id"], "unit_id": unit["id"], "quantity": 1,
            "imei": unit["imei"] or "", "serial": unit["serial"] or "",
            "color": unit["color"] or "",
        }
    product = conn.execute("SELECT * FROM products WHERE barcode = ?", (q,)).fetchone()
    if not product:
        clause, params = product_search_sql(q)
        product = conn.execute(
            f"SELECT * FROM products p WHERE 1=1 {clause} LIMIT 1", params
        ).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="Товар или IMEI не найден")
    if int(product["track_units"] or 0):
        raise HTTPException(status_code=400, detail="Для телефона выберите IMEI из списка или отсканируйте")
    stock = get_warehouse_stock(conn, warehouse_id, product["id"])
    if stock <= 0:
        raise HTTPException(status_code=400, detail="Товар не числится на этом складе")
    return {
        "product_id": product["id"], "unit_id": None, "quantity": 1,
        "imei": "", "serial": "", "color": product["color"] or "",
    }


def _stocktake_hit_from_unit(conn: sqlite3.Connection, warehouse_id: int, unit_id: int) -> dict[str, Any]:
    unit = conn.execute(
        """
        SELECT u.*, p.name AS product_name, p.color
        FROM product_units u
        JOIN products p ON p.id = u.product_id
        WHERE u.id = ? AND u.warehouse_id = ? AND u.status = 'in_stock'
        """,
        (unit_id, warehouse_id),
    ).fetchone()
    if not unit:
        raise HTTPException(status_code=404, detail="Устройство не найдено на этом складе")
    return {
        "product_id": unit["product_id"], "unit_id": unit["id"], "quantity": 1,
        "imei": unit["imei"] or "", "serial": unit["serial"] or "",
        "color": unit["color"] or "",
    }


@app.get("/api/stocktake/current")
async def stocktake_current(x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        row = conn.execute(
            "SELECT * FROM stocktake_sessions WHERE status = 'open' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if not row:
            return {"session": None}
        return _stocktake_summary(conn, row["id"])


@app.post("/api/stocktake/start")
async def stocktake_start(body: StocktakeStartIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    user = None
    with db() as conn:
        user = resolve_user(conn, x_pin)
        if conn.execute("SELECT id FROM stocktake_sessions WHERE status = 'open'").fetchone():
            raise HTTPException(status_code=400, detail="Уже есть открытая инвентаризация")
        resolve_warehouse_id(conn, body.warehouse_id)
        cur = conn.execute(
            """
            INSERT INTO stocktake_sessions (warehouse_id, user_id, user_name, notes, started_at, status)
            VALUES (?, ?, ?, ?, ?, 'open')
            """,
            (body.warehouse_id, user.get("id") if user else None, user.get("name", "") if user else "",
             body.notes, utc_now()),
        )
        session_id = cur.lastrowid
    with db() as conn:
        return _stocktake_summary(conn, session_id)


@app.post("/api/stocktake/{session_id}/scan")
async def stocktake_scan(session_id: int, body: StocktakeScanIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        session = conn.execute(
            "SELECT * FROM stocktake_sessions WHERE id = ? AND status = 'open'", (session_id,)
        ).fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Открытая инвентаризация не найдена")
        if body.unit_id:
            hit = _stocktake_hit_from_unit(conn, session["warehouse_id"], body.unit_id)
        elif body.product_id and not (body.q or "").strip():
            product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
            if not product:
                raise HTTPException(status_code=404, detail="Товар не найден")
            if int(product["track_units"] or 0):
                raise HTTPException(status_code=400, detail="Для телефона выберите конкретный IMEI")
            hit = {
                "product_id": product["id"], "unit_id": None, "quantity": 1,
                "imei": "", "serial": "", "color": product["color"] or "",
            }
        else:
            hit = _resolve_stocktake_scan(conn, session["warehouse_id"], body.q)
        if hit.get("unit_id"):
            dup = conn.execute(
                "SELECT id FROM stocktake_lines WHERE session_id = ? AND unit_id = ?",
                (session_id, hit["unit_id"]),
            ).fetchone()
            if dup:
                raise HTTPException(status_code=400, detail="Это устройство уже отсканировано")
        conn.execute(
            """
            INSERT INTO stocktake_lines
            (session_id, product_id, unit_id, quantity, imei, serial, color, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (session_id, hit["product_id"], hit.get("unit_id"), hit["quantity"],
             hit["imei"], hit["serial"], hit["color"], utc_now()),
        )
        return _stocktake_summary(conn, session_id)


@app.post("/api/stocktake/{session_id}/count")
async def stocktake_count(session_id: int, body: StocktakeCountIn, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        session = conn.execute(
            "SELECT * FROM stocktake_sessions WHERE id = ? AND status = 'open'", (session_id,)
        ).fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Открытая инвентаризация не найдена")
        product = conn.execute("SELECT * FROM products WHERE id = ?", (body.product_id,)).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="Товар не найден")
        if int(product["track_units"] or 0):
            raise HTTPException(status_code=400, detail="Для телефонов сканируйте IMEI по одному")
        conn.execute(
            """
            INSERT INTO stocktake_lines
            (session_id, product_id, unit_id, quantity, imei, serial, color, created_at)
            VALUES (?, ?, NULL, ?, '', '', ?, ?)
            """,
            (session_id, body.product_id, body.quantity, product["color"] or "", utc_now()),
        )
        return _stocktake_summary(conn, session_id)


@app.delete("/api/stocktake/{session_id}/lines/{line_id}")
async def stocktake_undo_line(session_id: int, line_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        session = conn.execute(
            "SELECT id FROM stocktake_sessions WHERE id = ? AND status = 'open'", (session_id,)
        ).fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Открытая инвентаризация не найдена")
        conn.execute(
            "DELETE FROM stocktake_lines WHERE id = ? AND session_id = ?", (line_id, session_id)
        )
        return _stocktake_summary(conn, session_id)


@app.post("/api/stocktake/{session_id}/complete")
async def stocktake_complete(session_id: int, x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        session = conn.execute(
            "SELECT * FROM stocktake_sessions WHERE id = ? AND status = 'open'", (session_id,)
        ).fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Открытая инвентаризация не найдена")
        summary = _stocktake_summary(conn, session_id)
        wh = session["warehouse_id"]
        for v in summary["variances"]:
            if v["track_units"]:
                continue
            diff = v["difference"]
            if diff != 0:
                adjust_warehouse_stock(
                    conn, wh, v["product_id"], diff, "stocktake",
                    notes=f"Инвентаризация #{session_id}: {v['counted']} из {v['expected']}",
                )
        conn.execute(
            "UPDATE stocktake_sessions SET status = 'closed', completed_at = ? WHERE id = ?",
            (utc_now(), session_id),
        )
        summary["session"]["status"] = "closed"
        return summary


@app.get("/api/stocktake/history")
async def stocktake_history(limit: int = Query(default=20, ge=1, le=100), x_pin: str | None = Header(default=None, alias="X-Pin")):
    check_pin(x_pin, min_role="warehouse")
    with db() as conn:
        rows = conn.execute(
            """
            SELECT s.*, w.name AS warehouse_name
            FROM stocktake_sessions s
            JOIN warehouses w ON w.id = s.warehouse_id
            ORDER BY s.started_at DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [row_to_dict(r) for r in rows]


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app:app", host="0.0.0.0", port=settings.port, reload=False)
